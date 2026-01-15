"""
Tests unitarios para funciones helper refactorizadas de reports.py.

Cubre las 9 funciones helper creadas durante la refactorización de Días 5-6:
- Excel Import helpers (5 funciones)
- PDF helpers (2 funciones)
- Excel Format helpers (2 funciones)
"""
import pytest
from datetime import date, datetime, timedelta
from io import BytesIO
from unittest.mock import Mock, patch, MagicMock
from decimal import Decimal
import openpyxl
from openpyxl import Workbook

from core.views.reports import (
    _validate_and_load_excel,
    _extract_row_data,
    _process_all_row_dates,
    _update_existing_equipment,
    _create_new_equipment,
    _get_pdf_file_url,
    _get_pdf_image_data,
    _add_excel_header,
    _add_excel_resumen_section,
)
from core.models import Equipo, Empresa
from core.constants import ESTADO_ACTIVO, ESTADO_INACTIVO


# =============================================================================
# TESTS: EXCEL IMPORT HELPERS (5 funciones)
# =============================================================================

@pytest.mark.django_db
class TestValidateAndLoadExcel:
    """Tests para _validate_and_load_excel"""

    def test_validate_and_load_excel_success(self):
        """Test carga exitosa de archivo Excel válido"""
        # Crear un Excel simple en memoria
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Test'

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        # Ejecutar función
        result = _validate_and_load_excel(excel_buffer)

        # Verificar resultado
        assert result['error'] is None
        assert result['workbook'] is not None
        assert result['sheet'] is not None
        assert result['column_mapping'] is not None
        assert 'A' in result['column_mapping']
        assert result['column_mapping']['A'] == 'codigo_interno'

    def test_validate_and_load_excel_invalid_file(self):
        """Test manejo de archivo Excel inválido"""
        # Crear un archivo que no es Excel
        invalid_file = BytesIO(b"Este no es un Excel valido")

        # Ejecutar función
        result = _validate_and_load_excel(invalid_file)

        # Verificar que detecta el error
        assert result['error'] is not None
        assert 'Error cargando archivo Excel' in result['error']
        assert result['workbook'] is None

    def test_validate_and_load_excel_column_mapping_complete(self):
        """Test que el mapeo de columnas está completo"""
        workbook = Workbook()
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        result = _validate_and_load_excel(excel_buffer)

        # Verificar columnas esenciales
        assert 'codigo_interno' in result['column_mapping'].values()
        assert 'nombre' in result['column_mapping'].values()
        assert 'empresa_nombre' in result['column_mapping'].values()
        assert 'fecha_adquisicion' in result['column_mapping'].values()
        assert 'frecuencia_calibracion_meses' in result['column_mapping'].values()


@pytest.mark.django_db
class TestExtractRowData:
    """Tests para _extract_row_data"""

    def test_extract_row_data_basic(self):
        """Test extracción básica de datos de una fila"""
        workbook = Workbook()
        sheet = workbook.active

        # Datos de prueba en fila 1
        sheet['A1'] = 'EQ001'
        sheet['B1'] = 'Equipo Test'
        sheet['C1'] = 'Empresa Test'

        column_mapping = {
            'A': 'codigo_interno',
            'B': 'nombre',
            'C': 'empresa_nombre'
        }

        result = _extract_row_data(sheet, 1, column_mapping)

        assert result['codigo_interno'] == 'EQ001'
        assert result['nombre'] == 'Equipo Test'
        assert result['empresa_nombre'] == 'Empresa Test'

    def test_extract_row_data_strips_whitespace(self):
        """Test que elimina espacios en blanco de strings"""
        workbook = Workbook()
        sheet = workbook.active

        sheet['A1'] = '  EQ001  '
        sheet['B1'] = ' Nombre con espacios '

        column_mapping = {
            'A': 'codigo_interno',
            'B': 'nombre'
        }

        result = _extract_row_data(sheet, 1, column_mapping)

        assert result['codigo_interno'] == 'EQ001'
        assert result['nombre'] == 'Nombre con espacios'

    def test_extract_row_data_ignores_none_values(self):
        """Test que ignora valores None pero mantiene strings vacíos"""
        workbook = Workbook()
        sheet = workbook.active

        sheet['A1'] = 'EQ001'
        sheet['B1'] = None
        sheet['C1'] = ''

        column_mapping = {
            'A': 'codigo_interno',
            'B': 'nombre',
            'C': 'marca'
        }

        result = _extract_row_data(sheet, 1, column_mapping)

        assert 'codigo_interno' in result
        assert 'nombre' not in result  # None no debe incluirse
        assert 'marca' in result  # String vacío sí se incluye
        assert result['marca'] == ''

    def test_extract_row_data_handles_numbers(self):
        """Test manejo de valores numéricos"""
        workbook = Workbook()
        sheet = workbook.active

        sheet['A1'] = 'EQ001'
        sheet['B1'] = 12.5
        sheet['C1'] = 100

        column_mapping = {
            'A': 'codigo_interno',
            'B': 'frecuencia_calibracion_meses',
            'C': 'cantidad'
        }

        result = _extract_row_data(sheet, 1, column_mapping)

        assert result['codigo_interno'] == 'EQ001'
        assert result['frecuencia_calibracion_meses'] == 12.5
        assert result['cantidad'] == 100


@pytest.mark.django_db
class TestProcessAllRowDates:
    """Tests para _process_all_row_dates"""

    def test_process_dates_all_valid(self):
        """Test procesamiento de fechas válidas"""
        row_data = {
            'fecha_adquisicion': '2024-01-15',
            'fecha_version_formato': '2024-02-20',
            'fecha_ultima_calibracion': '2024-03-10',
            'fecha_ultimo_mantenimiento': '2024-04-05',
            'fecha_ultima_comprobacion': '2024-05-01'
        }

        result = _process_all_row_dates(row_data, 10)

        assert len(result['errors']) == 0
        assert result['dates']['fecha_adquisicion'] == date(2024, 1, 15)
        assert result['dates']['fecha_version_formato'] == date(2024, 2, 20)
        assert result['dates']['fecha_ultima_calibracion'] == date(2024, 3, 10)
        assert result['dates']['fecha_ultimo_mantenimiento'] == date(2024, 4, 5)
        assert result['dates']['fecha_ultima_comprobacion'] == date(2024, 5, 1)

    def test_process_dates_multiple_formats(self):
        """Test diferentes formatos de fecha"""
        row_data = {
            'fecha_adquisicion': '15/01/2024',  # DD/MM/YYYY
            'fecha_version_formato': '2024-02-20',  # YYYY-MM-DD
            'fecha_ultima_calibracion': '10-03-2024'  # DD-MM-YYYY
        }

        result = _process_all_row_dates(row_data, 5)

        assert len(result['errors']) == 0
        assert result['dates']['fecha_adquisicion'] == date(2024, 1, 15)
        assert result['dates']['fecha_version_formato'] == date(2024, 2, 20)
        assert result['dates']['fecha_ultima_calibracion'] == date(2024, 3, 10)

    def test_process_dates_invalid_date(self):
        """Test manejo de fecha inválida"""
        row_data = {
            'fecha_adquisicion': 'fecha-invalida',
            'fecha_version_formato': '2024-02-20'
        }

        result = _process_all_row_dates(row_data, 8)

        assert len(result['errors']) > 0
        assert 'Fila 8' in result['errors'][0]
        assert result['dates']['fecha_adquisicion'] is None
        assert result['dates']['fecha_version_formato'] == date(2024, 2, 20)

    def test_process_dates_empty_data(self):
        """Test sin fechas proporcionadas"""
        row_data = {}

        result = _process_all_row_dates(row_data, 1)

        assert len(result['errors']) == 0
        assert result['dates']['fecha_adquisicion'] is None
        assert result['dates']['fecha_version_formato'] is None
        assert result['dates']['fecha_ultima_calibracion'] is None


@pytest.mark.django_db
class TestUpdateExistingEquipment:
    """Tests para _update_existing_equipment"""

    def test_update_basic_fields(self, sample_empresa):
        """Test actualización de campos básicos"""
        # Crear equipo existente
        equipo = Equipo.objects.create(
            codigo_interno='EQ001',
            nombre='Equipo Original',
            empresa=sample_empresa,
            marca='Marca Original',
            estado=ESTADO_ACTIVO
        )

        # Datos para actualizar
        row_data = {
            'nombre': 'Equipo Actualizado',
            'marca': 'Marca Nueva',
            'modelo': 'Modelo V2'
        }
        dates_dict = {}

        # Ejecutar actualización
        campos = _update_existing_equipment(equipo, row_data, dates_dict)

        # Verificar cambios
        equipo.refresh_from_db()
        assert equipo.nombre == 'Equipo Actualizado'
        assert equipo.marca == 'Marca Nueva'
        assert equipo.modelo == 'Modelo V2'
        assert 'nombre' in campos
        assert 'marca' in campos
        assert 'modelo' in campos

    def test_update_only_provided_fields(self, sample_empresa):
        """Test que solo actualiza campos proporcionados"""
        equipo = Equipo.objects.create(
            codigo_interno='EQ002',
            nombre='Equipo Test',
            empresa=sample_empresa,
            marca='Marca Original',
            modelo='Modelo Original'
        )

        # Solo actualizar nombre
        row_data = {'nombre': 'Nuevo Nombre'}
        dates_dict = {}

        campos = _update_existing_equipment(equipo, row_data, dates_dict)

        equipo.refresh_from_db()
        assert equipo.nombre == 'Nuevo Nombre'
        assert equipo.marca == 'Marca Original'  # No cambió
        assert equipo.modelo == 'Modelo Original'  # No cambió
        assert len(campos) == 1

    def test_update_with_dates(self, sample_empresa):
        """Test actualización con fechas"""
        equipo = Equipo.objects.create(
            codigo_interno='EQ003',
            nombre='Equipo Test',
            empresa=sample_empresa
        )

        row_data = {'nombre': 'Nuevo Nombre'}
        dates_dict = {
            'fecha_adquisicion': date(2024, 1, 15),
            'fecha_ultima_calibracion': date(2024, 6, 1)
        }

        campos = _update_existing_equipment(equipo, row_data, dates_dict)

        equipo.refresh_from_db()
        assert equipo.fecha_adquisicion == date(2024, 1, 15)
        assert equipo.fecha_ultima_calibracion == date(2024, 6, 1)
        assert 'fecha_adquisicion' in campos
        assert 'fecha_ultima_calibracion' in campos

    def test_update_no_changes_returns_empty(self, sample_empresa):
        """Test sin cambios no actualiza nada"""
        equipo = Equipo.objects.create(
            codigo_interno='EQ004',
            nombre='Equipo Test',
            empresa=sample_empresa
        )

        row_data = {}  # Sin datos
        dates_dict = {}

        campos = _update_existing_equipment(equipo, row_data, dates_dict)

        assert len(campos) == 0


@pytest.mark.django_db
class TestCreateNewEquipment:
    """Tests para _create_new_equipment"""

    def test_create_basic_equipment(self, sample_empresa):
        """Test creación básica de equipo"""
        row_data = {
            'codigo_interno': 'EQ-NEW-001',
            'nombre': 'Equipo Nuevo'
        }
        dates_dict = {}

        equipo = _create_new_equipment(row_data, sample_empresa, dates_dict)

        assert equipo.pk is not None
        assert equipo.codigo_interno == 'EQ-NEW-001'
        assert equipo.nombre == 'Equipo Nuevo'
        assert equipo.empresa == sample_empresa
        assert equipo.estado == 'Activo'  # Default

    def test_create_equipment_with_all_fields(self, sample_empresa):
        """Test creación con todos los campos"""
        row_data = {
            'codigo_interno': 'EQ-FULL-001',
            'nombre': 'Equipo Completo',
            'tipo_equipo': 'Equipo de Medición',
            'marca': 'Test Brand',
            'modelo': 'Model X',
            'numero_serie': 'SN123456',
            'ubicacion_nombre': 'Laboratorio',
            'responsable': 'Juan Pérez',
            'estado': ESTADO_ACTIVO,
            'rango_medida': '0-100',
            'resolucion': '0.01',
            'frecuencia_calibracion_meses': 12
        }
        dates_dict = {
            'fecha_adquisicion': date(2024, 1, 1),
            'fecha_ultima_calibracion': date(2024, 6, 1)
        }

        equipo = _create_new_equipment(row_data, sample_empresa, dates_dict)

        assert equipo.codigo_interno == 'EQ-FULL-001'
        assert equipo.marca == 'Test Brand'
        assert equipo.modelo == 'Model X'
        assert equipo.numero_serie == 'SN123456'
        assert equipo.ubicacion == 'Laboratorio'
        assert equipo.responsable == 'Juan Pérez'
        assert equipo.fecha_adquisicion == date(2024, 1, 1)
        assert equipo.fecha_ultima_calibracion == date(2024, 6, 1)

    def test_create_equipment_with_defaults(self, sample_empresa):
        """Test valores por defecto cuando no se proporcionan"""
        row_data = {
            'codigo_interno': 'EQ-DEF-001',
            'nombre': 'Equipo Básico'
        }
        dates_dict = {}

        equipo = _create_new_equipment(row_data, sample_empresa, dates_dict)

        assert equipo.marca == ''
        assert equipo.modelo == ''
        assert equipo.numero_serie == ''
        assert equipo.ubicacion == ''
        assert equipo.tipo_equipo == 'Equipo de Medición'


# =============================================================================
# TESTS: PDF HELPERS (2 funciones)
# =============================================================================

@pytest.mark.django_db
class TestGetPdfFileUrl:
    """Tests para _get_pdf_file_url"""

    def test_get_pdf_file_url_none_field(self):
        """Test con campo None"""
        mock_request = Mock()

        result = _get_pdf_file_url(mock_request, None)

        assert result is None

    def test_get_pdf_file_url_empty_name(self):
        """Test con campo sin nombre de archivo"""
        mock_request = Mock()
        mock_file = Mock()
        mock_file.name = ''

        result = _get_pdf_file_url(mock_request, mock_file)

        assert result is None

    @patch('core.views.reports.default_storage')
    def test_get_pdf_file_url_local_storage(self, mock_storage):
        """Test con FileSystemStorage local"""
        mock_storage.__class__.__name__ = 'FileSystemStorage'
        mock_storage.url.return_value = '/media/test.pdf'

        mock_request = Mock()
        mock_request.build_absolute_uri.return_value = 'http://localhost/media/test.pdf'

        mock_file = Mock()
        mock_file.name = 'test.pdf'

        result = _get_pdf_file_url(mock_request, mock_file)

        # Verificar que retorna URL (puede ser relativa o absoluta)
        assert 'test.pdf' in result
        assert result.startswith(('/media', 'http://', 'https://'))

    @patch('core.views.reports.default_storage')
    def test_get_pdf_file_url_s3_storage(self, mock_storage):
        """Test con S3 storage"""
        mock_storage.__class__.__name__ = 'S3Boto3Storage'
        mock_storage.exists.return_value = True
        mock_storage.url.return_value = '//s3.amazonaws.com/bucket/test.pdf'

        mock_request = Mock()
        mock_file = Mock()
        mock_file.name = 'test.pdf'

        result = _get_pdf_file_url(mock_request, mock_file)

        assert result.startswith('https://')
        assert 'test.pdf' in result


@pytest.mark.django_db
class TestGetPdfImageData:
    """Tests para _get_pdf_image_data"""

    def test_get_pdf_image_data_none_field(self):
        """Test con campo None"""
        result = _get_pdf_image_data(None)
        assert result is None

    def test_get_pdf_image_data_empty_name(self):
        """Test con campo sin nombre"""
        mock_file = Mock()
        mock_file.name = ''

        result = _get_pdf_image_data(mock_file)
        assert result is None

    def test_get_pdf_image_data_non_image_extension(self):
        """Test con archivo que no es imagen"""
        mock_file = Mock()
        mock_file.name = 'document.pdf'

        result = _get_pdf_image_data(mock_file)
        assert result is None

    @patch('core.views.reports.default_storage')
    def test_get_pdf_image_data_too_large(self, mock_storage):
        """Test con imagen muy grande (>1MB)"""
        mock_storage.size.return_value = 2 * 1024 * 1024  # 2MB

        mock_file = Mock()
        mock_file.name = 'large_image.jpg'

        result = _get_pdf_image_data(mock_file)

        assert result is None

    @patch('core.views.reports.default_storage')
    def test_get_pdf_image_data_success(self, mock_storage):
        """Test conversión exitosa a base64"""
        # Simular imagen pequeña
        mock_storage.size.return_value = 100  # 100 bytes
        mock_storage.open.return_value.__enter__.return_value.read.return_value = b'fake_image_data'

        mock_file = Mock()
        mock_file.name = 'test_image.jpg'

        result = _get_pdf_image_data(mock_file)

        assert result is not None
        assert result.startswith('data:image/jpeg;base64,')


# =============================================================================
# TESTS: EXCEL FORMAT HELPERS (2 funciones)
# =============================================================================

@pytest.mark.django_db
class TestAddExcelHeader:
    """Tests para _add_excel_header"""

    def test_add_excel_header_basic(self, sample_empresa):
        """Test creación de encabezado básico"""
        workbook = Workbook()
        sheet = workbook.active

        row = _add_excel_header(sheet, sample_empresa)

        # Verificar que retorna fila correcta
        assert row == 6

        # Verificar título
        assert 'SAM METROLOGÍA' in sheet['A1'].value

        # Verificar empresa
        assert sample_empresa.nombre.upper() in sheet['A3'].value

    def test_add_excel_header_styling(self, sample_empresa):
        """Test que aplica estilos correctamente"""
        workbook = Workbook()
        sheet = workbook.active

        _add_excel_header(sheet, sample_empresa)

        # Verificar estilos del título
        assert sheet['A1'].font.bold is True
        assert sheet['A1'].font.size == 20
        # Verificar color (puede variar formato RGB)
        assert '1f4e79' in sheet['A1'].fill.start_color.rgb

    def test_add_excel_header_row_heights(self, sample_empresa):
        """Test altura de filas configurada"""
        workbook = Workbook()
        sheet = workbook.active

        _add_excel_header(sheet, sample_empresa)

        assert sheet.row_dimensions[1].height == 40
        assert sheet.row_dimensions[2].height == 25


@pytest.mark.django_db
class TestAddExcelResumenSection:
    """Tests para _add_excel_resumen_section"""

    def test_add_resumen_section_basic(self, sample_empresa):
        """Test creación de sección resumen"""
        # Crear equipos de prueba
        equipos = [
            Equipo.objects.create(
                codigo_interno=f'EQ{i:03d}',
                nombre=f'Equipo {i}',
                empresa=sample_empresa,
                estado=ESTADO_ACTIVO if i % 2 == 0 else ESTADO_INACTIVO
            )
            for i in range(1, 6)  # 5 equipos
        ]

        workbook = Workbook()
        sheet = workbook.active

        next_row = _add_excel_resumen_section(sheet, equipos, 6)

        # Verificar que retorna siguiente fila
        assert next_row > 6

        # Verificar título de sección
        assert 'RESUMEN GENERAL' in sheet['A6'].value

    def test_add_resumen_section_calculations(self, sample_empresa):
        """Test cálculos de estadísticas"""
        # 3 activos, 2 inactivos
        equipos = [
            Equipo.objects.create(
                codigo_interno=f'EQ{i:03d}',
                nombre=f'Equipo {i}',
                empresa=sample_empresa,
                estado=ESTADO_ACTIVO if i <= 3 else ESTADO_INACTIVO
            )
            for i in range(1, 6)
        ]

        workbook = Workbook()
        sheet = workbook.active

        _add_excel_resumen_section(sheet, equipos, 6)

        # Verificar datos (fila 9 es donde empiezan los datos después de headers)
        # Fila 9: Total de Equipos, 5, 100%
        assert sheet.cell(row=9, column=2).value == 5
        assert sheet.cell(row=9, column=3).value == '100%'

        # Fila 10: Equipos Activos, 3, 60%
        assert sheet.cell(row=10, column=2).value == 3
        assert '60' in sheet.cell(row=10, column=3).value

    def test_add_resumen_section_empty_list(self):
        """Test con lista vacía de equipos"""
        workbook = Workbook()
        sheet = workbook.active

        next_row = _add_excel_resumen_section(sheet, [], 6)

        # Debería manejar división por cero
        assert next_row > 6
        assert sheet.cell(row=9, column=2).value == 0
        assert '0%' in sheet.cell(row=10, column=3).value
