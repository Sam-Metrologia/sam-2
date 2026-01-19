# core/views/export_financiero.py
# Funciones de exportación para análisis financiero

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from datetime import date
from io import BytesIO

from ..utils.analisis_financiero import (
    calcular_analisis_financiero_empresa,
    calcular_proyeccion_costos_empresa,
    calcular_metricas_financieras_sam,
    calcular_presupuesto_mensual_detallado
)
from ..models import Empresa


@login_required
def exportar_analisis_financiero_excel(request):
    """
    Exporta el análisis financiero completo a Excel
    - Vista EMPRESA: Análisis de costos y proyección
    - Vista SAM: Métricas del negocio multi-empresa
    """
    user = request.user
    today = date.today()
    current_year = today.year

    # Obtener año de proyección desde parámetro GET (por defecto: año siguiente)
    año_proyeccion = int(request.GET.get('año_proyeccion', current_year + 1))

    # Crear workbook
    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    subheader_font = Font(bold=True, color="1F2937")
    subheader_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    currency_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    percentage_format = '0.0%'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if user.is_superuser:
        # VISTA SAM: Exportar métricas del negocio
        ws = wb.active
        ws.title = "Análisis Financiero SAM"

        # Filtrado por empresa
        selected_company_id = request.GET.get('empresa_id')
        empresas_disponibles = Empresa.objects.filter(is_deleted=False).order_by('nombre')

        if selected_company_id:
            empresas_queryset = empresas_disponibles.filter(id=selected_company_id)
        else:
            empresas_queryset = empresas_disponibles

        # Calcular métricas SAM
        metricas_sam = calcular_metricas_financieras_sam(empresas_queryset, current_year, current_year - 1)

        # Header principal
        ws.merge_cells('A1:F1')
        ws['A1'] = f"📈 Análisis Financiero del Negocio SAM - {current_year}"
        ws['A1'].font = Font(bold=True, size=16, color="1F2937")
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Generado: {today.strftime('%d/%m/%Y')} | Empresas analizadas: {empresas_queryset.count()}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # Métricas principales
        row = 4
        headers = ['Métrica', 'Valor Actual', 'Comparativo', 'Crecimiento', 'Estado', 'Observaciones']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Datos de métricas
        metricas_data = [
            ['Ingreso Total YTD', metricas_sam['ingreso_ytd_actual'], metricas_sam['ingreso_año_anterior'],
             f"{metricas_sam['crecimiento_ingresos_porcentaje']}%",
             'Positivo' if metricas_sam['crecimiento_ingresos_porcentaje'] >= 0 else 'Negativo',
             f"Acumulado año {current_year}"],
            ['Empresas Activas', metricas_sam['empresas_activas_actual'], metricas_sam['empresas_activas_anterior'],
             f"{metricas_sam['crecimiento_empresas_porcentaje']}%",
             'Expansión' if metricas_sam['crecimiento_empresas_porcentaje'] >= 0 else 'Contracción',
             'Base de clientes'],
            ['Proyección Fin Año', metricas_sam['proyeccion_fin_año'], metricas_sam['ingreso_ytd_actual'],
             f"{((metricas_sam['proyeccion_fin_año'] - metricas_sam['ingreso_ytd_actual']) / max(metricas_sam['ingreso_ytd_actual'], 1)) * 100:.1f}%",
             'Estimado', 'Proyección lineal'],
        ]

        for row_data in metricas_data:
            row += 1
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [2, 3]:  # Valores monetarios
                    cell.number_format = currency_format
                elif col == 4 and '%' in str(value):  # Porcentajes
                    cell.number_format = percentage_format

        filename = f"analisis_financiero_sam_{current_year}_{today.strftime('%Y%m%d')}.xlsx"

    elif getattr(user, 'rol_usuario', None) == 'GERENCIA' and user.empresa:
        # VISTA EMPRESA: Exportar análisis de costos CON PRESUPUESTO CALENDARIO
        empresa = user.empresa

        # Calcular análisis financiero
        analisis_financiero = calcular_analisis_financiero_empresa(empresa, current_year, today)
        proyeccion_costos = calcular_proyeccion_costos_empresa(empresa, año_proyeccion)
        presupuesto_calendario = calcular_presupuesto_mensual_detallado(empresa, año_proyeccion)

        # HOJA 1: RESUMEN EJECUTIVO
        ws = wb.active
        ws.title = "Resumen Ejecutivo"

        # Header principal
        ws.merge_cells('A1:F1')
        ws['A1'] = f"💰 Análisis Financiero Metrológico - {empresa.nombre}"
        ws['A1'].font = Font(bold=True, size=16, color="1F2937")
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Período: {current_year} | Generado: {today.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # Sección 1: Gasto YTD
        row = 4
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "📊 TRAZABILIDAD DE COSTOS YTD"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        row += 2
        headers = ['Tipo de Actividad', 'Costo YTD', 'Porcentaje', 'Descripción']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Datos de costos YTD
        costos_data = [
            ['Calibraciones', analisis_financiero['costos_calibracion_ytd'],
             f"{analisis_financiero['porcentaje_calibracion']}%", 'Usualmente externo'],
            ['Mantenimientos', analisis_financiero['costos_mantenimiento_ytd'],
             f"{analisis_financiero['porcentaje_mantenimiento']}%", 'Usualmente interno'],
            ['Comprobaciones', analisis_financiero['costos_comprobacion_ytd'],
             f"{analisis_financiero['porcentaje_comprobacion']}%", 'Usualmente interno'],
            ['TOTAL', analisis_financiero['gasto_ytd_total'], '100%', 'Gasto real incurrido'],
        ]

        for row_data in costos_data:
            row += 1
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 2:  # Valores monetarios
                    cell.number_format = currency_format
                if row_data[0] == 'TOTAL':  # Fila total en negrita
                    cell.font = Font(bold=True)

        # Sección 2: Proyección
        row += 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"🔮 PROYECCIÓN DE COSTOS {año_proyeccion}"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        row += 2
        headers = ['Métrica', 'Valor Proyectado', 'Base Cálculo', 'Observaciones']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Datos de proyección
        variacion = ((proyeccion_costos['proyeccion_gasto_proximo_año'] - analisis_financiero['gasto_ytd_total']) / max(analisis_financiero['gasto_ytd_total'], 1)) * 100

        proyeccion_data = [
            ['Gasto Estimado Total', proyeccion_costos['proyeccion_gasto_proximo_año'],
             f"{proyeccion_costos['actividades_proyectadas_total']} actividades",
             'Basado en homogeneidad equipos'],
            ['Variación vs YTD', f"{variacion:.1f}%",
             f"${analisis_financiero['gasto_ytd_total']:,.0f} actual",
             'Incremento/reducción esperado'],
        ]

        for row_data in proyeccion_data:
            row += 1
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 2 and '$' in str(value):  # Valores monetarios
                    cell.number_format = currency_format

        # Nota importante
        row += 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "⚠️ NOTA: Esta proyección no incluye costos por mantenimientos correctivos no programados"
        ws[f'A{row}'].font = Font(italic=True, color="D97706")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # HOJA 2: PRESUPUESTO CALENDARIO DETALLADO
        ws_calendario = wb.create_sheet(title=f"Presupuesto {año_proyeccion}")

        # Estilos específicos para calendario
        calibracion_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # Azul claro
        mantenimiento_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Verde claro
        comprobacion_fill = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")  # Naranja claro
        mes_header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")  # Índigo
        mes_header_font = Font(bold=True, size=12, color="FFFFFF")
        subtotal_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # Gris claro
        subtotal_font = Font(bold=True, size=10)

        # Header principal
        ws_calendario.merge_cells('A1:E1')
        ws_calendario['A1'] = f"📅 PRESUPUESTO CALENDARIO {año_proyeccion} - {empresa.nombre}"
        ws_calendario['A1'].font = Font(bold=True, size=14, color="1F2937")
        ws_calendario['A1'].alignment = Alignment(horizontal='center')

        # Resumen ejecutivo del presupuesto
        resumen = presupuesto_calendario['resumen']
        ws_calendario.merge_cells('A2:E2')
        ws_calendario['A2'] = f"Total Anual: ${resumen['total_anual']:,.0f} COP | {resumen['count_total_actividades']} actividades programadas"
        ws_calendario['A2'].font = Font(bold=True, size=10, color="059669")
        ws_calendario['A2'].alignment = Alignment(horizontal='center')

        ws_calendario.merge_cells('A3:E3')
        ws_calendario['A3'] = f"Calibraciones: {resumen['count_calibraciones']} | Mantenimientos: {resumen['count_mantenimientos']} | Comprobaciones: {resumen['count_comprobaciones']}"
        ws_calendario['A3'].font = Font(size=9, color="6B7280")
        ws_calendario['A3'].alignment = Alignment(horizontal='center')

        row_cal = 5

        # Headers comunes para todas las secciones (calibraciones, mantenimientos, comprobaciones)
        headers_cal = ['Código', 'Nombre Equipo', 'Marca', 'Modelo', 'Costo']

        # Procesar cada mes
        for mes_data in presupuesto_calendario['presupuesto_por_mes']:
            # Si el mes no tiene actividades, saltar
            if mes_data['total_mes'] == 0:
                continue

            # HEADER DEL MES
            ws_calendario.merge_cells(f'A{row_cal}:E{row_cal}')
            cell_mes = ws_calendario[f'A{row_cal}']
            cell_mes.value = f"🗓️  {mes_data['nombre_mes'].upper()} {año_proyeccion} - Total: ${mes_data['total_mes']:,.0f}"
            cell_mes.font = mes_header_font
            cell_mes.fill = mes_header_fill
            cell_mes.alignment = Alignment(horizontal='center', vertical='center')
            cell_mes.border = border
            ws_calendario.row_dimensions[row_cal].height = 25
            row_cal += 1

            # CALIBRACIONES
            if mes_data['calibraciones']:
                # Subheader
                ws_calendario.merge_cells(f'A{row_cal}:E{row_cal}')
                cell_cal_header = ws_calendario[f'A{row_cal}']
                cell_cal_header.value = f"🔵 CALIBRACIONES ({len(mes_data['calibraciones'])})"
                cell_cal_header.font = Font(bold=True, size=10, color="1E40AF")
                cell_cal_header.fill = calibracion_fill
                cell_cal_header.border = border
                row_cal += 1

                # Headers de columnas
                for col, header in enumerate(headers_cal, 1):
                    cell = ws_calendario.cell(row=row_cal, column=col, value=header)
                    cell.font = Font(bold=True, size=9)
                    cell.fill = subheader_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                row_cal += 1

                # Datos de calibraciones
                for cal in mes_data['calibraciones']:
                    ws_calendario.cell(row=row_cal, column=1, value=cal['codigo']).border = border
                    ws_calendario.cell(row=row_cal, column=2, value=cal['nombre']).border = border
                    ws_calendario.cell(row=row_cal, column=3, value=cal['marca']).border = border
                    ws_calendario.cell(row=row_cal, column=4, value=cal['modelo']).border = border
                    cost_cell = ws_calendario.cell(row=row_cal, column=5, value=cal['costo'])
                    cost_cell.number_format = currency_format
                    cost_cell.border = border
                    row_cal += 1

                # Subtotal calibraciones
                ws_calendario.merge_cells(f'A{row_cal}:D{row_cal}')
                subtotal_cell = ws_calendario[f'A{row_cal}']
                subtotal_cell.value = "Subtotal Calibraciones"
                subtotal_cell.font = subtotal_font
                subtotal_cell.fill = subtotal_fill
                subtotal_cell.border = border
                subtotal_cell.alignment = Alignment(horizontal='right')

                subtotal_valor = ws_calendario.cell(row=row_cal, column=5, value=mes_data['total_calibraciones'])
                subtotal_valor.number_format = currency_format
                subtotal_valor.font = subtotal_font
                subtotal_valor.fill = subtotal_fill
                subtotal_valor.border = border
                row_cal += 2

            # MANTENIMIENTOS
            if mes_data['mantenimientos']:
                # Subheader
                ws_calendario.merge_cells(f'A{row_cal}:E{row_cal}')
                cell_mant_header = ws_calendario[f'A{row_cal}']
                cell_mant_header.value = f"🟢 MANTENIMIENTOS ({len(mes_data['mantenimientos'])})"
                cell_mant_header.font = Font(bold=True, size=10, color="047857")
                cell_mant_header.fill = mantenimiento_fill
                cell_mant_header.border = border
                row_cal += 1

                # Headers de columnas
                for col, header in enumerate(headers_cal, 1):
                    cell = ws_calendario.cell(row=row_cal, column=col, value=header)
                    cell.font = Font(bold=True, size=9)
                    cell.fill = subheader_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                row_cal += 1

                # Datos de mantenimientos
                for mant in mes_data['mantenimientos']:
                    ws_calendario.cell(row=row_cal, column=1, value=mant['codigo']).border = border
                    ws_calendario.cell(row=row_cal, column=2, value=mant['nombre']).border = border
                    ws_calendario.cell(row=row_cal, column=3, value=mant['marca']).border = border
                    ws_calendario.cell(row=row_cal, column=4, value=mant['modelo']).border = border
                    cost_cell = ws_calendario.cell(row=row_cal, column=5, value=mant['costo'])
                    cost_cell.number_format = currency_format
                    cost_cell.border = border
                    row_cal += 1

                # Subtotal mantenimientos
                ws_calendario.merge_cells(f'A{row_cal}:D{row_cal}')
                subtotal_cell = ws_calendario[f'A{row_cal}']
                subtotal_cell.value = "Subtotal Mantenimientos"
                subtotal_cell.font = subtotal_font
                subtotal_cell.fill = subtotal_fill
                subtotal_cell.border = border
                subtotal_cell.alignment = Alignment(horizontal='right')

                subtotal_valor = ws_calendario.cell(row=row_cal, column=5, value=mes_data['total_mantenimientos'])
                subtotal_valor.number_format = currency_format
                subtotal_valor.font = subtotal_font
                subtotal_valor.fill = subtotal_fill
                subtotal_valor.border = border
                row_cal += 2

            # COMPROBACIONES
            if mes_data['comprobaciones']:
                # Subheader
                ws_calendario.merge_cells(f'A{row_cal}:E{row_cal}')
                cell_comp_header = ws_calendario[f'A{row_cal}']
                cell_comp_header.value = f"🟠 COMPROBACIONES ({len(mes_data['comprobaciones'])})"
                cell_comp_header.font = Font(bold=True, size=10, color="C2410C")
                cell_comp_header.fill = comprobacion_fill
                cell_comp_header.border = border
                row_cal += 1

                # Headers de columnas
                for col, header in enumerate(headers_cal, 1):
                    cell = ws_calendario.cell(row=row_cal, column=col, value=header)
                    cell.font = Font(bold=True, size=9)
                    cell.fill = subheader_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                row_cal += 1

                # Datos de comprobaciones
                for comp in mes_data['comprobaciones']:
                    ws_calendario.cell(row=row_cal, column=1, value=comp['codigo']).border = border
                    ws_calendario.cell(row=row_cal, column=2, value=comp['nombre']).border = border
                    ws_calendario.cell(row=row_cal, column=3, value=comp['marca']).border = border
                    ws_calendario.cell(row=row_cal, column=4, value=comp['modelo']).border = border
                    cost_cell = ws_calendario.cell(row=row_cal, column=5, value=comp['costo'])
                    cost_cell.number_format = currency_format
                    cost_cell.border = border
                    row_cal += 1

                # Subtotal comprobaciones
                ws_calendario.merge_cells(f'A{row_cal}:D{row_cal}')
                subtotal_cell = ws_calendario[f'A{row_cal}']
                subtotal_cell.value = "Subtotal Comprobaciones"
                subtotal_cell.font = subtotal_font
                subtotal_cell.fill = subtotal_fill
                subtotal_cell.border = border
                subtotal_cell.alignment = Alignment(horizontal='right')

                subtotal_valor = ws_calendario.cell(row=row_cal, column=5, value=mes_data['total_comprobaciones'])
                subtotal_valor.number_format = currency_format
                subtotal_valor.font = subtotal_font
                subtotal_valor.fill = subtotal_fill
                subtotal_valor.border = border
                row_cal += 2

            # TOTAL DEL MES
            ws_calendario.merge_cells(f'A{row_cal}:D{row_cal}')
            total_mes_cell = ws_calendario[f'A{row_cal}']
            total_mes_cell.value = f"💰 TOTAL {mes_data['nombre_mes'].upper()}"
            total_mes_cell.font = Font(bold=True, size=11, color="FFFFFF")
            total_mes_cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            total_mes_cell.border = border
            total_mes_cell.alignment = Alignment(horizontal='right')

            total_mes_valor = ws_calendario.cell(row=row_cal, column=5, value=mes_data['total_mes'])
            total_mes_valor.number_format = currency_format
            total_mes_valor.font = Font(bold=True, size=11, color="FFFFFF")
            total_mes_valor.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            total_mes_valor.border = border
            ws_calendario.row_dimensions[row_cal].height = 20

            row_cal += 3  # Espacio entre meses

        # Ajustar anchos de columnas para la hoja de calendario
        ws_calendario.column_dimensions['A'].width = 15
        ws_calendario.column_dimensions['B'].width = 30
        ws_calendario.column_dimensions['C'].width = 20
        ws_calendario.column_dimensions['D'].width = 20
        ws_calendario.column_dimensions['E'].width = 18

        filename = f"analisis_financiero_{empresa.nombre.replace(' ', '_')}_{current_year}_{today.strftime('%Y%m%d')}.xlsx"
    else:
        # Usuario sin permisos
        return HttpResponse("Sin permisos para exportar análisis financiero", status=403)

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Preparar respuesta HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response