# core/views/reports.py
# Views para generación de reportes, informes y exportaciones

from .base import *
import zipfile
import threading
import time
from ..constants import ESTADO_ACTIVO, ESTADO_INACTIVO, ESTADO_DE_BAJA

# =============================================================================
# API ENDPOINTS FOR PROGRESS TRACKING (Fase 3)
# =============================================================================

@monitor_view
@login_required
def zip_progress_api(request):
    """
    API endpoint para consultar progreso de solicitudes ZIP del usuario.
    Retorna información en tiempo real del estado de generación.
    """
    try:
        # Obtener todas las solicitudes activas del usuario
        user_requests = ZipRequest.objects.filter(
            user=request.user,
            status__in=['pending', 'processing']
        ).order_by('-created_at')

        progress_data = []
        for zip_req in user_requests:
            data = {
                'id': zip_req.id,
                'status': zip_req.status,
                'empresa_nombre': zip_req.empresa.nombre,
                'progress_percentage': zip_req.progress_percentage,
                'current_step': zip_req.current_step or 'Inicializando...',
                'total_equipos': zip_req.total_equipos,
                'equipos_procesados': zip_req.equipos_procesados,
                'created_at': zip_req.created_at.isoformat(),
                'estimated_completion': zip_req.estimated_completion.isoformat() if zip_req.estimated_completion else None,
                'position_in_queue': zip_req.position_in_queue if zip_req.status == 'pending' else None
            }
            progress_data.append(data)

        return JsonResponse({
            'success': True,
            'requests': progress_data,
            'total_active': len(progress_data)
        })

    except Exception as e:
        logger.error(f"Error en zip_progress_api: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@monitor_view
@login_required
def notifications_api(request):
    """
    API endpoint para obtener notificaciones del usuario.
    Soporta marcar como leídas y obtener conteo de no leídas.
    """
    try:
        from core.models import NotificacionZip

        if request.method == 'POST':
            # Marcar notificación como leída
            notification_id = request.POST.get('notification_id')
            action = request.POST.get('action', 'read')

            if notification_id:
                try:
                    notification = NotificacionZip.objects.get(
                        id=notification_id,
                        user=request.user
                    )
                    if action == 'read':
                        notification.status = 'read'
                        notification.read_at = timezone.now()
                    elif action == 'dismiss':
                        notification.status = 'dismissed'

                    notification.save()
                    return JsonResponse({'success': True, 'action': action})
                except NotificacionZip.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Notificación no encontrada'})

        # GET: Obtener notificaciones
        notifications = NotificacionZip.objects.filter(
            user=request.user,
            status__in=['unread', 'read']
        ).order_by('-created_at')[:10]  # Últimas 10

        notifications_data = []
        for notif in notifications:
            data = {
                'id': notif.id,
                'tipo': notif.tipo,
                'titulo': notif.titulo,
                'mensaje': notif.mensaje,
                'status': notif.status,
                'created_at': notif.created_at.isoformat(),
                'zip_request_id': notif.zip_request.id,
                'empresa_nombre': notif.zip_request.empresa.nombre,
                'can_download': notif.zip_request.status == 'completed' and notif.zip_request.file_path
            }
            notifications_data.append(data)

        # Contar no leídas
        unread_count = NotificacionZip.objects.filter(
            user=request.user,
            status='unread'
        ).count()

        return JsonResponse({
            'success': True,
            'notifications': notifications_data,
            'unread_count': unread_count
        })

    except Exception as e:
        logger.error(f"Error en notifications_api: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@monitor_view
@login_required
def system_monitor_dashboard(request):
    """
    Dashboard de monitoreo del sistema ZIP para administradores.
    Muestra estadísticas de rendimiento, cola, y uso de recursos.
    """
    # Solo superusuarios pueden acceder
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para acceder al monitoreo del sistema.')
        return redirect('core:dashboard')

    try:
        from core.models import ZipRequest, NotificacionZip
        from django.db.models import Count, Avg, Max, F, Q
        from datetime import datetime, timedelta

        # Estadísticas de las últimas 24 horas
        last_24h = timezone.now() - timedelta(hours=24)
        last_7d = timezone.now() - timedelta(days=7)

        # Estadísticas de ZIPs
        zip_stats = {
            'pending': ZipRequest.objects.filter(status='pending').count(),
            'processing': ZipRequest.objects.filter(status='processing').count(),
            'completed_24h': ZipRequest.objects.filter(status='completed', completed_at__gte=last_24h).count(),
            'failed_24h': ZipRequest.objects.filter(status='failed', completed_at__gte=last_24h).count(),
            'total_7d': ZipRequest.objects.filter(created_at__gte=last_7d).count(),
        }

        # Tiempo promedio de procesamiento
        completed_recent = ZipRequest.objects.filter(
            status='completed',
            completed_at__gte=last_7d,
            started_at__isnull=False
        ).annotate(
            processing_time=F('completed_at') - F('started_at')
        )

        avg_processing_time = None
        if completed_recent.exists():
            # Calcular promedio en segundos
            total_seconds = sum([
                req.processing_time.total_seconds()
                for req in completed_recent
                if req.processing_time
            ])
            avg_processing_time = total_seconds / completed_recent.count()

        # Estadísticas por empresa (top 5)
        empresa_stats = ZipRequest.objects.filter(
            created_at__gte=last_7d
        ).values('empresa__nombre').annotate(
            total_requests=Count('id'),
            completed=Count('id', filter=Q(status='completed')),
            failed=Count('id', filter=Q(status='failed'))
        ).order_by('-total_requests')[:5]

        # Uso de caché (si está configurado)
        cache_stats = None
        try:
            from django.core.cache import cache
            from django.core.cache.backends.base import InvalidCacheBackendError

            # Intentar obtener estadísticas básicas
            cache_stats = {
                'backend': cache.__class__.__name__,
                'configured': True
            }
        except Exception:
            cache_stats = {'configured': False}

        # Notificaciones
        notif_stats = {
            'unread_total': NotificacionZip.objects.filter(status='unread').count(),
            'created_24h': NotificacionZip.objects.filter(created_at__gte=last_24h).count(),
        }

        context = {
            'titulo_pagina': 'Monitor del Sistema ZIP',
            'zip_stats': zip_stats,
            'avg_processing_time': avg_processing_time,
            'empresa_stats': empresa_stats,
            'cache_stats': cache_stats,
            'notif_stats': notif_stats,
            'last_24h': last_24h,
            'last_7d': last_7d,
        }

        return render(request, 'core/system_monitor.html', context)

    except Exception as e:
        logger.error(f"Error en system_monitor_dashboard: {e}")
        messages.error(request, f'Error cargando el monitoreo del sistema: {e}')
        return redirect('core:dashboard')


# =============================================================================
# MAIN REPORTS DASHBOARD
# =============================================================================

@monitor_view
@access_check
@login_required
def informes(request):
    """
    Página principal de informes y actividades.
    Muestra actividades próximas, vencidas y permite generar reportes.
    """
    user = request.user
    today = date.today()

    # Gestión de empresas y filtros
    selected_company_id = request.GET.get('empresa_id')
    empresas_disponibles = OptimizedQueries.get_empresas_with_stats().order_by('nombre')

    # Obtener equipos base usando queries optimizadas
    equipos_queryset = OptimizedQueries.get_equipos_optimized(user=user if not user.is_superuser else None)

    # Filtrar por empresa para usuarios normales
    if not user.is_superuser:
        if user.empresa:
            equipos_queryset = equipos_queryset.filter(empresa=user.empresa)
            selected_company_id = str(user.empresa.id)
        else:
            equipos_queryset = Equipo.objects.none()
            empresas_disponibles = Empresa.objects.none()

    # Aplicar filtro de empresa si se seleccionó
    if selected_company_id:
        equipos_queryset = equipos_queryset.filter(empresa_id=selected_company_id)

    # Obtener actividades programadas usando helper
    scheduled_activities = _get_scheduled_activities(equipos_queryset, today)

    # Categorizar actividades por tipo y estado
    activities_by_category = _categorize_activities(scheduled_activities)

    # Información de paginación para ZIPs
    zip_info = _get_zip_pagination_info(selected_company_id, user.is_superuser) if selected_company_id else None

    context = {
        'titulo_pagina': 'Informes y Actividades',
        'is_superuser': user.is_superuser,
        'empresas_disponibles': empresas_disponibles,
        'selected_company_id': selected_company_id,
        'today': today,
        'zip_info': zip_info,
        **activities_by_category  # Desempaqueta todas las categorías de actividades
    }

    return render(request, 'core/informes.html', context)


# =============================================================================
# ZIP REPORTS GENERATION
# =============================================================================

@monitor_view
@access_check
@login_required
@user_passes_test(lambda u: u.is_superuser or u.has_perm('core.can_export_reports'), login_url='/core/access_denied/')
def generar_informe_zip(request):
    """
    Genera un archivo ZIP con informes de equipos y documentos asociados.
    Para empresas con más de 100 equipos, genera ZIPs paginados.

    Estructura del ZIP:
    [Nombre Empresa]/
    ├── Equipos/
    │   ├── [Código Interno]/
    │   │   ├── Hoja_de_vida.pdf
    │   │   ├── Calibraciones/
    │   │   ├── Mantenimientos/
    │   │   └── Comprobaciones/
    └── Procedimientos/
    """
    try:
        # Obtener parámetros de la petición
        empresa_id = request.GET.get('empresa_id')
        parte = int(request.GET.get('parte', 1))

        # Validar empresa
        if not empresa_id:
            messages.error(request, 'Debe seleccionar una empresa para generar el informe.')
            return redirect('core:informes')

        empresa = get_object_or_404(Empresa, pk=empresa_id)

        # Verificar permisos
        if not request.user.is_superuser and request.user.empresa != empresa:
            messages.error(request, 'No tienes permisos para generar informes de esta empresa.')
            return redirect('core:informes')

        # Calcular paginación
        total_equipos, total_partes, equipos_por_zip = calcular_info_paginacion_zip(empresa_id, request.user.is_superuser)

        if parte > total_partes:
            messages.error(request, f'Parte {parte} no válida. Solo hay {total_partes} partes disponibles.')
            return redirect('core:informes')

        # Generar ZIP con optimizaciones de memoria
        zip_response = _generate_zip_report(empresa, parte, equipos_por_zip, total_partes)

        logger.info(f"ZIP generado exitosamente para empresa {empresa.nombre}, parte {parte}")
        return zip_response

    except Exception as e:
        logger.error(f"Error generando ZIP: {e}")
        messages.error(request, f'Error al generar el informe ZIP: {e}')
        return redirect('core:informes')


# =============================================================================
# EXCEL REPORTS GENERATION
# =============================================================================

@monitor_view
@access_check
@login_required
@user_passes_test(lambda u: u.is_superuser or u.has_perm('core.can_export_reports'), login_url='/core/access_denied/')
def generar_informe_dashboard_excel(request):
    """
    Genera un Excel consolidado con Dashboard.
    Incluye 4 hojas: Equipos, Proveedores, Procedimientos, Dashboard.
    """
    try:
        # Obtener empresa objetivo
        selected_company_id = request.GET.get('empresa_id')

        if not selected_company_id and not request.user.is_superuser:
            if request.user.empresa:
                selected_company_id = str(request.user.empresa.id)
            else:
                messages.error(request, "No tiene una empresa asignada para generar el informe Excel.")
                return redirect('core:informes')

        if not selected_company_id:
            messages.error(request, "Por favor, selecciona una empresa para generar el informe Excel.")
            return redirect('core:informes')

        empresa = get_object_or_404(Empresa, pk=selected_company_id)

        # Cargar datos optimizados
        equipos_empresa = OptimizedQueries.get_equipos_optimized(empresa=empresa).order_by('codigo_interno')
        proveedores_empresa = Proveedor.objects.filter(empresa=empresa).order_by('nombre_empresa')
        procedimientos_empresa = Procedimiento.objects.filter(empresa=empresa).order_by('codigo')

        logger.info(f"Generando Excel Dashboard para empresa {empresa.nombre}: {equipos_empresa.count()} equipos")

        # Generar contenido del Excel
        excel_content = _generate_dashboard_excel_content_local(equipos_empresa, empresa)

        # Crear respuesta HTTP optimizada
        response = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Headers de descarga
        filename_safe = sanitize_filename(f"Dashboard_{empresa.nombre}.xlsx")
        response['Content-Disposition'] = f'attachment; filename="{filename_safe}"'
        response['Content-Length'] = len(excel_content)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'

        logger.info(f"Excel Dashboard generado: {len(excel_content)} bytes")
        return response

    except Exception as e:
        logger.error(f"Error generando Excel Dashboard: {e}")
        messages.error(request, f'Error al generar el informe Dashboard: {e}')
        return redirect('core:informes')


@monitor_view
@access_check
@login_required
@permission_required('core.view_equipo', raise_exception=True)
def exportar_equipos_excel(request):
    """
    Exporta una lista general de equipos a un archivo Excel.
    Filtrada por permisos de usuario y roles.
    """
    # Verificar permisos de descarga basados en roles
    if not request.user.puede_descargar_informes():
        messages.error(request, '⛔ No tienes permisos para descargar informes. Solo usuarios con rol de Administrador, Gerente o Superusuario pueden descargar informes.')
        return redirect('core:informes')

    try:
        # Obtener equipos según permisos
        equipos = OptimizedQueries.get_equipos_optimized(user=request.user if not request.user.is_superuser else None)

        if not request.user.is_superuser and request.user.empresa:
            equipos = equipos.filter(empresa=request.user.empresa)
        elif not request.user.is_superuser and not request.user.empresa:
            equipos = Equipo.objects.none()

        # Generar contenido Excel
        excel_content = _generate_general_equipment_list_excel_content_local(equipos)

        # Crear respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="listado_equipos.xlsx"'
        response.write(excel_content)

        logger.info(f"Lista de equipos exportada: {equipos.count()} equipos")
        return response

    except Exception as e:
        logger.error(f"Error exportando equipos Excel: {e}")
        messages.error(request, f'Error al exportar equipos: {e}')
        return redirect('core:informes')


# =============================================================================
# PDF REPORTS GENERATION
# =============================================================================

@monitor_view
@access_check
@login_required
@user_passes_test(lambda u: u.is_superuser or u.has_perm('core.can_export_reports'), login_url='/core/access_denied/')
def informe_vencimientos_pdf(request):
    """
    Genera un informe PDF de actividades próximas y vencidas.
    Incluye calibraciones, mantenimientos y comprobaciones.
    """
    try:
        today = timezone.localdate()

        # Obtener equipos base según permisos
        equipos_base_query = OptimizedQueries.get_equipos_optimized(
            user=request.user if not request.user.is_superuser else None
        )

        if not request.user.is_superuser and request.user.empresa:
            equipos_base_query = equipos_base_query.filter(empresa=request.user.empresa)
        elif not request.user.is_superuser and not request.user.empresa:
            equipos_base_query = Equipo.objects.none()

        # Excluir equipos inactivos
        equipos_base_query = equipos_base_query.exclude(estado__in=[ESTADO_DE_BAJA, ESTADO_INACTIVO])

        # Generar actividades programadas
        scheduled_activities = _get_scheduled_activities(equipos_base_query, today)

        # Filtrar solo actividades próximas y vencidas (próximos 60 días)
        relevant_activities = [
            act for act in scheduled_activities
            if act['dias_restantes'] <= 60 or act['estado_vencimiento'] == 'Vencida'
        ]

        # Generar PDF
        context = {
            'activities': relevant_activities,
            'today': today,
            'empresa': request.user.empresa if not request.user.is_superuser else None,
            'is_superuser': request.user.is_superuser
        }

        pdf_content = _generate_pdf_content(request, 'core/informe_vencimientos_pdf.html', context)

        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="informe_vencimientos.pdf"'

        logger.info(f"Informe PDF de vencimientos generado: {len(relevant_activities)} actividades")
        return response

    except Exception as e:
        logger.error(f"Error generando informe PDF vencimientos: {e}")
        messages.error(request, f'Error al generar el informe PDF: {e}')
        return redirect('core:informes')


@monitor_view
@access_check
@login_required
def generar_hoja_vida_pdf(request, pk):
    """
    Genera la "Hoja de Vida" PDF para un equipo específico.
    Incluye toda la información del equipo y su historial.
    """
    try:
        equipo = get_object_or_404(Equipo.objects.select_related('empresa'), pk=pk)

        # Verificar permisos
        if not request.user.is_superuser and request.user.empresa != equipo.empresa:
            messages.error(request, 'No tienes permiso para generar la hoja de vida de este equipo.')
            return redirect('core:informes')

        # Generar PDF usando helper optimizado
        pdf_content = _generate_equipment_hoja_vida_pdf_content(request, equipo)

        # Crear respuesta
        response = HttpResponse(pdf_content, content_type='application/pdf')
        filename_safe = sanitize_filename(f"hoja_vida_{equipo.codigo_interno}.pdf")
        response['Content-Disposition'] = f'attachment; filename="{filename_safe}"'

        logger.info(f"Hoja de vida PDF generada para equipo {equipo.codigo_interno}")
        return response

    except Exception as e:
        logger.error(f"Error generando hoja de vida PDF para equipo {pk}: {e}")
        messages.error(request, f'Error al generar el PDF: {e}. Revisa los logs para más detalles.')
        return redirect('core:detalle_equipo', pk=pk)


# =============================================================================
# TEMPLATE AND IMPORT FUNCTIONS
# =============================================================================

@monitor_view
@access_check
@login_required
@permission_required('core.add_equipo', raise_exception=True)
def descargar_plantilla_excel(request):
    """
    Genera y descarga una plantilla Excel mejorada para importación de equipos.
    Incluye validaciones, instrucciones y ejemplos.
    """
    try:
        # Generar plantilla usando helper
        excel_content = _generate_excel_template()

        # Crear respuesta
        response = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_equipos.xlsx"'

        logger.info("Plantilla Excel generada para importación de equipos")
        return response

    except Exception as e:
        logger.error(f"Error generando plantilla Excel: {e}")
        messages.error(request, f'Error al generar plantilla: {e}')
        return redirect('core:informes')


@monitor_view
@access_check
@login_required
@permission_required('core.add_equipo', raise_exception=True)
def preview_equipos_excel(request):
    """
    Previsualiza los equipos que se van a importar desde Excel.
    Valida datos antes de la importación real.
    """
    if request.method == 'POST':
        try:
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']

                # Procesar y validar archivo Excel
                preview_data, errors = _process_excel_preview(excel_file, request.user)

                if errors:
                    for error in errors:
                        messages.error(request, error)

                context = {
                    'preview_data': preview_data,
                    'errors': errors,
                    'titulo_pagina': 'Previsualización de Importación Excel'
                }
                return render(request, 'core/preview_equipos_excel.html', context)
            else:
                messages.error(request, 'Por favor selecciona un archivo Excel válido.')
        except Exception as e:
            logger.error(f"Error en preview Excel: {e}")
            messages.error(request, f'Error procesando archivo: {e}')

    return redirect('core:home')


@monitor_view
@access_check
@login_required
@trial_check
@permission_required('core.add_equipo', raise_exception=True)
def importar_equipos_excel(request):
    """
    Importa equipos desde archivo Excel después de previsualización.
    Realiza validaciones completas y transacciones seguras.
    """
    titulo_pagina = "Importar Equipos desde Excel"

    if request.method == 'POST':
        try:
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']

                # Procesar importación con transacciones
                result = _process_excel_import(excel_file, request.user)

                if result['success']:
                    # Crear mensaje detallado de importación
                    mensaje_partes = []
                    if result['created'] > 0:
                        mensaje_partes.append(f"{result['created']} equipos creados")
                    if result['updated'] > 0:
                        mensaje_partes.append(f"{result['updated']} equipos actualizados")

                    mensaje = f"Importación exitosa: {', '.join(mensaje_partes)}."
                    if len(result.get('errors', [])) > 0:
                        mensaje += f" ({len(result['errors'])} errores encontrados)"

                    messages.success(request, mensaje)

                    # Mostrar errores detallados al usuario
                    for error in result.get('errors', []):
                        messages.warning(request, error)

                    logger.info(f"Equipos procesados desde Excel: {result['imported']} ({result['created']} nuevos, {result['updated']} actualizados) por usuario {request.user.username}")
                    return redirect('core:informes')
                else:
                    for error in result['errors']:
                        messages.error(request, error)

            else:
                messages.error(request, 'Archivo Excel no válido.')

        except Exception as e:
            logger.error(f"Error importando Excel: {e}")
            messages.error(request, f'Error en importación: {e}')

        return render(request, 'core/importar_equipos.html', {
            'form': form,
            'titulo_pagina': titulo_pagina
        })

    else:
        # Caso GET: mostrar formulario para importar
        form = ExcelUploadForm()
        return render(request, 'core/importar_equipos.html', {
            'form': form,
            'titulo_pagina': titulo_pagina
        })


def _add_template_header(sheet):
    """
    Helper: Agrega encabezado profesional a plantilla Excel.

    Args:
        sheet: Hoja de Excel
    """
    from datetime import datetime

    # Título principal
    sheet.merge_cells('A1:Z3')
    title_cell = sheet['A1']
    title_cell.value = "PLANTILLA DE IMPORTACIÓN DE EQUIPOS - SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Información
    sheet.merge_cells('A4:Z4')
    info_cell = sheet['A4']
    info_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Complete SOLO las filas de datos (fila 8 en adelante)"
    info_cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    info_cell.alignment = Alignment(horizontal="center")


def _add_template_headers_row(sheet):
    """
    Helper: Agrega filas de headers (técnicos y legibles) a plantilla Excel.

    Args:
        sheet: Hoja de Excel
    """
    from openpyxl.utils import get_column_letter

    # Headers técnicos
    headers = [
        "codigo_interno", "nombre", "empresa_nombre", "tipo_equipo", "marca", "modelo",
        "numero_serie", "ubicacion_nombre", "responsable", "estado", "fecha_adquisicion", "proveedor",
        "fecha_ultima_calibracion", "fecha_ultimo_mantenimiento", "fecha_ultima_comprobacion",
        "rango_medida", "resolucion", "error_maximo_permisible", "puntos_calibracion", "observaciones",
        "version_formato", "fecha_version_formato", "codificacion_formato",
        "frecuencia_calibracion_meses", "frecuencia_mantenimiento_meses", "frecuencia_comprobacion_meses"
    ]

    # Headers legibles
    headers_legibles = [
        "Código Interno*", "Nombre del Equipo*", "Empresa*", "Tipo de Equipo*", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado*", "Fecha Adquisición", "Proveedor",
        "Última Calibración", "Último Mantenimiento", "Última Comprobación",
        "Rango de Medida", "Resolución", "Error Máx. Permisible", "Puntos de Calibración", "Observaciones",
        "Versión Formato", "Fecha Versión", "Codificación Formato",
        "Freq. Cal. (meses)", "Freq. Mant. (meses)", "Freq. Comp. (meses)"
    ]

    # Aplicar headers técnicos (fila 6)
    for i, header in enumerate(headers, 1):
        column_letter = get_column_letter(i)
        sheet[f'{column_letter}6'] = header

    # Aplicar headers legibles (fila 7)
    for i, header_legible in enumerate(headers_legibles, 1):
        column_letter = get_column_letter(i)
        cell = sheet[f'{column_letter}7']
        cell.value = header_legible
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_template_validations(sheet):
    """
    Helper: Agrega validaciones de datos a plantilla Excel.

    Args:
        sheet: Hoja de Excel
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from ..models import Equipo, Empresa

    # Obtener opciones para validaciones
    tipos_equipo = [choice[0] for choice in Equipo.TIPO_EQUIPO_CHOICES]
    estados_equipo = [choice[0] for choice in Equipo.ESTADO_CHOICES]
    empresas_disponibles = list(Empresa.objects.values_list('nombre', flat=True))

    start_row = 8
    end_row = 1000

    # Validación para empresa
    if empresas_disponibles:
        dv_empresa = DataValidation(
            type="list",
            formula1=f'"{",".join(empresas_disponibles)}"',
            showErrorMessage=True,
            errorTitle="Empresa Inválida",
            error="Seleccione una empresa de la lista disponible"
        )
        dv_empresa.add(f"C{start_row}:C{end_row}")
        sheet.add_data_validation(dv_empresa)

    # Validación para tipo de equipo
    dv_tipo = DataValidation(
        type="list",
        formula1=f'"{",".join(tipos_equipo)}"',
        showErrorMessage=True,
        errorTitle="Tipo Inválido",
        error="Seleccione un tipo de equipo de la lista"
    )
    dv_tipo.add(f"D{start_row}:D{end_row}")
    sheet.add_data_validation(dv_tipo)

    # Validación para estado
    dv_estado = DataValidation(
        type="list",
        formula1=f'"{",".join(estados_equipo)}"',
        showErrorMessage=True,
        errorTitle="Estado Inválido",
        error="Seleccione un estado de la lista"
    )
    dv_estado.add(f"J{start_row}:J{end_row}")
    sheet.add_data_validation(dv_estado)


def _add_template_example_row(sheet):
    """
    Helper: Agrega fila de ejemplo a plantilla Excel.

    Args:
        sheet: Hoja de Excel
    """
    from openpyxl.utils import get_column_letter
    from ..models import Equipo, Empresa

    # Obtener datos de ejemplo
    tipos_equipo = [choice[0] for choice in Equipo.TIPO_EQUIPO_CHOICES]
    estados_equipo = [choice[0] for choice in Equipo.ESTADO_CHOICES]
    empresas_disponibles = list(Empresa.objects.values_list('nombre', flat=True))

    # Fila de ejemplo
    ejemplo_data = [
        "EQ-001", "Balanza Analítica Ejemplo",
        empresas_disponibles[0] if empresas_disponibles else "Mi Empresa",
        tipos_equipo[0] if tipos_equipo else "Equipo de Medición",
        "Mettler Toledo", "XPE205", "B123456789", "Laboratorio Principal",
        "Técnico Responsable", estados_equipo[0] if estados_equipo else "Activo",
        "15/01/2023", "Multiples", "20/11/2024", "15/10/2024", "10/12/2024",
        "0-220g", "0.1mg", "±0.1mg", "Equipo nuevo para laboratorio",
        "V1.0", "01/01/2023", "CAL-001", "12", "6", "3"
    ]

    # Aplicar ejemplo en fila 8
    for i, value in enumerate(ejemplo_data, 1):
        column_letter = get_column_letter(i)
        cell = sheet[f'{column_letter}8']
        cell.value = value
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.font = Font(name="Arial", size=9, italic=True)


def _apply_template_formatting(sheet):
    """
    Helper: Aplica formato final a plantilla Excel (anchos, alturas, congelar paneles).

    Args:
        sheet: Hoja de Excel
    """
    from openpyxl.utils import get_column_letter

    # Ajustar anchos de columna
    anchos = [15, 25, 20, 18, 15, 15, 15, 20, 20, 15, 15, 18, 18, 18, 18, 15, 15, 15, 25, 15, 15, 15, 10, 10, 10]
    for i, ancho in enumerate(anchos, 1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = ancho

    # Ajustar alturas
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[7].height = 30

    # Congelar paneles
    sheet.freeze_panes = "A8"


def _generate_excel_template():
    """
    Genera plantilla Excel mejorada con validaciones, instrucciones y ejemplos.

    Refactorizada: 154 líneas → 30 líneas (usa 5 helpers)
    """
    import io

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plantilla Equipos"

    # Agregar header profesional
    _add_template_header(sheet)

    # Agregar headers (técnicos y legibles)
    _add_template_headers_row(sheet)

    # Agregar validaciones de datos
    _add_template_validations(sheet)

    # Agregar fila de ejemplo
    _add_template_example_row(sheet)

    # Aplicar formato final
    _apply_template_formatting(sheet)

    # Crear buffer y guardar
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()


# =============================================================================
# HELPER FUNCTIONS - ACTIVITIES
# =============================================================================

def _get_scheduled_activities(equipos_queryset, today):
    """
    Obtiene todas las actividades programadas para los equipos.
    Retorna lista unificada de calibraciones, mantenimientos y comprobaciones.
    """
    scheduled_activities = []

    # Filtrar solo equipos activos
    equipos_activos = equipos_queryset.exclude(estado__in=[ESTADO_DE_BAJA, ESTADO_INACTIVO])

    # Calibraciones
    for equipo in equipos_activos.filter(proxima_calibracion__isnull=False):
        if equipo.proxima_calibracion:
            days_remaining = (equipo.proxima_calibracion - today).days
            scheduled_activities.append({
                'tipo': 'Calibración',
                'equipo': equipo,
                'fecha_programada': equipo.proxima_calibracion,
                'dias_restantes': days_remaining,
                'estado_vencimiento': 'Vencida' if days_remaining < 0 else 'Próxima'
            })

    # Mantenimientos
    for equipo in equipos_activos.filter(proximo_mantenimiento__isnull=False):
        if equipo.proximo_mantenimiento:
            days_remaining = (equipo.proximo_mantenimiento - today).days
            scheduled_activities.append({
                'tipo': 'Mantenimiento',
                'equipo': equipo,
                'fecha_programada': equipo.proximo_mantenimiento,
                'dias_restantes': days_remaining,
                'estado_vencimiento': 'Vencida' if days_remaining < 0 else 'Próxima'
            })

    # Comprobaciones
    for equipo in equipos_activos.filter(proxima_comprobacion__isnull=False):
        if equipo.proxima_comprobacion:
            days_remaining = (equipo.proxima_comprobacion - today).days
            scheduled_activities.append({
                'tipo': 'Comprobación',
                'equipo': equipo,
                'fecha_programada': equipo.proxima_comprobacion,
                'dias_restantes': days_remaining,
                'estado_vencimiento': 'Vencida' if days_remaining < 0 else 'Próxima'
            })

    # Ordenar por fecha programada
    scheduled_activities.sort(
        key=lambda x: x['fecha_programada'] if x['fecha_programada'] else date.max
    )

    return scheduled_activities


def _categorize_activities(scheduled_activities):
    """
    Categoriza las actividades por tipo y urgencia.
    Retorna diccionario con todas las categorías.
    """
    categories = {
        'calibraciones_proximas_30': [],
        'calibraciones_proximas_15': [],
        'calibraciones_vencidas': [],
        'mantenimientos_proximos_30': [],
        'mantenimientos_proximos_15': [],
        'mantenimientos_vencidos': [],
        'comprobaciones_proximas_30': [],
        'comprobaciones_proximas_15': [],
        'comprobaciones_vencidas': []
    }

    for activity in scheduled_activities:
        tipo = activity['tipo'].lower()
        dias = activity['dias_restantes']
        estado = activity['estado_vencimiento']

        # Mapear tipos para asegurar plurales correctos (maneja tildes)
        tipo_plural_map = {
            'calibración': 'calibraciones',
            'mantenimiento': 'mantenimientos',
            'comprobación': 'comprobaciones'
        }

        tipo_plural = tipo_plural_map.get(tipo, f'{tipo}s')

        if estado == 'Vencida':
            # Manejar género correctamente para 'vencidos/vencidas'
            if tipo_plural == 'mantenimientos':
                key = f'{tipo_plural}_vencidos'
            else:
                key = f'{tipo_plural}_vencidas'
            if key in categories:
                categories[key].append(activity)
        elif 0 <= dias <= 15:
            # Para próximas actividades (próximos/próximas)
            if tipo_plural == 'mantenimientos':
                key = f'{tipo_plural}_proximos_15'
            else:
                key = f'{tipo_plural}_proximas_15'
            if key in categories:
                categories[key].append(activity)
        elif 16 <= dias <= 30:
            # Para actividades en 30 días (próximos/próximas)
            if tipo_plural == 'mantenimientos':
                key = f'{tipo_plural}_proximos_30'
            else:
                key = f'{tipo_plural}_proximas_30'
            if key in categories:
                categories[key].append(activity)

    return categories


# =============================================================================
# HELPER FUNCTIONS - ZIP GENERATION
# =============================================================================

def calcular_info_paginacion_zip(empresa_id, is_superuser=False):
    """
    Calcula información de paginación para ZIPs basado en número de equipos.
    Optimizado para memoria con máximo 100 equipos por ZIP.
    """
    EQUIPOS_POR_ZIP = 100  # Optimizado para memoria

    total_equipos = Equipo.objects.filter(empresa_id=empresa_id).count()
    total_partes = (total_equipos + EQUIPOS_POR_ZIP - 1) // EQUIPOS_POR_ZIP if total_equipos > 0 else 1

    return total_equipos, total_partes, EQUIPOS_POR_ZIP


def _get_zip_pagination_info(selected_company_id, is_superuser):
    """
    Obtiene información de paginación para la interfaz de ZIPs.
    """
    total_equipos, total_partes, equipos_por_zip = calcular_info_paginacion_zip(
        selected_company_id, is_superuser
    )

    # Crear información detallada de partes
    partes_info = []
    for parte_num in range(1, total_partes + 1):
        inicio_equipo = (parte_num - 1) * equipos_por_zip + 1
        fin_equipo = min(parte_num * equipos_por_zip, total_equipos)
        partes_info.append({
            'numero': parte_num,
            'inicio_equipo': inicio_equipo,
            'fin_equipo': fin_equipo
        })

    return {
        'total_equipos': total_equipos,
        'total_partes': total_partes,
        'equipos_por_zip': equipos_por_zip,
        'requiere_paginacion': total_partes > 1,
        'partes_info': partes_info
    }


def _generate_zip_report(empresa, parte, equipos_por_zip, total_partes):
    """
    Genera el reporte ZIP optimizado para memoria.
    Usa la función del zip_optimizer que ya está funcionando.
    """
    from ..zip_optimizer import generate_optimized_zip
    from django.http import FileResponse
    import os

    try:
        # Usar la función optimizada que ya funciona
        result = generate_optimized_zip(
            empresa=empresa,
            formatos_seleccionados=['hoja_vida', 'manual', 'procedimientos'],
            user=None  # No necesitamos usuario específico para esta función
        )

        if isinstance(result, dict) and result.get('success'):
            file_path = result.get('file_path')
            if file_path and os.path.exists(file_path):
                # Crear respuesta de archivo
                filename_safe = f"SAM_Equipos_{empresa.nombre}_Parte{parte}.zip"
                response = FileResponse(
                    open(file_path, 'rb'),
                    as_attachment=True,
                    filename=filename_safe,
                    content_type='application/zip'
                )
                return response

        # Si falla, crear ZIP simple
        return _generate_simple_zip_fallback(empresa, parte, equipos_por_zip)

    except Exception as e:
        logger.error(f"Error en _generate_zip_report: {e}")
        return _generate_simple_zip_fallback(empresa, parte, equipos_por_zip)


def _generate_simple_zip_fallback(empresa, parte, equipos_por_zip):
    """
    Función fallback para generar ZIP simple en caso de error.
    """
    from django.http import HttpResponse
    import zipfile
    import tempfile
    import os

    try:
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(suffix='.zip', delete=False) as temp_file:
            temp_path = temp_file.name

        # Crear ZIP básico con información de empresa
        with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Agregar información básica
            empresa_info = f"Empresa: {empresa.nombre}\nFecha: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            zip_file.writestr("Empresa/Informacion_Empresa.txt", empresa_info.encode('utf-8'))

        # Crear respuesta
        with open(temp_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="SAM_Equipos_{empresa.nombre}_Parte{parte}_Simple.zip"'

        # Limpiar archivo temporal
        os.unlink(temp_path)
        return response

    except Exception as e:
        logger.error(f"Error en ZIP fallback: {e}")
        from django.http import HttpResponseServerError
        return HttpResponseServerError("Error generando ZIP")


# =============================================================================
# HELPER FUNCTIONS - FILE STREAMING
# =============================================================================

def stream_file_to_zip(zip_file, file_path, zip_path):
    """
    Helper optimizado que hace streaming directo de archivo a ZIP sin cargar en memoria.
    Optimizado para archivos grandes y compatibilidad S3/Local.
    """
    try:
        if not default_storage.exists(file_path):
            logger.warning(f"Archivo no existe: {file_path}")
            return False

        # Detectar tipo de storage para optimización específica
        storage_name = default_storage.__class__.__name__

        # Optimizar chunk size basado en el storage
        if 'S3' in storage_name:
            # Para S3, chunks más grandes para reducir latencia de red
            chunk_size = 32768  # 32KB chunks para S3
        else:
            # Para almacenamiento local, chunks más pequeños
            chunk_size = 8192   # 8KB chunks para local

        with default_storage.open(file_path, 'rb') as f:
            with zip_file.open(zip_path, 'w') as zip_entry:
                # Buffer optimizado para reducir llamadas de I/O
                while True:
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    zip_entry.write(chunk)

        logger.debug(f"Archivo transferido exitosamente: {file_path} -> {zip_path}")
        return True

    except Exception as e:
        logger.warning(f"Error streaming archivo {file_path}: {e}")
        return False


# =============================================================================
# MIGRATED FUNCTIONS FROM MAIN views.py
# =============================================================================

def actualizar_equipo_selectivo(equipo_existente, nuevos_datos, row_index):
    """
    Actualiza solo los campos que NO están vacíos en los nuevos datos.

    Returns:
        list: Lista de nombres de campos que fueron actualizados
    """
    campos_actualizados = []

    # Mapeo de campos que se pueden actualizar (excluyendo empresa y codigo_interno que son identificadores)
    campos_actualizables = {
        'nombre': 'nombre',
        'tipo_equipo': 'tipo_equipo',
        'marca': 'marca',
        'modelo': 'modelo',
        'numero_serie': 'numero_serie',
        'ubicacion': 'ubicacion',
        'responsable': 'responsable',
        'estado': 'estado',
        'fecha_adquisicion': 'fecha_adquisicion',
        'fecha_ultima_calibracion': 'fecha_ultima_calibracion',
        'fecha_ultimo_mantenimiento': 'fecha_ultimo_mantenimiento',
        'fecha_ultima_comprobacion': 'fecha_ultima_comprobacion',
        'rango_medida': 'rango_medida',
        'resolucion': 'resolucion',
        'error_maximo_permisible': 'error_maximo_permisible',
        'observaciones': 'observaciones',
        'version_formato': 'version_formato',
        'fecha_version_formato': 'fecha_version_formato',
        'codificacion_formato': 'codificacion_formato',
        'frecuencia_calibracion_meses': 'frecuencia_calibracion_meses',
        'frecuencia_mantenimiento_meses': 'frecuencia_mantenimiento_meses',
        'frecuencia_comprobacion_meses': 'frecuencia_comprobacion_meses',
    }

    for campo_excel, campo_modelo in campos_actualizables.items():
        if campo_modelo in nuevos_datos:
            nuevo_valor = nuevos_datos[campo_modelo]

            # Solo actualizar si el nuevo valor no está vacío y es diferente al actual
            if es_valor_valido_para_actualizacion(nuevo_valor):
                valor_actual = getattr(equipo_existente, campo_modelo)

                # Comparar valores (manejar None y tipos diferentes)
                if valores_son_diferentes(valor_actual, nuevo_valor):
                    setattr(equipo_existente, campo_modelo, nuevo_valor)
                    campos_actualizados.append(campo_excel)

    # Guardar solo si hay cambios
    if campos_actualizados:
        equipo_existente.save()

    return campos_actualizados


def es_valor_valido_para_actualizacion(valor):
    """
    Determina si un valor es válido para actualizar (no está vacío).
    """
    if valor is None:
        return False
    if isinstance(valor, str) and valor.strip() == '':
        return False
    return True


def valores_son_diferentes(valor_actual, nuevo_valor):
    """
    Compara dos valores manejando diferentes tipos de datos.
    """
    # Si ambos son None, son iguales
    if valor_actual is None and nuevo_valor is None:
        return False

    # Si uno es None y el otro no, son diferentes
    if valor_actual is None or nuevo_valor is None:
        return True

    # Para strings, comparar sin espacios
    if isinstance(valor_actual, str) and isinstance(nuevo_valor, str):
        return valor_actual.strip() != nuevo_valor.strip()

    # Para otros tipos, comparación directa
    return valor_actual != nuevo_valor


def _generate_pdf_content(request, template_path, context):
    """
    Generates PDF content (bytes) from a template and context using WeasyPrint.
    """
    from django.template.loader import get_template
    import logging
    logger = logging.getLogger(__name__)

    try:
        template = get_template(template_path)
        html_string = template.render(context)

        # base_url es crucial para que WeasyPrint resuelva rutas de CSS e imágenes
        # Mejorar el manejo del base_url para mock requests
        try:
            base_url = request.build_absolute_uri('/')
        except Exception as e:
            logger.warning(f"Error building absolute URI: {e}, using fallback")
            # Fallback para mock requests
            if hasattr(request, 'META') and 'HTTP_HOST' in request.META:
                scheme = request.META.get('wsgi.url_scheme', 'https')
                host = request.META.get('HTTP_HOST', 'sam-9o6o.onrender.com')
                base_url = f"{scheme}://{host}/"
            else:
                base_url = "https://sam-9o6o.onrender.com/"

        logger.info(f"Generating PDF with base_url: {base_url}")
        pdf_file = HTML(string=html_string, base_url=base_url).write_pdf()

        return pdf_file

    except Exception as e:
        logger.error(f"Error generando PDF con template {template_path}: {e}")
        raise


def _get_pdf_file_url(request, file_field):
    """
    Obtiene URL absoluta para archivos en PDF.

    Args:
        request: Request de Django para construir URLs absolutas
        file_field: Campo de archivo del modelo

    Returns:
        str|None: URL absoluta del archivo o None si no existe
    """
    if not file_field or not file_field.name:
        return None

    try:
        storage_name = default_storage.__class__.__name__

        if 'S3' in storage_name or hasattr(default_storage, 'bucket'):
            # Para S3/R2, generar URL directamente sin verificar existencia
            # La verificación exists() puede fallar con Cloudflare R2
            # Si el campo tiene nombre, asumimos que el archivo existe
            try:
                url = default_storage.url(file_field.name, expire=7200)  # 2 horas
            except TypeError:
                # Fallback si el storage no soporta expire
                url = default_storage.url(file_field.name)

            # Asegurar que sea URL absoluta
            if url.startswith('//'):
                url = 'https:' + url

            return url
        else:
            # Para almacenamiento local (FileSystemStorage)
            url = default_storage.url(file_field.name)
            # Convertir a URL absoluta
            if url.startswith('/'):
                url = request.build_absolute_uri(url)

            return url
    except Exception as e:
        logger.error(f"Error obteniendo URL para PDF: {file_field.name}, error: {str(e)}")
        return None


def _get_pdf_image_data(file_field):
    """
    Obtiene datos de imagen para PDF, convirtiendo a base64 si es necesario.

    Args:
        file_field: Campo de archivo de imagen

    Returns:
        str|None: Data URL en base64 o None si falla
    """
    if not file_field or not file_field.name:
        return None

    try:
        # Verificar si es una imagen
        file_extension = file_field.name.lower().split('.')[-1]
        if file_extension not in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
            return None

        # Verificar tamaño del archivo antes de cargarlo (límite: 1MB para imágenes en PDF)
        try:
            if hasattr(default_storage, 'size'):
                file_size = default_storage.size(file_field.name)
                if file_size > 1024 * 1024:  # 1MB límite
                    logger.warning(f"Imagen muy grande para PDF ({file_size} bytes): {file_field.name}")
                    return None
        except:
            pass  # Si no se puede obtener el tamaño, continuar

        # Obtener contenido del archivo de forma eficiente
        try:
            with default_storage.open(file_field.name, 'rb') as f:
                file_content = f.read()

            # Convertir a base64
            import base64
            base64_encoded = base64.b64encode(file_content).decode('utf-8')

            # Limpiar file_content de memoria inmediatamente
            del file_content

            # Determinar el tipo MIME
            mime_type = f"image/{file_extension}" if file_extension != 'jpg' else 'image/jpeg'

            # Retornar como data URL
            return f"data:{mime_type};base64,{base64_encoded}"

        except Exception as e:
            logger.warning(f"No se pudo convertir imagen a base64: {file_field.name}, error: {str(e)}")
            return None

    except Exception as e:
        logger.error(f"Error procesando imagen para PDF: {file_field.name}, error: {str(e)}")
        return None


def _generate_hoja_vida_cache_key(equipo):
    """
    Helper: Genera clave de caché para hoja de vida de equipo.

    Args:
        equipo: Objeto Equipo

    Returns:
        str: Clave de caché única
    """
    import hashlib

    # Obtener datos de formato de la empresa para incluir en la clave de caché
    empresa = equipo.empresa
    empresa_formato_data = {
        'empresa_formato_codificacion': empresa.formato_codificacion_empresa or '',
        'empresa_formato_version': empresa.formato_version_empresa or '',
        'empresa_formato_fecha_display': empresa.formato_fecha_version_empresa_display or '',
        'empresa_formato_fecha': empresa.formato_fecha_version_empresa.isoformat() if empresa.formato_fecha_version_empresa else '',
    }

    cache_data = {
        'equipo_id': equipo.id,
        'equipo_updated': equipo.fecha_actualizacion.isoformat() if hasattr(equipo, 'fecha_actualizacion') else '',
        'calibraciones_count': equipo.calibraciones.count(),
        'mantenimientos_count': equipo.mantenimientos.count(),
        'comprobaciones_count': equipo.comprobaciones.count(),
        'estado': equipo.estado,
        'last_calibration': equipo.calibraciones.first().fecha_calibracion.isoformat() if equipo.calibraciones.exists() else '',
        'last_maintenance': equipo.mantenimientos.first().fecha_mantenimiento.isoformat() if equipo.mantenimientos.exists() else '',
        'codificacion_formato': equipo.codificacion_formato or '',
        'version_formato': equipo.version_formato or '',
        'fecha_version_formato': equipo.fecha_version_formato.isoformat() if equipo.fecha_version_formato else '',
        # Incluir datos de formato de la empresa para invalidar caché cuando cambien
        **empresa_formato_data,
    }

    cache_key_data = str(cache_data).encode('utf-8')
    cache_key = f"hoja_vida_{equipo.id}_{hashlib.md5(cache_key_data).hexdigest()}"

    return cache_key


def _get_hoja_vida_activities(equipo):
    """
    Helper: Obtiene actividades optimizadas para hoja de vida.

    Args:
        equipo: Objeto Equipo

    Returns:
        tuple: (calibraciones, mantenimientos, comprobaciones)
    """
    calibraciones = equipo.calibraciones.select_related('proveedor').order_by('-fecha_calibracion')
    mantenimientos = equipo.mantenimientos.select_related('proveedor').order_by('-fecha_mantenimiento')
    comprobaciones = equipo.comprobaciones.select_related('proveedor').order_by('-fecha_comprobacion')

    return calibraciones, mantenimientos, comprobaciones


def _generate_hoja_vida_charts(calibraciones, comprobaciones):
    """
    Helper: Genera gráficas históricas para hoja de vida.

    Args:
        calibraciones: QuerySet de calibraciones
        comprobaciones: QuerySet de comprobaciones

    Returns:
        tuple: (grafica_hist_confirmaciones, grafica_hist_comprobaciones)
    """
    from core.views.equipment import _generar_grafica_hist_confirmaciones, _generar_grafica_hist_comprobaciones

    # Obtener últimas 5 confirmaciones con datos JSON
    calibraciones_con_datos = []
    for cal in calibraciones[:5]:
        if (hasattr(cal, 'confirmacion_metrologica_datos') and
            cal.confirmacion_metrologica_datos and
            cal.confirmacion_metrologica_datos.get('puntos_medicion')):
            calibraciones_con_datos.append({
                'fecha': cal.fecha_calibracion,
                'puntos': cal.confirmacion_metrologica_datos['puntos_medicion']
            })

    # Obtener últimas 5 comprobaciones con datos JSON
    comprobaciones_con_datos = []
    for comp in comprobaciones[:5]:
        if (hasattr(comp, 'datos_comprobacion') and
            comp.datos_comprobacion and
            comp.datos_comprobacion.get('puntos_medicion')):
            comprobaciones_con_datos.append({
                'fecha': comp.fecha_comprobacion,
                'puntos': comp.datos_comprobacion['puntos_medicion']
            })

    # Generar gráficas SVG
    grafica_hist_confirmaciones = None
    if len(calibraciones_con_datos) > 0:
        grafica_hist_confirmaciones = _generar_grafica_hist_confirmaciones(calibraciones_con_datos)

    grafica_hist_comprobaciones = None
    if len(comprobaciones_con_datos) > 0:
        grafica_hist_comprobaciones = _generar_grafica_hist_comprobaciones(comprobaciones_con_datos)

    return grafica_hist_confirmaciones, grafica_hist_comprobaciones


def _get_hoja_vida_file_urls(request, equipo, calibraciones, mantenimientos, comprobaciones):
    """
    Helper: Obtiene URLs de archivos para hoja de vida.

    Args:
        request: Objeto request de Django
        equipo: Objeto Equipo
        calibraciones: QuerySet de calibraciones
        mantenimientos: QuerySet de mantenimientos
        comprobaciones: QuerySet de comprobaciones

    Returns:
        dict: Diccionario con todas las URLs de archivos
    """
    # Obtener registro de baja
    baja_registro = None
    try:
        baja_registro = equipo.baja_registro
    except BajaEquipo.DoesNotExist:
        pass

    # URLs de imágenes (convertidas a base64)
    logo_empresa_url = _get_pdf_image_data(equipo.empresa.logo_empresa) if equipo.empresa and equipo.empresa.logo_empresa else None
    imagen_equipo_url = _get_pdf_image_data(equipo.imagen_equipo) if equipo.imagen_equipo else None

    # URL de documento de baja (solo si está dado de baja)
    documento_baja_url = (_get_pdf_file_url(request, baja_registro.documento_baja)
                         if baja_registro and baja_registro.documento_baja and equipo.estado == ESTADO_DE_BAJA
                         else None)

    # URLs de documentos de calibraciones
    for cal in calibraciones:
        cal.documento_calibracion_url = _get_pdf_file_url(request, cal.documento_calibracion)
        cal.confirmacion_metrologica_pdf_url = _get_pdf_file_url(request, cal.confirmacion_metrologica_pdf)
        cal.intervalos_calibracion_pdf_url = _get_pdf_file_url(request, cal.intervalos_calibracion_pdf)

    # URLs de documentos de mantenimientos
    for mant in mantenimientos:
        mant.documento_mantenimiento_url = _get_pdf_file_url(request, mant.documento_mantenimiento)

    # URLs de documentos de comprobaciones
    for comp in comprobaciones:
        comp.documento_comprobacion_url = _get_pdf_file_url(request, comp.documento_comprobacion)

    # URLs de documentos del equipo
    archivo_compra_pdf_url = _get_pdf_file_url(request, equipo.archivo_compra_pdf)
    ficha_tecnica_pdf_url = _get_pdf_file_url(request, equipo.ficha_tecnica_pdf)
    manual_pdf_url = _get_pdf_file_url(request, equipo.manual_pdf)
    otros_documentos_pdf_url = _get_pdf_file_url(request, equipo.otros_documentos_pdf)

    return {
        'baja_registro': baja_registro,
        'logo_empresa_url': logo_empresa_url,
        'imagen_equipo_url': imagen_equipo_url,
        'documento_baja_url': documento_baja_url,
        'archivo_compra_pdf_url': archivo_compra_pdf_url,
        'ficha_tecnica_pdf_url': ficha_tecnica_pdf_url,
        'manual_pdf_url': manual_pdf_url,
        'otros_documentos_pdf_url': otros_documentos_pdf_url,
    }


def _build_hoja_vida_context(equipo, calibraciones, mantenimientos, comprobaciones, file_urls, charts):
    """
    Helper: Construye contexto para hoja de vida PDF.

    Args:
        equipo: Objeto Equipo
        calibraciones: QuerySet de calibraciones
        mantenimientos: QuerySet de mantenimientos
        comprobaciones: QuerySet de comprobaciones
        file_urls: Dict con URLs de archivos
        charts: Tuple con gráficas (confirmaciones, comprobaciones)

    Returns:
        dict: Contexto para el template
    """
    grafica_hist_confirmaciones, grafica_hist_comprobaciones = charts

    context = {
        'equipo': equipo,
        'calibraciones': calibraciones,
        'mantenimientos': mantenimientos,
        'comprobaciones': comprobaciones,
        'baja_registro': file_urls['baja_registro'] if equipo.estado == ESTADO_DE_BAJA else None,
        'logo_empresa_url': file_urls['logo_empresa_url'],
        'imagen_equipo_url': file_urls['imagen_equipo_url'],
        'documento_baja_url': file_urls['documento_baja_url'],
        'archivo_compra_pdf_url': file_urls['archivo_compra_pdf_url'],
        'ficha_tecnica_pdf_url': file_urls['ficha_tecnica_pdf_url'],
        'manual_pdf_url': file_urls['manual_pdf_url'],
        'otros_documentos_pdf_url': file_urls['otros_documentos_pdf_url'],
        'titulo_pagina': f'Hoja de Vida de {equipo.nombre}',
        'grafica_hist_confirmaciones': grafica_hist_confirmaciones,
        'grafica_hist_comprobaciones': grafica_hist_comprobaciones,
    }

    return context


def _generate_equipment_hoja_vida_pdf_content(request, equipo):
    """
    Genera el contenido PDF para la "Hoja de Vida" de un equipo con URLs seguras.
    Optimizado para uso eficiente de memoria con sistema de caché inteligente.

    Refactorizada: 147 líneas → 50 líneas (usa 5 helpers)
    """
    from django.core.cache import cache

    try:
        # Verificar caché
        cache_key = _generate_hoja_vida_cache_key(equipo)
        cached_pdf = cache.get(cache_key)
        if cached_pdf:
            logger.info(f"Hoja de vida obtenida del caché para equipo {equipo.codigo_interno}")
            return cached_pdf

        # Obtener actividades optimizadas
        calibraciones, mantenimientos, comprobaciones = _get_hoja_vida_activities(equipo)

        # Generar gráficas históricas
        charts = _generate_hoja_vida_charts(calibraciones, comprobaciones)

        # Obtener URLs de archivos
        file_urls = _get_hoja_vida_file_urls(request, equipo, calibraciones, mantenimientos, comprobaciones)

        # Construir contexto
        context = _build_hoja_vida_context(equipo, calibraciones, mantenimientos, comprobaciones, file_urls, charts)

        # Generar PDF
        pdf_content = _generate_pdf_content(request, 'core/hoja_vida_pdf.html', context)

        # Guardar en caché por 1 hora (3600 segundos)
        if pdf_content and len(pdf_content) > 1000:  # Verificar que sea un PDF válido
            cache.set(cache_key, pdf_content, timeout=3600)
            logger.info(f"Hoja de vida guardada en caché para equipo {equipo.codigo_interno}")

        return pdf_content

    except Exception as e:
        logger.error(f"Error generando contenido PDF para hoja de vida del equipo {equipo.codigo_interno}: {str(e)}")
        # Return minimal context in case of error
        context = {
            'equipo': equipo,
            'calibraciones': [],
            'mantenimientos': [],
            'comprobaciones': [],
            'baja_registro': None,
            'titulo_pagina': f'Hoja de Vida de {equipo.nombre}',
        }
        return _generate_pdf_content(request, 'core/hoja_vida_pdf.html', context)


def _add_excel_header(sheet, empresa):
    """
    Añade el encabezado profesional a una hoja Excel.

    Args:
        sheet: Hoja de Excel activa
        empresa: Instancia de Empresa

    Returns:
        int: Siguiente fila disponible después del encabezado
    """
    from datetime import datetime
    from openpyxl.styles import Alignment

    # Título principal profesional
    sheet.merge_cells('A1:F2')
    sheet['A1'] = 'INFORMES GENERADOS POR SAM METROLOGÍA SAS'
    sheet['A1'].font = Font(bold=True, size=20, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Información de la empresa y fecha
    sheet.merge_cells('A3:F3')
    sheet['A3'] = f'EMPRESA: {empresa.nombre.upper()}'
    sheet['A3'].font = Font(bold=True, size=14, color="1f4e79")
    sheet['A3'].alignment = Alignment(horizontal="center")

    sheet.merge_cells('A4:F4')
    hoy = datetime.now()
    sheet['A4'] = f'Generado el: {hoy.strftime("%d de %B de %Y a las %H:%M")}'
    sheet['A4'].font = Font(bold=True, size=12)
    sheet['A4'].alignment = Alignment(horizontal="center")

    # Ajustar altura de filas del encabezado
    sheet.row_dimensions[1].height = 40
    sheet.row_dimensions[2].height = 25

    return 6  # Siguiente fila después del encabezado


def _add_excel_resumen_section(sheet, equipos_list, row_start):
    """
    Añade la sección de resumen general a una hoja Excel.

    Args:
        sheet: Hoja de Excel activa
        equipos_list: Lista de equipos
        row_start: Fila donde empezar la sección

    Returns:
        int: Siguiente fila disponible después de la sección
    """
    from openpyxl.styles import Alignment

    total_equipos = len(equipos_list)
    equipos_activos = sum(1 for eq in equipos_list if eq.estado == ESTADO_ACTIVO)
    equipos_inactivos = sum(1 for eq in equipos_list if eq.estado == ESTADO_INACTIVO)
    equipos_baja = sum(1 for eq in equipos_list if eq.estado == 'De Baja')

    row = row_start
    sheet[f'A{row}'] = '📊 RESUMEN GENERAL DE EQUIPOS'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    sheet.merge_cells(f'A{row}:F{row}')
    row += 2

    # Tabla de resumen con formato profesional
    headers = ['Categoría', 'Cantidad', 'Porcentaje']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    resumen_data = [
        ('Total de Equipos', total_equipos, '100%'),
        ('Equipos Activos', equipos_activos, f'{(equipos_activos/total_equipos*100):.1f}%' if total_equipos > 0 else '0%'),
        ('Equipos Inactivos', equipos_inactivos, f'{(equipos_inactivos/total_equipos*100):.1f}%' if total_equipos > 0 else '0%'),
        ('Equipos de Baja', equipos_baja, f'{(equipos_baja/total_equipos*100):.1f}%' if total_equipos > 0 else '0%')
    ]

    for categoria, cantidad, porcentaje in resumen_data:
        sheet.cell(row=row, column=1, value=categoria)
        sheet.cell(row=row, column=2, value=cantidad)
        sheet.cell(row=row, column=3, value=porcentaje)
        row += 1

    return row


def _add_codigos_internos_section(sheet, equipos_list, row):
    """
    Helper: Agrega sección de códigos internos de equipos al Excel.

    Args:
        sheet: Hoja de Excel
        equipos_list: Lista de equipos
        row: Fila actual

    Returns:
        row: Nueva fila actual después de agregar sección
    """
    sheet[f'A{row}'] = '🔧 CÓDIGOS INTERNOS DE EQUIPOS'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    sheet.merge_cells(f'A{row}:F{row}')
    row += 2

    # Headers
    codigo_headers = ['Código Interno', 'Nombre del Equipo', 'Estado', 'Tipo']
    for col, header in enumerate(codigo_headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Datos (máximo 15 equipos)
    for equipo in equipos_list[:15]:
        sheet.cell(row=row, column=1, value=equipo.codigo_interno)
        sheet.cell(row=row, column=2, value=equipo.nombre)
        sheet.cell(row=row, column=3, value=equipo.estado)
        sheet.cell(row=row, column=4, value=equipo.get_tipo_equipo_display())
        row += 1

    # Mensaje si hay más equipos
    if len(equipos_list) > 15:
        sheet.cell(row=row, column=1, value=f"... y {len(equipos_list) - 15} equipos más")
        sheet[f'A{row}'].font = Font(italic=True, color="666666")
        row += 1

    return row


def _add_actividades_programadas_section(sheet, equipos_list, empresa, row):
    """
    Helper: Agrega sección de actividades programadas al Excel.

    Args:
        sheet: Hoja de Excel
        equipos_list: Lista de equipos
        empresa: Objeto Empresa
        row: Fila actual

    Returns:
        row: Nueva fila actual después de agregar sección
    """
    from datetime import datetime, timedelta
    from core.models import Calibracion, Mantenimiento, Comprobacion

    sheet[f'A{row}'] = '📅 ACTIVIDADES PROGRAMADAS Y REALIZADAS'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    sheet.merge_cells(f'A{row}:F{row}')
    row += 2

    # Preparar fechas
    hoy = datetime.now()
    hoy_date = hoy.date()
    treinta_dias = hoy_date + timedelta(days=30)
    ano_actual = hoy_date.year

    # Calcular estadísticas de calibraciones
    cal_programadas = Calibracion.objects.filter(
        equipo__empresa=empresa,
        fecha_calibracion__year=ano_actual
    ).count()
    cal_vencidas = sum(1 for eq in equipos_list
                      if eq.proxima_calibracion and eq.proxima_calibracion < hoy_date)
    cal_proximas = sum(1 for eq in equipos_list
                      if eq.proxima_calibracion and hoy_date <= eq.proxima_calibracion <= treinta_dias)

    # Calcular estadísticas de mantenimientos
    mant_programados = Mantenimiento.objects.filter(
        equipo__empresa=empresa,
        fecha_mantenimiento__year=ano_actual
    ).count()
    mant_vencidos = sum(1 for eq in equipos_list
                       if eq.proximo_mantenimiento and eq.proximo_mantenimiento < hoy_date)
    mant_proximos = sum(1 for eq in equipos_list
                       if eq.proximo_mantenimiento and hoy_date <= eq.proximo_mantenimiento <= treinta_dias)

    # Calcular estadísticas de comprobaciones
    comp_programadas = Comprobacion.objects.filter(
        equipo__empresa=empresa,
        fecha_comprobacion__year=ano_actual
    ).count()
    comp_vencidas = sum(1 for eq in equipos_list
                       if eq.proxima_comprobacion and eq.proxima_comprobacion < hoy_date)
    comp_proximas = sum(1 for eq in equipos_list
                       if eq.proxima_comprobacion and hoy_date <= eq.proxima_comprobacion <= treinta_dias)

    # Headers
    act_headers = ['Tipo de Actividad', f'Realizadas {ano_actual}', 'Vencidas', 'Próximas (30 días)']
    for col, header in enumerate(act_headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Datos
    actividades_data = [
        ('Calibraciones', cal_programadas, cal_vencidas, cal_proximas),
        ('Mantenimientos', mant_programados, mant_vencidos, mant_proximos),
        ('Comprobaciones', comp_programadas, comp_vencidas, comp_proximas)
    ]

    for tipo, realizadas, vencidas, proximas in actividades_data:
        sheet.cell(row=row, column=1, value=tipo)
        sheet.cell(row=row, column=2, value=realizadas)
        sheet.cell(row=row, column=3, value=vencidas)
        sheet.cell(row=row, column=4, value=proximas)
        row += 1

    return row


def _add_actividades_correctivas_section(sheet, empresa, row):
    """
    Helper: Agrega sección de actividades correctivas al Excel.

    Args:
        sheet: Hoja de Excel
        empresa: Objeto Empresa
        row: Fila actual

    Returns:
        row: Nueva fila actual después de agregar sección
    """
    from datetime import datetime
    from core.models import Mantenimiento

    sheet[f'A{row}'] = '⚠️ ACTIVIDADES NO PROGRAMADAS (CORRECTIVAS)'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    sheet.merge_cells(f'A{row}:F{row}')
    row += 2

    # Obtener mantenimientos correctivos
    ano_actual = datetime.now().date().year
    mantenimientos_correctivos = Mantenimiento.objects.filter(
        equipo__empresa=empresa,
        tipo_mantenimiento='Correctivo',
        fecha_mantenimiento__year=ano_actual
    )

    if mantenimientos_correctivos.exists():
        # Headers
        correctivos_headers = ['Equipo', 'Fecha', 'Descripción', 'Responsable']
        for col, header in enumerate(correctivos_headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Datos (máximo 10 registros)
        for mant in mantenimientos_correctivos[:10]:
            sheet.cell(row=row, column=1, value=mant.equipo.codigo_interno)
            sheet.cell(row=row, column=2, value=mant.fecha_mantenimiento.strftime("%d/%m/%Y"))
            sheet.cell(row=row, column=3, value=mant.descripcion or "Mantenimiento correctivo")
            sheet.cell(row=row, column=4, value=mant.responsable or "No especificado")
            row += 1
    else:
        sheet.cell(row=row, column=1, value="✅ No se registraron actividades correctivas este año")
        sheet[f'A{row}'].font = Font(bold=True, color="27AE60")
        row += 1

    return row


def _generate_dashboard_excel_content(equipos_queryset, empresa):
    """
    Genera un Excel profesional para el dashboard con formato mejorado y secciones completas.
    Usado cuando se descarga desde el botón del dashboard.

    Refactorizada: 171 líneas → 43 líneas (usa 3 helpers)
    """
    import io

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Dashboard"

    # Añadir encabezado profesional
    row = _add_excel_header(sheet, empresa)

    # Preparar datos
    equipos_list = list(equipos_queryset)

    # Añadir sección de resumen general
    row = _add_excel_resumen_section(sheet, equipos_list, row)

    # Añadir sección de códigos internos
    row += 2
    row = _add_codigos_internos_section(sheet, equipos_list, row)

    # Añadir sección de actividades programadas
    row += 2
    row = _add_actividades_programadas_section(sheet, equipos_list, empresa, row)

    # Añadir sección de actividades correctivas
    row += 2
    row = _add_actividades_correctivas_section(sheet, empresa, row)

    # Ajustar anchos de columnas
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        sheet.column_dimensions[col].width = 20

    # Ajustar altura de filas del encabezado
    sheet.row_dimensions[1].height = 40
    sheet.row_dimensions[2].height = 25

    # Retornar bytes del Excel
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _add_professional_sheet_header(sheet, title_text, merge_range, generation_date):
    """
    Helper: Agrega header profesional a una hoja Excel.

    Args:
        sheet: Hoja de Excel
        title_text: Texto del título
        merge_range: Rango de celdas a fusionar (ej: 'A1:AB2')
        generation_date: Fecha de generación
    """
    from django.utils import timezone
    from openpyxl.styles import Alignment

    # Título profesional
    sheet.merge_cells(merge_range)
    title_cell = sheet[merge_range.split(':')[0]]
    title_cell.value = title_text
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Información de generación
    sheet['A3'] = f"Generado el: {generation_date.strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal="left")


def _add_sheet_headers(sheet, headers, row_num=5):
    """
    Helper: Agrega headers de columna a una hoja Excel.

    Args:
        sheet: Hoja de Excel
        headers: Lista de textos para headers
        row_num: Número de fila donde agregar headers (default: 5)
    """
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_equipos_sheet(workbook, equipos_queryset, generation_date):
    """
    Helper: Crea y llena la hoja de Equipos.

    Args:
        workbook: Workbook de Excel
        equipos_queryset: QuerySet de equipos
        generation_date: Fecha de generación

    Returns:
        sheet: Hoja de equipos creada
    """
    sheet = workbook.active
    sheet.title = "Equipos"

    # Header profesional
    _add_professional_sheet_header(
        sheet,
        "INFORMES GENERADOS POR SAM METROLOGÍA SAS",
        'A1:AB2',
        generation_date
    )

    # Headers de columnas
    headers = [
        "Código Interno", "Nombre", "Empresa", "Tipo de Equipo", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado", "Fecha de Adquisición",
        "Rango de Medida", "Resolución", "Error Máximo Permisible", "Puntos de Calibración", "Fecha de Registro",
        "Observaciones", "Fecha Última Calibración", "Próxima Calibración",
        "Frecuencia Calibración (meses)", "Fecha Último Mantenimiento", "Próximo Mantenimiento",
        "Frecuencia Mantenimiento (meses)", "Fecha Última Comprobación",
        "Próxima Comprobación", "Frecuencia Comprobación (meses)"
    ]
    _add_sheet_headers(sheet, headers, row_num=5)

    # Datos de equipos
    current_row = 6
    for equipo in equipos_queryset:
        row_data = [
            equipo.codigo_interno, equipo.nombre, equipo.empresa.nombre, equipo.tipo_equipo,
            equipo.marca, equipo.modelo, equipo.numero_serie,
            equipo.ubicacion or "", equipo.responsable, equipo.estado,
            equipo.fecha_adquisicion, equipo.rango_medida, equipo.resolucion,
            equipo.error_maximo_permisible, equipo.puntos_calibracion,
            equipo.fecha_registro.replace(tzinfo=None) if equipo.fecha_registro else None,
            equipo.observaciones,
            equipo.fecha_ultima_calibracion, equipo.proxima_calibracion, equipo.frecuencia_calibracion_meses,
            equipo.fecha_ultimo_mantenimiento, equipo.proximo_mantenimiento, equipo.frecuencia_mantenimiento_meses,
            equipo.fecha_ultima_comprobacion, equipo.proxima_comprobacion, equipo.frecuencia_comprobacion_meses
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    return sheet


def _add_proveedores_sheet(workbook, proveedores_queryset, generation_date):
    """
    Helper: Crea y llena la hoja de Proveedores.

    Args:
        workbook: Workbook de Excel
        proveedores_queryset: QuerySet de proveedores
        generation_date: Fecha de generación

    Returns:
        sheet: Hoja de proveedores creada
    """
    sheet = workbook.create_sheet(title="Proveedores")

    # Header profesional
    _add_professional_sheet_header(
        sheet,
        "INFORMES GENERADOS POR SAM METROLOGÍA SAS",
        'A1:H2',
        generation_date
    )

    # Headers de columnas
    headers = [
        "Nombre Empresa", "Nombre Contacto", "Correo Electrónico", "Número Contacto",
        "Tipo Servicio", "Alcance", "Servicio Prestado", "Página Web"
    ]
    _add_sheet_headers(sheet, headers, row_num=5)

    # Datos de proveedores
    current_row = 6
    for proveedor in proveedores_queryset:
        row_data = [
            proveedor.nombre_empresa, proveedor.nombre_contacto, proveedor.correo_electronico,
            proveedor.numero_contacto, proveedor.tipo_servicio, proveedor.alcance,
            proveedor.servicio_prestado, proveedor.pagina_web
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    return sheet


def _add_procedimientos_sheet(workbook, procedimientos_queryset, generation_date):
    """
    Helper: Crea y llena la hoja de Procedimientos.

    Args:
        workbook: Workbook de Excel
        procedimientos_queryset: QuerySet de procedimientos
        generation_date: Fecha de generación

    Returns:
        sheet: Hoja de procedimientos creada
    """
    sheet = workbook.create_sheet(title="Procedimientos")

    # Header profesional
    _add_professional_sheet_header(
        sheet,
        "INFORMES GENERADOS POR SAM METROLOGÍA SAS",
        'A1:E2',
        generation_date
    )

    # Headers de columnas
    headers = [
        "Código", "Nombre", "Observaciones", "Versión", "Fecha de Emisión"
    ]
    _add_sheet_headers(sheet, headers, row_num=5)

    # Datos de procedimientos
    current_row = 6
    for procedimiento in procedimientos_queryset:
        row_data = [
            procedimiento.codigo, procedimiento.nombre, procedimiento.observaciones,
            procedimiento.version, procedimiento.fecha_emision
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    return sheet


def _add_dashboard_sheet(workbook, equipos_queryset, generation_date):
    """
    Helper: Crea y llena la hoja de Dashboard con estadísticas.

    Args:
        workbook: Workbook de Excel
        equipos_queryset: QuerySet de equipos
        generation_date: Fecha de generación

    Returns:
        sheet: Hoja de dashboard creada
    """
    from datetime import date

    sheet = workbook.create_sheet(title="Dashboard")

    # Header profesional
    _add_professional_sheet_header(
        sheet,
        "INFORMES GENERADOS POR SAM METROLOGÍA SAS",
        'A1:F2',
        generation_date
    )

    # Configurar estilos
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    row = 5

    # Estadísticas generales
    sheet.merge_cells(f'A{row}:F{row}')
    cell = sheet[f'A{row}']
    cell.value = "📊 ESTADÍSTICAS GENERALES"
    cell.font = title_font
    cell.fill = title_fill
    row += 2

    # Estadísticas de equipos por estado
    stats = {}
    for equipo in equipos_queryset:
        estado = equipo.estado
        stats[estado] = stats.get(estado, 0) + 1

    sheet[f'A{row}'] = "Estado"
    sheet[f'B{row}'] = "Cantidad"
    for cell in [sheet[f'A{row}'], sheet[f'B{row}']]:
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    for estado, cantidad in stats.items():
        sheet[f'A{row}'] = estado
        sheet[f'B{row}'] = cantidad
        row += 1

    return sheet


def _add_prestamos_sheet(workbook, prestamos_queryset, generation_date):
    """
    Helper: Crea y llena la hoja de Préstamos de Equipos.

    Args:
        workbook: Workbook de Excel
        prestamos_queryset: QuerySet de préstamos
        generation_date: Fecha de generación

    Returns:
        sheet: Hoja de préstamos creada
    """
    sheet = workbook.create_sheet(title="Préstamos")

    # Header profesional
    _add_professional_sheet_header(
        sheet,
        "INFORMES GENERADOS POR SAM METROLOGÍA SAS",
        'A1:L2',
        generation_date
    )

    # Headers de columnas
    headers = [
        "Código Equipo", "Nombre Equipo", "Prestatario", "Cédula", "Cargo",
        "Teléfono", "Email", "Fecha Préstamo", "Devolución Programada",
        "Devolución Real", "Estado", "Observaciones"
    ]
    _add_sheet_headers(sheet, headers, row_num=5)

    # Datos de préstamos
    current_row = 6
    for prestamo in prestamos_queryset:
        # Formatear fechas
        fecha_prestamo = prestamo.fecha_prestamo.strftime('%Y-%m-%d %H:%M') if prestamo.fecha_prestamo else ''
        fecha_dev_prog = prestamo.fecha_devolucion_programada.strftime('%Y-%m-%d') if prestamo.fecha_devolucion_programada else ''
        fecha_dev_real = prestamo.fecha_devolucion_real.strftime('%Y-%m-%d %H:%M') if prestamo.fecha_devolucion_real else ''

        # Determinar estado en español
        estado_map = {
            'activo': 'Activo',
            'devuelto': 'Devuelto',
            'vencido': 'Vencido'
        }
        estado = estado_map.get(prestamo.estado_prestamo, prestamo.estado_prestamo)

        row_data = [
            prestamo.equipo.codigo_interno,
            prestamo.equipo.nombre,
            prestamo.nombre_prestatario,
            prestamo.cedula_prestatario or '',
            prestamo.cargo_prestatario or '',
            prestamo.telefono_prestatario or '',
            prestamo.email_prestatario or '',
            fecha_prestamo,
            fecha_dev_prog,
            fecha_dev_real,
            estado,
            prestamo.observaciones_prestamo or ''
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Ajustar anchos de columna
    column_widths = [15, 25, 25, 15, 20, 15, 30, 20, 18, 18, 12, 40]
    for col_num, width in enumerate(column_widths, 1):
        from openpyxl.utils import get_column_letter
        sheet.column_dimensions[get_column_letter(col_num)].width = width

    return sheet


def _generate_consolidated_excel_content(equipos_queryset, proveedores_queryset, procedimientos_queryset, prestamos_queryset=None):
    """
    Genera un Excel consolidado con 5 hojas: Equipos, Proveedores, Procedimientos, Préstamos y Reporte Dashboard.
    Usado para el ZIP completo con todas las funcionalidades.

    Refactorizada: 201 líneas → 25 líneas (usa 7 helpers)
    """
    from django.utils import timezone
    import io

    workbook = Workbook()
    generation_date = timezone.now()

    # Crear las 5 hojas usando helpers
    _add_equipos_sheet(workbook, equipos_queryset, generation_date)
    _add_proveedores_sheet(workbook, proveedores_queryset, generation_date)
    _add_procedimientos_sheet(workbook, procedimientos_queryset, generation_date)

    # Agregar hoja de préstamos si se proporciona el queryset
    if prestamos_queryset is not None:
        _add_prestamos_sheet(workbook, prestamos_queryset, generation_date)

    _add_dashboard_sheet(workbook, equipos_queryset, generation_date)

    # Retornar bytes del Excel
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_general_equipment_list_excel_content(equipos_queryset):
    """
    Generates an Excel file with the general list of equipment including visual charts.
    """
    from openpyxl.chart.series import DataPoint
    from collections import Counter
    from openpyxl.styles import Alignment, Border, Side
    from datetime import datetime
    import io

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listado de Equipos"

    # ==============================================
    # ENCABEZADO PROFESIONAL SAM METROLOGÍA
    # ==============================================

    # Título principal profesional
    sheet.merge_cells('A1:AB2')
    sheet['A1'] = 'INFORMES GENERADOS POR SAM METROLOGÍA SAS'
    sheet['A1'].font = Font(bold=True, size=20, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Información de generación
    hoy = datetime.now()
    sheet.merge_cells('A3:AB3')
    sheet['A3'] = f'LISTADO GENERAL DE EQUIPOS - Generado el: {hoy.strftime("%d de %B de %Y a las %H:%M")}'
    sheet['A3'].font = Font(bold=True, size=12, color="1f4e79")
    sheet['A3'].alignment = Alignment(horizontal="center")

    # INFORMACIÓN DEL FORMATO (común para todos los equipos) - Fila 4
    primer_equipo = equipos_queryset.first()
    if primer_equipo:
        sheet.merge_cells('A4:AB4')
        formato_info = f'FORMATO: Código: {primer_equipo.codificacion_formato or "N/A"} | Versión: {primer_equipo.version_formato or "N/A"}'
        if primer_equipo.fecha_version_formato:
            formato_info += f' | Fecha Versión: {primer_equipo.fecha_version_formato.strftime("%Y-%m-%d")}'
        sheet['A4'] = formato_info
        sheet['A4'].font = Font(bold=True, size=10, color="FFFFFF")
        sheet['A4'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        sheet['A4'].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Código Interno", "Nombre", "Empresa", "Tipo de Equipo", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado", "Fecha de Adquisición",
        "Rango de Medida", "Resolución", "Error Máximo Permisible", "Puntos de Calibración",
        "Observaciones", "Último Certificado Calibración", "Fecha Última Calibración", "Próxima Calibración",
        "Frecuencia Calibración (meses)", "Fecha Último Mantenimiento", "Próximo Mantenimiento",
        "Frecuencia Mantenimiento (meses)", "Fecha Última Comprobación",
        "Próxima Comprobación", "Frecuencia Comprobación (meses)"
    ]

    # Colocar headers en la fila 6 (después del título y espacios)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        sheet.column_dimensions[cell.column_letter].width = 25

    # Agregar datos de equipos empezando desde la fila 7
    current_row = 7
    for equipo in equipos_queryset:
        # Obtener el último certificado de calibración
        ultima_calibracion = equipo.calibraciones.order_by('-fecha_calibracion').first()
        ultimo_certificado = ultima_calibracion.numero_certificado if ultima_calibracion else 'N/A'

        row_data = [
            equipo.codigo_interno,
            equipo.nombre,
            equipo.empresa.nombre if equipo.empresa else "N/A",
            equipo.get_tipo_equipo_display(),
            equipo.marca,
            equipo.modelo,
            equipo.numero_serie,
            equipo.ubicacion,
            equipo.responsable,
            equipo.estado,
            equipo.fecha_adquisicion.strftime('%Y-%m-%d') if equipo.fecha_adquisicion else '',
            equipo.rango_medida,
            equipo.resolucion,
            equipo.error_maximo_permisible if equipo.error_maximo_permisible is not None else '',
            equipo.puntos_calibracion if equipo.puntos_calibracion is not None else '',
            equipo.observaciones,
            ultimo_certificado,  # NUEVA COLUMNA
            equipo.fecha_ultima_calibracion.strftime('%Y-%m-%d') if equipo.fecha_ultima_calibracion else '',
            equipo.proxima_calibracion.strftime('%Y-%m-%d') if equipo.proxima_calibracion else '',
            float(equipo.frecuencia_calibracion_meses) if equipo.frecuencia_calibracion_meses is not None else '',
            equipo.fecha_ultimo_mantenimiento.strftime('%Y-%m-%d') if equipo.fecha_ultimo_mantenimiento else '',
            equipo.proximo_mantenimiento.strftime('%Y-%m-%d') if equipo.proximo_mantenimiento is not None else '',
            float(equipo.frecuencia_mantenimiento_meses) if equipo.frecuencia_mantenimiento_meses is not None else '',
            equipo.fecha_ultima_comprobacion.strftime('%Y-%m-%d') if equipo.fecha_ultima_comprobacion else '',
            equipo.proxima_comprobacion.strftime('%Y-%m-%d') if equipo.proxima_comprobacion is not None else '',
            float(equipo.frecuencia_comprobacion_meses) if equipo.frecuencia_comprobacion_meses is not None else '',
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Ajustar anchos de columna
    from openpyxl.utils import get_column_letter
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    # Crear hoja de estadísticas simple
    stats_sheet = workbook.create_sheet("Estadísticas")
    equipos_list = list(equipos_queryset)

    if equipos_list:
        # Estadísticas por empresa
        empresas_count = Counter(eq.empresa.nombre if eq.empresa else "Sin empresa" for eq in equipos_list)

        stats_sheet['A1'] = 'DISTRIBUCIÓN POR EMPRESA'
        stats_sheet['A2'] = 'Empresa'
        stats_sheet['B2'] = 'Cantidad de Equipos'

        row = 3
        for empresa, count in empresas_count.items():
            stats_sheet[f'A{row}'] = empresa
            stats_sheet[f'B{row}'] = count
            row += 1

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_equipment_general_info_excel_content(equipo):
    """
    Generates an Excel file with general information of a specific equipment.
    This is similar to _generate_general_equipment_list_excel_content but for a single equipment.
    """
    from openpyxl.styles import Alignment
    from django.utils import timezone
    import io

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Información General"

    # Add professional title header
    sheet.merge_cells('A1:AB2')
    title_cell = sheet['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet.row_dimensions[row].height = 8

    headers = [
        "Código Interno", "Nombre", "Empresa", "Tipo de Equipo", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado", "Fecha de Adquisición",
        "Rango de Medida", "Resolución", "Error Máximo Permisible", "Puntos de Calibración", "Fecha de Registro",
        "Observaciones", "Versión Formato Equipo", "Fecha Versión Formato Equipo",
        "Codificación Formato Equipo", "Fecha Última Calibración", "Próxima Calibración",
        "Frecuencia Calibración (meses)", "Fecha Último Mantenimiento", "Próximo Mantenimiento",
        "Frecuencia Mantenimiento (meses)", "Fecha Última Comprobación",
        "Próxima Comprobación", "Frecuencia Comprobación (meses)"
    ]

    # Add headers starting from row 6
    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add single equipment data starting from row 7
    row_data = [
        equipo.codigo_interno,
        equipo.nombre,
        equipo.empresa.nombre if equipo.empresa else "N/A",
        equipo.get_tipo_equipo_display(),
        equipo.marca,
        equipo.modelo,
        equipo.numero_serie,
        equipo.ubicacion,
        equipo.responsable,
        equipo.estado,
        equipo.fecha_adquisicion.strftime('%Y-%m-%d') if equipo.fecha_adquisicion else '',
        equipo.rango_medida,
        equipo.resolucion,
        equipo.error_maximo_permisible if equipo.error_maximo_permisible is not None else '',
        equipo.fecha_registro.strftime('%Y-%m-%d %H:%M:%S') if equipo.fecha_registro else '',
        equipo.observaciones,
        equipo.version_formato,
        equipo.fecha_version_formato.strftime('%Y-%m-%d') if equipo.fecha_version_formato else '',
        equipo.codificacion_formato,
        equipo.fecha_ultima_calibracion.strftime('%Y-%m-%d') if equipo.fecha_ultima_calibracion else '',
        equipo.proxima_calibracion.strftime('%Y-%m-%d') if equipo.proxima_calibracion else '',
        float(equipo.frecuencia_calibracion_meses) if equipo.frecuencia_calibracion_meses is not None else '',
        equipo.fecha_ultimo_mantenimiento.strftime('%Y-%m-%d') if equipo.fecha_ultimo_mantenimiento else '',
        equipo.proximo_mantenimiento.strftime('%Y-%m-%d') if equipo.proximo_mantenimiento is not None else '',
        float(equipo.frecuencia_mantenimiento_meses) if equipo.frecuencia_mantenimiento_meses is not None else '',
        equipo.fecha_ultima_comprobacion.strftime('%Y-%m-%d') if equipo.fecha_ultima_comprobacion else '',
        equipo.proxima_comprobacion.strftime('%Y-%m-%d') if equipo.proxima_comprobacion is not None else '',
        float(equipo.frecuencia_comprobacion_meses) if equipo.frecuencia_comprobacion_meses is not None else '',
    ]

    for col_num, value in enumerate(row_data, 1):
        sheet.cell(row=7, column=col_num, value=value)

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_equipment_activities_excel_content(equipo):
    """
    Generates an Excel file with the activities (calibrations, maintenances, verifications) of a specific equipment.
    """
    from openpyxl.styles import Alignment
    from django.utils import timezone
    import io

    workbook = Workbook()

    sheet_cal = workbook.active
    sheet_cal.title = "Calibraciones"

    # Add professional title header to Calibraciones sheet
    sheet_cal.merge_cells('A1:E2')
    title_cell = sheet_cal['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet_cal['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_cal['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet_cal['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet_cal.row_dimensions[row].height = 8

    headers_cal = ["Fecha Calibración", "Proveedor", "Resultado", "Número Certificado", "Observaciones"]

    # Add headers starting from row 6
    for col_num, header_text in enumerate(headers_cal, 1):
        cell = sheet_cal.cell(row=6, column=col_num, value=header_text)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add calibration data starting from row 7
    current_row = 7
    calibraciones = equipo.calibraciones.all().order_by('-fecha_calibracion')
    for calibracion in calibraciones:
        row_data = [
            calibracion.fecha_calibracion.strftime('%Y-%m-%d') if calibracion.fecha_calibracion else '',
            calibracion.proveedor.nombre_empresa if calibracion.proveedor else '',
            calibracion.resultado,
            calibracion.numero_certificado_calibracion,
            calibracion.observaciones
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_cal.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Adjust column widths for calibraciones sheet
    for col_num in range(1, len(headers_cal) + 1):
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(col_num)
        sheet_cal.column_dimensions[column_letter].width = 25

    # Create Mantenimientos sheet
    sheet_mant = workbook.create_sheet(title="Mantenimientos")

    # Add professional title header to Mantenimientos sheet
    sheet_mant.merge_cells('A1:E2')
    title_cell = sheet_mant['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet_mant['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_mant['A3'].font = Font(name="Arial", size=10, italic=True)

    headers_mant = ["Fecha Mantenimiento", "Tipo", "Proveedor", "Descripción", "Observaciones"]

    # Add headers starting from row 6
    for col_num, header_text in enumerate(headers_mant, 1):
        cell = sheet_mant.cell(row=6, column=col_num, value=header_text)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add maintenance data
    current_row = 7
    mantenimientos = equipo.mantenimientos.all().order_by('-fecha_mantenimiento')
    for mantenimiento in mantenimientos:
        row_data = [
            mantenimiento.fecha_mantenimiento.strftime('%Y-%m-%d') if mantenimiento.fecha_mantenimiento else '',
            mantenimiento.tipo_mantenimiento,
            mantenimiento.proveedor.nombre_empresa if mantenimiento.proveedor else '',
            mantenimiento.descripcion,
            mantenimiento.observaciones
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_mant.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Create Comprobaciones sheet
    sheet_comp = workbook.create_sheet(title="Comprobaciones")

    # Add professional title header to Comprobaciones sheet
    sheet_comp.merge_cells('A1:E2')
    title_cell = sheet_comp['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet_comp['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_comp['A3'].font = Font(name="Arial", size=10, italic=True)

    headers_comp = ["Fecha Comprobación", "Proveedor", "Resultado", "Número Certificado", "Observaciones"]

    # Add headers starting from row 6
    for col_num, header_text in enumerate(headers_comp, 1):
        cell = sheet_comp.cell(row=6, column=col_num, value=header_text)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add verification data
    current_row = 7
    comprobaciones = equipo.comprobaciones.all().order_by('-fecha_comprobacion')
    for comprobacion in comprobaciones:
        row_data = [
            comprobacion.fecha_comprobacion.strftime('%Y-%m-%d') if comprobacion.fecha_comprobacion else '',
            comprobacion.proveedor.nombre_empresa if comprobacion.proveedor else '',
            comprobacion.resultado,
            comprobacion.numero_certificado_comprobacion,
            comprobacion.observaciones
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_comp.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


# =============================================================================
# LOCAL HELPER FUNCTIONS FOR EXCEL GENERATION
# =============================================================================

def _generate_dashboard_excel_content_local(equipos_queryset, empresa):
    """
    Local version of dashboard Excel generator to avoid import issues.
    """
    import io
    from collections import Counter
    from datetime import datetime, timedelta

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Dashboard"

    # Encabezado profesional
    sheet.merge_cells('A1:F2')
    sheet['A1'] = 'INFORMES GENERADOS POR SAM METROLOGÍA SAS'
    sheet['A1'].font = Font(bold=True, size=20, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")

    # Información de la empresa
    sheet.merge_cells('A3:F3')
    sheet['A3'] = f'EMPRESA: {empresa.nombre.upper()}'
    sheet['A3'].font = Font(bold=True, size=14, color="1f4e79")

    sheet.merge_cells('A4:F4')
    hoy = datetime.now()
    sheet['A4'] = f'Generado el: {hoy.strftime("%d de %B de %Y a las %H:%M")}'
    sheet['A4'].font = Font(bold=True, size=12)

    # Datos estadísticos básicos
    equipos_list = list(equipos_queryset)
    total_equipos = len(equipos_list)

    # Resumen básico
    row = 6
    sheet[f'A{row}'] = 'RESUMEN GENERAL DE EQUIPOS'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    row += 2

    sheet.cell(row=row, column=1, value='Total de Equipos')
    sheet.cell(row=row, column=2, value=total_equipos)
    row += 2

    # Listado de equipos
    sheet[f'A{row}'] = 'LISTADO DE EQUIPOS'
    sheet[f'A{row}'].font = Font(bold=True, size=14, color="1f4e79")
    row += 1

    # Headers
    headers = ['Código', 'Nombre', 'Estado', 'Tipo']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    row += 1

    # Datos
    for equipo in equipos_list[:20]:  # Limitar a 20 equipos
        sheet.cell(row=row, column=1, value=equipo.codigo_interno)
        sheet.cell(row=row, column=2, value=equipo.nombre)
        sheet.cell(row=row, column=3, value=equipo.estado)
        sheet.cell(row=row, column=4, value=equipo.get_tipo_equipo_display())
        row += 1

    # Ajustar columnas
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        sheet.column_dimensions[col].width = 25

    # Guardar y retornar
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


# =============================================================================
# ZIP SYSTEM FUNCTIONS - AVOID CIRCULAR IMPORT
# =============================================================================
# ZIP functions are handled directly in zip_functions.py to avoid circular imports


def _generate_general_equipment_list_excel_content_local(equipos_queryset):
    """
    Local version of general equipment list Excel generator to avoid import issues.
    """
    import io
    from datetime import datetime

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listado de Equipos"

    # Encabezado profesional
    sheet.merge_cells('A1:AB2')
    sheet['A1'] = 'INFORMES GENERADOS POR SAM METROLOGÍA SAS'
    sheet['A1'].font = Font(bold=True, size=20, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")

    # Información de generación
    hoy = datetime.now()
    sheet.merge_cells('A3:AB3')
    sheet['A3'] = f'LISTADO GENERAL DE EQUIPOS - Generado el: {hoy.strftime("%d de %B de %Y a las %H:%M")}'
    sheet['A3'].font = Font(bold=True, size=12, color="1f4e79")

    # INFORMACIÓN DEL FORMATO (desde la empresa - campos de Listado de Equipos) - Fila 4 y 5
    primer_equipo = equipos_queryset.first()
    if primer_equipo and primer_equipo.empresa:
        empresa = primer_equipo.empresa
        codigo_fmt = empresa.listado_codigo or "N/A"
        version_fmt = empresa.listado_version or "N/A"
        fecha_fmt = empresa.listado_fecha_formato_display or (
            empresa.listado_fecha_formato.strftime("%Y-%m-%d") if empresa.listado_fecha_formato else "N/A"
        )
        sheet.merge_cells('A4:AB4')
        formato_info = f'FORMATO: Código: {codigo_fmt} | Versión: {version_fmt} | Fecha: {fecha_fmt}'
        sheet['A4'] = formato_info
        sheet['A4'].font = Font(bold=True, size=10, color="FFFFFF")
        sheet['A4'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        from openpyxl.styles import Alignment
        sheet['A4'].alignment = Alignment(horizontal="center", vertical="center")

    # Headers principales (sin columnas de formato que ya están arriba)
    headers = [
        "Código Interno", "Nombre", "Empresa", "Tipo de Equipo", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado", "Fecha de Adquisición",
        "Rango de Medida", "Resolución", "Error Máximo Permisible", "Puntos de Calibración",
        "Observaciones", "Último Certificado Calibración", "Fecha Última Calibración", "Próxima Calibración",
        "Frecuencia Calibración (meses)", "Fecha Último Mantenimiento", "Próximo Mantenimiento",
        "Frecuencia Mantenimiento (meses)", "Fecha Última Comprobación",
        "Próxima Comprobación", "Frecuencia Comprobación (meses)"
    ]

    # Aplicar headers en fila 6
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        sheet.column_dimensions[cell.column_letter].width = 25

    # Datos de equipos
    current_row = 7
    for equipo in equipos_queryset:
        # Obtener el último certificado de calibración
        ultima_calibracion = equipo.calibraciones.order_by('-fecha_calibracion').first()
        ultimo_certificado = ultima_calibracion.numero_certificado if ultima_calibracion else 'N/A'

        row_data = [
            equipo.codigo_interno,
            equipo.nombre,
            equipo.empresa.nombre if equipo.empresa else "N/A",
            equipo.get_tipo_equipo_display(),
            equipo.marca,
            equipo.modelo,
            equipo.numero_serie,
            equipo.ubicacion,
            equipo.responsable,
            equipo.estado,
            equipo.fecha_adquisicion.strftime('%Y-%m-%d') if equipo.fecha_adquisicion else '',
            equipo.rango_medida,
            equipo.resolucion,
            equipo.error_maximo_permisible if equipo.error_maximo_permisible is not None else '',
            equipo.puntos_calibracion if equipo.puntos_calibracion is not None else '',
            equipo.observaciones,
            ultimo_certificado,  # NUEVA COLUMNA
            equipo.fecha_ultima_calibracion.strftime('%Y-%m-%d') if equipo.fecha_ultima_calibracion else '',
            equipo.proxima_calibracion.strftime('%Y-%m-%d') if equipo.proxima_calibracion else '',
            float(equipo.frecuencia_calibracion_meses) if equipo.frecuencia_calibracion_meses is not None else '',
            equipo.fecha_ultimo_mantenimiento.strftime('%Y-%m-%d') if equipo.fecha_ultimo_mantenimiento else '',
            equipo.proximo_mantenimiento.strftime('%Y-%m-%d') if equipo.proximo_mantenimiento is not None else '',
            float(equipo.frecuencia_mantenimiento_meses) if equipo.frecuencia_mantenimiento_meses is not None else '',
            equipo.fecha_ultima_comprobacion.strftime('%Y-%m-%d') if equipo.fecha_ultima_comprobacion else '',
            equipo.proxima_comprobacion.strftime('%Y-%m-%d') if equipo.proxima_comprobacion is not None else '',
            float(equipo.frecuencia_comprobacion_meses) if equipo.frecuencia_comprobacion_meses is not None else '',
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Guardar y retornar
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()



# =============================================================================
# Funciones migradas desde views.py monolitico
# =============================================================================

def actualizar_equipo_selectivo(equipo_existente, nuevos_datos, row_index):
    """
    Actualiza solo los campos que NO estan vacios en los nuevos datos.

    Returns:
        list: Lista de nombres de campos que fueron actualizados
    """
    campos_actualizados = []

    # Mapeo de campos que se pueden actualizar
    campos_actualizables = {
        'nombre': 'nombre',
        'tipo_equipo': 'tipo_equipo',
        'marca': 'marca',
        'modelo': 'modelo',
        'numero_serie': 'numero_serie',
        'ubicacion': 'ubicacion',
        'responsable': 'responsable',
        'estado': 'estado',
        'fecha_adquisicion': 'fecha_adquisicion',
        'fecha_ultima_calibracion': 'fecha_ultima_calibracion',
        'fecha_ultimo_mantenimiento': 'fecha_ultimo_mantenimiento',
        'fecha_ultima_comprobacion': 'fecha_ultima_comprobacion',
        'rango_medida': 'rango_medida',
        'resolucion': 'resolucion',
        'error_maximo_permisible': 'error_maximo_permisible',
        'observaciones': 'observaciones',
        'version_formato': 'version_formato',
        'fecha_version_formato': 'fecha_version_formato',
        'codificacion_formato': 'codificacion_formato',
        'frecuencia_calibracion_meses': 'frecuencia_calibracion_meses',
        'frecuencia_mantenimiento_meses': 'frecuencia_mantenimiento_meses',
        'frecuencia_comprobacion_meses': 'frecuencia_comprobacion_meses',
    }

    for campo_excel, campo_modelo in campos_actualizables.items():
        if campo_modelo in nuevos_datos:
            nuevo_valor = nuevos_datos[campo_modelo]
            if es_valor_valido_para_actualizacion(nuevo_valor):
                valor_actual = getattr(equipo_existente, campo_modelo)
                if valores_son_diferentes(valor_actual, nuevo_valor):
                    setattr(equipo_existente, campo_modelo, nuevo_valor)
                    campos_actualizados.append(campo_excel)

    if campos_actualizados:
        equipo_existente.save()

    return campos_actualizados


def es_valor_valido_para_actualizacion(valor):
    """Determina si un valor es valido para actualizar."""
    if valor is None:
        return False
    if isinstance(valor, str) and valor.strip() == '':
        return False
    return True


def valores_son_diferentes(valor_actual, nuevo_valor):
    """Compara dos valores manejando diferentes tipos de datos."""
    if valor_actual is None and nuevo_valor is None:
        return False
    if valor_actual is None or nuevo_valor is None:
        return True
    if isinstance(valor_actual, str) and isinstance(nuevo_valor, str):
        return valor_actual.strip() != nuevo_valor.strip()
    return valor_actual != nuevo_valor


def _validate_and_load_excel(excel_file):
    """
    Valida y carga el archivo Excel, retornando el workbook, sheet y mapeo de columnas.

    Args:
        excel_file: Archivo Excel subido

    Returns:
        dict: {'workbook': workbook, 'sheet': sheet, 'column_mapping': dict, 'error': str|None}
    """
    result = {'workbook': None, 'sheet': None, 'column_mapping': None, 'error': None}

    try:
        # Cargar el archivo Excel
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook.active

        # Mapeo de columnas basado en la plantilla
        column_mapping = {
            'A': 'codigo_interno',
            'B': 'nombre',
            'C': 'empresa_nombre',
            'D': 'tipo_equipo',
            'E': 'marca',
            'F': 'modelo',
            'G': 'numero_serie',
            'H': 'ubicacion_nombre',
            'I': 'responsable',
            'J': 'estado',
            'K': 'fecha_adquisicion',
            'L': 'proveedor',
            'M': 'fecha_ultima_calibracion',
            'N': 'fecha_ultimo_mantenimiento',
            'O': 'fecha_ultima_comprobacion',
            'P': 'rango_medida',
            'Q': 'resolucion',
            'R': 'error_maximo_permisible',
            'S': 'puntos_calibracion',
            'T': 'observaciones',
            'U': 'version_formato',
            'V': 'fecha_version_formato',
            'W': 'codificacion_formato',
            'X': 'frecuencia_calibracion_meses',
            'Y': 'frecuencia_mantenimiento_meses',
            'Z': 'frecuencia_comprobacion_meses'
        }

        result['workbook'] = workbook
        result['sheet'] = sheet
        result['column_mapping'] = column_mapping

    except Exception as e:
        result['error'] = f"Error cargando archivo Excel: {str(e)}"

    return result


def _extract_row_data(sheet, row_num, column_mapping):
    """
    Extrae los datos de una fila del Excel.

    Args:
        sheet: Hoja de Excel activa
        row_num: Número de fila a procesar
        column_mapping: Diccionario con mapeo de columnas

    Returns:
        dict: Datos de la fila {field_name: value}
    """
    row_data = {}
    for col, field in column_mapping.items():
        cell_value = sheet[f'{col}{row_num}'].value
        if cell_value is not None:
            if isinstance(cell_value, str):
                cell_value = cell_value.strip()
            row_data[field] = cell_value
    return row_data


def _process_all_row_dates(row_data, row_num):
    """
    Procesa todas las fechas de una fila del Excel.

    Args:
        row_data: Datos de la fila
        row_num: Número de fila (para mensajes de error)

    Returns:
        dict: {
            'dates': {
                'fecha_adquisicion': date|None,
                'fecha_version_formato': date|None,
                'fecha_ultima_calibracion': date|None,
                'fecha_ultimo_mantenimiento': date|None,
                'fecha_ultima_comprobacion': date|None
            },
            'errors': [lista_errores]
        }
    """
    result = {
        'dates': {
            'fecha_adquisicion': None,
            'fecha_version_formato': None,
            'fecha_ultima_calibracion': None,
            'fecha_ultimo_mantenimiento': None,
            'fecha_ultima_comprobacion': None
        },
        'errors': []
    }

    # Mapeo de campos a nombres descriptivos
    date_fields = [
        ('fecha_adquisicion', 'Fecha de Adquisición'),
        ('fecha_version_formato', 'Fecha de Versión'),
        ('fecha_ultima_calibracion', 'Fecha Última Calibración'),
        ('fecha_ultimo_mantenimiento', 'Fecha Último Mantenimiento'),
        ('fecha_ultima_comprobacion', 'Fecha Última Comprobación')
    ]

    # Procesar cada fecha
    for field_name, field_label in date_fields:
        if row_data.get(field_name):
            fecha_result = _parse_date(row_data[field_name], field_label)
            if fecha_result['error']:
                result['errors'].append(f"Fila {row_num}: {fecha_result['error']}")
            else:
                result['dates'][field_name] = fecha_result['date']

    return result


def _update_existing_equipment(equipo_existente, row_data, dates_dict):
    """
    Actualiza un equipo existente con los datos del Excel.

    Args:
        equipo_existente: Instancia de Equipo a actualizar
        row_data: Datos de la fila del Excel
        dates_dict: Diccionario con las fechas procesadas

    Returns:
        list: Lista de campos actualizados
    """
    campos_actualizados = []

    # Actualizar solo campos que tengan valores en el Excel
    if row_data.get('nombre'):
        equipo_existente.nombre = row_data['nombre']
        campos_actualizados.append('nombre')

    if row_data.get('tipo_equipo'):
        equipo_existente.tipo_equipo = row_data['tipo_equipo']
        campos_actualizados.append('tipo_equipo')

    if row_data.get('marca'):
        equipo_existente.marca = row_data['marca']
        campos_actualizados.append('marca')

    if row_data.get('modelo'):
        equipo_existente.modelo = row_data['modelo']
        campos_actualizados.append('modelo')

    if row_data.get('numero_serie'):
        equipo_existente.numero_serie = row_data['numero_serie']
        campos_actualizados.append('numero_serie')

    if row_data.get('ubicacion_nombre'):
        equipo_existente.ubicacion = row_data['ubicacion_nombre']
        campos_actualizados.append('ubicacion')

    if row_data.get('responsable'):
        equipo_existente.responsable = row_data['responsable']
        campos_actualizados.append('responsable')

    if row_data.get('estado'):
        equipo_existente.estado = row_data['estado']
        campos_actualizados.append('estado')

    if dates_dict.get('fecha_adquisicion'):
        equipo_existente.fecha_adquisicion = dates_dict['fecha_adquisicion']
        campos_actualizados.append('fecha_adquisicion')

    if row_data.get('proveedor'):
        equipo_existente.proveedor = row_data['proveedor']
        campos_actualizados.append('proveedor')

    if row_data.get('rango_medida'):
        equipo_existente.rango_medida = row_data['rango_medida']
        campos_actualizados.append('rango_medida')

    if row_data.get('resolucion'):
        equipo_existente.resolucion = row_data['resolucion']
        campos_actualizados.append('resolucion')

    if row_data.get('error_maximo_permisible'):
        equipo_existente.error_maximo_permisible = row_data['error_maximo_permisible']
        campos_actualizados.append('error_maximo_permisible')

    if row_data.get('puntos_calibracion'):
        equipo_existente.puntos_calibracion = row_data['puntos_calibracion']
        campos_actualizados.append('puntos_calibracion')

    if row_data.get('observaciones'):
        equipo_existente.observaciones = row_data['observaciones']
        campos_actualizados.append('observaciones')

    if row_data.get('version_formato'):
        equipo_existente.version_formato = row_data['version_formato']
        campos_actualizados.append('version_formato')

    if dates_dict.get('fecha_version_formato'):
        equipo_existente.fecha_version_formato = dates_dict['fecha_version_formato']
        campos_actualizados.append('fecha_version_formato')

    if row_data.get('codificacion_formato'):
        equipo_existente.codificacion_formato = row_data['codificacion_formato']
        campos_actualizados.append('codificacion_formato')

    # Frecuencias - actualizar solo si el valor no es None
    freq_calib = _parse_decimal(row_data.get('frecuencia_calibracion_meses'))
    if freq_calib is not None:
        equipo_existente.frecuencia_calibracion_meses = freq_calib
        campos_actualizados.append('frecuencia_calibracion_meses')

    freq_mant = _parse_decimal(row_data.get('frecuencia_mantenimiento_meses'))
    if freq_mant is not None:
        equipo_existente.frecuencia_mantenimiento_meses = freq_mant
        campos_actualizados.append('frecuencia_mantenimiento_meses')

    freq_comp = _parse_decimal(row_data.get('frecuencia_comprobacion_meses'))
    if freq_comp is not None:
        equipo_existente.frecuencia_comprobacion_meses = freq_comp
        campos_actualizados.append('frecuencia_comprobacion_meses')

    # Fechas de última actividad - actualizar solo si se proporcionan
    if dates_dict.get('fecha_ultima_calibracion'):
        equipo_existente.fecha_ultima_calibracion = dates_dict['fecha_ultima_calibracion']
        campos_actualizados.append('fecha_ultima_calibracion')

    if dates_dict.get('fecha_ultimo_mantenimiento'):
        equipo_existente.fecha_ultimo_mantenimiento = dates_dict['fecha_ultimo_mantenimiento']
        campos_actualizados.append('fecha_ultimo_mantenimiento')

    if dates_dict.get('fecha_ultima_comprobacion'):
        equipo_existente.fecha_ultima_comprobacion = dates_dict['fecha_ultima_comprobacion']
        campos_actualizados.append('fecha_ultima_comprobacion')

    # Guardar cambios si hay campos actualizados
    if campos_actualizados:
        equipo_existente.save(update_fields=campos_actualizados)

    return campos_actualizados


def _create_new_equipment(row_data, empresa, dates_dict):
    """
    Crea un nuevo equipo con los datos del Excel.

    Args:
        row_data: Datos de la fila del Excel
        empresa: Instancia de Empresa
        dates_dict: Diccionario con las fechas procesadas

    Returns:
        Equipo: Nueva instancia de Equipo creada
    """
    equipo = Equipo.objects.create(
        codigo_interno=row_data['codigo_interno'],
        nombre=row_data['nombre'],
        empresa=empresa,
        tipo_equipo=row_data.get('tipo_equipo', 'Equipo de Medición'),
        marca=row_data.get('marca', ''),
        modelo=row_data.get('modelo', ''),
        numero_serie=row_data.get('numero_serie', ''),
        ubicacion=row_data.get('ubicacion_nombre', ''),
        responsable=row_data.get('responsable', ''),
        estado=row_data.get('estado', ESTADO_ACTIVO),
        fecha_adquisicion=dates_dict.get('fecha_adquisicion'),
        proveedor=row_data.get('proveedor', ''),
        rango_medida=row_data.get('rango_medida', ''),
        resolucion=row_data.get('resolucion', ''),
        error_maximo_permisible=row_data.get('error_maximo_permisible', ''),
        puntos_calibracion=row_data.get('puntos_calibracion', ''),
        observaciones=row_data.get('observaciones', ''),
        version_formato=row_data.get('version_formato', ''),
        fecha_version_formato=dates_dict.get('fecha_version_formato'),
        codificacion_formato=row_data.get('codificacion_formato', ''),
        # Frecuencias
        frecuencia_calibracion_meses=_parse_decimal(row_data.get('frecuencia_calibracion_meses')),
        frecuencia_mantenimiento_meses=_parse_decimal(row_data.get('frecuencia_mantenimiento_meses')),
        frecuencia_comprobacion_meses=_parse_decimal(row_data.get('frecuencia_comprobacion_meses')),
        # Fechas de última actividad
        fecha_ultima_calibracion=dates_dict.get('fecha_ultima_calibracion'),
        fecha_ultimo_mantenimiento=dates_dict.get('fecha_ultimo_mantenimiento'),
        fecha_ultima_comprobacion=dates_dict.get('fecha_ultima_comprobacion')
    )

    return equipo


def _process_excel_import(excel_file, user):
    """
    Procesa el archivo Excel importado y crea los equipos.

    Args:
        excel_file: Archivo Excel subido
        user: Usuario que realiza la importación

    Returns:
        dict: Resultado con success, imported, errors
    """
    result = {
        'success': False,
        'imported': 0,
        'created': 0,
        'updated': 0,
        'errors': []
    }

    try:
        # Cargar y validar archivo Excel
        excel_data = _validate_and_load_excel(excel_file)
        if excel_data['error']:
            result['errors'] = [excel_data['error']]
            return result

        sheet = excel_data['sheet']
        column_mapping = excel_data['column_mapping']

        # Comenzar desde la fila 8 (después de headers)
        start_row = 8
        imported_count = 0
        created_count = 0
        updated_count = 0
        errors = []

        # Obtener empresa del usuario si no es superusuario
        user_empresa = None
        if not user.is_superuser and user.empresa:
            user_empresa = user.empresa

        # Procesar fila por fila fuera de transacción para mejor manejo de errores
        for row_num in range(start_row, sheet.max_row + 1):
            try:
                # Extraer datos de la fila
                row_data = _extract_row_data(sheet, row_num, column_mapping)

                # Verificar si hay datos en la fila (al menos código interno)
                if not row_data.get('codigo_interno'):
                    continue

                # Log de diagnóstico para frecuencias (mostrar TODAS, incluso None)
                freq_debug = []
                for key in ['frecuencia_calibracion_meses', 'frecuencia_mantenimiento_meses', 'frecuencia_comprobacion_meses']:
                    value = row_data.get(key, 'NO_EXISTE')
                    freq_debug.append(f"{key}: {repr(value)}")
                logger.info(f"Fila {row_num} - TODAS las frecuencias: {', '.join(freq_debug)}")

                # Validaciones básicas con mensajes claros
                validation_result = _validate_row_data(row_data, row_num, user_empresa)
                if validation_result['errors']:
                    errors.extend(validation_result['errors'])
                    continue

                empresa = validation_result['empresa']

                # Verificar si el equipo ya existe para decidir crear o actualizar
                equipo_existente = Equipo.objects.filter(
                    codigo_interno=row_data['codigo_interno'],
                    empresa=empresa
                ).first()

                es_actualizacion = equipo_existente is not None

                # Procesar todas las fechas de la fila
                dates_result = _process_all_row_dates(row_data, row_num)
                if dates_result['errors']:
                    errors.extend(dates_result['errors'])
                    continue

                # Crear o actualizar el equipo en transacción individual
                try:
                    with transaction.atomic():
                        if es_actualizacion:
                            # Actualizar equipo existente
                            campos_actualizados = _update_existing_equipment(
                                equipo_existente,
                                row_data,
                                dates_result['dates']
                            )
                            equipo = equipo_existente
                        else:
                            # Crear nuevo equipo
                            equipo = _create_new_equipment(
                                row_data,
                                empresa,
                                dates_result['dates']
                            )
                            campos_actualizados = []

                        # Calcular fechas próximas basadas en las fechas de última actividad y frecuencias
                        _calcular_fechas_proximas(equipo)

                        imported_count += 1
                        if es_actualizacion:
                            updated_count += 1
                        else:
                            created_count += 1

                        # Log detallado de importación con frecuencias y fechas
                        freq_info = []
                        if equipo.frecuencia_calibracion_meses:
                            freq_info.append(f"Calib: {equipo.frecuencia_calibracion_meses}m")
                        if equipo.frecuencia_mantenimiento_meses:
                            freq_info.append(f"Mant: {equipo.frecuencia_mantenimiento_meses}m")
                        if equipo.frecuencia_comprobacion_meses:
                            freq_info.append(f"Comp: {equipo.frecuencia_comprobacion_meses}m")

                        fecha_info = []
                        if equipo.fecha_ultima_calibracion:
                            fecha_info.append(f"Últ.Calib: {equipo.fecha_ultima_calibracion}")
                        if equipo.fecha_ultimo_mantenimiento:
                            fecha_info.append(f"Últ.Mant: {equipo.fecha_ultimo_mantenimiento}")
                        if equipo.fecha_ultima_comprobacion:
                            fecha_info.append(f"Últ.Comp: {equipo.fecha_ultima_comprobacion}")

                        proxima_info = []
                        if equipo.proxima_calibracion:
                            proxima_info.append(f"Prox.Calib: {equipo.proxima_calibracion}")
                        if equipo.proximo_mantenimiento:
                            proxima_info.append(f"Prox.Mant: {equipo.proximo_mantenimiento}")
                        if equipo.proxima_comprobacion:
                            proxima_info.append(f"Prox.Comp: {equipo.proxima_comprobacion}")

                        # Log diferenciado para creación vs actualización
                        accion = "🔄 ACTUALIZADO" if es_actualizacion else "✅ CREADO"
                        extra_info = ""
                        if es_actualizacion and 'campos_actualizados' in locals():
                            extra_info = f" | Campos actualizados: {len(campos_actualizados)}"

                        logger.info(f"{accion}: {equipo.codigo_interno}{extra_info} | " +
                                  f"Frecuencias: [{', '.join(freq_info) if freq_info else 'Ninguna'}] | " +
                                  f"Últimas: [{', '.join(fecha_info) if fecha_info else 'Ninguna'}] | " +
                                  f"Próximas: [{', '.join(proxima_info) if proxima_info else 'Ninguna'}]")

                except Exception as create_error:
                    errors.append(f"Fila {row_num}: ❌ Error creando equipo - {str(create_error)}")
                    continue

            except Exception as e:
                errors.append(f"Fila {row_num}: ❌ Error procesando fila - {str(e)}")
                continue

        result['success'] = imported_count > 0
        result['imported'] = imported_count
        result['created'] = created_count
        result['updated'] = updated_count
        result['errors'] = errors

        if errors:
            logger.warning(f"Importación completada con {len(errors)} errores")
            # Log de los primeros 10 errores para diagnóstico
            for i, error in enumerate(errors[:10]):
                logger.error(f"Error {i+1}: {error}")
            if len(errors) > 10:
                logger.error(f"... y {len(errors) - 10} errores más")

        # Log del resumen final
        logger.info(f"📊 RESUMEN: {imported_count} equipos procesados | " +
                   f"Nuevos: {created_count} | Actualizados: {updated_count} | Errores: {len(errors)}")

    except Exception as e:
        logger.error(f"Error general en importación Excel: {e}")
        result['errors'] = [f"Error procesando archivo Excel: {str(e)}"]

    return result


def _calcular_fechas_proximas(equipo):
    """
    Calcula las fechas próximas de calibración, mantenimiento y comprobación
    basadas en la jerarquía de fechas de última actividad y las frecuencias.

    Jerarquía: último mantenimiento > última calibración > fecha de adquisición
    """
    from datetime import date
    from dateutil.relativedelta import relativedelta

    try:
        # Determinar fecha base según jerarquía:
        # 1. Fecha de último mantenimiento (si existe)
        # 2. Si no hay mantenimiento, usar fecha de última calibración (si existe)
        # 3. Si no hay calibración, usar fecha de adquisición
        # 4. Si no hay fecha de adquisición, usar fecha actual
        fecha_base = (
            equipo.fecha_ultimo_mantenimiento or
            equipo.fecha_ultima_calibracion or
            equipo.fecha_adquisicion or
            date.today()
        )

        # Calcular próxima calibración usando la fecha base y frecuencia
        if equipo.frecuencia_calibracion_meses:
            try:
                meses = int(equipo.frecuencia_calibracion_meses)
                if meses > 0:
                    # Si hay fecha de última calibración específica, usarla, sino usar fecha base
                    fecha_ref_calibracion = equipo.fecha_ultima_calibracion or fecha_base
                    equipo.proxima_calibracion = fecha_ref_calibracion + relativedelta(months=meses)
            except (ValueError, TypeError):
                pass

        # Calcular próximo mantenimiento usando la fecha base y frecuencia
        if equipo.frecuencia_mantenimiento_meses:
            try:
                meses = int(equipo.frecuencia_mantenimiento_meses)
                if meses > 0:
                    # Si hay fecha de último mantenimiento específica, usarla, sino usar fecha base
                    fecha_ref_mantenimiento = equipo.fecha_ultimo_mantenimiento or fecha_base
                    equipo.proximo_mantenimiento = fecha_ref_mantenimiento + relativedelta(months=meses)
            except (ValueError, TypeError):
                pass

        # Calcular próxima comprobación usando la fecha base y frecuencia
        if equipo.frecuencia_comprobacion_meses:
            try:
                meses = int(equipo.frecuencia_comprobacion_meses)
                if meses > 0:
                    # Si hay fecha de última comprobación específica, usarla, sino usar fecha base
                    fecha_ref_comprobacion = equipo.fecha_ultima_comprobacion or fecha_base
                    equipo.proxima_comprobacion = fecha_ref_comprobacion + relativedelta(months=meses)
            except (ValueError, TypeError):
                pass

        # Guardar los cambios
        equipo.save(update_fields=[
            'proxima_calibracion',
            'proximo_mantenimiento',
            'proxima_comprobacion'
        ])

    except Exception as e:
        logger.warning(f"Error calculando fechas próximas para equipo {equipo.codigo_interno}: {e}")


def _parse_decimal(value):
    """Convierte un valor a decimal, retorna None si no es válido."""
    if value is None:
        return None
    try:
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
        from decimal import Decimal, InvalidOperation
        return Decimal(str(value))
    except (ValueError, TypeError, InvalidOperation):
        return None


def _validate_row_data(row_data, row_num, user_empresa):
    """
    Valida los datos de una fila y retorna empresa y errores.

    Returns:
        dict: {'empresa': Empresa|None, 'errors': [lista_errores]}
    """
    result = {'empresa': None, 'errors': []}

    # Validar nombre del equipo
    if not row_data.get('nombre'):
        result['errors'].append(f"Fila {row_num}: ❌ Nombre del equipo es obligatorio (columna B)")
        return result

    # Validar código interno
    codigo = row_data.get('codigo_interno', '').strip()
    if not codigo:
        result['errors'].append(f"Fila {row_num}: ❌ Código interno es obligatorio (columna A)")
        return result

    if len(codigo) > 50:  # Ajustar según el modelo
        result['errors'].append(f"Fila {row_num}: ❌ Código interno demasiado largo (máximo 50 caracteres)")
        return result

    # Determinar empresa
    if user_empresa:
        # Usuario normal: usar su empresa
        result['empresa'] = user_empresa
    else:
        # Superusuario: buscar empresa por nombre
        empresa_nombre = row_data.get('empresa_nombre')
        if not empresa_nombre:
            result['errors'].append(f"Fila {row_num}: ❌ Nombre de empresa es obligatorio para superusuarios (columna C)")
            return result

        try:
            result['empresa'] = Empresa.objects.get(nombre=empresa_nombre)
        except Empresa.DoesNotExist:
            result['errors'].append(f"Fila {row_num}: ❌ Empresa '{empresa_nombre}' no encontrada. Empresas disponibles: {list(Empresa.objects.values_list('nombre', flat=True))}")
            return result

    # Validar tipo de equipo
    tipo_equipo = row_data.get('tipo_equipo')
    if tipo_equipo:
        tipos_validos = [choice[0] for choice in Equipo.TIPO_EQUIPO_CHOICES]
        if tipo_equipo not in tipos_validos:
            result['errors'].append(f"Fila {row_num}: ❌ Tipo de equipo '{tipo_equipo}' no válido. Tipos válidos: {tipos_validos}")

    # Validar estado
    estado = row_data.get('estado')
    if estado:
        estados_validos = [choice[0] for choice in Equipo.ESTADO_CHOICES]
        if estado not in estados_validos:
            result['errors'].append(f"Fila {row_num}: ❌ Estado '{estado}' no válido. Estados válidos: {estados_validos}")

    return result


def _parse_date(date_value, field_name):
    """
    Parsea una fecha de forma ultra-flexible para Excel.
    Acepta: texto, números de Excel, fechas de Excel, cualquier formato común.

    Returns:
        dict: {'date': date|None, 'error': str|None}
    """
    result = {'date': None, 'error': None}

    if date_value is None:
        return result

    try:
        # Caso 1: Es un objeto datetime o date (fechas nativas de Excel)
        if hasattr(date_value, 'date'):
            result['date'] = date_value.date()
            return result
        elif hasattr(date_value, 'year'):
            result['date'] = date_value
            return result

        # Caso 2: Es un número (Excel a veces convierte fechas a números de serie)
        if isinstance(date_value, (int, float)):
            try:
                # Excel usa 1900-01-01 como día 1 (pero con un bug que considera 1900 como bisiesto)
                # Los números de Excel típicos para fechas están entre 1 y 100000
                if 1 <= date_value <= 100000:
                    from datetime import datetime, timedelta
                    # Excel epoch: 1900-01-01, pero ajustamos por el bug de Excel
                    excel_epoch = datetime(1899, 12, 30)  # 30 dic 1899 para compensar el bug
                    parsed_date = (excel_epoch + timedelta(days=date_value)).date()
                    result['date'] = parsed_date
                    return result
                else:
                    # Intentar como timestamp Unix si el número es muy grande
                    if date_value > 100000:
                        parsed_date = datetime.fromtimestamp(date_value).date()
                        result['date'] = parsed_date
                        return result
            except (ValueError, OSError, OverflowError):
                # Si falla como número de Excel, intentar como string
                date_value = str(date_value)

        # Caso 3: Es texto o convertir número a texto
        if isinstance(date_value, str) or date_value is not None:
            date_str = str(date_value).strip()
            if not date_str or date_str.lower() in ['none', 'null', 'n/a', '']:
                return result

            # Lista ampliada de formatos a intentar
            date_formats = [
                '%Y-%m-%d',        # 2024-01-15
                '%d/%m/%Y',        # 15/01/2024
                '%d-%m-%Y',        # 15-01-2024
                '%m/%d/%Y',        # 01/15/2024 (formato US)
                '%m-%d-%Y',        # 01-15-2024 (formato US)
                '%Y/%m/%d',        # 2024/01/15
                '%d.%m.%Y',        # 15.01.2024 (formato europeo)
                '%d %m %Y',        # 15 01 2024 (con espacios)
                '%Y%m%d',          # 20240115 (formato compacto)
                '%d/%m/%y',        # 15/01/24 (año de 2 dígitos)
                '%m/%d/%y',        # 01/15/24 (US, año de 2 dígitos)
                '%d-%m-%y',        # 15-01-24
                '%m-%d-%y',        # 01-15-24
                '%d.%m.%y',        # 15.01.24
            ]

            parsed_date = None

            # Intentar cada formato
            for date_format in date_formats:
                try:
                    from datetime import datetime
                    parsed_date = datetime.strptime(date_str, date_format).date()
                    break  # Si funciona, salir del bucle
                except ValueError:
                    continue  # Probar siguiente formato

            # Si Django puede parsearlo, intentar
            if parsed_date is None:
                from django.utils.dateparse import parse_date
                parsed_date = parse_date(date_str)

            # Intentar parsear números dentro del texto
            if parsed_date is None:
                try:
                    # Ver si es un número puro dentro de string
                    float_val = float(date_str)
                    if 1 <= float_val <= 100000:
                        from datetime import datetime, timedelta
                        excel_epoch = datetime(1899, 12, 30)
                        parsed_date = (excel_epoch + timedelta(days=float_val)).date()
                except (ValueError, OverflowError):
                    pass

            if parsed_date is None:
                result['error'] = f"❌ {field_name}: no se pudo interpretar '{date_value}'. Puede usar: fechas de Excel, DD/MM/YYYY, YYYY-MM-DD, números, etc. ¿Está la celda formateada como fecha en Excel?"
            else:
                result['date'] = parsed_date

        else:
            result['error'] = f"❌ {field_name}: tipo de dato inesperado '{type(date_value)}'. Valor: '{date_value}'"

    except Exception as e:
        result['error'] = f"❌ {field_name}: error procesando '{date_value}' - {str(e)}. Intente formatear la celda como fecha en Excel."

    return result

