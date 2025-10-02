# core/views/admin.py
# Views administrativas: usuarios, proveedores, procedimientos, autenticación

from .base import *

# =============================================================================
# AUTHENTICATION VIEWS
# =============================================================================

@monitor_view
def cambiar_password(request):
    """
    Vista para cambio de contraseña del usuario actual.
    """
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
            logger.info(f"Usuario {request.user.username} cambió su contraseña")
            return redirect('core:password_change_done')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'core/cambiar_password.html', {
        'form': form,
        'titulo_pagina': 'Cambiar Contraseña'
    })


@monitor_view
def password_change_done(request):
    """
    Vista de confirmación de cambio de contraseña.
    """
    return render(request, 'core/password_change_done.html', {
        'titulo_pagina': 'Contraseña Cambiada'
    })


@monitor_view
@access_check
@login_required
@superuser_required
def redirect_to_change_password(request, pk):
    """
    Redirección temporal para compatibilidad con URLs antiguas.
    """
    return redirect('core:change_user_password', pk=pk)


@monitor_view
@access_check
@login_required
@superuser_required
@require_POST
@csrf_exempt
def toggle_download_permission(request):
    """
    Alterna permisos de descarga de un usuario vía AJAX.
    """
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        can_download = data.get('can_download') or data.get('grant_permission')  # Compatibilidad

        if user_id is None or can_download is None:
            return JsonResponse({'status': 'error', 'message': 'Datos incompletos'}, status=400)

        usuario = get_object_or_404(CustomUser, pk=user_id)

        # Asignar o remover permiso de descarga
        permission = Permission.objects.get(codename='can_export_reports')

        if can_download:
            usuario.user_permissions.add(permission)
        else:
            usuario.user_permissions.remove(permission)

        # Verificar que el cambio se aplicó correctamente
        has_permission_now = usuario.user_permissions.filter(codename='can_export_reports').exists()

        estado_texto = 'otorgados' if can_download else 'revocados'
        logger.info(f"Permisos de descarga {estado_texto} para usuario {usuario.username} por {request.user.username} - Estado final: {has_permission_now}")

        return JsonResponse({
            'status': 'success',
            'message': f'Permisos de descarga {estado_texto} exitosamente',
            'can_download': can_download,
            'has_permission_now': has_permission_now  # Para debug
        })

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido'}, status=400)
    except Exception as e:
        logger.error(f"Error en toggle_download_permission: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@monitor_view
def user_login(request):
    """
    Vista de login personalizada con redirección inteligente.
    """
    if request.user.is_authenticated:
        return redirect('core:dashboard')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)

            if user is not None:
                login(request, user)
                logger.info(f"Usuario {username} inició sesión exitosamente")
                return redirect('core:dashboard')
            else:
                messages.error(request, 'Credenciales inválidas.')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = AuthenticationForm()

    return render(request, 'registration/login.html', {
        'form': form,
        'titulo_pagina': 'Iniciar Sesión'
    })


@monitor_view
@login_required
def user_logout(request):
    """
    Vista de logout con mensaje de confirmación.
    """
    username = request.user.username
    logout(request)
    messages.success(request, f'Has cerrado sesión exitosamente, {username}.')
    logger.info(f"Usuario {username} cerró sesión")
    return redirect('core:login')


@monitor_view
@access_check
@login_required
def perfil_usuario(request):
    """
    Vista del perfil de usuario con edición de información básica.
    """
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
            logger.info(f"Usuario {request.user.username} actualizó su perfil")
            return redirect('core:perfil_usuario')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = UserProfileForm(instance=request.user)

    context = {
        'form': form,
        'titulo_pagina': 'Mi Perfil'
    }
    return render(request, 'core/perfil_usuario.html', context)


# =============================================================================
# USER MANAGEMENT VIEWS
# =============================================================================

@monitor_view
@access_check
@login_required
@superuser_required
def listar_usuarios(request):
    """
    Lista todos los usuarios del sistema (solo superusuarios).
    Incluye filtros y paginación.
    """
    query = request.GET.get('q', '')
    usuarios_list = CustomUser.objects.all().select_related('empresa')

    # Aplicar filtro de búsqueda
    if query:
        usuarios_list = usuarios_list.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )

    # Ordenar por fecha de última conexión
    usuarios_list = usuarios_list.order_by('-last_login', 'username')

    # Paginación
    paginator = Paginator(usuarios_list, 25)
    page_number = request.GET.get('page')
    try:
        usuarios = paginator.page(page_number)
    except PageNotAnInteger:
        usuarios = paginator.page(1)
    except EmptyPage:
        usuarios = paginator.page(paginator.num_pages)

    context = {
        'usuarios': usuarios,
        'query': query,
        'titulo_pagina': 'Gestión de Usuarios'
    }
    return render(request, 'core/listar_usuarios.html', context)


@monitor_view
@access_check
@login_required
@superuser_required
def añadir_usuario(request, empresa_pk=None):
    """
    Añade un nuevo usuario al sistema.
    Opcionalmente se puede especificar una empresa.
    """
    empresa = None
    if empresa_pk:
        empresa = get_object_or_404(Empresa, pk=empresa_pk)

    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            try:
                usuario = form.save(commit=False)
                if empresa:
                    usuario.empresa = empresa
                usuario.save()

                messages.success(request, f'Usuario "{usuario.username}" creado exitosamente.')
                logger.info(f"Usuario creado: {usuario.username} por {request.user.username}")

                if empresa:
                    return redirect('core:detalle_empresa', pk=empresa.pk)
                else:
                    return redirect('core:listar_usuarios')

            except Exception as e:
                messages.error(request, f'Error al crear usuario: {e}')
                logger.error(f"Error creando usuario: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = CustomUserCreationForm()

    context = {
        'form': form,
        'empresa': empresa,
        'titulo_pagina': f'Añadir Usuario{"" if not empresa else f" a {empresa.nombre}"}'
    }
    return render(request, 'core/añadir_usuario.html', context)


@monitor_view
@access_check
@login_required
@superuser_required
def detalle_usuario(request, pk):
    """
    Muestra los detalles de un usuario específico.
    """
    usuario = get_object_or_404(CustomUser, pk=pk)

    # Obtener estadísticas del usuario
    actividad_reciente = {
        'ultimo_login': usuario.last_login,
        'fecha_registro': usuario.date_joined,
        'activo': usuario.is_active,
        'empresa': usuario.empresa,
    }

    context = {
        'usuario': usuario,
        'usuario_a_ver': usuario,  # Para compatibilidad con el template
        'actividad_reciente': actividad_reciente,
        'titulo_pagina': f'Detalle de Usuario: {usuario.username}'
    }
    return render(request, 'core/detalle_usuario.html', context)


@monitor_view
@access_check
@login_required
@superuser_required
def editar_usuario(request, pk):
    """
    Edita la información de un usuario existente.
    """
    usuario = get_object_or_404(CustomUser, pk=pk)

    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=usuario)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f'Usuario "{usuario.username}" actualizado exitosamente.')
                logger.info(f"Usuario actualizado: {usuario.username} por {request.user.username}")
                return redirect('core:detalle_usuario', pk=usuario.pk)
            except Exception as e:
                messages.error(request, f'Error al actualizar usuario: {e}')
                logger.error(f"Error actualizando usuario {usuario.pk}: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = CustomUserChangeForm(instance=usuario)

    context = {
        'form': form,
        'usuario': usuario,
        'usuario_a_editar': usuario,  # Para compatibilidad con el template
        'titulo_pagina': f'Editar Usuario: {usuario.username}'
    }
    return render(request, 'core/editar_usuario.html', context)


@monitor_view
@access_check
@login_required
@superuser_required
def eliminar_usuario(request, pk):
    """
    Elimina un usuario del sistema (con confirmación).
    """
    usuario = get_object_or_404(CustomUser, pk=pk)

    # Prevenir auto-eliminación
    if usuario == request.user:
        messages.error(request, 'No puedes eliminar tu propia cuenta.')
        return redirect('core:listar_usuarios')

    if request.method == 'POST':
        try:
            username = usuario.username
            logger.info(f"Eliminando usuario: {username} por {request.user.username}")
            usuario.delete()
            messages.success(request, f'Usuario "{username}" eliminado exitosamente.')
            return redirect('core:listar_usuarios')
        except Exception as e:
            messages.error(request, f'Error al eliminar usuario: {e}')
            logger.error(f"Error eliminando usuario {usuario.pk}: {e}")
            return redirect('core:listar_usuarios')

    # Contexto para confirmación
    context = {
        'object_name': f'el usuario "{usuario.username}"',
        'return_url_name': 'core:listar_usuarios',
        'return_url_pk': None,
        'titulo_pagina': f'Eliminar Usuario: {usuario.username}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


@monitor_view
@access_check
@login_required
@superuser_required
def change_user_password(request, pk):
    """
    Cambia la contraseña de un usuario específico.
    """
    usuario = get_object_or_404(CustomUser, pk=pk)

    if request.method == 'POST':
        form = PasswordChangeForm(usuario, request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f'Contraseña cambiada exitosamente para "{usuario.username}".')
                logger.info(f"Contraseña cambiada para usuario {usuario.username} por {request.user.username}")
                return redirect('core:detalle_usuario', pk=usuario.pk)
            except Exception as e:
                messages.error(request, f'Error al cambiar contraseña: {e}')
                logger.error(f"Error cambiando contraseña usuario {usuario.pk}: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = PasswordChangeForm(usuario)

    context = {
        'form': form,
        'usuario': usuario,
        'user_to_change': usuario,  # Para compatibilidad con el template
        'titulo_pagina': f'Cambiar Contraseña: {usuario.username}'
    }
    return render(request, 'core/change_user_password.html', context)


@monitor_view
@access_check
@login_required
@superuser_required
@require_POST
@csrf_exempt
def toggle_user_active_status(request):
    """
    Alterna el estado 'is_active' de un usuario vía AJAX.
    """
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        is_active = data.get('is_active')

        if user_id is None or is_active is None:
            return JsonResponse({'status': 'error', 'message': 'Datos incompletos'}, status=400)

        usuario = get_object_or_404(CustomUser, pk=user_id)

        # Prevenir auto-desactivación
        if usuario == request.user:
            return JsonResponse({'status': 'error', 'message': 'No puedes desactivar tu propia cuenta'}, status=400)

        usuario.is_active = is_active
        usuario.save()

        estado_texto = 'activado' if is_active else 'desactivado'
        logger.info(f"Usuario {usuario.username} {estado_texto} por {request.user.username}")

        return JsonResponse({
            'status': 'success',
            'message': f'Usuario {estado_texto} exitosamente',
            'is_active': is_active
        })

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido'}, status=400)
    except Exception as e:
        logger.error(f"Error en toggle_user_active_status: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# =============================================================================
# PROCEDIMIENTOS MANAGEMENT
# =============================================================================

@monitor_view
@access_check
@login_required
@permission_required('core.view_procedimiento', raise_exception=True)
def listar_procedimientos(request):
    """
    Lista todos los procedimientos.
    Filtrados por empresa para usuarios no-superusuarios.
    """
    procedimientos = Procedimiento.objects.all().select_related('empresa')

    # Filtrar por empresa si no es superusuario
    if not request.user.is_superuser and request.user.empresa:
        procedimientos = procedimientos.filter(empresa=request.user.empresa)
    elif not request.user.is_superuser and not request.user.empresa:
        procedimientos = Procedimiento.objects.none()

    procedimientos = procedimientos.order_by('codigo')

    context = {
        'procedimientos': procedimientos,
        'titulo_pagina': 'Listado de Procedimientos'
    }
    return render(request, 'core/listar_procedimientos.html', context)


@monitor_view
@access_check
@login_required
@permission_required('core.add_procedimiento', raise_exception=True)
def añadir_procedimiento(request):
    """
    Añade un nuevo procedimiento.
    """
    if request.method == 'POST':
        form = ProcedimientoForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                procedimiento = form.save(commit=False)

                # Asignar empresa automáticamente para usuarios no-superusuarios
                if not request.user.is_superuser and not procedimiento.empresa:
                    procedimiento.empresa = request.user.empresa

                procedimiento.save()
                messages.success(request, 'Procedimiento añadido exitosamente.')
                logger.info(f"Procedimiento creado: {procedimiento.codigo} por {request.user.username}")
                return redirect('core:listar_procedimientos')
            except Exception as e:
                messages.error(request, f'Error al añadir procedimiento: {e}')
                logger.error(f"Error añadiendo procedimiento: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = ProcedimientoForm()

    context = {
        'form': form,
        'titulo_pagina': 'Añadir Nuevo Procedimiento'
    }
    return render(request, 'core/añadir_procedimiento.html', context)


@monitor_view
@access_check
@login_required
@permission_required('core.change_procedimiento', raise_exception=True)
def editar_procedimiento(request, pk):
    """
    Edita un procedimiento existente.
    """
    procedimiento = get_object_or_404(Procedimiento, pk=pk)

    # Verificar permisos por empresa
    if not request.user.is_superuser and request.user.empresa != procedimiento.empresa:
        messages.error(request, 'No tienes permiso para editar este procedimiento.')
        return redirect('core:listar_procedimientos')

    if request.method == 'POST':
        form = ProcedimientoForm(request.POST, request.FILES, instance=procedimiento)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Procedimiento actualizado exitosamente.')
                logger.info(f"Procedimiento actualizado: {procedimiento.codigo} por {request.user.username}")
                return redirect('core:listar_procedimientos')
            except Exception as e:
                messages.error(request, f'Error al actualizar procedimiento: {e}')
                logger.error(f"Error actualizando procedimiento {procedimiento.pk}: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = ProcedimientoForm(instance=procedimiento)

    context = {
        'form': form,
        'procedimiento': procedimiento,
        'titulo_pagina': f'Editar Procedimiento: {procedimiento.codigo}'
    }
    return render(request, 'core/editar_procedimiento.html', context)


@monitor_view
@access_check
@login_required
@permission_required('core.delete_procedimiento', raise_exception=True)
def eliminar_procedimiento(request, pk):
    """
    Elimina un procedimiento (con confirmación).
    """
    procedimiento = get_object_or_404(Procedimiento, pk=pk)

    # Verificar permisos por empresa
    if not request.user.is_superuser and request.user.empresa != procedimiento.empresa:
        messages.error(request, 'No tienes permiso para eliminar este procedimiento.')
        return redirect('core:listar_procedimientos')

    if request.method == 'POST':
        try:
            codigo_procedimiento = procedimiento.codigo
            logger.info(f"Eliminando procedimiento: {codigo_procedimiento} por {request.user.username}")
            procedimiento.delete()
            messages.success(request, f'Procedimiento "{codigo_procedimiento}" eliminado exitosamente.')
            return redirect('core:listar_procedimientos')
        except Exception as e:
            messages.error(request, f'Error al eliminar procedimiento: {e}')
            logger.error(f"Error eliminando procedimiento {procedimiento.pk}: {e}")
            return redirect('core:listar_procedimientos')

    # Contexto para confirmación
    context = {
        'object_name': f'el procedimiento "{procedimiento.codigo}"',
        'return_url_name': 'core:listar_procedimientos',
        'return_url_pk': None,
        'titulo_pagina': f'Eliminar Procedimiento: {procedimiento.codigo}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


# =============================================================================
# PROVEEDORES MANAGEMENT
# =============================================================================

@monitor_view
@access_check
@login_required
@permission_required('core.view_proveedor', raise_exception=True)
def listar_proveedores(request):
    """
    Lista todos los proveedores.
    Filtrados por empresa para usuarios no-superusuarios.
    """
    query = request.GET.get('q', '')
    proveedores = Proveedor.objects.all().select_related('empresa')

    # Filtrar por empresa si no es superusuario
    if not request.user.is_superuser and request.user.empresa:
        proveedores = proveedores.filter(empresa=request.user.empresa)
    elif not request.user.is_superuser and not request.user.empresa:
        proveedores = Proveedor.objects.none()

    # Aplicar filtro de búsqueda
    if query:
        proveedores = proveedores.filter(
            Q(nombre_empresa__icontains=query) |
            Q(nit__icontains=query) |
            Q(contacto__icontains=query) |
            Q(email__icontains=query)
        )

    proveedores = proveedores.order_by('nombre_empresa')

    # Paginación
    paginator = Paginator(proveedores, 20)
    page_number = request.GET.get('page')
    try:
        proveedores_page = paginator.page(page_number)
    except PageNotAnInteger:
        proveedores_page = paginator.page(1)
    except EmptyPage:
        proveedores_page = paginator.page(paginator.num_pages)

    context = {
        'proveedores': proveedores_page,
        'query': query,
        'titulo_pagina': 'Listado de Proveedores'
    }
    return render(request, 'core/listar_proveedores.html', context)


@monitor_view
@access_check
@login_required
@permission_required('core.add_proveedor', raise_exception=True)
def añadir_proveedor(request):
    """
    Añade un nuevo proveedor.
    """
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            try:
                proveedor = form.save(commit=False)

                # Asignar empresa automáticamente para usuarios no-superusuarios
                if not request.user.is_superuser and not proveedor.empresa:
                    proveedor.empresa = request.user.empresa

                proveedor.save()
                messages.success(request, 'Proveedor añadido exitosamente.')
                logger.info(f"Proveedor creado: {proveedor.nombre_empresa} por {request.user.username}")
                return redirect('core:listar_proveedores')
            except Exception as e:
                messages.error(request, f'Error al añadir proveedor: {e}')
                logger.error(f"Error añadiendo proveedor: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = ProveedorForm()

    context = {
        'form': form,
        'titulo_pagina': 'Añadir Nuevo Proveedor'
    }
    return render(request, 'core/añadir_proveedor.html', context)


@monitor_view
@access_check
@login_required
@permission_required('core.change_proveedor', raise_exception=True)
def editar_proveedor(request, pk):
    """
    Edita un proveedor existente.
    """
    proveedor = get_object_or_404(Proveedor, pk=pk)

    # Verificar permisos por empresa
    if not request.user.is_superuser and request.user.empresa != proveedor.empresa:
        messages.error(request, 'No tienes permiso para editar este proveedor.')
        return redirect('core:listar_proveedores')

    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Proveedor actualizado exitosamente.')
                logger.info(f"Proveedor actualizado: {proveedor.nombre_empresa} por {request.user.username}")
                return redirect('core:listar_proveedores')
            except Exception as e:
                messages.error(request, f'Error al actualizar proveedor: {e}')
                logger.error(f"Error actualizando proveedor {proveedor.pk}: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = ProveedorForm(instance=proveedor)

    context = {
        'form': form,
        'proveedor': proveedor,
        'titulo_pagina': f'Editar Proveedor: {proveedor.nombre_empresa}'
    }
    return render(request, 'core/editar_proveedor.html', context)


@monitor_view
@access_check
@login_required
@permission_required('core.delete_proveedor', raise_exception=True)
def eliminar_proveedor(request, pk):
    """
    Elimina un proveedor (con confirmación).
    """
    proveedor = get_object_or_404(Proveedor, pk=pk)

    # Verificar permisos por empresa
    if not request.user.is_superuser and request.user.empresa != proveedor.empresa:
        messages.error(request, 'No tienes permiso para eliminar este proveedor.')
        return redirect('core:listar_proveedores')

    if request.method == 'POST':
        try:
            nombre_proveedor = proveedor.nombre_empresa
            logger.info(f"Eliminando proveedor: {nombre_proveedor} por {request.user.username}")
            proveedor.delete()
            messages.success(request, f'Proveedor "{nombre_proveedor}" eliminado exitosamente.')
            return redirect('core:listar_proveedores')
        except Exception as e:
            messages.error(request, f'Error al eliminar proveedor: {e}')
            logger.error(f"Error eliminando proveedor {proveedor.pk}: {e}")
            return redirect('core:listar_proveedores')

    # Contexto para confirmación
    context = {
        'object_name': f'el proveedor "{proveedor.nombre_empresa}"',
        'return_url_name': 'core:listar_proveedores',
        'return_url_pk': None,
        'titulo_pagina': f'Eliminar Proveedor: {proveedor.nombre_empresa}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


@monitor_view
@access_check
@login_required
@permission_required('core.view_proveedor', raise_exception=True)
def detalle_proveedor(request, pk):
    """
    Muestra los detalles de un proveedor específico.
    """
    proveedor = get_object_or_404(Proveedor, pk=pk)

    # Verificar permisos por empresa
    if not request.user.is_superuser and request.user.empresa != proveedor.empresa:
        messages.error(request, 'No tienes permiso para ver este proveedor.')
        return redirect('core:listar_proveedores')

    # Obtener estadísticas del proveedor
    calibraciones_count = proveedor.calibraciones_realizadas.count()
    mantenimientos_count = proveedor.mantenimientos_realizados.count()
    comprobaciones_count = proveedor.comprobaciones_realizadas.count()

    context = {
        'proveedor': proveedor,
        'estadisticas': {
            'calibraciones': calibraciones_count,
            'mantenimientos': mantenimientos_count,
            'comprobaciones': comprobaciones_count,
            'total_servicios': calibraciones_count + mantenimientos_count + comprobaciones_count
        },
        'titulo_pagina': f'Detalle de Proveedor: {proveedor.nombre_empresa}'
    }
    return render(request, 'core/detalle_proveedor.html', context)


# =============================================================================
# UTILITY VIEWS
# =============================================================================

@monitor_view
@access_check
@login_required
def contact_us(request):
    """
    Renderiza la página de contacto.
    """
    return render(request, 'core/contact_us.html', {
        'titulo_pagina': 'Contacto'
    })


@monitor_view
@access_check
@login_required
def subir_pdf(request):
    """
    Vista para subir un archivo PDF y registrarlo en la base de datos.
    """
    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES, request=request)
        archivo_subido = request.FILES.get('archivo')

        if form.is_valid() and archivo_subido:
            try:
                nombre_archivo = archivo_subido.name
                ruta_s3 = f'pdfs/{nombre_archivo}'

                # Subir archivo usando función auxiliar
                subir_archivo(nombre_archivo, archivo_subido)

                # Guardar registro en base de datos
                documento = form.save(commit=False)
                documento.nombre_archivo = nombre_archivo
                documento.archivo_s3_path = ruta_s3
                documento.subido_por = request.user

                if not request.user.is_superuser and request.user.empresa:
                    documento.empresa = request.user.empresa

                documento.save()

                messages.success(request, f'Archivo "{nombre_archivo}" subido y registrado exitosamente.')
                logger.info(f"Documento subido: {nombre_archivo} por {request.user.username}")
                return redirect('core:home')

            except Exception as e:
                messages.error(request, f'Error al subir o registrar el archivo: {e}')
                logger.error(f'Error subiendo archivo {nombre_archivo}: {e}')
        else:
            messages.error(request, 'Por favor, corrige los errores del formulario y asegúrate de seleccionar un archivo.')
    else:
        form = DocumentoForm(request=request)

    context = {
        'form': form,
        'titulo_pagina': 'Subir Documento PDF'
    }
    return render(request, 'core/subir_pdf.html', context)


# ===== DIAGNÓSTICOS DEL SISTEMA =====

@monitor_view
@login_required
def cache_diagnostics(request):
    """Vista de diagnóstico temporal para verificar el estado del cache."""
    from django.db import connection
    from django.core.cache import cache
    from django.conf import settings
    from datetime import datetime

    # Verificar si la tabla existe
    table_exists = False
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables
                    WHERE table_name = 'sam_cache_table'
                );
            """)
            table_exists = cursor.fetchone()[0]
    except Exception:
        table_exists = False

    # Test de cache
    cache_works = False
    cache_error = None
    try:
        test_key = 'diagnostics_test_key'
        test_value = 'diagnostics_test_value_123'
        cache.set(test_key, test_value, 60)
        retrieved = cache.get(test_key)
        cache_works = (retrieved == test_value)
        if cache_works:
            cache.delete(test_key)
    except Exception as e:
        cache_error = str(e)

    # Información de configuración
    cache_config = settings.CACHES.get('default', {})
    cache_backend = cache_config.get('BACKEND', 'Unknown')
    cache_location = cache_config.get('LOCATION', 'Unknown')

    context = {
        'table_exists': table_exists,
        'cache_works': cache_works,
        'cache_error': cache_error,
        'cache_backend': cache_backend,
        'cache_location': cache_location,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return render(request, 'core/cache_diagnostics.html', context)


def _generate_general_proveedor_list_excel_content(proveedores_queryset):
    """
    Generates an Excel file with the general list of providers.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listado de Proveedores"

    # Add professional title header
    sheet.merge_cells('A1:I2')
    title_cell = sheet['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet.row_dimensions[row].height = 8

    headers = [
        "Nombre de la Empresa Proveedora", "Empresa Cliente", "Tipo de Servicio", "Nombre de Contacto",
        "Número de Contacto", "Correo Electrónico", "Página Web",
        "Alcance", "Servicio Prestado"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        sheet.column_dimensions[cell.column_letter].width = 25

    # Add provider data starting from row 7
    current_row = 7
    for proveedor in proveedores_queryset:
        row_data = [
            proveedor.nombre_empresa,
            proveedor.empresa.nombre if proveedor.empresa else "N/A",
            proveedor.get_tipo_servicio_display(),
            proveedor.nombre_contacto,
            proveedor.numero_contacto,
            proveedor.correo_electronico,
            proveedor.pagina_web,
            proveedor.alcance,
            proveedor.servicio_prestado,
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Ajustar anchos de columna evitando problemas con celdas fusionadas
    from openpyxl.utils import get_column_letter
    headers = [
        "Nombre de la Empresa Proveedora", "Empresa Cliente", "Tipo de Servicio", "Nombre de Contacto",
        "Número de Contacto", "Correo Electrónico", "Página Web",
        "Alcance", "Servicio Prestado"
    ]
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_procedimiento_info_excel_content(procedimientos_queryset):
    """
    Generates an Excel file with the general list of procedures.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listado de Procedimientos"

    # Add professional title header
    sheet.merge_cells('A1:G2')
    title_cell = sheet['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet.row_dimensions[row].height = 8

    headers = [
        "Nombre", "Código", "Versión", "Fecha de Emisión", "Empresa", "Observaciones", "Documento PDF"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        sheet.column_dimensions[cell.column_letter].width = 25

    # Add procedure data starting from row 7
    current_row = 7
    for proc in procedimientos_queryset:
        row_data = [
            proc.nombre,
            proc.codigo,
            proc.version,
            proc.fecha_emision.strftime('%Y-%m-%d') if proc.fecha_emision else '',
            proc.empresa.nombre if proc.empresa else "N/A",
            proc.observaciones,
            proc.documento_pdf.url if proc.documento_pdf else 'N/A',
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Ajustar anchos de columna evitando problemas con celdas fusionadas
    from openpyxl.utils import get_column_letter
    headers = [
        "Nombre", "Código", "Versión", "Fecha de Emisión", "Empresa", "Observaciones", "Documento PDF"
    ]
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()