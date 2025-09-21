# core/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.contrib import messages
from django.db.models import Q, Count, Prefetch
from django.db import models
from datetime import date, timedelta, datetime
import calendar
import io
import json
import os
import zipfile
from collections import defaultdict
import decimal
from functools import wraps # Importar wraps para decoradores

# IMPORTACIONES ADICIONALES PARA LA IMPORTACIÓN DE EXCEL
import openpyxl
from django.db import transaction
from django.utils.dateparse import parse_date

from django.http import HttpResponse, JsonResponse, Http404, HttpResponseRedirect
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage

from dateutil.relativedelta import relativedelta
from django.urls import reverse # Importar reverse

from django.conf import settings
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.utils.html import mark_safe

# Importar para PDF
from django.template.loader import get_template
from weasyprint import HTML

# Logging para reemplazar prints de debug
import logging
logger = logging.getLogger(__name__)

# Importar validadores de almacenamiento
from .storage_validators import StorageLimitValidator

# Importar los formularios de tu aplicación (asegúrate de que todos estos existan en .forms)
from .forms import (
    AuthenticationForm,
    CalibracionForm, MantenimientoForm, ComprobacionForm, EquipoForm,
    BajaEquipoForm, UbicacionForm, ProcedimientoForm, ProveedorForm,
    ExcelUploadForm,
    CustomUserCreationForm, CustomUserChangeForm, UserProfileForm, EmpresaForm, EmpresaFormatoForm,
    DocumentoForm # Asegúrate de que DocumentoForm esté importado aquí
)
# Importar modelos
from .models import (
    Equipo, Calibracion, Mantenimiento, Comprobacion, BajaEquipo, Empresa,
    CustomUser, Ubicacion, Procedimiento, Proveedor, Documento # ASEGÚRATE de que Documento esté importado aquí
)

# Importar para autenticación y grupos
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import Group, Permission # Importar los modelos Group y Permission
from django.core.exceptions import ValidationError

# Importar para default_storage (manejo de archivos S3/local)
from django.core.files.storage import default_storage
from storages.backends.s3boto3 import S3Boto3Storage
# Importar el módulo forms de Django para las excepciones
from django import forms

# =============================================================================
# IMPORTAR SERVICIOS MEJORADOS Y UTILIDADES DE SEGURIDAD
# =============================================================================

from .services_new import file_upload_service, equipment_service, cache_manager
from .security import StorageQuotaManager
from .templatetags.file_tags import secure_file_url, pdf_image_url

# Configurar logger
import logging
logger = logging.getLogger('core')

# =============================================================================
# UTILIDADES MEJORADAS PARA GESTIÓN DE ARCHIVOS E IMÁGENES
# =============================================================================

def get_secure_file_url(file_field, expire_seconds=3600):
    """Obtiene URL segura para archivos, manejando tanto local como S3"""
    if not file_field:
        return None

    try:
        # Si es un campo de archivo de Django
        if hasattr(file_field, 'url'):
            if hasattr(default_storage, 'url'):
                # Para S3 o storage personalizado
                return default_storage.url(file_field.name, expire=expire_seconds)
            else:
                # Para almacenamiento local
                return file_field.url
        return None
    except Exception as e:
        logger.error(f"Error obteniendo URL de archivo: {str(e)}")
        return None

def get_empresa_logo_url(empresa, expire_seconds=3600):
    """Obtiene URL del logo de empresa de forma segura"""
    if not empresa or not empresa.logo_empresa:
        return None
    return get_secure_file_url(empresa.logo_empresa, expire_seconds)

def get_equipo_imagen_url(equipo, expire_seconds=3600):
    """Obtiene URL de imagen de equipo de forma segura"""
    if not equipo or not equipo.imagen_equipo:
        return None
    return get_secure_file_url(equipo.imagen_equipo, expire_seconds)


# =============================================================================
# Decoradores personalizados
# =============================================================================

def trial_check(view_func):
    """
    Decorador que verifica si la empresa del usuario tiene acceso a funciones avanzadas
    y bloquea ciertas acciones si no tiene un plan activo.
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        # Los superusuarios siempre tienen acceso
        if request.user.is_superuser:
            return view_func(request, *args, **kwargs)

        if request.user.is_authenticated and hasattr(request.user, 'empresa'):
            empresa = request.user.empresa

            # Verificar estado de la suscripción usando el sistema unificado
            estado_plan = empresa.get_estado_suscripcion_display()

            # Si el plan está expirado, bloquear ciertas acciones
            if empresa and estado_plan in ["Plan Expirado", "Período de Prueba Expirado"]:
                # Lista de vistas que requieren suscripción activa
                protected_views = [
                    'añadir_equipo', 'editar_equipo', 'añadir_calibracion', 'añadir_mantenimiento',
                    'añadir_comprobacion', 'generar_informe_zip', 'solicitar_zip',
                    'importar_equipos_excel', 'exportar_equipos_excel'
                ]

                # Verificar si es una vista protegida
                view_name = view_func.__name__
                if view_name in protected_views:
                    dias_restantes = empresa.get_dias_restantes_plan()

                    if estado_plan == "Período de Prueba Expirado":
                        mensaje = f'Su período de prueba ha expirado. Para continuar usando SAM Metrología, contacte con soporte para activar su suscripción.'
                    else:
                        mensaje = f'Su plan ha expirado. Para continuar usando estas funciones, contacte con soporte para renovar su suscripción.'

                    messages.error(request, mensaje)
                    return redirect('core:dashboard')

        return view_func(request, *args, **kwargs)
    return wrapper

def access_check(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            # Si no está autenticado, redirigir a login
            return redirect(settings.LOGIN_URL)

        # Si es superusuario, siempre permitir el acceso (superusuario puede ver todo)
        if request.user.is_superuser:
            return view_func(request, *args, **kwargs)

        # Para usuarios normales, verificar estado de usuario y empresa
        user_empresa = request.user.empresa
        
        # 1. Verificar si el usuario mismo está activo
        if not request.user.is_active:
            messages.error(request, 'Tu cuenta de usuario ha sido desactivada. Contacta al administrador.')
            return redirect('core:access_denied')

        # 2. Verificar el estado de la empresa si el usuario tiene una asociada
        if user_empresa:
            # Usar el método actualizado del modelo Empresa para verificar el estado de la suscripción
            # Consideramos que el estado 'Activo' o 'Período de Prueba' permite el acceso
            # Otros estados como 'Expirado' o 'Cancelado' deniegan el acceso
            estado_suscripcion_empresa = user_empresa.get_estado_suscripcion_display()
            
            # Acceso manual anula otras restricciones, si está activo
            if user_empresa.acceso_manual_activo:
                return view_func(request, *args, **kwargs)

            # Denegar acceso si el plan está expirado o ha sido cancelado
            if "Expirado" in estado_suscripcion_empresa or user_empresa.estado_suscripcion == 'Cancelado':
                messages.error(request, f'El acceso para tu empresa ({user_empresa.nombre}) ha expirado o ha sido cancelado. Contacta al administrador.')
                return redirect('core:access_denied')
            
            # Si pasa las verificaciones, continuar con la vista
            return view_func(request, *args, **kwargs)
        else:
            # Si el usuario normal no tiene empresa asignada, denegar acceso
            messages.warning(request, 'Tu cuenta no está asociada a ninguna empresa. Contacta al administrador.')
            return redirect('core:access_denied')
    return _wrapped_view

# =============================================================================
# Funciones de utilidad (helpers)
# =============================================================================

def es_miembro_empresa(user, empresa_id):
    """Verifica si el usuario pertenece a la empresa especificada."""
    return user.is_authenticated and user.empresa and user.empresa.pk == empresa_id

def sanitize_filename(filename):
    """
    Sanitiza el nombre del archivo para prevenir path traversal y caracteres peligrosos.
    """
    import re
    import os
    
    # Obtener solo el nombre del archivo, sin ruta
    filename = os.path.basename(filename)
    
    # Reemplazar caracteres peligrosos
    filename = filename.replace('..', '_')
    filename = filename.replace('/', '_')
    filename = filename.replace('\\', '_')
    filename = filename.replace(':', '_')
    
    # Remover caracteres no alfanuméricos excepto puntos, guiones y guiones bajos
    filename = re.sub(r'[^\w\-_\.]', '_', filename)
    
    # Limitar longitud
    if len(filename) > 100:
        name, ext = os.path.splitext(filename)
        filename = name[:95] + ext
    
    return filename

def subir_archivo(nombre_archivo, contenido):
    """
    Sube un archivo a AWS S3 usando el almacenamiento configurado en Django.
    :param nombre_archivo: Nombre con el que se guardará el archivo en S3.
    :param contenido: El objeto de archivo (ej. InMemoryUploadedFile de request.FILES).
    """
    # Sanitizar el nombre del archivo
    nombre_archivo_seguro = sanitize_filename(nombre_archivo)
    
    storage = default_storage
    ruta_s3 = f'pdfs/{nombre_archivo_seguro}'
    storage.save(ruta_s3, contenido)
    logger.info(f'Archivo subido a: {ruta_s3}')

# --- Función auxiliar para proyectar actividades y categorizarlas (para las gráficas de torta) ---
def get_projected_activities_for_year(equipment_queryset, activity_type, current_year, today):
    """
    Generates a list of projected activities for the current year for a given activity type.
    Each projected activity will have a 'date' and 'status' (realized, overdue, pending).
    This function is primarily for the annual summary (pie charts).
    Applies to Calibracion and Comprobacion.
    
    Excludes equipment that is 'De Baja' or 'Inactivo'.
    """
    projected_activities = []
    
    # Filtrar equipos para excluir los que están "De Baja" o "Inactivo" de las proyecciones
    # APLICACIÓN CLAVE: Excluir equipos inactivos o de baja de las proyecciones de las gráficas de línea.
    equipment_queryset = equipment_queryset.exclude(estado__in=['De Baja', 'Inactivo'])

    # Fetch all realized activities for the current year for quick lookup
    realized_activities_for_year = []
    if activity_type == 'calibracion':
        realized_activities_for_year = Calibracion.objects.filter(
            equipo__in=equipment_queryset, # Filtrar por el queryset ya depurado
            fecha_calibracion__year=current_year
        ).values_list('equipo_id', 'fecha_calibracion')
        freq_attr = 'frecuencia_calibracion_meses'
    elif activity_type == 'comprobacion':
        realized_activities_for_year = Comprobacion.objects.filter(
            equipo__in=equipment_queryset, # Filtrar por el queryset ya depurado
            fecha_comprobacion__year=current_year
        ).values_list('equipo_id', 'fecha_comprobacion')
        freq_attr = 'frecuencia_comprobacion_meses'
    else:
        return []

    # Create a set of (equipo_id, year, month) for realized activities for quick lookup
    realized_set = set()
    for eq_id, date_obj in realized_activities_for_year:
        realized_set.add((eq_id, date_obj.year, date_obj.month))

    for equipo in equipment_queryset: # Iterar sobre los equipos ya filtrados
        freq_months = getattr(equipo, freq_attr)

        if freq_months is None or freq_months <= 0:
            continue

        # Determine the effective start date for planning this activity type for this equipment
        plan_start_date = equipo.fecha_adquisicion if equipo.fecha_adquisicion else \
                          (equipo.fecha_registro.date() if equipo.fecha_registro else date(current_year, 1, 1))

        # Calculate the first projected date *relevant to the current year* based on the frequency
        delta_years = current_year - plan_start_date.year
        delta_months = (delta_years * 12) + (1 - plan_start_date.month)

        num_intervals_to_reach_year = 0
        if freq_months > 0:
            num_intervals_to_reach_year = max(0, (delta_months + freq_months - 1) // freq_months)
        
        current_projection_date = plan_start_date + relativedelta(months=int(num_intervals_to_reach_year * freq_months))


        # Now, project activities for the entire current year
        for _ in range(int(12 / freq_months) + 2 if freq_months > 0 else 0): # Project slightly beyond 12 months to catch edge cases
            if current_projection_date.year == current_year:
                # Check if this specific projected activity (for this equipment, in this year and month) was realized
                is_realized = (equipo.id, current_projection_date.year, current_projection_date.month) in realized_set

                status = ''
                if is_realized:
                    status = 'Realizado'
                elif current_projection_date < today:
                    status = 'No Cumplido' # Overdue and not realized
                else:
                    status = 'Pendiente/Programado' # Future and not realized

                projected_activities.append({
                    'equipo_id': equipo.id,
                    'date': current_projection_date,
                    'status': status
                })
            elif current_projection_date.year > current_year:
                break # Stop projecting if we've gone past the current year

            try:
                current_projection_date += relativedelta(months=int(freq_months))
            except OverflowError: # Prevent extremely large dates causing errors
                break
    return projected_activities

def get_projected_maintenance_compliance_for_year(equipment_queryset, current_year, today):
    """
    Generates a list of projected maintenance activities for the current year,
    categorized by compliance status (realized, overdue, pending).
    This specifically targets 'Preventivo' and 'Predictivo' maintenance as they are typically scheduled.
    
    Excludes equipment that is 'De Baja' or 'Inactivo'.
    """
    projected_activities = []
    
    # Filtrar equipos para excluir los que están "De Baja" o "Inactivo" de las proyecciones
    # APLICACIÓN CLAVE: Excluir equipos inactivos o de baja de las proyecciones de las gráficas de línea.
    equipment_queryset = equipment_queryset.exclude(estado__in=['De Baja', 'Inactivo'])

    # Fetch all realized *scheduled* maintenance activities for the current year for quick lookup
    # Only consider 'Preventivo' and 'Predictivo' as scheduled
    realized_scheduled_maintenance_for_year = Mantenimiento.objects.filter(
        equipo__in=equipment_queryset, # Filtrar por el queryset ya depurado
        fecha_mantenimiento__year=current_year,
        tipo_mantenimiento__in=['Preventivo', 'Predictivo']
    ).values_list('equipo_id', 'fecha_mantenimiento')
    
    realized_set = set()
    for eq_id, date_obj in realized_scheduled_maintenance_for_year:
        realized_set.add((eq_id, date_obj.year, date_obj.month))

    for equipo in equipment_queryset: # Iterar sobre los equipos ya filtrados
        freq_months = equipo.frecuencia_mantenimiento_meses

        if freq_months is None or freq_months <= 0:
            continue

        # Determine the effective start date for planning this activity type for this equipment
        plan_start_date = equipo.fecha_adquisicion if equipo.fecha_adquisicion else \
                          (equipo.fecha_registro.date() if equipo.fecha_registro else date(current_year, 1, 1))

        # Calculate the first projected date *relevant to the current year* based on the frequency
        delta_years = current_year - plan_start_date.year
        delta_months = (delta_years * 12) + (1 - plan_start_date.month)

        num_intervals_to_reach_year = 0
        if freq_months > 0:
            num_intervals_to_reach_year = max(0, (delta_months + freq_months - 1) // freq_months)
        
        current_projection_date = plan_start_date + relativedelta(months=int(num_intervals_to_reach_year * freq_months))

        # Now, project activities for the entire current year
        for _ in range(int(12 / freq_months) + 2 if freq_months > 0 else 0):
            if current_projection_date.year == current_year:
                is_realized = (equipo.id, current_projection_date.year, current_projection_date.month) in realized_set

                status = ''
                if is_realized:
                    status = 'Realizado'
                elif current_projection_date < today:
                    status = 'No Cumplido' # Overdue and not realized
                else:
                    status = 'Pendiente/Programado' # Future and not realized

                projected_activities.append({
                    'equipo_id': equipo.id,
                    'date': current_projection_date,
                    'status': status
                })
            elif current_projection_date.year > current_year:
                break
            try:
                current_projection_date += relativedelta(months=int(freq_months))
            except OverflowError:
                break
    return projected_activities


# --- Funciones Auxiliares para Generación de PDF (se mantienen para Hoja de Vida y Listado General) ---

def _generate_pdf_content(request, template_path, context):
    """
    Generates PDF content (bytes) from a template and context using WeasyPrint.
    """
    from django.template.loader import get_template
    import logging
    logger = logging.getLogger(__name__)

    try:
        template = get_template(template_path)
        html_string = template.render(context)

        # base_url es crucial para que WeasyPrint resuelva rutas de CSS e imágenes
        # Mejorar el manejo del base_url para mock requests
        try:
            base_url = request.build_absolute_uri('/')
        except Exception as e:
            logger.warning(f"Error building absolute URI: {e}, using fallback")
            # Fallback para mock requests
            if hasattr(request, 'META') and 'HTTP_HOST' in request.META:
                scheme = request.META.get('wsgi.url_scheme', 'https')
                host = request.META.get('HTTP_HOST', 'sam-9o6o.onrender.com')
                base_url = f"{scheme}://{host}/"
            else:
                base_url = "https://sam-9o6o.onrender.com/"

        logger.info(f"Generating PDF with base_url: {base_url}")
        pdf_file = HTML(string=html_string, base_url=base_url).write_pdf()

        return pdf_file

    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}", exc_info=True)
        raise

def _generate_equipment_hoja_vida_pdf_content(request, equipo):
    """
    Genera el contenido PDF para la "Hoja de Vida" de un equipo con URLs seguras.
    """
    try:
        # Optimizar consultas con select_related y prefetch_related
        calibraciones = equipo.calibraciones.select_related('proveedor').order_by('-fecha_calibracion')
        mantenimientos = equipo.mantenimientos.select_related('proveedor').order_by('-fecha_mantenimiento')
        comprobaciones = equipo.comprobaciones.select_related('proveedor').order_by('-fecha_comprobacion')

        baja_registro = None
        try:
            baja_registro = equipo.baja_registro
        except BajaEquipo.DoesNotExist:
            pass

        # Función helper mejorada para obtener URLs de archivos para PDF
        def get_pdf_file_url(file_field):
            """Obtiene URL absoluta para archivos en PDF"""
            if not file_field or not file_field.name:
                return None

            try:
                storage_name = default_storage.__class__.__name__

                if 'S3' in storage_name or hasattr(default_storage, 'bucket'):
                    # Para S3, verificar que el archivo existe antes de generar URL
                    try:
                        if not default_storage.exists(file_field.name):
                            logger.warning(f"Archivo no existe en S3: {file_field.name}")
                            return None
                    except Exception:
                        pass  # Si falla la verificación, intentar generar URL anyway

                    # Para S3, usar URL con mayor tiempo de expiración para PDFs
                    try:
                        url = default_storage.url(file_field.name, expire=7200)  # 2 horas
                    except TypeError:
                        # Fallback si el storage no soporta expire
                        url = default_storage.url(file_field.name)

                    # Asegurar que sea URL absoluta
                    if url.startswith('//'):
                        url = 'https:' + url

                    return url
                else:
                    # Para almacenamiento local (FileSystemStorage)
                    url = default_storage.url(file_field.name)
                    # Convertir a URL absoluta
                    if url.startswith('/'):
                        url = request.build_absolute_uri(url)

                    return url
            except Exception as e:
                logger.error(f"Error obteniendo URL para PDF: {file_field.name}, error: {str(e)}")
                return None

        # Función especial para imágenes en PDF - convierte a base64 si es necesario
        def get_pdf_image_data(file_field):
            """Obtiene datos de imagen para PDF, convirtiendo a base64 si es necesario"""
            if not file_field or not file_field.name:
                return None

            try:
                # Verificar si es una imagen
                file_extension = file_field.name.lower().split('.')[-1]
                if file_extension not in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
                    return None

                # Intentar obtener el contenido del archivo
                try:
                    file_content = default_storage.open(file_field.name).read()

                    # Convertir a base64
                    import base64
                    base64_encoded = base64.b64encode(file_content).decode('utf-8')

                    # Determinar el tipo MIME
                    mime_type = f"image/{file_extension}" if file_extension != 'jpg' else 'image/jpeg'

                    # Retornar como data URL
                    return f"data:{mime_type};base64,{base64_encoded}"

                except Exception as e:
                    logger.warning(f"No se pudo convertir imagen a base64: {file_field.name}, error: {str(e)}")
                    # Fallback a URL normal
                    return get_pdf_file_url(file_field)

            except Exception as e:
                logger.error(f"Error procesando imagen para PDF: {file_field.name}, error: {str(e)}")
                return None

        # Obtener URLs de archivos con el nuevo sistema seguro
        # Para imágenes, usar la función especializada que convierte a base64
        logo_empresa_url = get_pdf_image_data(equipo.empresa.logo_empresa) if equipo.empresa and equipo.empresa.logo_empresa else None
        imagen_equipo_url = get_pdf_image_data(equipo.imagen_equipo) if equipo.imagen_equipo else None
        documento_baja_url = get_pdf_file_url(baja_registro.documento_baja) if baja_registro and baja_registro.documento_baja else None

        for cal in calibraciones:
            cal.documento_calibracion_url = get_pdf_file_url(cal.documento_calibracion)
            cal.confirmacion_metrologica_pdf_url = get_pdf_file_url(cal.confirmacion_metrologica_pdf)
            cal.intervalos_calibracion_pdf_url = get_pdf_file_url(cal.intervalos_calibracion_pdf)

        for mant in mantenimientos:
            mant.documento_mantenimiento_url = get_pdf_file_url(mant.documento_mantenimiento)
        for comp in comprobaciones:
            comp.documento_comprobacion_url = get_pdf_file_url(comp.documento_comprobacion)

        archivo_compra_pdf_url = get_pdf_file_url(equipo.archivo_compra_pdf)
        ficha_tecnica_pdf_url = get_pdf_file_url(equipo.ficha_tecnica_pdf)
        manual_pdf_url = get_pdf_file_url(equipo.manual_pdf)
        otros_documentos_pdf_url = get_pdf_file_url(equipo.otros_documentos_pdf)

        context = {
            'equipo': equipo,
            'calibraciones': calibraciones,
            'mantenimientos': mantenimientos,
            'comprobaciones': comprobaciones,
            'baja_registro': baja_registro,
            'logo_empresa_url': logo_empresa_url,
            'imagen_equipo_url': imagen_equipo_url,
            'documento_baja_url': documento_baja_url,
            'archivo_compra_pdf_url': archivo_compra_pdf_url,
            'ficha_tecnica_pdf_url': ficha_tecnica_pdf_url,
            'manual_pdf_url': manual_pdf_url,
            'otros_documentos_pdf_url': otros_documentos_pdf_url,
            'titulo_pagina': f'Hoja de Vida de {equipo.nombre}',
        }
        return _generate_pdf_content(request, 'core/hoja_vida_pdf.html', context)

    except Exception as e:
        logger.error(f"Error generando contenido PDF para hoja de vida del equipo {equipo.codigo_interno}: {str(e)}")
        # Return minimal context in case of error
        context = {
            'equipo': equipo,
            'calibraciones': [],
            'mantenimientos': [],
            'comprobaciones': [],
            'baja_registro': None,
            'titulo_pagina': f'Hoja de Vida de {equipo.nombre}',
        }
        return _generate_pdf_content(request, 'core/hoja_vida_pdf.html', context)


# --- Funciones Auxiliares para Generación de Excel ---

def _generate_dashboard_excel_content(equipos_queryset, empresa):
    """
    Genera un Excel profesional para el dashboard con formato mejorado y secciones completas.
    Usado cuando se descarga desde el botón del dashboard.
    """
    from collections import Counter
    from datetime import datetime, timedelta
    from openpyxl.styles import Alignment

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Dashboard"

    # ==============================================
    # ENCABEZADO PROFESIONAL SAM METROLOGÍA
    # ==============================================

    # Título principal profesional
    sheet.merge_cells('A1:F2')
    sheet['A1'] = 'INFORMES GENERADOS POR SAM METROLOGÍA SAS'
    sheet['A1'].font = Font(bold=True, size=20, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Información de la empresa y fecha
    sheet.merge_cells('A3:F3')
    sheet['A3'] = f'EMPRESA: {empresa.nombre.upper()}'
    sheet['A3'].font = Font(bold=True, size=14, color="1f4e79")
    sheet['A3'].alignment = Alignment(horizontal="center")

    sheet.merge_cells('A4:F4')
    hoy = datetime.now()
    sheet['A4'] = f'Generado el: {hoy.strftime("%d de %B de %Y a las %H:%M")}'
    sheet['A4'].font = Font(bold=True, size=12)
    sheet['A4'].alignment = Alignment(horizontal="center")

    # ==============================================
    # DATOS ESTADÍSTICOS
    # ==============================================

    equipos_list = list(equipos_queryset)
    total_equipos = len(equipos_list)
    equipos_activos = sum(1 for eq in equipos_list if eq.estado == 'Activo')
    equipos_inactivos = sum(1 for eq in equipos_list if eq.estado == 'Inactivo')
    equipos_baja = sum(1 for eq in equipos_list if eq.estado == 'De Baja')

    # Estadísticas por tipo
    tipos_count = Counter(eq.get_tipo_equipo_display() for eq in equipos_list)

    # Fechas para análisis
    hoy_date = hoy.date()
    treinta_dias = hoy_date + timedelta(days=30)
    ano_actual = hoy_date.year

    # ==============================================
    # SECCIÓN: RESUMEN GENERAL
    # ==============================================

    row = 6
    sheet[f'A{row}'] = '📊 RESUMEN GENERAL DE EQUIPOS'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    sheet.merge_cells(f'A{row}:F{row}')
    row += 2

    # Tabla de resumen con formato profesional
    headers = ['Categoría', 'Cantidad', 'Porcentaje']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    resumen_data = [
        ('Total de Equipos', total_equipos, '100%'),
        ('Equipos Activos', equipos_activos, f'{(equipos_activos/total_equipos*100):.1f}%' if total_equipos > 0 else '0%'),
        ('Equipos Inactivos', equipos_inactivos, f'{(equipos_inactivos/total_equipos*100):.1f}%' if total_equipos > 0 else '0%'),
        ('Equipos de Baja', equipos_baja, f'{(equipos_baja/total_equipos*100):.1f}%' if total_equipos > 0 else '0%')
    ]

    for categoria, cantidad, porcentaje in resumen_data:
        sheet.cell(row=row, column=1, value=categoria)
        sheet.cell(row=row, column=2, value=cantidad)
        sheet.cell(row=row, column=3, value=porcentaje)
        row += 1

    # ==============================================
    # SECCIÓN: CÓDIGOS INTERNOS DE EQUIPOS
    # ==============================================

    row += 2
    sheet[f'A{row}'] = '🔧 CÓDIGOS INTERNOS DE EQUIPOS'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    sheet.merge_cells(f'A{row}:F{row}')
    row += 2

    # Headers para códigos
    codigo_headers = ['Código Interno', 'Nombre del Equipo', 'Estado', 'Tipo']
    for col, header in enumerate(codigo_headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Datos de equipos (limitado a primeros 15 para no sobrecargar)
    for equipo in equipos_list[:15]:
        sheet.cell(row=row, column=1, value=equipo.codigo_interno)
        sheet.cell(row=row, column=2, value=equipo.nombre)
        sheet.cell(row=row, column=3, value=equipo.estado)
        sheet.cell(row=row, column=4, value=equipo.get_tipo_equipo_display())
        row += 1

    if len(equipos_list) > 15:
        sheet.cell(row=row, column=1, value=f"... y {len(equipos_list) - 15} equipos más")
        sheet[f'A{row}'].font = Font(italic=True, color="666666")
        row += 1

    # ==============================================
    # SECCIÓN: ACTIVIDADES PROGRAMADAS
    # ==============================================

    row += 2
    sheet[f'A{row}'] = '📅 ACTIVIDADES PROGRAMADAS Y REALIZADAS'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    sheet.merge_cells(f'A{row}:F{row}')
    row += 2

    # Calcular estadísticas de actividades
    from core.models import Calibracion, Mantenimiento, Comprobacion

    # Calibraciones
    cal_programadas = Calibracion.objects.filter(
        equipo__empresa=empresa,
        fecha_calibracion__year=ano_actual
    ).count()
    cal_vencidas = sum(1 for eq in equipos_list
                      if eq.proxima_calibracion and eq.proxima_calibracion < hoy_date)
    cal_proximas = sum(1 for eq in equipos_list
                      if eq.proxima_calibracion and hoy_date <= eq.proxima_calibracion <= treinta_dias)

    # Mantenimientos
    mant_programados = Mantenimiento.objects.filter(
        equipo__empresa=empresa,
        fecha_mantenimiento__year=ano_actual
    ).count()
    mant_vencidos = sum(1 for eq in equipos_list
                       if eq.proximo_mantenimiento and eq.proximo_mantenimiento < hoy_date)
    mant_proximos = sum(1 for eq in equipos_list
                       if eq.proximo_mantenimiento and hoy_date <= eq.proximo_mantenimiento <= treinta_dias)

    # Comprobaciones
    comp_programadas = Comprobacion.objects.filter(
        equipo__empresa=empresa,
        fecha_comprobacion__year=ano_actual
    ).count()
    comp_vencidas = sum(1 for eq in equipos_list
                       if eq.proxima_comprobacion and eq.proxima_comprobacion < hoy_date)
    comp_proximas = sum(1 for eq in equipos_list
                       if eq.proxima_comprobacion and hoy_date <= eq.proxima_comprobacion <= treinta_dias)

    # Tabla de actividades
    act_headers = ['Tipo de Actividad', f'Realizadas {ano_actual}', 'Vencidas', 'Próximas (30 días)']
    for col, header in enumerate(act_headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    actividades_data = [
        ('Calibraciones', cal_programadas, cal_vencidas, cal_proximas),
        ('Mantenimientos', mant_programados, mant_vencidos, mant_proximos),
        ('Comprobaciones', comp_programadas, comp_vencidas, comp_proximas)
    ]

    for tipo, realizadas, vencidas, proximas in actividades_data:
        sheet.cell(row=row, column=1, value=tipo)
        sheet.cell(row=row, column=2, value=realizadas)
        sheet.cell(row=row, column=3, value=vencidas)
        sheet.cell(row=row, column=4, value=proximas)
        row += 1

    # ==============================================
    # SECCIÓN: ACTIVIDADES NO PROGRAMADAS
    # ==============================================

    row += 2
    sheet[f'A{row}'] = '⚠️ ACTIVIDADES NO PROGRAMADAS (CORRECTIVAS)'
    sheet[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
    sheet.merge_cells(f'A{row}:F{row}')
    row += 2

    # Mantenimientos correctivos (no programados)
    mantenimientos_correctivos = Mantenimiento.objects.filter(
        equipo__empresa=empresa,
        tipo_mantenimiento='Correctivo',
        fecha_mantenimiento__year=ano_actual
    )

    if mantenimientos_correctivos.exists():
        # Headers para correctivos
        correctivos_headers = ['Equipo', 'Fecha', 'Descripción', 'Responsable']
        for col, header in enumerate(correctivos_headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for mant in mantenimientos_correctivos[:10]:  # Limitar a 10
            sheet.cell(row=row, column=1, value=mant.equipo.codigo_interno)
            sheet.cell(row=row, column=2, value=mant.fecha_mantenimiento.strftime("%d/%m/%Y"))
            sheet.cell(row=row, column=3, value=mant.descripcion or "Mantenimiento correctivo")
            sheet.cell(row=row, column=4, value=mant.responsable or "No especificado")
            row += 1
    else:
        sheet.cell(row=row, column=1, value="✅ No se registraron actividades correctivas este año")
        sheet[f'A{row}'].font = Font(bold=True, color="27AE60")
        row += 1

    # Ajustar anchos de columnas
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        sheet.column_dimensions[col].width = 20

    # Ajustar altura de filas del encabezado
    sheet.row_dimensions[1].height = 40
    sheet.row_dimensions[2].height = 25

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_consolidated_excel_content(equipos_queryset, proveedores_queryset, procedimientos_queryset):
    """
    Genera un Excel consolidado con 4 hojas: Equipos, Proveedores, Procedimientos y Reporte Dashboard.
    Usado para el ZIP completo con todas las funcionalidades.
    """
    from django.utils import timezone
    from openpyxl.styles import Alignment
    workbook = Workbook()

    # === HOJA 1: EQUIPOS ===
    sheet_equipos = workbook.active
    sheet_equipos.title = "Equipos"

    # Agregar título profesional a hoja de Equipos
    sheet_equipos.merge_cells('A1:AB2')
    title_cell = sheet_equipos['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Información de generación
    sheet_equipos['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_equipos['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet_equipos['A3'].alignment = Alignment(horizontal="left")

    headers_equipos = [
        "Código Interno", "Nombre", "Empresa", "Tipo de Equipo", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado", "Fecha de Adquisición",
        "Rango de Medida", "Resolución", "Error Máximo Permisible", "Fecha de Registro",
        "Observaciones", "Fecha Última Calibración", "Próxima Calibración",
        "Frecuencia Calibración (meses)", "Fecha Último Mantenimiento", "Próximo Mantenimiento",
        "Frecuencia Mantenimiento (meses)", "Fecha Última Comprobación",
        "Próxima Comprobación", "Frecuencia Comprobación (meses)"
    ]

    # Agregar headers en fila 5
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, header_text in enumerate(headers_equipos, 1):
        cell = sheet_equipos.cell(row=5, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos de equipos empezando desde la fila 6
    current_row = 6
    for equipo in equipos_queryset:
        row_data = [
            equipo.codigo_interno, equipo.nombre, equipo.empresa.nombre, equipo.tipo_equipo,
            equipo.marca, equipo.modelo, equipo.numero_serie,
            equipo.ubicacion or "", equipo.responsable, equipo.estado,
            equipo.fecha_adquisicion, equipo.rango_medida, equipo.resolucion,
            equipo.error_maximo_permisible, equipo.fecha_registro.replace(tzinfo=None) if equipo.fecha_registro else None, equipo.observaciones,
            equipo.fecha_ultima_calibracion, equipo.proxima_calibracion, equipo.frecuencia_calibracion_meses,
            equipo.fecha_ultimo_mantenimiento, equipo.proximo_mantenimiento, equipo.frecuencia_mantenimiento_meses,
            equipo.fecha_ultima_comprobacion, equipo.proxima_comprobacion, equipo.frecuencia_comprobacion_meses
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_equipos.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # === HOJA 2: PROVEEDORES ===
    sheet_proveedores = workbook.create_sheet(title="Proveedores")

    # Agregar título profesional a hoja de Proveedores
    sheet_proveedores.merge_cells('A1:H2')
    title_cell = sheet_proveedores['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Información de generación
    sheet_proveedores['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_proveedores['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet_proveedores['A3'].alignment = Alignment(horizontal="left")

    headers_proveedores = [
        "Nombre Empresa", "Nombre Contacto", "Correo Electrónico", "Número Contacto",
        "Tipo Servicio", "Alcance", "Servicio Prestado", "Página Web"
    ]

    # Agregar headers en fila 5
    for col_num, header_text in enumerate(headers_proveedores, 1):
        cell = sheet_proveedores.cell(row=5, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos de proveedores empezando desde la fila 6
    current_row = 6
    for proveedor in proveedores_queryset:
        row_data = [
            proveedor.nombre_empresa, proveedor.nombre_contacto, proveedor.correo_electronico,
            proveedor.numero_contacto, proveedor.tipo_servicio, proveedor.alcance, proveedor.servicio_prestado,
            proveedor.pagina_web
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_proveedores.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # === HOJA 3: PROCEDIMIENTOS ===
    sheet_procedimientos = workbook.create_sheet(title="Procedimientos")

    # Agregar título profesional a hoja de Procedimientos
    sheet_procedimientos.merge_cells('A1:E2')
    title_cell = sheet_procedimientos['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Información de generación
    sheet_procedimientos['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_procedimientos['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet_procedimientos['A3'].alignment = Alignment(horizontal="left")

    headers_procedimientos = [
        "Código", "Nombre", "Observaciones", "Versión", "Fecha de Emisión"
    ]

    # Agregar headers en fila 5
    for col_num, header_text in enumerate(headers_procedimientos, 1):
        cell = sheet_procedimientos.cell(row=5, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos de procedimientos empezando desde la fila 6
    current_row = 6
    for procedimiento in procedimientos_queryset:
        row_data = [
            procedimiento.codigo, procedimiento.nombre, procedimiento.observaciones,
            procedimiento.version, procedimiento.fecha_emision
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_procedimientos.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # === HOJA 4: DASHBOARD DETALLADO ===
    sheet_dashboard = workbook.create_sheet(title="Dashboard")

    # Agregar título profesional a hoja de Dashboard
    sheet_dashboard.merge_cells('A1:F2')
    title_cell = sheet_dashboard['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Información de generación
    sheet_dashboard['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_dashboard['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet_dashboard['A3'].alignment = Alignment(horizontal="left")

    from datetime import date, timedelta
    today = date.today()

    # Configurar estilos
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    row = 5

    # === ESTADÍSTICAS GENERALES ===
    sheet_dashboard.merge_cells(f'A{row}:F{row}')
    cell = sheet_dashboard[f'A{row}']
    cell.value = "📊 ESTADÍSTICAS GENERALES"
    cell.font = title_font
    cell.fill = title_fill
    row += 2

    # Estadísticas de equipos por estado
    stats = {}
    for equipo in equipos_queryset:
        estado = equipo.estado
        stats[estado] = stats.get(estado, 0) + 1

    sheet_dashboard[f'A{row}'] = "Estado"
    sheet_dashboard[f'B{row}'] = "Cantidad"
    for cell in [sheet_dashboard[f'A{row}'], sheet_dashboard[f'B{row}']]:
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    for estado, cantidad in stats.items():
        sheet_dashboard[f'A{row}'] = estado
        sheet_dashboard[f'B{row}'] = cantidad
        row += 1

    row += 2

    # === EQUIPOS INACTIVOS ===
    equipos_inactivos = [e for e in equipos_queryset if e.estado == 'Inactivo']
    sheet_dashboard.merge_cells(f'A{row}:D{row}')
    cell = sheet_dashboard[f'A{row}']
    cell.value = f"⚠️ EQUIPOS INACTIVOS ({len(equipos_inactivos)})"
    cell.font = title_font
    cell.fill = title_fill
    row += 2

    if equipos_inactivos:
        headers = ["Código Interno", "Nombre", "Ubicación", "Responsable"]
        for i, header in enumerate(headers):
            cell = sheet_dashboard.cell(row=row, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for equipo in equipos_inactivos:
            sheet_dashboard[f'A{row}'] = equipo.codigo_interno
            sheet_dashboard[f'B{row}'] = equipo.nombre
            sheet_dashboard[f'C{row}'] = equipo.ubicacion or ""
            sheet_dashboard[f'D{row}'] = equipo.responsable
            row += 1
    else:
        sheet_dashboard[f'A{row}'] = "✅ No hay equipos inactivos"
        row += 1

    row += 2

    # === EQUIPOS DADOS DE BAJA ===
    equipos_baja = [e for e in equipos_queryset if e.estado == 'De Baja']
    sheet_dashboard.merge_cells(f'A{row}:E{row}')
    cell = sheet_dashboard[f'A{row}']
    cell.value = f"🔴 EQUIPOS DADOS DE BAJA ({len(equipos_baja)})"
    cell.font = title_font
    cell.fill = title_fill
    row += 2

    if equipos_baja:
        headers = ["Código Interno", "Nombre", "Ubicación", "Responsable", "Fecha Baja"]
        for i, header in enumerate(headers):
            cell = sheet_dashboard.cell(row=row, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for equipo in equipos_baja:
            sheet_dashboard[f'A{row}'] = equipo.codigo_interno
            sheet_dashboard[f'B{row}'] = equipo.nombre
            sheet_dashboard[f'C{row}'] = equipo.ubicacion or ""
            sheet_dashboard[f'D{row}'] = equipo.responsable
            # Buscar fecha de baja
            try:
                baja_registro = equipo.baja_registro
                fecha_baja = baja_registro.fecha_baja if baja_registro else "N/A"
            except:
                fecha_baja = "N/A"
            sheet_dashboard[f'E{row}'] = fecha_baja
            row += 1
    else:
        sheet_dashboard[f'A{row}'] = "✅ No hay equipos dados de baja"
        row += 1

    row += 3

    # === PRÓXIMAS CALIBRACIONES ===
    sheet_dashboard.merge_cells(f'A{row}:E{row}')
    cell = sheet_dashboard[f'A{row}']
    cell.value = "🔧 PRÓXIMAS CALIBRACIONES"
    cell.font = title_font
    cell.fill = title_fill
    row += 2

    calibraciones_proximas = []
    for equipo in equipos_queryset:
        if equipo.proxima_calibracion:
            dias_restantes = (equipo.proxima_calibracion - today).days
            mes = equipo.proxima_calibracion.strftime("%B %Y")
            calibraciones_proximas.append({
                'codigo': equipo.codigo_interno,
                'fecha': equipo.proxima_calibracion,
                'mes': mes,
                'dias_restantes': dias_restantes,
                'estado': 'Vencida' if dias_restantes < 0 else 'Próxima'
            })

    calibraciones_proximas.sort(key=lambda x: x['fecha'])

    if calibraciones_proximas:
        headers = ["Fecha", "Actividad", "Mes", "Código Equipo", "Estado"]
        for i, header in enumerate(headers):
            cell = sheet_dashboard.cell(row=row, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for cal in calibraciones_proximas:
            sheet_dashboard[f'A{row}'] = cal['fecha'].strftime("%d/%m/%Y")
            sheet_dashboard[f'B{row}'] = "Calibración"
            sheet_dashboard[f'C{row}'] = cal['mes']
            sheet_dashboard[f'D{row}'] = cal['codigo']
            sheet_dashboard[f'E{row}'] = cal['estado']
            row += 1
    else:
        sheet_dashboard[f'A{row}'] = "✅ No hay calibraciones programadas"
        row += 1

    row += 3

    # === PRÓXIMOS MANTENIMIENTOS ===
    sheet_dashboard.merge_cells(f'A{row}:E{row}')
    cell = sheet_dashboard[f'A{row}']
    cell.value = "🔨 PRÓXIMOS MANTENIMIENTOS"
    cell.font = title_font
    cell.fill = title_fill
    row += 2

    mantenimientos_proximos = []
    for equipo in equipos_queryset:
        if equipo.proximo_mantenimiento:
            dias_restantes = (equipo.proximo_mantenimiento - today).days
            mes = equipo.proximo_mantenimiento.strftime("%B %Y")
            mantenimientos_proximos.append({
                'codigo': equipo.codigo_interno,
                'fecha': equipo.proximo_mantenimiento,
                'mes': mes,
                'dias_restantes': dias_restantes,
                'estado': 'Vencido' if dias_restantes < 0 else 'Próximo'
            })

    mantenimientos_proximos.sort(key=lambda x: x['fecha'])

    if mantenimientos_proximos:
        headers = ["Fecha", "Actividad", "Mes", "Código Equipo", "Estado"]
        for i, header in enumerate(headers):
            cell = sheet_dashboard.cell(row=row, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for mant in mantenimientos_proximos:
            sheet_dashboard[f'A{row}'] = mant['fecha'].strftime("%d/%m/%Y")
            sheet_dashboard[f'B{row}'] = "Mantenimiento"
            sheet_dashboard[f'C{row}'] = mant['mes']
            sheet_dashboard[f'D{row}'] = mant['codigo']
            sheet_dashboard[f'E{row}'] = mant['estado']
            row += 1
    else:
        sheet_dashboard[f'A{row}'] = "✅ No hay mantenimientos programados"
        row += 1

    row += 3

    # === PRÓXIMAS COMPROBACIONES ===
    sheet_dashboard.merge_cells(f'A{row}:E{row}')
    cell = sheet_dashboard[f'A{row}']
    cell.value = "✅ PRÓXIMAS COMPROBACIONES"
    cell.font = title_font
    cell.fill = title_fill
    row += 2

    comprobaciones_proximas = []
    for equipo in equipos_queryset:
        if equipo.proxima_comprobacion:
            dias_restantes = (equipo.proxima_comprobacion - today).days
            mes = equipo.proxima_comprobacion.strftime("%B %Y")
            comprobaciones_proximas.append({
                'codigo': equipo.codigo_interno,
                'fecha': equipo.proxima_comprobacion,
                'mes': mes,
                'dias_restantes': dias_restantes,
                'estado': 'Vencida' if dias_restantes < 0 else 'Próxima'
            })

    comprobaciones_proximas.sort(key=lambda x: x['fecha'])

    if comprobaciones_proximas:
        headers = ["Fecha", "Actividad", "Mes", "Código Equipo", "Estado"]
        for i, header in enumerate(headers):
            cell = sheet_dashboard.cell(row=row, column=i+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for comp in comprobaciones_proximas:
            sheet_dashboard[f'A{row}'] = comp['fecha'].strftime("%d/%m/%Y")
            sheet_dashboard[f'B{row}'] = "Comprobación"
            sheet_dashboard[f'C{row}'] = comp['mes']
            sheet_dashboard[f'D{row}'] = comp['codigo']
            sheet_dashboard[f'E{row}'] = comp['estado']
            row += 1
    else:
        sheet_dashboard[f'A{row}'] = "✅ No hay comprobaciones programadas"
        row += 1

    # Ajustar ancho de columnas en todas las hojas
    for sheet in [sheet_equipos, sheet_proveedores, sheet_procedimientos, sheet_dashboard]:
        for column in sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

    # === HOJA 4: REPORTE DASHBOARD (DESHABILITADA PARA OPTIMIZAR RENDIMIENTO) ===
    # Hoja de dashboard temporalmente deshabilitada para optimizar memoria
    # empresa = equipos_queryset.first().empresa if equipos_queryset.exists() else None

    if False:  # Deshabilitado temporalmente
        sheet_dashboard = workbook.create_sheet(title="Reporte Dashboard")

        # Reutilizar la lógica del dashboard pero adaptada para hoja adicional
        from collections import Counter
        from datetime import datetime, timedelta
        from openpyxl.styles import Alignment

        # Configurar datos
        equipos_list = list(equipos_queryset)
        total_equipos = len(equipos_list)
        equipos_activos = sum(1 for eq in equipos_list if eq.estado == 'Activo')
        equipos_inactivos = sum(1 for eq in equipos_list if eq.estado == 'Inactivo')
        equipos_baja = sum(1 for eq in equipos_list if eq.estado == 'De Baja')

        hoy = datetime.now()
        hoy_date = hoy.date()
        ano_actual = hoy_date.year

        # Encabezado profesional
        sheet_dashboard.merge_cells('A1:F2')
        sheet_dashboard['A1'] = 'INFORMES GENERADOS POR SAM METROLOGÍA SAS'
        sheet_dashboard['A1'].font = Font(bold=True, size=20, color="FFFFFF")
        sheet_dashboard['A1'].fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
        sheet_dashboard['A1'].alignment = Alignment(horizontal="center", vertical="center")

        sheet_dashboard.merge_cells('A3:F3')
        sheet_dashboard['A3'] = f'EMPRESA: {empresa.nombre.upper()}'
        sheet_dashboard['A3'].font = Font(bold=True, size=14, color="1f4e79")
        sheet_dashboard['A3'].alignment = Alignment(horizontal="center")

        sheet_dashboard.merge_cells('A4:F4')
        sheet_dashboard['A4'] = f'Generado el: {hoy.strftime("%d de %B de %Y a las %H:%M")}'
        sheet_dashboard['A4'].font = Font(bold=True, size=12)
        sheet_dashboard['A4'].alignment = Alignment(horizontal="center")

        # Resumen general
        row = 6
        sheet_dashboard[f'A{row}'] = '📊 RESUMEN GENERAL DE EQUIPOS'
        sheet_dashboard[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
        sheet_dashboard.merge_cells(f'A{row}:F{row}')
        row += 2

        # Tabla de resumen
        headers = ['Categoría', 'Cantidad', 'Porcentaje']
        for col, header in enumerate(headers, 1):
            cell = sheet_dashboard.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row += 1

        resumen_data = [
            ('Total de Equipos', total_equipos, '100%'),
            ('Equipos Activos', equipos_activos, f'{(equipos_activos/total_equipos*100):.1f}%' if total_equipos > 0 else '0%'),
            ('Equipos Inactivos', equipos_inactivos, f'{(equipos_inactivos/total_equipos*100):.1f}%' if total_equipos > 0 else '0%'),
            ('Equipos de Baja', equipos_baja, f'{(equipos_baja/total_equipos*100):.1f}%' if total_equipos > 0 else '0%')
        ]

        for categoria, cantidad, porcentaje in resumen_data:
            sheet_dashboard.cell(row=row, column=1, value=categoria)
            sheet_dashboard.cell(row=row, column=2, value=cantidad)
            sheet_dashboard.cell(row=row, column=3, value=porcentaje)
            row += 1

        # Estadísticas de actividades
        row += 2
        sheet_dashboard[f'A{row}'] = '📅 ACTIVIDADES REALIZADAS'
        sheet_dashboard[f'A{row}'].font = Font(bold=True, size=16, color="1f4e79")
        sheet_dashboard.merge_cells(f'A{row}:F{row}')
        row += 2

        # Importar modelos para calcular actividades
        from core.models import Calibracion, Mantenimiento, Comprobacion

        cal_realizadas = Calibracion.objects.filter(equipo__empresa=empresa, fecha_calibracion__year=ano_actual).count()
        mant_realizados = Mantenimiento.objects.filter(equipo__empresa=empresa, fecha_mantenimiento__year=ano_actual).count()
        comp_realizadas = Comprobacion.objects.filter(equipo__empresa=empresa, fecha_comprobacion__year=ano_actual).count()

        # Tabla de actividades
        act_headers = ['Tipo de Actividad', f'Realizadas {ano_actual}']
        for col, header in enumerate(act_headers, 1):
            cell = sheet_dashboard.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row += 1

        actividades_data = [
            ('Calibraciones', cal_realizadas),
            ('Mantenimientos', mant_realizados),
            ('Comprobaciones', comp_realizadas)
        ]

        for tipo, realizadas in actividades_data:
            sheet_dashboard.cell(row=row, column=1, value=tipo)
            sheet_dashboard.cell(row=row, column=2, value=realizadas)
            row += 1

        # Ajustar anchos de columnas
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            sheet_dashboard.column_dimensions[col].width = 20

        # Ajustar altura de filas del encabezado
        sheet_dashboard.row_dimensions[1].height = 40

    # Guardar y retornar contenido
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_general_equipment_list_excel_content(equipos_queryset):
    """
    Generates an Excel file with the general list of equipment including visual charts.
    """
    from openpyxl.chart.series import DataPoint
    from collections import Counter
    from openpyxl.styles import Alignment
    from datetime import datetime

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listado de Equipos"

    # ==============================================
    # ENCABEZADO PROFESIONAL SAM METROLOGÍA
    # ==============================================

    # Título principal profesional
    sheet.merge_cells('A1:AB2')
    sheet['A1'] = 'INFORMES GENERADOS POR SAM METROLOGÍA SAS'
    sheet['A1'].font = Font(bold=True, size=20, color="FFFFFF")
    sheet['A1'].fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Información de generación
    hoy = datetime.now()
    sheet.merge_cells('A3:AB3')
    sheet['A3'] = f'LISTADO GENERAL DE EQUIPOS - Generado el: {hoy.strftime("%d de %B de %Y a las %H:%M")}'
    sheet['A3'].font = Font(bold=True, size=12, color="1f4e79")
    sheet['A3'].alignment = Alignment(horizontal="center")

    # Espacio antes de headers (filas 4 y 5 vacías)
    # No usamos append() para evitar conflictos con celdas fusionadas

    headers = [
        "Código Interno", "Nombre", "Empresa", "Tipo de Equipo", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado", "Fecha de Adquisición",
        "Rango de Medida", "Resolución", "Error Máximo Permisible", "Fecha de Registro",
        "Observaciones", "Versión Formato Equipo", "Fecha Versión Formato Equipo",
        "Codificación Formato Equipo", "Fecha Última Calibración", "Próxima Calibración",
        "Frecuencia Calibración (meses)", "Fecha Último Mantenimiento", "Próximo Mantenimiento",
        "Frecuencia Mantenimiento (meses)", "Fecha Última Comprobación",
        "Próxima Comprobación", "Frecuencia Comprobación (meses)"
    ]

    # Colocar headers en la fila 6 (después del título y espacios)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        sheet.column_dimensions[cell.column_letter].width = 25

    # Agregar datos de equipos empezando desde la fila 7
    current_row = 7
    for equipo in equipos_queryset:
        row_data = [
            equipo.codigo_interno,
            equipo.nombre,
            equipo.empresa.nombre if equipo.empresa else "N/A",
            equipo.get_tipo_equipo_display(),
            equipo.marca,
            equipo.modelo,
            equipo.numero_serie,
            equipo.ubicacion,
            equipo.responsable,
            equipo.estado,
            equipo.fecha_adquisicion.strftime('%Y-%m-%d') if equipo.fecha_adquisicion else '',
            equipo.rango_medida,
            equipo.resolucion,
            equipo.error_maximo_permisible if equipo.error_maximo_permisible is not None else '',
            equipo.fecha_registro.strftime('%Y-%m-%d %H:%M:%S') if equipo.fecha_registro else '',
            equipo.observaciones,
            equipo.version_formato,
            equipo.fecha_version_formato.strftime('%Y-%m-%d') if equipo.fecha_version_formato else '',
            equipo.codificacion_formato,
            equipo.fecha_ultima_calibracion.strftime('%Y-%m-%d') if equipo.fecha_ultima_calibracion else '',
            equipo.proxima_calibracion.strftime('%Y-%m-%d') if equipo.proxima_calibracion else '',
            float(equipo.frecuencia_calibracion_meses) if equipo.frecuencia_calibracion_meses is not None else '',
            equipo.fecha_ultimo_mantenimiento.strftime('%Y-%m-%d') if equipo.fecha_ultimo_mantenimiento else '', # CORREGIDO
            equipo.proximo_mantenimiento.strftime('%Y-%m-%d') if equipo.proximo_mantenimiento is not None else '',
            float(equipo.frecuencia_mantenimiento_meses) if equipo.frecuencia_mantenimiento_meses is not None else '',
            equipo.fecha_ultima_comprobacion.strftime('%Y-%m-%d') if equipo.fecha_ultima_comprobacion else '',
            equipo.proxima_comprobacion.strftime('%Y-%m-%d') if equipo.proxima_comprobacion is not None else '',
            float(equipo.frecuencia_comprobacion_meses) if equipo.frecuencia_comprobacion_meses is not None else '',
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Ajustar anchos de columna evitando problemas con celdas fusionadas
    from openpyxl.utils import get_column_letter
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    # Crear hoja de estadísticas simple (sin gráficas para optimizar memoria)
    stats_sheet = workbook.create_sheet("Estadísticas")

    # Recopilar datos para estadísticas
    equipos_list = list(equipos_queryset)

    if equipos_list:
        # 1. Estadísticas por empresa
        empresas_count = Counter(eq.empresa.nombre if eq.empresa else "Sin empresa" for eq in equipos_list)

        stats_sheet['A1'] = 'DISTRIBUCIÓN POR EMPRESA'
        stats_sheet['A2'] = 'Empresa'
        stats_sheet['B2'] = 'Cantidad de Equipos'

        row = 3
        for empresa, count in empresas_count.items():
            stats_sheet[f'A{row}'] = empresa
            stats_sheet[f'B{row}'] = count
            row += 1

        # 2. Estadísticas por tipo
        tipos_count = Counter(eq.get_tipo_equipo_display() for eq in equipos_list)

        start_row = row + 2
        stats_sheet[f'A{start_row}'] = 'DISTRIBUCIÓN POR TIPO'
        stats_sheet[f'A{start_row+1}'] = 'Tipo de Equipo'
        stats_sheet[f'B{start_row+1}'] = 'Cantidad'

        row = start_row + 2
        for tipo, count in tipos_count.items():
            stats_sheet[f'A{row}'] = tipo
            stats_sheet[f'B{row}'] = count
            row += 1

        # 3. Estadísticas por estado
        estados_count = Counter(eq.estado for eq in equipos_list)

        start_row = row + 2
        stats_sheet[f'A{start_row}'] = 'DISTRIBUCIÓN POR ESTADO'
        stats_sheet[f'A{start_row+1}'] = 'Estado'
        stats_sheet[f'B{start_row+1}'] = 'Cantidad'

        row = start_row + 2
        for estado, count in estados_count.items():
            stats_sheet[f'A{row}'] = estado
            stats_sheet[f'B{row}'] = count
            row += 1

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_equipment_general_info_excel_content(equipo):
    """
    Generates an Excel file with general information of a specific equipment.
    This is similar to _generate_general_equipment_list_excel_content but for a single equipment.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Información General"

    # Add professional title header
    sheet.merge_cells('A1:AB2')
    title_cell = sheet['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet.row_dimensions[row].height = 8

    headers = [
        "Código Interno", "Nombre", "Empresa", "Tipo de Equipo", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado", "Fecha de Adquisición",
        "Rango de Medida", "Resolución", "Error Máximo Permisible", "Fecha de Registro",
        "Observaciones", "Versión Formato Equipo", "Fecha Versión Formato Equipo",
        "Codificación Formato Equipo", "Fecha Última Calibración", "Próxima Calibración",
        "Frecuencia Calibración (meses)", "Fecha Último Mantenimiento", "Próximo Mantenimiento",
        "Frecuencia Mantenimiento (meses)", "Fecha Última Comprobación",
        "Próxima Comprobación", "Frecuencia Comprobación (meses)"
    ]

    # Add headers starting from row 6
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        sheet.column_dimensions[cell.column_letter].width = 25

    # Procesar el equipo individual (parámetro de la función) - CORREGIDO
    row_data = [
        equipo.codigo_interno,
        equipo.nombre,
        equipo.empresa.nombre if equipo.empresa else "N/A",
        equipo.get_tipo_equipo_display(),
        equipo.marca,
        equipo.modelo,
        equipo.numero_serie,
        equipo.ubicacion,
        equipo.responsable,
        equipo.estado,
        equipo.fecha_adquisicion.strftime('%Y-%m-%d') if equipo.fecha_adquisicion else '',
        equipo.rango_medida,
        equipo.resolucion,
        equipo.error_maximo_permisible if equipo.error_maximo_permisible is not None else '',
        equipo.fecha_registro.strftime('%Y-%m-%d %H:%M:%S') if equipo.fecha_registro else '',
        equipo.observaciones,
        equipo.version_formato,
        equipo.fecha_version_formato.strftime('%Y-%m-%d') if equipo.fecha_version_formato else '',
        equipo.codificacion_formato,
        equipo.fecha_ultima_calibracion.strftime('%Y-%m-%d') if equipo.fecha_ultima_calibracion else '',
        equipo.proxima_calibracion.strftime('%Y-%m-%d') if equipo.proxima_calibracion else '',
        float(equipo.frecuencia_calibracion_meses) if equipo.frecuencia_calibracion_meses is not None else '',
        equipo.fecha_ultimo_mantenimiento.strftime('%Y-%m-%d') if equipo.fecha_ultimo_mantenimiento else '',
        equipo.proximo_mantenimiento.strftime('%Y-%m-%d') if equipo.proximo_mantenimiento is not None else '',
        float(equipo.frecuencia_mantenimiento_meses) if equipo.frecuencia_mantenimiento_meses is not None else '',
        equipo.fecha_ultima_comprobacion.strftime('%Y-%m-%d') if equipo.fecha_ultima_comprobacion else '',
        equipo.proxima_comprobacion.strftime('%Y-%m-%d') if equipo.proxima_comprobacion is not None else '',
        float(equipo.frecuencia_comprobacion_meses) if equipo.frecuencia_comprobacion_meses is not None else '',
    ]

    # Add data row starting from row 7
    for col_num, value in enumerate(row_data, 1):
        sheet.cell(row=7, column=col_num, value=value)

    # Ajustar anchos de columna evitando problemas con celdas fusionadas
    from openpyxl.utils import get_column_letter
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_equipment_activities_excel_content(equipo):
    """
    Generates an Excel file with the activities (calibrations, maintenances, verifications) of a specific equipment.
    """
    workbook = Workbook()

    sheet_cal = workbook.active
    sheet_cal.title = "Calibraciones"

    # Add professional title header to Calibraciones sheet
    sheet_cal.merge_cells('A1:E2')
    title_cell = sheet_cal['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet_cal['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_cal['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet_cal['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet_cal.row_dimensions[row].height = 8

    headers_cal = ["Fecha Calibración", "Proveedor", "Resultado", "Número Certificado", "Observaciones"]

    # Add headers starting from row 6
    for col_num, header_text in enumerate(headers_cal, 1):
        cell = sheet_cal.cell(row=6, column=col_num, value=header_text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add calibration data starting from row 7
    current_row = 7
    for cal in equipo.calibraciones.all().order_by('fecha_calibracion'):
        proveedor_nombre = cal.nombre_proveedor if cal.nombre_proveedor else ''
        row_data = [
            cal.fecha_calibracion.strftime('%Y-%m-%d') if cal.fecha_calibracion else '',
            proveedor_nombre,
            cal.resultado,
            cal.numero_certificado,
            cal.observaciones
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_cal.cell(row=current_row, column=col_num, value=value)
        current_row += 1
    # Ajustar anchos de columna para Calibraciones
    from openpyxl.utils import get_column_letter
    headers_cal = ["Fecha Calibración", "Proveedor", "Resultado", "Número Certificado", "Observaciones"]
    for col_num in range(1, len(headers_cal) + 1):
        column_letter = get_column_letter(col_num)
        sheet_cal.column_dimensions[column_letter].width = 25

    sheet_mant = workbook.create_sheet("Mantenimientos")

    # Add professional title header to Mantenimientos sheet
    sheet_mant.merge_cells('A1:F2')
    title_cell = sheet_mant['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet_mant['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_mant['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet_mant['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet_mant.row_dimensions[row].height = 8

    headers_mant = ["Fecha Mantenimiento", "Tipo", "Proveedor", "Responsable", "Costo", "Descripción"]

    # Add headers starting from row 6
    for col_num, header_text in enumerate(headers_mant, 1):
        cell = sheet_mant.cell(row=6, column=col_num, value=header_text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add maintenance data starting from row 7
    current_row = 7
    for mant in equipo.mantenimientos.all().order_by('fecha_mantenimiento'):
        proveedor_nombre = mant.nombre_proveedor if mant.nombre_proveedor else ''
        row_data = [
            mant.fecha_mantenimiento.strftime('%Y-%m-%d') if mant.fecha_mantenimiento else '',
            mant.get_tipo_mantenimiento_display(),
            proveedor_nombre,
            mant.responsable,
            float(mant.costo) if mant.costo is not None else '',
            mant.descripcion
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_mant.cell(row=current_row, column=col_num, value=value)
        current_row += 1
    # Ajustar anchos de columna para Mantenimientos
    headers_mant = ["Fecha Mantenimiento", "Tipo", "Proveedor", "Responsable", "Costo", "Descripción"]
    for col_num in range(1, len(headers_mant) + 1):
        column_letter = get_column_letter(col_num)
        sheet_mant.column_dimensions[column_letter].width = 25

    sheet_comp = workbook.create_sheet("Comprobaciones")

    # Add professional title header to Comprobaciones sheet
    sheet_comp.merge_cells('A1:E2')
    title_cell = sheet_comp['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet_comp['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet_comp['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet_comp['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet_comp.row_dimensions[row].height = 8

    headers_comp = ["Fecha Comprobación", "Proveedor", "Responsable", "Resultado", "Observaciones"]

    # Add headers starting from row 6
    for col_num, header_text in enumerate(headers_comp, 1):
        cell = sheet_comp.cell(row=6, column=col_num, value=header_text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add comprobacion data starting from row 7
    current_row = 7
    for comp in equipo.comprobaciones.all().order_by('fecha_comprobacion'):
        proveedor_nombre = comp.nombre_proveedor if comp.nombre_proveedor else ''
        row_data = [
            comp.fecha_comprobacion.strftime('%Y-%m-%d') if comp.fecha_comprobacion else '',
            proveedor_nombre,
            comp.responsable,
            comp.resultado,
            comp.observaciones
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet_comp.cell(row=current_row, column=col_num, value=value)
        current_row += 1
    # Ajustar anchos de columna para Comprobaciones
    headers_comp = ["Fecha Comprobación", "Proveedor", "Responsable", "Resultado", "Observaciones"]
    for col_num in range(1, len(headers_comp) + 1):
        column_letter = get_column_letter(col_num)
        sheet_comp.column_dimensions[column_letter].width = 25

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def _generate_general_proveedor_list_excel_content(proveedores_queryset):
    """
    Generates an Excel file with the general list of providers.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listado de Proveedores"

    # Add professional title header
    sheet.merge_cells('A1:I2')
    title_cell = sheet['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet.row_dimensions[row].height = 8

    headers = [
        "Nombre de la Empresa Proveedora", "Empresa Cliente", "Tipo de Servicio", "Nombre de Contacto",
        "Número de Contacto", "Correo Electrónico", "Página Web",
        "Alcance", "Servicio Prestado"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        sheet.column_dimensions[cell.column_letter].width = 25

    # Add provider data starting from row 7
    current_row = 7
    for proveedor in proveedores_queryset:
        row_data = [
            proveedor.nombre_empresa,
            proveedor.empresa.nombre if proveedor.empresa else "N/A",
            proveedor.get_tipo_servicio_display(),
            proveedor.nombre_contacto,
            proveedor.numero_contacto,
            proveedor.correo_electronico,
            proveedor.pagina_web,
            proveedor.alcance,
            proveedor.servicio_prestado,
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Ajustar anchos de columna evitando problemas con celdas fusionadas
    from openpyxl.utils import get_column_letter
    headers = [
        "Nombre de la Empresa Proveedora", "Empresa Cliente", "Tipo de Servicio", "Nombre de Contacto",
        "Número de Contacto", "Correo Electrónico", "Página Web",
        "Alcance", "Servicio Prestado"
    ]
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()

def _generate_procedimiento_info_excel_content(procedimientos_queryset):
    """
    Generates an Excel file with the general list of procedures.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listado de Procedimientos"

    # Add professional title header
    sheet.merge_cells('A1:G2')
    title_cell = sheet['A1']
    title_cell.value = "INFORMES GENERADOS POR SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generation timestamp
    sheet['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A3'].font = Font(name="Arial", size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal="left")

    # Add spacing
    for row in range(4, 6):
        sheet.row_dimensions[row].height = 8

    headers = [
        "Nombre", "Código", "Versión", "Fecha de Emisión", "Empresa", "Observaciones", "Documento PDF"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        sheet.column_dimensions[cell.column_letter].width = 25

    # Add procedure data starting from row 7
    current_row = 7
    for proc in procedimientos_queryset:
        row_data = [
            proc.nombre,
            proc.codigo,
            proc.version,
            proc.fecha_emision.strftime('%Y-%m-%d') if proc.fecha_emision else '',
            proc.empresa.nombre if proc.empresa else "N/A",
            proc.observaciones,
            proc.documento_pdf.url if proc.documento_pdf else 'N/A',
        ]
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1

    # Ajustar anchos de columna evitando problemas con celdas fusionadas
    from openpyxl.utils import get_column_letter
    headers = [
        "Nombre", "Código", "Versión", "Fecha de Emisión", "Empresa", "Observaciones", "Documento PDF"
    ]
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


# =============================================================================
# Vistas de Autenticación y Perfil de Usuario
# =============================================================================

def user_login(request):
    """Vista para el inicio de sesión de usuarios."""
    if request.user.is_authenticated:
        return redirect('core:dashboard') # Redirigir al dashboard si ya está logueado

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'¡Bienvenido, {username}!')
                return redirect('core:dashboard')
            else:
                messages.error(request, 'Nombre de usuario o contraseña incorrectos.')
        else:
            messages.error(request, 'Por favor, corrige los errores del formulario.')
    else:
        form = AuthenticationForm()
    return render(request, 'registration/login.html', {'form': form})

@login_required
def user_logout(request):
    """Vista para cerrar sesión de usuarios."""
    logout(request)
    messages.info(request, 'Has cerrado sesión exitosamente.')
    return redirect('core:login')

@access_check # APLICAR ESTE DECORADOR
@login_required
def cambiar_password(request):
    """Vista para que el usuario cambie su propia contraseña."""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Tu contraseña ha sido actualizada exitosamente!')
            return redirect('core:password_change_done')
        else:
            messages.error(request, 'Por favor corrige los errores a continuación.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'registration/password_change_form.html', {
        'form': form,
        'titulo_pagina': 'Cambiar Contraseña'
    })

@access_check # APLICAR ESTE DECORADOR
@login_required
def password_change_done(request):
    """Vista de confirmación de cambio de contraseña."""
    return render(request, 'core/password_change_done.html', {'titulo_pagina': 'Contraseña Cambiada'})

@access_check # APLICAR ESTE DECORADOR
@login_required
def perfil_usuario(request):
    """
    Handles user profile viewing and editing.
    """
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
            return redirect('core:perfil_usuario')
        else:
            messages.error(request, 'Error al actualizar tu perfil. Por favor, revisa los campos.')
    else:
        form = UserProfileForm(instance=request.user)
    return render(request, 'core/my_profile.html', {'form': form, 'usuario': request.user, 'titulo_pagina': 'Mi Perfil'})

# =============================================================================
# Vistas de Dashboard y Estadísticas
# =============================================================================

@access_check
@login_required
def dashboard(request):
    """
    Displays the dashboard with key metrics and charts.
    """
    user = request.user
    today = date.today()
    current_year = today.year
    
    # Filtrado por empresa para superusuarios
    selected_company_id = request.GET.get('empresa_id')
    empresas_disponibles = Empresa.objects.all().order_by('nombre')

    equipos_queryset = Equipo.objects.all()

    if not user.is_superuser:
        # Los usuarios normales solo ven datos de su propia empresa
        if user.empresa:
            equipos_queryset = equipos_queryset.filter(empresa=user.empresa)
            selected_company_id = str(user.empresa.id) # Asegurar que el filtro se aplique y muestre la empresa del usuario
        else:
            # Si un usuario normal no tiene empresa asignada, no ve equipos
            equipos_queryset = Equipo.objects.none()
            empresas_disponibles = Empresa.objects.none() # No hay empresas para filtrar

    if selected_company_id:
        equipos_queryset = equipos_queryset.filter(empresa_id=selected_company_id)

    # Excluir equipos "De Baja" de los cálculos del dashboard
    equipos_para_dashboard = equipos_queryset.exclude(estado='De Baja').exclude(estado='Inactivo') # Se añadió exclusión de 'Inactivo'

    # --- Indicadores Clave (Optimizado con una sola consulta) ---
    from django.db.models import Count, Case, When, IntegerField

    estadisticas_equipos = equipos_queryset.aggregate(
        total_equipos=Count('id'),
        equipos_activos=Count(Case(When(estado='Activo', then=1), output_field=IntegerField())),
        equipos_inactivos=Count(Case(When(estado='Inactivo', then=1), output_field=IntegerField())),
        equipos_de_baja=Count(Case(When(estado='De Baja', then=1), output_field=IntegerField()))
    )

    total_equipos = estadisticas_equipos['total_equipos']
    equipos_activos = estadisticas_equipos['equipos_activos']
    equipos_inactivos = estadisticas_equipos['equipos_inactivos']
    equipos_de_baja = estadisticas_equipos['equipos_de_baja']

    # --- Datos de Almacenamiento ---
    storage_usage_mb = 0
    storage_limit_mb = 0
    storage_percentage = 0
    storage_status_class = 'text-gray-700 bg-gray-100'
    storage_empresa_nombre = 'N/A'

    # --- Datos de Límites de Equipos ---
    equipos_limite = 0
    equipos_actuales_count = 0
    equipos_disponibles = 0
    equipos_limite_percentage = 0
    equipos_limite_warning = False
    equipos_limite_critical = False

    if user.is_superuser and selected_company_id:
        try:
            empresa_seleccionada = Empresa.objects.get(id=selected_company_id)
            storage_usage_mb = empresa_seleccionada.get_total_storage_used_mb()
            storage_limit_mb = empresa_seleccionada.limite_almacenamiento_mb
            storage_percentage = empresa_seleccionada.get_storage_usage_percentage()
            storage_status_class = empresa_seleccionada.get_storage_status_class()
            storage_empresa_nombre = empresa_seleccionada.nombre

            # Datos de límite de equipos
            equipos_limite = empresa_seleccionada.get_limite_equipos()
            equipos_actuales_count = Equipo.objects.filter(empresa=empresa_seleccionada).count()
            if equipos_limite != float('inf'):
                equipos_disponibles = max(0, equipos_limite - equipos_actuales_count)
                equipos_limite_percentage = (equipos_actuales_count / equipos_limite) * 100 if equipos_limite > 0 else 0
                equipos_limite_warning = equipos_limite_percentage >= 80
                equipos_limite_critical = equipos_actuales_count >= equipos_limite
        except Empresa.DoesNotExist:
            pass
    elif user.empresa:
        storage_usage_mb = user.empresa.get_total_storage_used_mb()
        storage_limit_mb = user.empresa.limite_almacenamiento_mb
        storage_percentage = user.empresa.get_storage_usage_percentage()
        storage_status_class = user.empresa.get_storage_status_class()
        storage_empresa_nombre = user.empresa.nombre

        # Datos de límite de equipos
        equipos_limite = user.empresa.get_limite_equipos()
        equipos_actuales_count = Equipo.objects.filter(empresa=user.empresa).count()
        if equipos_limite != float('inf'):
            equipos_disponibles = max(0, equipos_limite - equipos_actuales_count)
            equipos_limite_percentage = (equipos_actuales_count / equipos_limite) * 100 if equipos_limite > 0 else 0
            equipos_limite_warning = equipos_limite_percentage >= 80
            equipos_limite_critical = equipos_actuales_count >= equipos_limite

    # Detección de actividades vencidas y próximas (Optimizado con una sola consulta)
    # Se basan directamente en los campos proxima_X del modelo Equipo, excluyendo Inactivos y De Baja
    from datetime import timedelta
    fecha_limite_proximas = today + timedelta(days=30)

    estadisticas_actividades = equipos_para_dashboard.aggregate(
        calibraciones_vencidas=Count(
            Case(When(
                proxima_calibracion__isnull=False,
                proxima_calibracion__lt=today,
                then=1
            ), output_field=IntegerField())
        ),
        calibraciones_proximas=Count(
            Case(When(
                proxima_calibracion__isnull=False,
                proxima_calibracion__gte=today,
                proxima_calibracion__lte=fecha_limite_proximas,
                then=1
            ), output_field=IntegerField())
        ),
        mantenimientos_vencidos=Count(
            Case(When(
                proximo_mantenimiento__isnull=False,
                proximo_mantenimiento__lt=today,
                then=1
            ), output_field=IntegerField())
        ),
        mantenimientos_proximas=Count(
            Case(When(
                proximo_mantenimiento__isnull=False,
                proximo_mantenimiento__gte=today,
                proximo_mantenimiento__lte=fecha_limite_proximas,
                then=1
            ), output_field=IntegerField())
        ),
        comprobaciones_vencidas=Count(
            Case(When(
                proxima_comprobacion__isnull=False,
                proxima_comprobacion__lt=today,
                then=1
            ), output_field=IntegerField())
        ),
        comprobaciones_proximas=Count(
            Case(When(
                proxima_comprobacion__isnull=False,
                proxima_comprobacion__gte=today,
                proxima_comprobacion__lte=fecha_limite_proximas,
                then=1
            ), output_field=IntegerField())
        )
    )

    calibraciones_vencidas = estadisticas_actividades['calibraciones_vencidas']
    calibraciones_proximas = estadisticas_actividades['calibraciones_proximas']
    mantenimientos_vencidos = estadisticas_actividades['mantenimientos_vencidos']
    mantenimientos_proximas = estadisticas_actividades['mantenimientos_proximas']
    comprobaciones_vencidas = estadisticas_actividades['comprobaciones_vencidas']
    comprobaciones_proximas = estadisticas_actividades['comprobaciones_proximas']

    # Obtener los códigos de equipos vencidos para mostrar en el dashboard (Optimizado)
    vencidos_calibracion_codigos = list(equipos_para_dashboard.filter(
        proxima_calibracion__lt=today
    ).values_list('codigo_interno', flat=True))

    vencidos_mantenimiento_codigos = list(equipos_para_dashboard.filter(
        proximo_mantenimiento__lt=today
    ).values_list('codigo_interno', flat=True))

    vencidos_comprobacion_codigos = list(equipos_para_dashboard.filter(
        proxima_comprobacion__lt=today
    ).values_list('codigo_interno', flat=True))


    # --- Datos para Gráficas de Línea (Programadas vs Realizadas por Mes) ---
    line_chart_labels = []
    
    # Initialize programmed data arrays
    programmed_calibrations_line_data = [0] * 12
    programmed_mantenimientos_line_data = [0] * 12
    programmed_comprobaciones_line_data = [0] * 12

    realized_calibrations_line_data = [0] * 12
    realized_preventive_mantenimientos_line_data = [0] * 12
    realized_corrective_mantenimientos_line_data = [0] * 12
    realized_other_mantenimientos_line_data = [0] * 12
    realized_predictive_mantenimientos_line_data = [0] * 12
    realized_inspection_mantenimientos_line_data = [0] * 12
    realized_comprobaciones_line_data = [0] * 12

    # Calcular el primer mes del rango (6 meses antes del actual)
    start_date_range = today - relativedelta(months=6)
    # Ajustar al primer día del mes
    start_date_range = start_date_range.replace(day=1)

    for i in range(12):
        target_date = start_date_range + relativedelta(months=i)
        line_chart_labels.append(f"{calendar.month_abbr[target_date.month]}. {target_date.year}")

    # Datos "Realizadas" (basado en registros de actividad) - Solo para equipos que no estén de baja o inactivos
    calibraciones_realizadas_period = Calibracion.objects.filter(
        equipo__in=equipos_para_dashboard, # Filtrar por equipos elegibles
        fecha_calibracion__gte=start_date_range,
        fecha_calibracion__lte=start_date_range + relativedelta(months=12, days=-1)
    )
    mantenimientos_realizados_period = Mantenimiento.objects.filter(
        equipo__in=equipos_para_dashboard, # Filtrar por equipos elegibles
        fecha_mantenimiento__gte=start_date_range,
        fecha_mantenimiento__lte=start_date_range + relativedelta(months=12, days=-1)
    )
    comprobaciones_realizadas_period = Comprobacion.objects.filter(
        equipo__in=equipos_para_dashboard, # Filtrar por equipos elegibles
        fecha_comprobacion__gte=start_date_range,
        fecha_comprobacion__lte=start_date_range + relativedelta(months=12, days=-1)
    )

    for cal in calibraciones_realizadas_period:
        month_index = ((cal.fecha_calibracion.year - start_date_range.year) * 12 + cal.fecha_calibracion.month - start_date_range.month)
        if 0 <= month_index < 12:
            realized_calibrations_line_data[month_index] += 1
    
    for mant in mantenimientos_realizados_period:
        month_index = ((mant.fecha_mantenimiento.year - start_date_range.year) * 12 + mant.fecha_mantenimiento.month - start_date_range.month)
        if 0 <= month_index < 12:
            if mant.tipo_mantenimiento == 'Preventivo':
                realized_preventive_mantenimientos_line_data[month_index] += 1
            elif mant.tipo_mantenimiento == 'Correctivo':
                realized_corrective_mantenimientos_line_data[month_index] += 1
            elif mant.tipo_mantenimiento == 'Predictivo':
                realized_predictive_mantenimientos_line_data[month_index] += 1
            elif mant.tipo_mantenimiento == 'Inspección':
                realized_inspection_mantenimientos_line_data[month_index] += 1
            else:
                realized_other_mantenimientos_line_data[month_index] += 1

    for comp in comprobaciones_realizadas_period:
        month_index = ((comp.fecha_comprobacion.year - start_date_range.year) * 12 + comp.fecha_comprobacion.month - start_date_range.month)
        if 0 <= month_index < 12:
            realized_comprobaciones_line_data[month_index] += 1

    # Datos "Programadas" (basado en un plan fijo anual desde la fecha de adquisición/registro)
    # Usar equipos_para_dashboard para la programación (ya excluye De Baja e Inactivo)
    for equipo in equipos_para_dashboard:
        # Determinar la fecha de inicio del plan para este equipo (fecha de adquisición o registro)
        plan_start_date = equipo.fecha_adquisicion if equipo.fecha_adquisicion else \
                          (equipo.fecha_registro.date() if equipo.fecha_registro else date(current_year, 1, 1))

        # Calibraciones Programadas
        if equipo.frecuencia_calibracion_meses is not None and equipo.frecuencia_calibracion_meses > 0:
            freq = int(equipo.frecuencia_calibracion_meses)
            
            diff_months = (start_date_range.year - plan_start_date.year) * 12 + (start_date_range.month - plan_start_date.month)
            num_intervals = 0
            if freq > 0:
                num_intervals = max(0, (diff_months + freq - 1) // freq)

            current_plan_date = plan_start_date + relativedelta(months=num_intervals * freq)
            
            for _ in range(12 + freq): # Iterar lo suficiente para cubrir el rango de 12 meses
                if start_date_range <= current_plan_date < start_date_range + relativedelta(months=12):
                    month_index = ((current_plan_date.year - start_date_range.year) * 12 + current_plan_date.month - start_date_range.month)
                    if 0 <= month_index < 12:
                        programmed_calibrations_line_data[month_index] += 1
                
                if current_plan_date >= start_date_range + relativedelta(months=12 + freq):
                    break
                
                try:
                    current_plan_date += relativedelta(months=freq)
                except OverflowError:
                    break


        # Mantenimientos Programados (misma lógica optimizada)
        if equipo.frecuencia_mantenimiento_meses is not None and equipo.frecuencia_mantenimiento_meses > 0:
            freq = int(equipo.frecuencia_mantenimiento_meses)
            
            diff_months = (start_date_range.year - plan_start_date.year) * 12 + (start_date_range.month - plan_start_date.month)
            num_intervals = 0
            if freq > 0:
                num_intervals = max(0, (diff_months + freq - 1) // freq)

            current_plan_date = plan_start_date + relativedelta(months=num_intervals * freq)

            for _ in range(12 + freq):
                if start_date_range <= current_plan_date < start_date_range + relativedelta(months=12):
                    month_index = ((current_plan_date.year - start_date_range.year) * 12 + current_plan_date.month - start_date_range.month)
                    if 0 <= month_index < 12:
                        programmed_mantenimientos_line_data[month_index] += 1
                
                if current_plan_date >= start_date_range + relativedelta(months=12 + freq):
                    break
                
                try:
                    current_plan_date += relativedelta(months=freq)
                except OverflowError:
                    break

        # Comprobaciones Programadas (misma lógica optimizada)
        if equipo.frecuencia_comprobacion_meses is not None and equipo.frecuencia_comprobacion_meses > 0:
            freq = int(equipo.frecuencia_comprobacion_meses)
            
            diff_months = (start_date_range.year - plan_start_date.year) * 12 + (start_date_range.month - plan_start_date.month)
            num_intervals = 0
            if freq > 0:
                num_intervals = max(0, (diff_months + freq - 1) // freq)

            current_plan_date = plan_start_date + relativedelta(months=num_intervals * freq)

            for _ in range(12 + freq):
                if start_date_range <= current_plan_date < start_date_range + relativedelta(months=12):
                    month_index = ((current_plan_date.year - start_date_range.year) * 12 + current_plan_date.month - start_date_range.month)
                    if 0 <= month_index < 12:
                        programmed_comprobaciones_line_data[month_index] += 1
                
                if current_plan_date >= start_date_range + relativedelta(months=12 + freq):
                    break
                
                try:
                    current_plan_date += relativedelta(months=freq)
                except OverflowError:
                    break

    # --- Datos para Gráficas de Torta (Cumplimiento Anual) ---
    # Colores para las gráficas de torta
    pie_chart_colors_cal = ['#28a745', '#dc3545', '#007bff'] # Verde (Realizado), Rojo (No Cumplido), Azul (Pendiente/Programado)
    pie_chart_colors_comp = ['#28a745', '#dc3545', '#007bff'] # Mismos colores para comprobaciones
    pie_chart_colors_equipos = ['#28a745', '#ffc107', '#dc3545', '#17a2b8', '#6c757d', '#8672cb'] # Activo, En Mantenimiento, De Baja, En Calibración, En Comprobación, Inactivo (usar solo los relevantes)

    # Estado de Equipos (Torta) - Usar el queryset general (incluye De Baja, Inactivo)
    estado_equipos_counts = defaultdict(int)
    for equipo in equipos_queryset.all(): # NOTA: Este queryset original incluye todos los estados
        estado_equipos_counts[equipo.estado] += 1
    
    estado_equipos_labels = list(estado_equipos_counts.keys())
    estado_equipos_data = list(estado_equipos_counts.values())


    # Calibraciones
    projected_calibraciones = get_projected_activities_for_year(equipos_queryset, 'calibracion', current_year, today)
    
    cal_total_programmed_anual_display = len(projected_calibraciones)
    cal_realized_anual_display = sum(1 for act in projected_calibraciones if act['status'] == 'Realizado')
    cal_no_cumplido_anual_display = sum(1 for act in projected_calibraciones if act['status'] == 'No Cumplido')
    cal_pendiente_anual_display = sum(1 for act in projected_calibraciones if act['status'] == 'Pendiente/Programado')

    # Handle case where no activities are programmed for the year
    if cal_total_programmed_anual_display == 0:
        calibraciones_torta_labels = ['Sin Actividades']
        calibraciones_torta_data = [1] # A small value to render the chart
        pie_chart_colors_cal_display = ['#cccccc'] # Grey for no activities
        cal_realized_anual_percent = 0
        cal_no_cumplido_anual_percent = 0
        cal_pendiente_anual_percent = 0
    else:
        calibraciones_torta_labels = ['Realizado', 'No Cumplido', 'Pendiente/Programado']
        calibraciones_torta_data = [
            cal_realized_anual_display,
            cal_no_cumplido_anual_display,
            cal_pendiente_anual_display
        ]
        pie_chart_colors_cal_display = pie_chart_colors_cal # Use predefined colors
        
        # Asegurarse de que el denominador no sea cero antes de calcular el porcentaje
        if cal_total_programmed_anual_display > 0:
            cal_realized_anual_percent = (cal_realized_anual_display / cal_total_programmed_anual_display * 100)
            cal_no_cumplido_anual_percent = (cal_no_cumplido_anual_display / cal_total_programmed_anual_display * 100)
            cal_pendiente_anual_percent = (cal_pendiente_anual_display / cal_total_programmed_anual_display * 100)
        else:
            cal_realized_anual_percent = 0
            cal_no_cumplido_anual_percent = 0
            cal_pendiente_anual_percent = 0
            

    # Comprobaciones (similar logic)
    projected_comprobaciones = get_projected_activities_for_year(equipos_queryset, 'comprobacion', current_year, today)

    comp_total_programmed_anual_display = len(projected_comprobaciones)
    comp_realized_anual_display = sum(1 for act in projected_comprobaciones if act['status'] == 'Realizado')
    comp_no_cumplido_anual_display = sum(1 for act in projected_comprobaciones if act['status'] == 'No Cumplido')
    comp_pendiente_anual_display = sum(1 for act in projected_comprobaciones if act['status'] == 'Pendiente/Programado')

    if comp_total_programmed_anual_display == 0:
        comprobaciones_torta_labels = ['Sin Actividades']
        comprobaciones_torta_data = [1]
        pie_chart_colors_comp_display = ['#cccccc']
        comp_realized_anual_percent = 0
        comp_no_cumplido_anual_percent = 0
        comp_pendiente_anual_percent = 0
    else:
        comprobaciones_torta_labels = ['Realizado', 'No Cumplido', 'Pendiente/Programado']
        comprobaciones_torta_data = [
            comp_realized_anual_display,
            comp_no_cumplido_anual_display,
            comp_pendiente_anual_display
        ]
        pie_chart_colors_comp_display = pie_chart_colors_comp
        
        # Asegurarse de que el denominador no sea cero antes de calcular el porcentaje
        if comp_total_programmed_anual_display > 0:
            comp_realized_anual_percent = (comp_realized_anual_display / comp_total_programmed_anual_display * 100)
            comp_no_cumplido_anual_percent = (comp_no_cumplido_anual_display / comp_total_programmed_anual_display * 100)
            comp_pendiente_anual_percent = (comp_pendiente_anual_display / comp_total_programmed_anual_display * 100)
        else:
            comp_realized_anual_percent = 0
            comp_no_cumplido_anual_percent = 0
            comp_pendiente_anual_percent = 0

    # NUEVA LÓGICA: Mantenimientos por Cumplimiento (Torta)
    projected_mantenimientos = get_projected_maintenance_compliance_for_year(equipos_queryset, current_year, today)
    
    mant_total_programmed_anual_display = len(projected_mantenimientos)
    mant_realized_anual_display = sum(1 for act in projected_mantenimientos if act['status'] == 'Realizado')
    mant_no_cumplido_anual_display = sum(1 for act in projected_mantenimientos if act['status'] == 'No Cumplido')
    mant_pendiente_anual_display = sum(1 for act in projected_mantenimientos if act['status'] == 'Pendiente/Programado')

    if mant_total_programmed_anual_display == 0:
        mantenimientos_cumplimiento_torta_labels = ['Sin Actividades Programadas']
        mantenimientos_cumplimiento_torta_data = [1]
        pie_chart_colors_mant_compliance_display = ['#cccccc']
        mant_realized_anual_percent = 0
        mant_no_cumplido_anual_percent = 0
        mant_pendiente_anual_percent = 0
    else:
        mantenimientos_cumplimiento_torta_labels = ['Realizado', 'No Cumplido', 'Pendiente/Programado']
        mantenimientos_cumplimiento_torta_data = [
            mant_realized_anual_display,
            mant_no_cumplido_anual_display,
            mant_pendiente_anual_display
        ]
        # Colores específicos para cumplimiento de mantenimiento (similar a calibración/comprobación)
        pie_chart_colors_mant_compliance_display = ['#28a745', '#dc3545', '#007bff'] 
        
        if mant_total_programmed_anual_display > 0:
            mant_realized_anual_percent = (mant_realized_anual_display / mant_total_programmed_anual_display * 100)
            mant_no_cumplido_anual_percent = (mant_no_cumplido_anual_display / mant_total_programmed_anual_display * 100)
            mant_pendiente_anual_percent = (mant_pendiente_anual_display / mant_total_programmed_anual_display * 100)
        else:
            mant_realized_anual_percent = 0
            mant_no_cumplido_anual_percent = 0
            mant_pendiente_anual_percent = 0


    # Mantenimientos por Tipo (Torta) - Esta ya existía y es para todos los mantenimientos realizados
    mantenimientos_tipo_counts = defaultdict(int)
    # Excluir equipos de baja o inactivos de este conteo también
    for mant in Mantenimiento.objects.filter(equipo__in=equipos_para_dashboard):
        mantenimientos_tipo_counts[mant.tipo_mantenimiento] += 1
    
    mantenimientos_tipo_labels = list(mantenimientos_tipo_counts.keys())
    mantenimientos_tipo_data = list(mantenimientos_tipo_counts.values())
    pie_chart_colors_mant_types = ['#ffc107', '#dc3545', '#17a2b8', '#6c757d', '#8672cb'] # Preventivo, Correctivo, Predictivo, Inspección, Otro


    # --- NUEVA LÓGICA PARA MANTENIMIENTOS CORRECTIVOS ---
    latest_corrective_maintenances_query = Mantenimiento.objects.filter(tipo_mantenimiento='Correctivo').order_by('-fecha_mantenimiento')
    
    if not user.is_superuser:
        if user.empresa:
            latest_corrective_maintenances_query = latest_corrective_maintenances_query.filter(equipo__empresa=user.empresa)
        else:
            latest_corrective_maintenances_query = Mantenimiento.objects.none()
    elif selected_company_id:
        latest_corrective_maintenances_query = latest_corrective_maintenances_query.filter(equipo__empresa_id=selected_company_id)

    latest_corrective_maintenances = latest_corrective_maintenances_query[:5] # Limitar a los 5 más recientes

    # Convertir datos a JSON para pasarlos a JavaScript
    context = {
        'titulo_pagina': 'Panel de Control de Metrología',
        'today': today,
        'is_superuser': user.is_superuser,
        'empresas_disponibles': empresas_disponibles,
        'selected_company_id': selected_company_id,

        'total_equipos': total_equipos,
        'equipos_activos': equipos_activos,
        'equipos_inactivos': equipos_inactivos,
        'equipos_de_baja': equipos_de_baja,

        'calibraciones_vencidas': calibraciones_vencidas,
        'calibraciones_proximas': calibraciones_proximas,
        'mantenimientos_vencidos': mantenimientos_vencidos,
        'mantenimientos_proximas': mantenimientos_proximas,
        'comprobaciones_vencidas': comprobaciones_vencidas,
        'comprobaciones_proximas': comprobaciones_proximas,

        'vencidos_calibracion_codigos': vencidos_calibracion_codigos,
        'vencidos_mantenimiento_codigos': vencidos_mantenimiento_codigos,
        'vencidos_comprobacion_codigos': vencidos_comprobacion_codigos,

        # Datos para tablas de resumen de tortas (ya calculados)
        'cal_total_programmed_anual': cal_total_programmed_anual_display,
        'cal_realized_anual': cal_realized_anual_display,
        'cal_no_cumplido_anual': cal_no_cumplido_anual_display,
        'cal_pendiente_anual': cal_pendiente_anual_display,
        'cal_realized_anual_percent': round(cal_realized_anual_percent),
        'cal_no_cumplido_anual_percent': round(cal_no_cumplido_anual_percent),
        'cal_pendiente_anual_percent': round(cal_pendiente_anual_percent),

        'comp_total_programmed_anual': comp_total_programmed_anual_display,
        'comp_realized_anual': comp_realized_anual_display,
        'comp_no_cumplido_anual': comp_no_cumplido_anual_display,
        'comp_pendiente_anual': comp_pendiente_anual_display,
        'comp_realized_anual_percent': round(comp_realized_anual_percent),
        'comp_no_cumplido_anual_percent': round(comp_no_cumplido_anual_percent),
        'comp_pendiente_anual_percent': round(comp_pendiente_anual_percent),

        # NUEVOS DATOS PARA CUMPLIMIENTO DE MANTENIMIENTO
        'mant_total_programmed_anual': mant_total_programmed_anual_display,
        'mant_realized_anual': mant_realized_anual_display,
        'mant_no_cumplido_anual': mant_no_cumplido_anual_display,
        'mant_pendiente_anual': mant_pendiente_anual_display,
        'mant_realized_anual_percent': round(mant_realized_anual_percent),
        'mant_no_cumplido_anual_percent': round(mant_no_cumplido_anual_percent),
        'mant_pendiente_anual_percent': round(mant_pendiente_anual_percent),


        # Datos para gráficas de línea
        'line_chart_labels_json': mark_safe(json.dumps(line_chart_labels)),
        'programmed_calibrations_line_data_json': mark_safe(json.dumps(programmed_calibrations_line_data)),
        'realized_calibrations_line_data_json': mark_safe(json.dumps(realized_calibrations_line_data)),
        'programmed_mantenimientos_line_data_json': mark_safe(json.dumps(programmed_mantenimientos_line_data)),
        'realized_preventive_mantenimientos_line_data_json': mark_safe(json.dumps(realized_preventive_mantenimientos_line_data)),
        'realized_corrective_mantenimientos_line_data_json': mark_safe(json.dumps(realized_corrective_mantenimientos_line_data)),
        'realized_other_mantenimientos_line_data_json': mark_safe(json.dumps(realized_other_mantenimientos_line_data)),
        'realized_predictive_mantenimientos_line_data_json': mark_safe(json.dumps(realized_predictive_mantenimientos_line_data)),
        'realized_inspection_mantenimientos_line_data_json': mark_safe(json.dumps(realized_inspection_mantenimientos_line_data)),
        'programmed_comprobaciones_line_data_json': mark_safe(json.dumps(programmed_comprobaciones_line_data)),
        'realized_comprobaciones_line_data_json': mark_safe(json.dumps(realized_comprobaciones_line_data)),
        
        # Datos para gráficas de torta
        'estado_equipos_labels_json': mark_safe(json.dumps(estado_equipos_labels)),
        'estado_equipos_data_json': mark_safe(json.dumps(estado_equipos_data)),
        'pie_chart_colors_equipos_json': mark_safe(json.dumps(pie_chart_colors_equipos)),
        'calibraciones_torta_labels_json': mark_safe(json.dumps(calibraciones_torta_labels)),
        'calibraciones_torta_data_json': mark_safe(json.dumps(calibraciones_torta_data)),
        'pie_chart_colors_cal_json': mark_safe(json.dumps(pie_chart_colors_cal_display)), # Usar la variable con el sufijo _display
        'comprobaciones_torta_labels_json': mark_safe(json.dumps(comprobaciones_torta_labels)),
        'comprobaciones_torta_data_json': mark_safe(json.dumps(comprobaciones_torta_data)),
        'pie_chart_colors_comp_json': mark_safe(json.dumps(pie_chart_colors_comp_display)), # Usar la variable con el sufijo _display
        
        # NUEVOS DATOS JSON PARA CUMPLIMIENTO DE MANTENIMIENTO
        'mantenimientos_cumplimiento_torta_labels_json': mark_safe(json.dumps(mantenimientos_cumplimiento_torta_labels)),
        'mantenimientos_cumplimiento_torta_data_json': mark_safe(json.dumps(mantenimientos_cumplimiento_torta_data)),
        'pie_chart_colors_mant_compliance_json': mark_safe(json.dumps(pie_chart_colors_mant_compliance_display)),

        # EXISTENTE: Datos para mantenimientos por tipo
        'mantenimientos_tipo_labels_json': mark_safe(json.dumps(mantenimientos_tipo_labels)),
        'mantenimientos_tipo_data_json': mark_safe(json.dumps(mantenimientos_tipo_data)),
        'pie_chart_colors_mant_types_json': mark_safe(json.dumps(pie_chart_colors_mant_types)),

        # Datos para el cuadro de mantenimientos correctivos
        'latest_corrective_maintenances': latest_corrective_maintenances,

        # Datos de almacenamiento
        'storage_usage_mb': storage_usage_mb,
        'storage_limit_mb': storage_limit_mb,
        'storage_percentage': storage_percentage,
        'storage_status_class': storage_status_class,
        'storage_empresa_nombre': storage_empresa_nombre,

        # Datos de límites de equipos
        'equipos_limite': equipos_limite,
        'equipos_actuales_count': equipos_actuales_count,
        'equipos_disponibles': equipos_disponibles,
        'equipos_limite_percentage': equipos_limite_percentage,
        'equipos_limite_warning': equipos_limite_warning,
        'equipos_limite_critical': equipos_limite_critical,

        # Información de plan actual (usando sistema unificado)
        'plan_info': {
            'plan_actual': user.empresa.get_plan_actual() if user.empresa else 'unknown',
            'estado_plan': user.empresa.get_estado_suscripcion_display() if user.empresa else 'Sin plan',
            'dias_restantes': user.empresa.get_dias_restantes_plan() if user.empresa else 0,
            'fecha_inicio': user.empresa.fecha_inicio_plan if user.empresa else None,
            'fecha_fin': user.empresa.get_fecha_fin_plan() if user.empresa else None,
            'limite_equipos_actual': user.empresa.get_limite_equipos() if user.empresa else 0,
            'limite_almacenamiento_actual': user.empresa.get_limite_almacenamiento() if user.empresa else 0,
        } if user.empresa else {},
    }
    return render(request, 'core/dashboard.html', context)


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/')
def activar_plan_pagado(request, empresa_id):
    """Vista para que superusuarios activen planes pagados."""
    empresa = get_object_or_404(Empresa, id=empresa_id)

    if request.method == 'POST':
        try:
            limite_equipos = int(request.POST.get('limite_equipos', 0))
            limite_almacenamiento_mb = int(request.POST.get('limite_almacenamiento_mb', 0))
            duracion_meses = request.POST.get('duracion_meses')
            duracion_meses = int(duracion_meses) if duracion_meses else None

            if limite_equipos <= 0 or limite_almacenamiento_mb <= 0:
                messages.error(request, 'Los límites deben ser mayores a 0')
                return redirect('core:detalle_empresa', pk=empresa_id)

            # Activar plan pagado
            empresa.activar_plan_pagado(
                limite_equipos=limite_equipos,
                limite_almacenamiento_mb=limite_almacenamiento_mb,
                duracion_meses=duracion_meses
            )

            messages.success(
                request,
                f'Plan pagado activado exitosamente para {empresa.nombre}. '
                f'Límites: {limite_equipos} equipos, {limite_almacenamiento_mb}MB. '
                f'{"Duración: " + str(duracion_meses) + " meses" if duracion_meses else "Sin límite de tiempo"}'
            )

        except ValueError as e:
            messages.error(request, f'Error en los datos proporcionados: {str(e)}')
        except Exception as e:
            messages.error(request, f'Error activando plan pagado: {str(e)}')

        return redirect('core:detalle_empresa', pk=empresa_id)

    return redirect('core:detalle_empresa', pk=empresa_id)

@access_check # APLICAR ESTE DECORADOR
@login_required
def contact_us(request):
    """
    Renders the contact us page.
    """
    return render(request, 'core/contact_us.html')

# --- Vistas de Equipos ---

@access_check # APLICAR ESTE DECORADOR
@login_required
def subir_pdf(request):
    """
    Vista para subir un archivo PDF y registrarlo en la base de datos.
    """
    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES, request=request) # Pasar el request al form
        archivo_subido = request.FILES.get('archivo') # Obtener el archivo directamente del request.FILES

        if form.is_valid() and archivo_subido: # Asegurarse de que el archivo también esté presente
            nombre_archivo = archivo_subido.name
            ruta_s3 = f'pdfs/{nombre_archivo}' # La ruta que se guardará en el modelo

            try:
                # Sube el archivo a S3 usando la función auxiliar
                subir_archivo(nombre_archivo, archivo_subido)

                # Guarda el objeto Documento en la base de datos
                documento = form.save(commit=False)
                documento.nombre_archivo = nombre_archivo # El nombre real del archivo
                documento.archivo_s3_path = ruta_s3 # La ruta completa en S3
                documento.subido_por = request.user
                if not request.user.is_superuser and request.user.empresa:
                    documento.empresa = request.user.empresa # Asigna la empresa automáticamente
                documento.save()

                messages.success(request, f'Archivo "{nombre_archivo}" subido y registrado exitosamente.')
                return redirect('core:home') # O a una lista de documentos si creas una
            except Exception as e:
                messages.error(request, f'Error al subir o registrar el archivo: {e}')
                logger.error(f'Error al subir archivo {nombre_archivo}: {e}')
        else:
            messages.error(request, 'Por favor, corrige los errores del formulario y asegúrate de seleccionar un archivo.')
    else:
        form = DocumentoForm(request=request) # Pasa el request al inicializar
    return render(request, 'core/subir_pdf.html', {'form': form, 'titulo_pagina': 'Subir Documento PDF'})
    
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.view_equipo', raise_exception=True)
def home(request):
    """
    Lists all equipment, with filtering and pagination.
    """
    user = request.user
    query = request.GET.get('q')
    tipo_equipo_filter = request.GET.get('tipo_equipo')
    estado_filter = request.GET.get('estado')
    
    # --- INICIO: Lógica para el filtro de empresa para superusuarios y obtener info de formato ---
    selected_company_id = request.GET.get('empresa_id')
    empresas_disponibles = Empresa.objects.all().order_by('nombre')

    equipos_list = Equipo.objects.all().order_by('codigo_interno')

    current_company_format_info = None # Initialize to None

    if not user.is_superuser:
        # Los usuarios normales solo ven datos de su propia empresa
        if user.empresa:
            equipos_list = equipos_list.filter(empresa=user.empresa)
            selected_company_id = str(user.empresa.id) # Asegurar que el filtro se aplique y muestre la empresa del usuario
            current_company_format_info = user.empresa # Set company info for regular user
        else:
            # Si un usuario normal no tiene empresa asignada, no ve equipos
            equipos_list = Equipo.objects.none()
            empresas_disponibles = Empresa.objects.none() # No hay empresas para filtrar

    else: # If user is superuser
        if selected_company_id:
            try:
                # Get the selected company object for format info
                current_company_format_info = Empresa.objects.get(pk=selected_company_id)
                equipos_list = equipos_list.filter(empresa_id=selected_company_id)
            except Empresa.DoesNotExist:
                # Handle case where selected_company_id is invalid
                messages.error(request, 'La empresa seleccionada no existe.')
                equipos_list = Equipo.objects.none()
        # If superuser and no company selected, current_company_format_info remains None,
        # meaning no specific company format info will be displayed at the top.

    # --- INICIO: LÓGICA DE VALIDACIÓN DE LÍMITE DE EQUIPOS ---
    limite_alcanzado = False
    empresa_para_limite = None
    if user.is_authenticated and not user.is_superuser:
        empresa_para_limite = user.empresa
    elif user.is_superuser and selected_company_id:
        try:
            empresa_para_limite = Empresa.objects.get(pk=selected_company_id)
        except Empresa.DoesNotExist:
            empresa_para_limite = None

    if empresa_para_limite: # Asegurarse de que hay una empresa válida
        # Obtener el límite de equipos usando el método del modelo Empresa
        limite_equipos_empresa = empresa_para_limite.get_limite_equipos() 
        
        if limite_equipos_empresa is not None and limite_equipos_empresa != float('inf') and limite_equipos_empresa > 0:
            equipos_actuales = Equipo.objects.filter(empresa=empresa_para_limite).count()
            if equipos_actuales >= limite_equipos_empresa:
                limite_alcanzado = True
    # --- FIN: LÓGICA DE VALIDACIÓN DE LÍMITE DE EQUIPOS ---

    today = timezone.localdate() # Obtener la fecha actual con la zona horaria configurada

    # Filtrar por query de búsqueda
    if query:
        equipos_list = equipos_list.filter(
            Q(codigo_interno__icontains=query) |
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(modelo__icontains=query) |
            Q(numero_serie__icontains=query) |
            Q(responsable__icontains=query) |
            Q(ubicacion__icontains=query) # Filtrar por el campo de texto libre de ubicación
        )

    # Filtrar por tipo de equipo
    if tipo_equipo_filter:
        equipos_list = equipos_list.filter(tipo_equipo=tipo_equipo_filter)

    # Filtro por estado
    if estado_filter:
        equipos_list = equipos_list.filter(estado=estado_filter)
    else:
        # Por defecto, no mostrar "De Baja" a menos que se filtre explícitamente por él
        if not user.is_superuser or (user.is_superuser and not selected_company_id):
            equipos_list = equipos_list.exclude(estado='De Baja').exclude(estado='Inactivo') # Se añadió exclusión de 'Inactivo'

    # Añadir lógica para el estado de las fechas de próxima actividad
    for equipo in equipos_list:
        # Calibración
        # Excluir De Baja e Inactivo para la proyección de estado visual
        if equipo.proxima_calibracion and equipo.estado not in ['De Baja', 'Inactivo']:
            days_remaining = (equipo.proxima_calibracion - today).days
            if days_remaining < 0:
                equipo.proxima_calibracion_status = 'text-red-600 font-bold' # Vencido
            elif days_remaining <= 15:
                equipo.proxima_calibracion_status = 'text-yellow-600 font-bold' # Próximos 15 días
            elif days_remaining <= 30:
                equipo.proxima_calibracion_status = 'text-green-600' # Próximos 30 días
            else:
                equipo.proxima_calibracion_status = 'text-gray-900' # Más de 30 días o futuro lejano (negro)
        else:
            equipo.proxima_calibracion_status = 'text-gray-500' # N/A o sin fecha o de baja o inactivo

        # Comprobación
        if equipo.proxima_comprobacion and equipo.estado not in ['De Baja', 'Inactivo']:
            days_remaining = (equipo.proxima_comprobacion - today).days
            if days_remaining < 0:
                equipo.proxima_comprobacion_status = 'text-red-600 font-bold' # Vencido
            elif days_remaining <= 15:
                equipo.proxima_comprobacion_status = 'text-yellow-600 font-bold' # Próximos 15 días
            elif days_remaining <= 30:
                equipo.proxima_comprobacion_status = 'text-green-600' # Próximos 30 días
            else:
                equipo.proxima_comprobacion_status = 'text-gray-900' # Más de 30 días o futuro lejano (negro)
        else:
            equipo.proxima_comprobacion_status = 'text-gray-500' # N/A o sin fecha o de baja o inactivo

        # Mantenimiento
        if equipo.proximo_mantenimiento and equipo.estado not in ['De Baja', 'Inactivo']:
            days_remaining = (equipo.proximo_mantenimiento - today).days
            if days_remaining < 0:
                equipo.proximo_mantenimiento_status = 'text-red-600 font-bold' # Vencido
            elif days_remaining <= 15:
                equipo.proximo_mantenimiento_status = 'text-yellow-600 font-bold' # Próximos 15 días
            elif days_remaining <= 30:
                equipo.proximo_mantenimiento_status = 'text-green-600' # Próximos 30 días
            else:
                equipo.proximo_mantenimiento_status = 'text-gray-900' # Más de 30 días o futuro lejano (negro)
        else:
            equipo.proximo_mantenimiento_status = 'text-gray-500' # N/A o sin fecha o de baja o inactivo


    paginator = Paginator(equipos_list, 10)
    page_number = request.GET.get('page')
    try:
        equipos = paginator.page(page_number)
    except PageNotAnInteger:
        equipos = paginator.page(1)
    except EmptyPage:
        equipos = paginator.page(paginator.num_pages)

    tipo_equipo_choices = Equipo.TIPO_EQUIPO_CHOICES
    estado_choices = Equipo.ESTADO_CHOICES

    context = {
        'equipos': equipos,
        'query': query, # Pasar el query de búsqueda al contexto
        'tipo_equipo_choices': tipo_equipo_choices,
        'estado_choices': estado_choices,
        'titulo_pagina': 'Listado de Equipos',
        'is_superuser': user.is_superuser, # Pasar is_superuser al contexto
        'empresas_disponibles': empresas_disponibles, # Pasar empresas_disponibles al contexto
        'selected_company_id': selected_company_id, # Pasar selected_company_id al contexto
        'current_company_format_info': current_company_format_info, # Información de formato de la empresa actual
        'limite_alcanzado': limite_alcanzado, # <--- SE AÑADIÓ ESTA VARIABLE
    }
    return render(request, 'core/home.html', context)

@access_check # APLICAR ESTE DECORADOR
@login_required
@require_POST # Esta vista es solo para peticiones POST de AJAX
@csrf_exempt
def update_empresa_formato(request):
    """
    Updates the format information for a company via an AJAX POST request.
    This view is intended for dynamic updates from the dashboard or similar.
    """
    # This view remains a POST-only endpoint for AJAX.
    # The new editar_empresa_formato view will handle the GET/POST for a dedicated page.
    # ... (existing code for this view remains unchanged as it's for AJAX) ...
    # Determine which company to update
    company_to_update = None
    if request.user.is_superuser:
        # Superuser can update any company if 'empresa_id' is provided
        empresa_id = request.POST.get('empresa_id')
        if empresa_id:
            try:
                company_to_update = Empresa.objects.get(pk=empresa_id)
            except Empresa.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Empresa no encontrada.'}, status=404)
        else:
            return JsonResponse({'status': 'error', 'message': 'ID de empresa requerido para superusuario.'}, status=400)
    elif request.user.empresa:
        # Regular user can only update their own company
        company_to_update = request.user.empresa
    else:
        return JsonResponse({'status': 'error', 'message': 'Usuario no asociado a ninguna empresa.'}, status=403)

    form = EmpresaFormatoForm(request.POST, instance=company_to_update)
    if form.is_valid():
        form.save()
        return JsonResponse({
            'status': 'success',
            'message': 'Información de formato actualizada.',
            'version': company_to_update.formato_version_empresa,
            'fecha_version': company_to_update.formato_fecha_version_empresa.strftime('%d/%m/%Y') if company_to_update.formato_fecha_version_empresa else 'N/A',
            'codificacion': company_to_update.formato_codificacion_empresa,
        })
    else:
        errors = form.errors.as_json()
        return JsonResponse({'status': 'error', 'message': 'Errores de validación.', 'errors': errors}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)


# NUEVA VISTA: Para editar la información de formato de una empresa (GET y POST)
@access_check # APLICAR ESTE DECORADOR
@login_required
@require_http_methods(["GET", "POST"])
def editar_empresa_formato(request, pk):
    """
    Handles editing the format information for a specific company (dedicated page).
    Superusers can edit any company. Regular users can only edit their own company.
    """
    empresa = get_object_or_404(Empresa, pk=pk)

    # Permiso: Superusuario o usuario asociado a la empresa
    if not request.user.is_superuser and request.user.empresa != empresa:
        messages.error(request, 'No tienes permiso para editar la información de formato de esta empresa.')
        return redirect('core:home') # Volver al home si no tiene permisos

    if request.method == 'POST':
        form = EmpresaFormatoForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            messages.success(request, f'Información de formato para "{empresa.nombre}" actualizada exitosamente.')
            return redirect('core:home') # Volver al home después de editar formato
        else:
            messages.error(request, 'Hubo un error al actualizar el formato. Por favor, revisa los datos.')
    else:
        form = EmpresaFormatoForm(instance=empresa)

    context = {
        'form': form,
        'empresa': empresa,
        'titulo_pagina': f'Editar Formato de Empresa: {empresa.nombre}',
    }
    return render(request, 'core/editar_empresa_formato.html', context)


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_equipo', raise_exception=True)
@trial_check
def añadir_equipo(request):
    empresa_actual = None
    if request.user.is_authenticated and not request.user.is_superuser:
        empresa_actual = request.user.empresa
    
    # Validar límite de equipos
    limite_alcanzado = False
    if empresa_actual:
        limite_equipos_empresa = empresa_actual.get_limite_equipos() 
        if limite_equipos_empresa not in [None, float('inf')] and limite_equipos_empresa > 0:
            equipos_actuales = Equipo.objects.filter(empresa=empresa_actual).count()
            if equipos_actuales >= limite_equipos_empresa:
                limite_alcanzado = True

    if request.method == 'POST':
        form = EquipoForm(request.POST, request.FILES, request=request)
        if form.is_valid():
            try:
                equipo = form.save(commit=False)

                # Validar límite de equipos antes de crear
                try:
                    StorageLimitValidator.validate_equipment_limit(equipo.empresa)
                except ValidationError as e:
                    messages.error(request, str(e))
                    return render(request, 'core/añadir_equipo.html', {
                        'form': form,
                        'titulo_pagina': 'Añadir Nuevo Equipo',
                        'limite_alcanzado': True,
                    })

                # Calcular tamaño total de archivos a subir
                total_file_size = 0
                for campo_form in ['manual_pdf', 'archivo_compra_pdf', 'ficha_tecnica_pdf', 'otros_documentos_pdf', 'imagen_equipo']:
                    if campo_form in request.FILES:
                        archivo = request.FILES[campo_form]
                        total_file_size += archivo.size

                # Validar límite de almacenamiento
                try:
                    StorageLimitValidator.validate_storage_limit(equipo.empresa, total_file_size)
                except ValidationError as e:
                    messages.error(request, str(e))
                    return render(request, 'core/añadir_equipo.html', {
                        'form': form,
                        'titulo_pagina': 'Añadir Nuevo Equipo',
                        'limite_alcanzado': limite_alcanzado,
                    })

                # --- Subida manual de archivos ---
                archivos = {
                    'manual_pdf': 'pdfs',
                    'archivo_compra_pdf': 'pdfs',
                    'ficha_tecnica_pdf': 'pdfs',
                    'otros_documentos_pdf': 'pdfs',
                    'imagen_equipo': 'imagenes_equipos',
                }

                for campo_form, carpeta_destino in archivos.items():
                    if campo_form in request.FILES:
                        archivo_subido = request.FILES[campo_form]
                        nombre_archivo = sanitize_filename(archivo_subido.name)
                        # Subir archivo con carpeta específica
                        ruta_final = f"{carpeta_destino}/{nombre_archivo}"
                        default_storage.save(ruta_final, archivo_subido)
                        setattr(equipo, campo_form, ruta_final)

                equipo.save()

                messages.success(request, 'Equipo añadido exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except forms.ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                logger.error(f"Error general al guardar equipo: {str(e)}")
                messages.error(request, f'Error al guardar el equipo: {str(e)}')
        else:
            # Mostrar errores del formulario
            logger.warning(f"Errores del formulario: {form.errors}")
            logger.warning(f"Errores no de campo: {form.non_field_errors()}")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
            if form.non_field_errors():
                for error in form.non_field_errors():
                    messages.error(request, str(error))

    else:
        form = EquipoForm(request=request)
    
    context = {
        'form': form,
        'titulo_pagina': 'Añadir Nuevo Equipo',
        'limite_alcanzado': limite_alcanzado,
    }
    return render(request, 'core/añadir_equipo.html', context)


@access_check
@login_required
@permission_required('core.add_equipo', raise_exception=True)
def descargar_plantilla_excel(request):
    """
    Genera y descarga una plantilla Excel mejorada con validaciones, instrucciones y ejemplos.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Protection
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
    from datetime import datetime, date

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plantilla Equipos"

    # La protección se aplicará al final, después de configurar todo

    # ==============================================
    # ENCABEZADO PROFESIONAL
    # ==============================================

    # Título principal (Z es la columna 26, suficiente para las nuevas columnas)
    sheet.merge_cells('A1:Z3')
    title_cell = sheet['A1']
    title_cell.value = "PLANTILLA DE IMPORTACIÓN DE EQUIPOS - SAM METROLOGÍA SAS"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Información importante
    sheet.merge_cells('A4:Z4')
    info_cell = sheet['A4']
    info_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Complete SOLO las filas de datos (fila 8 en adelante)"
    info_cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    info_cell.alignment = Alignment(horizontal="center")

    # ==============================================
    # HEADERS CON VALIDACIONES
    # ==============================================

    # Headers principales (fila 6)
    headers = [
        "codigo_interno", "nombre", "empresa_nombre", "tipo_equipo", "marca", "modelo",
        "numero_serie", "ubicacion_nombre", "responsable", "estado", "fecha_adquisicion",
        "fecha_ultima_calibracion", "fecha_ultimo_mantenimiento", "fecha_ultima_comprobacion",
        "rango_medida", "resolucion", "error_maximo_permisible", "observaciones",
        "version_formato", "fecha_version_formato", "codificacion_formato",
        "frecuencia_calibracion_meses", "frecuencia_mantenimiento_meses", "frecuencia_comprobacion_meses"
    ]

    # Headers legibles (fila 7)
    headers_legibles = [
        "Código Interno", "Nombre del Equipo", "Nombre de Empresa", "Tipo de Equipo", "Marca", "Modelo",
        "Número de Serie", "Ubicación", "Responsable", "Estado", "Fecha Adquisición",
        "Última Calibración", "Último Mantenimiento", "Última Comprobación",
        "Rango de Medida", "Resolución", "Error Máximo", "Observaciones",
        "Versión Formato", "Fecha Versión", "Codificación",
        "Calibración (meses)", "Mantenimiento (meses)", "Comprobación (meses)"
    ]

    # Configurar headers
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Headers técnicos (ocultos para usuarios)
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=6, column=col, value=header)
        cell.font = Font(name="Arial", size=8, color="666666")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Headers legibles (visibles)
    for col, header in enumerate(headers_legibles, 1):
        cell = sheet.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # ==============================================
    # VALIDACIONES Y LISTAS DESPLEGABLES
    # ==============================================

    # Obtener datos dinámicos del sistema
    empresas_disponibles = []
    if request.user.is_superuser:
        empresas_disponibles = [emp.nombre for emp in Empresa.objects.all()]
    else:
        empresas_disponibles = [request.user.empresa.nombre] if request.user.empresa else []

    tipos_equipo = [choice[0] for choice in Equipo.TIPO_EQUIPO_CHOICES]
    estados_equipo = [choice[0] for choice in Equipo.ESTADO_CHOICES]

    # Crear hoja de datos de validación
    validation_sheet = workbook.create_sheet("Datos_Validacion")
    validation_sheet.sheet_state = 'hidden'  # Ocultar hoja

    # Escribir listas de validación
    for i, empresa in enumerate(empresas_disponibles, 1):
        validation_sheet[f'A{i}'] = empresa

    for i, tipo in enumerate(tipos_equipo, 1):
        validation_sheet[f'B{i}'] = tipo

    for i, estado in enumerate(estados_equipo, 1):
        validation_sheet[f'C{i}'] = estado

    # Aplicar validaciones (desde fila 8 hasta 1000)
    start_row = 8
    end_row = 1000

    # Validación para empresas (columna C)
    if empresas_disponibles:
        dv_empresa = DataValidation(
            type="list",
            formula1=f"Datos_Validacion!$A$1:$A${len(empresas_disponibles)}",
            showErrorMessage=True,
            errorTitle="Empresa Inválida",
            error="Seleccione una empresa de la lista desplegable"
        )
        dv_empresa.add(f"C{start_row}:C{end_row}")
        sheet.add_data_validation(dv_empresa)

    # Validación para tipos de equipo (columna D)
    dv_tipo = DataValidation(
        type="list",
        formula1=f"Datos_Validacion!$B$1:$B${len(tipos_equipo)}",
        showErrorMessage=True,
        errorTitle="Tipo de Equipo Inválido",
        error="Seleccione un tipo de equipo de la lista desplegable"
    )
    dv_tipo.add(f"D{start_row}:D{end_row}")
    sheet.add_data_validation(dv_tipo)

    # Validación para estados (columna J)
    dv_estado = DataValidation(
        type="list",
        formula1=f"Datos_Validacion!$C$1:$C${len(estados_equipo)}",
        showErrorMessage=True,
        errorTitle="Estado Inválido",
        error="Seleccione un estado de la lista desplegable"
    )
    dv_estado.add(f"J{start_row}:J{end_row}")
    sheet.add_data_validation(dv_estado)

    # Validación para fechas (columnas K y Q)
    dv_fecha = DataValidation(
        type="date",
        operator="between",
        formula1=date(2000, 1, 1),
        formula2=date(2099, 12, 31),
        showErrorMessage=True,
        errorTitle="Fecha Inválida",
        error="Ingrese una fecha válida en formato DD/MM/AAAA"
    )
    dv_fecha.add(f"K{start_row}:K{end_row}")  # fecha_adquisicion
    dv_fecha.add(f"Q{start_row}:Q{end_row}")  # fecha_version_formato
    sheet.add_data_validation(dv_fecha)

    # Validación para números decimales (frecuencias)
    dv_decimal = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1=0,
        showErrorMessage=True,
        errorTitle="Número Inválido",
        error="Ingrese un número mayor a 0 (puede usar decimales: 0.5, 1, 12, etc.)"
    )
    dv_decimal.add(f"S{start_row}:S{end_row}")  # frecuencia_calibracion_meses
    dv_decimal.add(f"T{start_row}:T{end_row}")  # frecuencia_mantenimiento_meses
    dv_decimal.add(f"U{start_row}:U{end_row}")  # frecuencia_comprobacion_meses
    sheet.add_data_validation(dv_decimal)

    # ==============================================
    # FILA DE EJEMPLO
    # ==============================================

    ejemplo_data = [
        "EQ-001",  # codigo_interno
        "Balanza Analítica Ejemplo",  # nombre
        empresas_disponibles[0] if empresas_disponibles else "Mi Empresa",  # empresa_nombre
        tipos_equipo[0] if tipos_equipo else "Equipo de Medición",  # tipo_equipo
        "Mettler Toledo",  # marca
        "XPE205",  # modelo
        "B123456789",  # numero_serie
        "Laboratorio Principal",  # ubicacion_nombre
        "Técnico Responsable",  # responsable
        estados_equipo[0] if estados_equipo else "Activo",  # estado
        "15/01/2023",  # fecha_adquisicion
        "20/11/2024",  # fecha_ultima_calibracion
        "15/10/2024",  # fecha_ultimo_mantenimiento
        "10/12/2024",  # fecha_ultima_comprobacion
        "0-220g",  # rango_medida
        "0.1mg",  # resolucion
        "±0.1mg",  # error_maximo_permisible
        "Equipo nuevo para laboratorio",  # observaciones
        "V1.0",  # version_formato
        "01/01/2023",  # fecha_version_formato
        "CAL-001",  # codificacion_formato
        "12",  # frecuencia_calibracion_meses
        "6",  # frecuencia_mantenimiento_meses
        "3"   # frecuencia_comprobacion_meses
    ]

    # Aplicar ejemplo con estilo diferente
    for col, valor in enumerate(ejemplo_data, 1):
        cell = sheet.cell(row=8, column=col, value=valor)
        cell.font = Font(name="Arial", size=10, italic=True, color="666666")
        cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    # ==============================================
    # COMENTARIOS E INSTRUCCIONES
    # ==============================================

    # Agregar comentarios a headers críticos
    comentarios = {
        1: "OBLIGATORIO: Código único del equipo en su empresa",
        2: "OBLIGATORIO: Nombre descriptivo del equipo",
        3: "OBLIGATORIO: Debe coincidir exactamente con una empresa existente",
        4: "OBLIGATORIO: Seleccione de la lista desplegable",
        7: "OBLIGATORIO: Número de serie único",
        8: "OBLIGATORIO: Ubicación física del equipo",
        10: "OBLIGATORIO: Seleccione de la lista desplegable",
        11: "OPCIONAL: Fecha adquisición - DD/MM/AAAA (ej: 15/01/2023)",
        12: "OPCIONAL: Fecha última calibración - DD/MM/AAAA (ej: 20/11/2024)",
        13: "OPCIONAL: Fecha último mantenimiento - DD/MM/AAAA (ej: 15/10/2024)",
        14: "OPCIONAL: Fecha última comprobación - DD/MM/AAAA (ej: 10/12/2024)",
        20: "Formato: DD/MM/AAAA (ej: 01/01/2023)",
        22: "Número de meses (puede ser decimal: 0.5, 1, 12)",
        23: "Número de meses (puede ser decimal: 0.5, 1, 12)",
        24: "Número de meses (puede ser decimal: 0.5, 1, 12)"
    }

    for col, texto in comentarios.items():
        comment = Comment(texto, "SAM Sistema")
        sheet.cell(row=7, column=col).comment = comment

    # ==============================================
    # CONFIGURACIÓN DE PROTECCIÓN (DESHABILITADA PARA FACILIDAD DE USO)
    # ==============================================

    # Nota: Protección deshabilitada para permitir edición libre
    # Las validaciones proporcionarán la guía necesaria al usuario

    # ==============================================
    # FORMATO Y AJUSTES FINALES
    # ==============================================

    # Ocultar fila de headers técnicos
    sheet.row_dimensions[6].hidden = True

    # Ajustar anchos de columna (actualizados para incluir las 3 nuevas fechas)
    anchos = [15, 25, 20, 18, 15, 15, 15, 20, 20, 15, 15, 18, 18, 18, 15, 15, 15, 25, 15, 15, 15, 10, 10, 10]
    for i, ancho in enumerate(anchos, 1):
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = ancho

    # Ajustar alturas
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[7].height = 30

    # Congelar paneles para que headers siempre sean visibles
    sheet.freeze_panes = "A8"

    # Configurar área de impresión
    sheet.print_area = "A1:U50"

    # ==============================================
    # PROTECCIÓN DESHABILITADA PARA FACILIDAD DE USO
    # ==============================================

    # Sin protección de hoja para permitir edición libre
    # Las validaciones de datos proporcionarán la guía necesaria

    # Crear respuesta HTTP
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Plantilla_Equipos_SAM_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@access_check
@login_required
@permission_required('core.add_equipo', raise_exception=True)
def preview_equipos_excel(request):
    """
    Vista para previsualizar equipos antes de importar (validación previa).
    """
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]

                # Detectar tipo de plantilla
                headers_row = 1
                if headers and headers[0] and "PLANTILLA DE IMPORTACIÓN" in str(headers[0]):
                    headers_row = 6
                    headers = [cell.value for cell in sheet[headers_row]]

                # Validar headers
                required_headers = ['codigo_interno', 'nombre', 'empresa_nombre', 'tipo_equipo']
                missing_headers = [h for h in required_headers if h not in headers]

                # Obtener preview de datos
                preview_data = []
                start_row = headers_row + 1 if headers_row > 1 else 2
                count = 0

                for row in sheet.iter_rows(min_row=start_row, values_only=True):
                    if count >= 10:  # Limitar preview a 10 filas
                        break
                    if any(cell for cell in row if cell is not None and str(cell).strip()):
                        row_dict = dict(zip(headers, row))
                        preview_data.append(row_dict)
                        count += 1

                context = {
                    'preview_data': preview_data,
                    'total_detected': count,
                    'missing_headers': missing_headers,
                    'headers': headers,
                    'can_import': len(missing_headers) == 0,
                }

                return render(request, 'core/preview_equipos.html', context)

            except Exception as e:
                messages.error(request, f'Error procesando archivo: {e}')
                return redirect('core:importar_equipos_excel')

    return redirect('core:importar_equipos_excel')


def actualizar_equipo_selectivo(equipo_existente, nuevos_datos, row_index):
    """
    Actualiza solo los campos que NO están vacíos en los nuevos datos.

    Returns:
        list: Lista de nombres de campos que fueron actualizados
    """
    campos_actualizados = []

    # Mapeo de campos que se pueden actualizar (excluyendo empresa y codigo_interno que son identificadores)
    campos_actualizables = {
        'nombre': 'nombre',
        'tipo_equipo': 'tipo_equipo',
        'marca': 'marca',
        'modelo': 'modelo',
        'numero_serie': 'numero_serie',
        'ubicacion': 'ubicacion',
        'responsable': 'responsable',
        'estado': 'estado',
        'fecha_adquisicion': 'fecha_adquisicion',
        'fecha_ultima_calibracion': 'fecha_ultima_calibracion',
        'fecha_ultimo_mantenimiento': 'fecha_ultimo_mantenimiento',
        'fecha_ultima_comprobacion': 'fecha_ultima_comprobacion',
        'rango_medida': 'rango_medida',
        'resolucion': 'resolucion',
        'error_maximo_permisible': 'error_maximo_permisible',
        'observaciones': 'observaciones',
        'version_formato': 'version_formato',
        'fecha_version_formato': 'fecha_version_formato',
        'codificacion_formato': 'codificacion_formato',
        'frecuencia_calibracion_meses': 'frecuencia_calibracion_meses',
        'frecuencia_mantenimiento_meses': 'frecuencia_mantenimiento_meses',
        'frecuencia_comprobacion_meses': 'frecuencia_comprobacion_meses',
    }

    for campo_excel, campo_modelo in campos_actualizables.items():
        if campo_modelo in nuevos_datos:
            nuevo_valor = nuevos_datos[campo_modelo]

            # Solo actualizar si el nuevo valor no está vacío y es diferente al actual
            if es_valor_valido_para_actualizacion(nuevo_valor):
                valor_actual = getattr(equipo_existente, campo_modelo)

                # Comparar valores (manejar None y tipos diferentes)
                if valores_son_diferentes(valor_actual, nuevo_valor):
                    setattr(equipo_existente, campo_modelo, nuevo_valor)
                    campos_actualizados.append(campo_excel)

    # Guardar solo si hay cambios
    if campos_actualizados:
        equipo_existente.save()

    return campos_actualizados


def es_valor_valido_para_actualizacion(valor):
    """
    Determina si un valor es válido para actualizar (no está vacío).
    """
    if valor is None:
        return False
    if isinstance(valor, str) and valor.strip() == '':
        return False
    return True


def valores_son_diferentes(valor_actual, nuevo_valor):
    """
    Compara dos valores manejando diferentes tipos de datos.
    """
    # Si ambos son None, son iguales
    if valor_actual is None and nuevo_valor is None:
        return False

    # Si uno es None y el otro no, son diferentes
    if valor_actual is None or nuevo_valor is None:
        return True

    # Para strings, comparar sin espacios
    if isinstance(valor_actual, str) and isinstance(nuevo_valor, str):
        return valor_actual.strip() != nuevo_valor.strip()

    # Para otros tipos, comparación directa
    return valor_actual != nuevo_valor


def calcular_proximas_fechas_personalizadas(equipo):
    """
    Aplica la lógica personalizada para calcular próximas fechas de actividades.

    Lógica CORRECTA por tipo de actividad:
    1. Su propia fecha anterior (fecha_ultimo_X para actividad X)
    2. fecha_ultima_calibracion (como segunda opción)
    3. fecha_adquisicion (como tercera opción)
    4. fecha_registro (como último recurso)
    """
    from datetime import date
    from dateutil.relativedelta import relativedelta

    def obtener_fecha_base_para_actividad(fecha_propia, equipo):
        """
        Obtiene la fecha base siguiendo la jerarquía correcta para cada actividad.
        """
        if fecha_propia:
            return fecha_propia
        elif equipo.fecha_ultima_calibracion:
            return equipo.fecha_ultima_calibracion
        elif equipo.fecha_adquisicion:
            return equipo.fecha_adquisicion
        else:
            return equipo.fecha_registro.date()

    # CALIBRACIÓN: Usar su propia fecha como primera opción
    if equipo.frecuencia_calibracion_meses and equipo.frecuencia_calibracion_meses > 0:
        fecha_base_calibracion = obtener_fecha_base_para_actividad(
            equipo.fecha_ultima_calibracion, equipo
        )
        equipo.proxima_calibracion = fecha_base_calibracion + relativedelta(
            months=int(equipo.frecuencia_calibracion_meses)
        )

    # MANTENIMIENTO: Usar jerarquía específica para mantenimiento
    if equipo.frecuencia_mantenimiento_meses and equipo.frecuencia_mantenimiento_meses > 0:
        fecha_base_mantenimiento = obtener_fecha_base_para_actividad(
            equipo.fecha_ultimo_mantenimiento, equipo
        )
        equipo.proximo_mantenimiento = fecha_base_mantenimiento + relativedelta(
            months=int(equipo.frecuencia_mantenimiento_meses)
        )

    # COMPROBACIÓN: Usar jerarquía específica para comprobación
    if equipo.frecuencia_comprobacion_meses and equipo.frecuencia_comprobacion_meses > 0:
        fecha_base_comprobacion = obtener_fecha_base_para_actividad(
            equipo.fecha_ultima_comprobacion, equipo
        )
        equipo.proxima_comprobacion = fecha_base_comprobacion + relativedelta(
            months=int(equipo.frecuencia_comprobacion_meses)
        )

    # Guardar solo los campos de fechas calculadas
    equipo.save(update_fields=[
        'proxima_calibracion',
        'proximo_mantenimiento',
        'proxima_comprobacion'
    ])


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_equipo', raise_exception=True)
def importar_equipos_excel(request):
    """
    Handles importing equipment from an Excel file.
    """
    titulo_pagina = "Importar Equipos desde Excel"
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            if not excel_file.name.endswith('.xlsx'):
                messages.error(request, 'El archivo debe ser un archivo Excel (.xlsx).')
                return render(request, 'core/importar_equipos.html', {'form': form, 'titulo_pagina': titulo_pagina})

            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                
                headers = [cell.value for cell in sheet[1]]
                
                column_mapping = {
                    'codigo_interno': 'codigo_interno',
                    'nombre': 'nombre',
                    'empresa_nombre': 'empresa',
                    'tipo_equipo': 'tipo_equipo',
                    'marca': 'marca',
                    'modelo': 'modelo',
                    'numero_serie': 'numero_serie',
                    'ubicacion_nombre': 'ubicacion',
                    'responsable': 'responsable',
                    'estado': 'estado',
                    'fecha_adquisicion': 'fecha_adquisicion',
                    'fecha_ultima_calibracion': 'fecha_ultima_calibracion',
                    'fecha_ultimo_mantenimiento': 'fecha_ultimo_mantenimiento',
                    'fecha_ultima_comprobacion': 'fecha_ultima_comprobacion',
                    'rango_medida': 'rango_medida',
                    'resolucion': 'resolucion',
                    'error_maximo_permisible': 'error_maximo_permisible',
                    'observaciones': 'observaciones',
                    'version_formato': 'version_formato',
                    'fecha_version_formato': 'fecha_version_formato',
                    'codificacion_formato': 'codificacion_formato',
                    'frecuencia_calibracion_meses': 'frecuencia_calibracion_meses',
                    'frecuencia_mantenimiento_meses': 'frecuencia_mantenimiento_meses',
                    'frecuencia_comprobacion_meses': 'frecuencia_comprobacion_meses',
                }

                # Validación mejorada de headers
                required_headers = ['codigo_interno', 'nombre', 'empresa_nombre', 'tipo_equipo', 'marca', 'modelo', 'numero_serie', 'ubicacion_nombre', 'responsable', 'estado', 'fecha_adquisicion']

                # Detectar automáticamente la fila de headers
                headers_row = 1
                headers_found = False

                # Buscar la fila que contiene "codigo_interno" para detectar headers (debe ser exacto)
                for row_num in range(1, 11):  # Buscar en las primeras 10 filas
                    try:
                        test_headers = [cell.value for cell in sheet[row_num]]

                        # Buscar exactamente "codigo_interno" (no variaciones)
                        if test_headers and any(str(h).strip() == 'codigo_interno' for h in test_headers if h):
                            headers_row = row_num
                            headers = test_headers
                            headers_found = True
                            break
                    except Exception:
                        continue

                if not headers_found:
                    # Debug: Mostrar las primeras filas para diagnosticar
                    debug_info = []
                    for row_num in range(1, 8):
                        try:
                            row_data = [cell.value for cell in sheet[row_num]]
                            debug_info.append(f"Fila {row_num}: {row_data[:3]}...")  # Primeras 3 columnas
                        except:
                            pass

                    messages.error(
                        request,
                        f'❌ No se encontró la fila de encabezados con "codigo_interno". '
                        f'Debug: {"; ".join(debug_info)}. '
                        'Descarga la plantilla mejorada para evitar errores.'
                    )
                    return render(request, 'core/importar_equipos.html', {'form': form, 'titulo_pagina': titulo_pagina})

                # Normalizar headers para mapeo flexible
                normalized_headers = []
                for h in headers:
                    if h:
                        # Convertir a string, quitar espacios, convertir a minúsculas, reemplazar guiones
                        normalized = str(h).strip().lower().replace(' ', '_').replace('-', '_')
                        normalized_headers.append(normalized)
                    else:
                        normalized_headers.append('')

                # Crear mapeo flexible de headers
                header_indices = {}
                for i, norm_header in enumerate(normalized_headers):
                    for excel_col, model_field in column_mapping.items():
                        if excel_col.lower().replace(' ', '_').replace('-', '_') == norm_header:
                            header_indices[excel_col] = i
                            break

                # Verificar headers obligatorios con mapeo flexible
                missing_headers = []
                for req_header in required_headers:
                    if req_header not in header_indices:
                        missing_headers.append(req_header)

                if missing_headers:
                    messages.error(
                        request,
                        f'❌ Faltan encabezados obligatorios: {", ".join(missing_headers)}. '
                        f'Encabezados encontrados: {", ".join([h for h in headers if h])}. '
                        f'Descarga la plantilla mejorada para evitar errores.'
                    )
                    return render(request, 'core/importar_equipos.html', {'form': form, 'titulo_pagina': titulo_pagina})

                # Contar filas con datos reales (saltar filas de headers y ejemplo)
                total_rows = 0
                start_row = max(headers_row + 1, 8)  # Siempre empezar desde la fila 8 o después de headers

                for row in sheet.iter_rows(min_row=start_row, values_only=True):
                    if any(cell for cell in row if cell is not None and str(cell).strip()):
                        # Saltar fila de ejemplo si contiene "EQ-001"
                        if row and "EQ-001" in str(row[0]):
                            continue
                        total_rows += 1

                if total_rows == 0:
                    messages.error(request, '❌ El archivo no contiene datos para importar. Agregue equipos en las filas después de los encabezados.')
                    return render(request, 'core/importar_equipos.html', {'form': form, 'titulo_pagina': titulo_pagina})

                messages.info(request, f'📋 Archivo validado: {total_rows} equipos detectados para procesar...')

                created_count = 0
                errors = []
                
                with transaction.atomic():
                    start_row = max(headers_row + 1, 8)  # Siempre empezar desde la fila 8 o después de headers

                    for row_index, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                        # Saltar fila de ejemplo si contiene "EQ-001" independientemente de la posición
                        if row and "EQ-001" in str(row[0]):
                            continue

                        # Verificar si la fila tiene datos
                        if not any(cell for cell in row if cell is not None and str(cell).strip()):
                            continue

                        # Saltar si la fila contiene headers legibles (caso de error de detección)
                        if row and any(header in str(row[0]).lower() for header in ['código', 'codigo', 'nombre del equipo']):
                            continue

                        # Usar mapeo flexible de headers para extraer datos
                        equipo_data = {}
                        row_errors = []

                        for excel_col, model_field in column_mapping.items():
                            value = None
                            if excel_col in header_indices:
                                col_index = header_indices[excel_col]
                                if col_index < len(row):
                                    value = row[col_index]
                            
                            if excel_col in required_headers and (value is None or (isinstance(value, str) and value.strip() == '')):
                                row_errors.append(f"'{excel_col}' es un campo obligatorio y está vacío.")
                                continue

                            if value is None or (isinstance(value, str) and value.strip() == ''):
                                equipo_data[model_field] = None
                                continue

                            if excel_col == 'empresa_nombre':
                                if value and str(value).strip():
                                    empresa_nombre = str(value).strip()
                                    empresa = None

                                    # Buscar empresa por nombre exacto primero
                                    try:
                                        empresa = Empresa.objects.get(nombre__iexact=empresa_nombre)
                                    except Empresa.DoesNotExist:
                                        # Buscar por coincidencia parcial si no se encuentra exacta
                                        empresas_parciales = Empresa.objects.filter(nombre__icontains=empresa_nombre)
                                        if empresas_parciales.count() == 1:
                                            empresa = empresas_parciales.first()
                                        elif empresas_parciales.count() > 1:
                                            empresas_nombres = [e.nombre for e in empresas_parciales]
                                            row_errors.append(f"Múltiples empresas encontradas para '{value}': {', '.join(empresas_nombres)}. Use el nombre exacto.")
                                        else:
                                            # Buscar en empresas disponibles para el usuario
                                            if request.user.is_superuser:
                                                disponibles = [e.nombre for e in Empresa.objects.all()[:10]]
                                            else:
                                                disponibles = [request.user.empresa.nombre] if request.user.empresa else []
                                            row_errors.append(f"Empresa '{value}' no encontrada. Empresas disponibles: {', '.join(disponibles)}")

                                    if empresa:
                                        if not request.user.is_superuser and request.user.empresa != empresa:
                                            row_errors.append(f"No tienes permiso para añadir equipos a la empresa '{value}'.")
                                        else:
                                            equipo_data['empresa'] = empresa
                                else:
                                    row_errors.append("El nombre de la empresa es obligatorio y no puede estar vacío.")
                            elif excel_col == 'ubicacion_nombre':
                                equipo_data['ubicacion'] = str(value).strip() if value is not None else ''
                            elif excel_col == 'tipo_equipo':
                                valid_choices = [choice[0] for choice in Equipo.TIPO_EQUIPO_CHOICES]
                                # Validación flexible para tipo de equipo
                                if value and str(value).strip():
                                    value_str = str(value).strip()
                                    # Buscar coincidencia exacta primero
                                    if value_str in valid_choices:
                                        equipo_data[model_field] = value_str
                                    else:
                                        # Buscar coincidencia insensible a mayúsculas
                                        value_lower = value_str.lower()
                                        matched = False
                                        for choice in valid_choices:
                                            if choice.lower() == value_lower:
                                                equipo_data[model_field] = choice
                                                matched = True
                                                break
                                        if not matched:
                                            row_errors.append(f"Tipo de equipo '{value}' no es válido. Opciones disponibles: {', '.join(valid_choices)}.")
                                else:
                                    equipo_data[model_field] = value
                            elif excel_col == 'estado':
                                valid_choices = [choice[0] for choice in Equipo.ESTADO_CHOICES]
                                # Validación flexible para estado
                                if value and str(value).strip():
                                    value_str = str(value).strip()
                                    # Buscar coincidencia exacta primero
                                    if value_str in valid_choices:
                                        equipo_data[model_field] = value_str
                                    else:
                                        # Buscar coincidencia insensible a mayúsculas
                                        value_lower = value_str.lower()
                                        matched = False
                                        for choice in valid_choices:
                                            if choice.lower() == value_lower:
                                                equipo_data[model_field] = choice
                                                matched = True
                                                break
                                        if not matched:
                                            row_errors.append(f"Estado '{value}' no es válido. Opciones disponibles: {', '.join(valid_choices)}.")
                                else:
                                    equipo_data[model_field] = value
                            elif excel_col in ['fecha_adquisicion', 'fecha_version_formato', 'fecha_ultima_calibracion', 'fecha_ultimo_mantenimiento', 'fecha_ultima_comprobacion']:
                                parsed_date = None

                                # Manejar diferentes tipos de fechas
                                if isinstance(value, datetime):
                                    parsed_date = value.date()
                                elif isinstance(value, date):
                                    parsed_date = value
                                elif isinstance(value, (int, float)):
                                    # Excel guarda fechas como números seriales
                                    try:
                                        # Convertir número serial de Excel a fecha
                                        from datetime import date as dt_date, timedelta
                                        excel_epoch = dt_date(1899, 12, 30)  # Época base de Excel
                                        parsed_date = excel_epoch + timedelta(days=int(value))

                                        # Validar que la fecha sea razonable (entre 1900 y 2100)
                                        if parsed_date.year < 1900 or parsed_date.year > 2100:
                                            row_errors.append(f"Fecha fuera de rango válido para '{excel_col}': '{value}' (año {parsed_date.year}). Use fechas entre 1900-2100.")
                                            parsed_date = None
                                    except (ValueError, OverflowError):
                                        row_errors.append(f"Número de fecha inválido para '{excel_col}': '{value}'. Use formato de fecha válido.")
                                elif value and str(value).strip():
                                    # Intentar parsear como string con múltiples formatos
                                    date_str = str(value).strip()
                                    date_formats = [
                                        '%d/%m/%Y',     # DD/MM/YYYY (formato español común)
                                        '%Y/%m/%d',     # YYYY/MM/DD
                                        '%Y-%m-%d',     # YYYY-MM-DD (ISO)
                                        '%d-%m-%Y',     # DD-MM-YYYY
                                        '%m/%d/%Y',     # MM/DD/YYYY (formato US)
                                        '%d.%m.%Y',     # DD.MM.YYYY (formato europeo)
                                        '%Y.%m.%d',     # YYYY.MM.DD
                                    ]

                                    for date_format in date_formats:
                                        try:
                                            parsed_date = datetime.strptime(date_str, date_format).date()
                                            break
                                        except ValueError:
                                            continue

                                    if parsed_date is None:
                                        row_errors.append(f"Formato de fecha inválido para '{excel_col}': '{value}'. Use DD/MM/YYYY, YYYY/MM/DD, YYYY-MM-DD, o similares.")

                                equipo_data[model_field] = parsed_date
                            elif excel_col in ['frecuencia_calibracion_meses', 'frecuencia_mantenimiento_meses', 'frecuencia_comprobacion_meses']:
                                try:
                                    if value is not None and str(value).strip() != '':
                                        # Manejar diferentes tipos numéricos flexiblemente
                                        if isinstance(value, (int, float)):
                                            equipo_data[model_field] = decimal.Decimal(str(value))
                                        else:
                                            # Limpiar el string para permitir decimales con coma
                                            value_str = str(value).strip().replace(',', '.')
                                            equipo_data[model_field] = decimal.Decimal(value_str)
                                    else:
                                        equipo_data[model_field] = None
                                except (ValueError, TypeError, decimal.InvalidOperation):
                                    row_errors.append(f"Valor numérico inválido para '{excel_col}': '{value}'. Use números enteros o decimales.")
                            elif excel_col == 'error_maximo_permisible':
                                if value is not None:
                                    if isinstance(value, (int, float)):
                                        equipo_data[model_field] = str(value)
                                    else:
                                        equipo_data[model_field] = str(value).strip()
                                else:
                                    equipo_data[model_field] = ''
                            else:
                                # Manejar campos de texto de manera flexible
                                if value is not None:
                                    if isinstance(value, (int, float)):
                                        equipo_data[model_field] = str(value)
                                    else:
                                        equipo_data[model_field] = str(value).strip()
                                else:
                                    equipo_data[model_field] = value

                        # Validación y creación/actualización de equipos
                        if 'empresa' in equipo_data and 'codigo_interno' in equipo_data and not row_errors:
                            codigo_interno = equipo_data['codigo_interno']
                            empresa = equipo_data['empresa']

                            # Verificar si ya existe un equipo con el mismo código en la misma empresa
                            equipo_existente = Equipo.objects.filter(
                                empresa=empresa,
                                codigo_interno__iexact=codigo_interno
                            ).first()

                            if equipo_existente:
                                # ACTUALIZACIÓN SELECTIVA: Actualizar solo campos no vacíos
                                campos_actualizados = actualizar_equipo_selectivo(equipo_existente, equipo_data, row_index)
                                if campos_actualizados:
                                    # Recalcular fechas después de la actualización
                                    calcular_proximas_fechas_personalizadas(equipo_existente)
                                    errors.append(f"ACTUALIZADO: Equipo '{codigo_interno}' en '{empresa.nombre}' - Campos: {', '.join(campos_actualizados)}")
                                else:
                                    errors.append(f"SALTADO: Equipo '{codigo_interno}' en '{empresa.nombre}' - Todos los campos estaban vacíos o iguales.")
                            else:
                                # Crear nuevo equipo
                                try:
                                    equipo = Equipo.objects.create(**equipo_data)
                                    calcular_proximas_fechas_personalizadas(equipo)
                                    created_count += 1
                                except Exception as e:
                                    errors.append(f"Fila {row_index}: Error al crear el equipo - {e}")
                                    raise

                        elif row_errors:
                            errors.append(f"Fila {row_index}: {'; '.join(row_errors)}")

                # Clasificar mensajes por tipo para mejor manejo
                errores_criticos = []
                equipos_saltados = []
                equipos_actualizados = []

                for error in errors:
                    if "SALTADO:" in error:
                        equipos_saltados.append(error)
                    elif "ACTUALIZADO:" in error:
                        equipos_actualizados.append(error)
                    else:
                        errores_criticos.append(error)

                # Mostrar resumen de la importación
                total_procesados = created_count + len(equipos_actualizados) + len(equipos_saltados)

                if total_procesados > 0 or errores_criticos:
                    # Mensajes principales de resumen
                    if created_count > 0:
                        messages.success(
                            request,
                            f'✅ {created_count} equipos nuevos creados exitosamente.'
                        )

                    if equipos_actualizados:
                        messages.success(
                            request,
                            f'🔄 {len(equipos_actualizados)} equipos existentes actualizados.'
                        )

                    if equipos_saltados:
                        messages.info(
                            request,
                            f'ℹ️ {len(equipos_saltados)} equipos saltados (sin cambios necesarios).'
                        )

                    if errores_criticos:
                        messages.warning(
                            request,
                            f'⚠️ {len(errores_criticos)} errores encontrados que requieren corrección.'
                        )

                        # Mostrar solo los primeros 6 errores críticos
                        for err in errores_criticos[:6]:
                            if not ("SALTADO:" in err or "ACTUALIZADO:" in err):
                                messages.error(request, err)

                        if len(errores_criticos) > 6:
                            messages.info(request, f'... y {len(errores_criticos) - 6} errores adicionales.')

                    # Mostrar detalles de actualizaciones (primeras 4)
                    for actualizado in equipos_actualizados[:4]:
                        messages.success(request, actualizado.replace("ACTUALIZADO: ", "✏️ "))

                    if len(equipos_actualizados) > 4:
                        messages.info(request, f'... y {len(equipos_actualizados) - 4} actualizaciones adicionales.')

                    # Mostrar algunos equipos saltados como información (primeros 2)
                    for saltado in equipos_saltados[:2]:
                        messages.info(request, saltado.replace("SALTADO: ", "⏭️ "))

                    if len(equipos_saltados) > 2:
                        messages.info(request, f'... y {len(equipos_saltados) - 2} equipos saltados adicionales.')

                    return render(request, 'core/importar_equipos.html', {'form': form, 'titulo_pagina': titulo_pagina})
                else:
                    messages.success(
                        request,
                        f'🎉 ¡Importación perfecta! Se crearon {created_count} equipos correctamente. '
                        f'Todos los datos fueron procesados sin errores.'
                    )
                    return redirect('core:home')
            
            except Exception as e: # Este catch capturará la excepción relanzada si hay un error atómico
                messages.error(request, f'Ocurrió un error inesperado al procesar el archivo o la transacción: {e}')
                logger.error(f"Error en importar_equipos_excel: {e}")
                return render(request, 'core/importar_equipos.html', {'form': form, 'titulo_pagina': titulo_pagina})
        else:
            messages.error(request, 'Por favor, corrige los errores del formulario de subida.')
    else:
        form = ExcelUploadForm()
    
    return render(request, 'core/importar_equipos.html', {'form': form, 'titulo_pagina': titulo_pagina})


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.view_equipo', raise_exception=True)
def detalle_equipo(request, pk):
    """
    Displays the details of a specific equipment.
    """
    equipo = get_object_or_404(Equipo, pk=pk)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para ver este equipo.')
        return redirect('core:home')

    calibraciones = equipo.calibraciones.all().order_by('-fecha_calibracion')
    for cal in calibraciones:
        if equipo.frecuencia_calibracion_meses and cal.fecha_calibracion:
            cal.proxima_actividad_para_este_registro = cal.fecha_calibracion + relativedelta(months=int(equipo.frecuencia_calibracion_meses))
        else:
            cal.proxima_actividad_para_este_registro = None

    mantenimientos = equipo.mantenimientos.all().order_by('-fecha_mantenimiento')
    for mant in mantenimientos:
        if equipo.frecuencia_mantenimiento_meses and mant.fecha_mantenimiento:
            mant.proxima_actividad_para_este_registro = mant.fecha_mantenimiento + relativedelta(months=int(equipo.frecuencia_mantenimiento_meses))
        else:
            mant.proxima_actividad_para_este_registro = None

    comprobaciones = equipo.comprobaciones.all().order_by('-fecha_comprobacion')
    for comp in comprobaciones:
        if equipo.frecuencia_comprobacion_meses and comp.fecha_comprobacion:
            comp.proxima_actividad_para_este_registro = comp.fecha_comprobacion + relativedelta(months=int(equipo.frecuencia_comprobacion_meses))
        else:
            comp.proxima_actividad_para_este_registro = None

    baja_registro = None
    try:
        baja_registro = equipo.baja_registro
    except BajaEquipo.DoesNotExist:
        pass

    # Utilizar default_storage y sanitización para obtener URL segura
    def get_file_url(file_field, carpeta="pdfs"):
        if file_field and file_field.name:
            try:
                nombre_archivo_seguro = sanitize_filename(file_field.name)
                ruta = f"{carpeta}/{nombre_archivo_seguro}"
                if default_storage.exists(ruta):
                    return default_storage.url(ruta)
            except Exception as e:
                logger.error(f"Error al obtener URL para {file_field.name}: {e}")
        return None

    # Archivos asociados - usar template tags seguros
    logo_empresa_url = secure_file_url(equipo.empresa.logo_empresa) if equipo.empresa and equipo.empresa.logo_empresa else None
    imagen_equipo_url = secure_file_url(equipo.imagen_equipo)
    documento_baja_url = pdf_image_url(baja_registro.documento_baja) if baja_registro and baja_registro.documento_baja else None

    for cal in calibraciones:
        cal.documento_calibracion_url = pdf_image_url(cal.documento_calibracion)
        cal.confirmacion_metrologica_pdf_url = pdf_image_url(cal.confirmacion_metrologica_pdf)
        cal.intervalos_calibracion_pdf_url = pdf_image_url(cal.intervalos_calibracion_pdf)

    for mant in mantenimientos:
        mant.documento_mantenimiento_url = pdf_image_url(mant.documento_mantenimiento)

    for comp in comprobaciones:
        comp.documento_comprobacion_url = pdf_image_url(comp.documento_comprobacion)

    archivo_compra_pdf_url = pdf_image_url(equipo.archivo_compra_pdf)
    ficha_tecnica_pdf_url = pdf_image_url(equipo.ficha_tecnica_pdf)
    manual_pdf_url = pdf_image_url(equipo.manual_pdf)
    otros_documentos_pdf_url = pdf_image_url(equipo.otros_documentos_pdf)

    return render(request, 'core/detalle_equipo.html', {
        'equipo': equipo,
        'calibraciones': calibraciones,
        'mantenimientos': mantenimientos,
        'comprobaciones': comprobaciones,
        'baja_registro': baja_registro,
        'logo_empresa_url': logo_empresa_url,
        'imagen_equipo_url': imagen_equipo_url,
        'documento_baja_url': documento_baja_url,
        'archivo_compra_pdf_url': archivo_compra_pdf_url,
        'ficha_tecnica_pdf_url': ficha_tecnica_pdf_url,
        'manual_pdf_url': manual_pdf_url,
        'otros_documentos_pdf_url': otros_documentos_pdf_url,
        'titulo_pagina': f'Detalle de {equipo.nombre}',
    })

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_equipo', raise_exception=True)
@trial_check
def editar_equipo(request, pk):
    """
    Handles editing an existing equipment.
    """
    equipo = get_object_or_404(Equipo, pk=pk)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para editar este equipo.')
        return redirect('core:home')

    if request.method == 'POST':
        form = EquipoForm(request.POST, request.FILES, instance=equipo, request=request)
        if form.is_valid():
            try:
                equipo = form.save(commit=False)

                # Calcular tamaño de archivos nuevos a subir
                total_new_file_size = 0
                for campo_form in ['manual_pdf', 'archivo_compra_pdf', 'ficha_tecnica_pdf', 'otros_documentos_pdf', 'imagen_equipo']:
                    if campo_form in request.FILES:
                        archivo = request.FILES[campo_form]
                        total_new_file_size += archivo.size

                # Validar límite de almacenamiento para archivos nuevos
                if total_new_file_size > 0:
                    try:
                        StorageLimitValidator.validate_storage_limit(equipo.empresa, total_new_file_size)
                    except ValidationError as e:
                        messages.error(request, str(e))
                        form = EquipoForm(instance=equipo, request=request)
                        return render(request, 'core/editar_equipo.html', {
                            'form': form,
                            'equipo': equipo,
                            'titulo_pagina': f'Editar {equipo.nombre}',
                        })

                # --- Subida manual de archivos ---
                archivos = {
                    'manual_pdf': 'pdfs',
                    'archivo_compra_pdf': 'pdfs',
                    'ficha_tecnica_pdf': 'pdfs',
                    'otros_documentos_pdf': 'pdfs',
                    'imagen_equipo': 'imagenes_equipos',
                }

                for campo_form, carpeta_destino in archivos.items():
                    if campo_form in request.FILES:
                        archivo_subido = request.FILES[campo_form]
                        nombre_archivo = sanitize_filename(archivo_subido.name)
                        # Construir ruta final según tipo de archivo
                        ruta_final = f"{carpeta_destino}/{nombre_archivo}"
                        default_storage.save(ruta_final, archivo_subido)
                        setattr(equipo, campo_form, ruta_final)

                if not request.user.is_superuser and not equipo.empresa:
                    equipo.empresa = request.user.empresa

                equipo.save()
                messages.success(request, 'Equipo actualizado exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except Exception as e:
                messages.error(request, f'Error al actualizar el equipo: {e}')
    else:
        form = EquipoForm(instance=equipo, request=request)

    return render(request, 'core/editar_equipo.html', {
        'form': form,
        'equipo': equipo,
        'titulo_pagina': f'Editar Equipo: {equipo.nombre}'
    })


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.delete_equipo', raise_exception=True)
def eliminar_equipo(request, pk):
    """
    Handles deleting an equipment.
    """
    equipo = get_object_or_404(Equipo, pk=pk)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para eliminar este equipo.')
        return redirect('core:home')

    if request.method == 'POST':
        try:
            equipo_nombre = equipo.nombre
            equipo.delete()
            messages.success(request, f'Equipo "{equipo_nombre}" eliminado exitosamente.')
            # Redirige a la página principal después de eliminar para evitar NoReverseMatch
            return redirect('core:home') 
        except Exception as e:
            messages.error(request, f'Error al eliminar el equipo: {e}')
            logger.error(f"Error al eliminar equipo {equipo.pk}: {e}") 
            # Si hay un error, redirige a home, ya que detalle_equipo podría no ser válido
            return redirect('core:home') 
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'el equipo "{equipo.nombre}" (Código: {equipo.codigo_interno})',
        'return_url_name': 'core:detalle_equipo', # URL a la que volver si se cancela
        'return_url_pk': equipo.pk, # PK para la URL de retorno si es un detalle, o None si es un listado
        'titulo_pagina': f'Eliminar Equipo: {equipo.nombre}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)

# --- Vistas de Calibraciones ---

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_calibracion', raise_exception=True)
def añadir_calibracion(request, equipo_pk):
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para añadir calibraciones a este equipo.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        form = CalibracionForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Calcular tamaño de archivos de calibración
                total_file_size = 0
                archivos_calibracion = ['documento_calibracion', 'confirmacion_metrologica_pdf', 'intervalos_calibracion_pdf']
                for campo in archivos_calibracion:
                    if campo in request.FILES:
                        archivo = request.FILES[campo]
                        total_file_size += archivo.size

                # Validar límite de almacenamiento
                if total_file_size > 0:
                    try:
                        StorageLimitValidator.validate_storage_limit(equipo.empresa, total_file_size)
                    except ValidationError as e:
                        messages.error(request, str(e))
                        form = CalibracionForm()
                        return render(request, 'core/añadir_calibracion.html', {
                            'form': form,
                            'equipo': equipo,
                            'titulo_pagina': f'Añadir Calibración a {equipo.nombre}',
                        })

                calibracion = Calibracion(
                    equipo=equipo,
                    fecha_calibracion=form.cleaned_data['fecha_calibracion'],
                    proveedor=form.cleaned_data['proveedor'],
                    nombre_proveedor=form.cleaned_data['nombre_proveedor'],
                    resultado=form.cleaned_data['resultado'],
                    numero_certificado=form.cleaned_data['numero_certificado'],
                    observaciones=form.cleaned_data['observaciones'],
                )

                # --- Manejo de archivos ---
                archivos = [
                    'documento_calibracion',
                    'confirmacion_metrologica_pdf',
                    'intervalos_calibracion_pdf',
                ]

                for campo in archivos:
                    if campo in request.FILES:
                        archivo_subido = request.FILES[campo]
                        nombre_archivo = sanitize_filename(archivo_subido.name)
                        ruta_final = f"pdfs/{nombre_archivo}"
                        default_storage.save(ruta_final, archivo_subido)
                        setattr(calibracion, campo, ruta_final)

                calibracion.save()

                messages.success(request, 'Calibración añadida exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except Exception as e:
                logger.error(f"ERROR al guardar calibración o archivo: {e}")
                messages.error(request, f'Hubo un error al guardar la calibración: {e}')

    else:
        form = CalibracionForm()

    return render(request, 'core/añadir_calibracion.html', {
        'form': form,
        'equipo': equipo,
        'titulo_pagina': f'Añadir Calibración para {equipo.nombre}'
    })
    
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_calibracion', raise_exception=True)
def editar_calibracion(request, equipo_pk, pk):
    """
    Handles editing an existing calibration.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    calibracion = get_object_or_404(Calibracion, pk=pk, equipo=equipo)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para editar esta calibración.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        form = CalibracionForm(request.POST, request.FILES, instance=calibracion)
        if form.is_valid():
            try:
                calibracion = form.save(commit=False)

                # --- Manejo de archivos (PDFs) ---
                archivos = [
                    'documento_calibracion',
                    'confirmacion_metrologica_pdf',
                    'intervalos_calibracion_pdf',
                ]

                for campo in archivos:
                    if campo in request.FILES:
                        archivo_subido = request.FILES[campo]
                        nombre_archivo = sanitize_filename(archivo_subido.name)
                        ruta_final = f"pdfs/{nombre_archivo}"
                        default_storage.save(ruta_final, archivo_subido)
                        setattr(calibracion, campo, ruta_final)

                calibracion.save()

                messages.success(request, 'Calibración actualizada exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except Exception as e:
                logger.error(f"ERROR al actualizar calibración o archivo: {e}")
                messages.error(request, f'Hubo un error al actualizar la calibración: {e}')
    else:
        form = CalibracionForm(instance=calibracion)

    return render(request, 'core/editar_calibracion.html', {
        'form': form,
        'equipo': equipo,
        'calibracion': calibracion,
        'titulo_pagina': f'Editar Calibración para {equipo.nombre}'
    })
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.delete_calibracion', raise_exception=True)
def eliminar_calibracion(request, equipo_pk, pk):
    """
    Handles deleting a calibration.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    calibracion = get_object_or_404(Calibracion, pk=pk, equipo=equipo)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para eliminar esta calibración.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        try:
            calibracion.delete()
            messages.success(request, 'Calibración eliminada exitosamente.')
            return redirect('core:detalle_equipo', pk=equipo.pk)
        except Exception as e:
            messages.error(request, f'Error al eliminar la calibración: {e}')
            logger.error(f"Error al eliminar calibración {calibracion.pk}: {e}") 
            return redirect('core:detalle_equipo', pk=equipo.pk)
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'la calibración de {equipo.nombre}',
        'return_url_name': 'core:detalle_equipo', # URL a la que volver si se cancela
        'return_url_pk': equipo.pk, # PK para la URL de retorno
        'titulo_pagina': f'Eliminar Calibración de {equipo.nombre}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


# --- Vistas de Mantenimientos ---

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_mantenimiento', raise_exception=True)
def añadir_mantenimiento(request, equipo_pk):
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para añadir mantenimientos a este equipo.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        form = MantenimientoForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Validar límite de almacenamiento para archivos de mantenimiento
                if 'documento_mantenimiento' in request.FILES:
                    archivo = request.FILES['documento_mantenimiento']
                    try:
                        StorageLimitValidator.validate_storage_limit(equipo.empresa, archivo.size)
                    except ValidationError as e:
                        messages.error(request, str(e))
                        form = MantenimientoForm()
                        return render(request, 'core/añadir_mantenimiento.html', {
                            'form': form,
                            'equipo': equipo,
                            'titulo_pagina': f'Añadir Mantenimiento a {equipo.nombre}',
                        })

                mantenimiento = Mantenimiento(
                    equipo=equipo,
                    fecha_mantenimiento=form.cleaned_data['fecha_mantenimiento'],
                    tipo_mantenimiento=form.cleaned_data['tipo_mantenimiento'],
                    proveedor=form.cleaned_data['proveedor'],
                    nombre_proveedor=form.cleaned_data['nombre_proveedor'],
                    responsable=form.cleaned_data['responsable'],
                    costo=form.cleaned_data['costo'],
                    descripcion=form.cleaned_data['descripcion'],
                    observaciones=form.cleaned_data['observaciones'],
                )

                # --- Manejo del archivo (PDF) ---
                if 'documento_mantenimiento' in request.FILES:
                    archivo_subido = request.FILES['documento_mantenimiento']
                    nombre_archivo = sanitize_filename(archivo_subido.name)
                    ruta_final = f"pdfs/{nombre_archivo}"
                    default_storage.save(ruta_final, archivo_subido)
                    mantenimiento.documento_mantenimiento = ruta_final

                mantenimiento.save()

                messages.success(request, 'Mantenimiento añadido exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except Exception as e:
                logger.error(f"ERROR al guardar mantenimiento o archivo: {e}")
                messages.error(request, f'Hubo un error al guardar el mantenimiento: {e}')

    else:
        form = MantenimientoForm()

    return render(request, 'core/añadir_mantenimiento.html', {
        'form': form,
        'equipo': equipo,
        'titulo_pagina': f'Añadir Mantenimiento para {equipo.nombre}'
    })

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_mantenimiento', raise_exception=True)
def editar_mantenimiento(request, equipo_pk, pk):
    """
    Handles editing an existing maintenance record.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    mantenimiento = get_object_or_404(Mantenimiento, pk=pk, equipo=equipo)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para editar este mantenimiento.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        form = MantenimientoForm(request.POST, request.FILES, instance=mantenimiento)
        if form.is_valid():
            try:
                mantenimiento = form.save(commit=False)

                # --- Manejo del archivo (PDF) ---
                if 'documento_mantenimiento' in request.FILES:
                    archivo_subido = request.FILES['documento_mantenimiento']
                    nombre_archivo = sanitize_filename(archivo_subido.name)
                    ruta_final = f"pdfs/{nombre_archivo}"
                    default_storage.save(ruta_final, archivo_subido)
                    mantenimiento.documento_mantenimiento = ruta_final

                mantenimiento.save()

                messages.success(request, 'Mantenimiento actualizado exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except Exception as e:
                logger.error(f"ERROR al actualizar mantenimiento o archivo: {e}")
                messages.error(request, f'Hubo un error al actualizar el mantenimiento: {e}')
    else:
        form = MantenimientoForm(instance=mantenimiento)

    return render(request, 'core/editar_mantenimiento.html', {
        'form': form,
        'equipo': equipo,
        'mantenimiento': mantenimiento,
        'titulo_pagina': f'Editar Mantenimiento para {equipo.nombre}'
    })
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.delete_mantenimiento', raise_exception=True)
def eliminar_mantenimiento(request, equipo_pk, pk):
    """
    Handles deleting a maintenance record.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    mantenimiento = get_object_or_404(Mantenimiento, pk=pk, equipo=equipo)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para eliminar este mantenimiento.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        try:
            mantenimiento.delete()
            messages.success(request, 'Mantenimiento eliminado exitosamente.')
            return redirect('core:detalle_equipo', pk=equipo.pk)
        except Exception as e:
            messages.error(request, f'Error al eliminar el mantenimiento: {e}')
            logger.error(f"Error al eliminar mantenimiento {mantenimiento.pk}: {e}") 
            return redirect('core:detalle_equipo', pk=equipo.pk)
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'el mantenimiento de {equipo.nombre}',
        'return_url_name': 'core:detalle_equipo', # URL a la que volver si se cancela
        'return_url_pk': equipo.pk, # PK para la URL de retorno
        'titulo_pagina': f'Eliminar Mantenimiento de {equipo.nombre}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.view_mantenimiento', raise_exception=True)
def detalle_mantenimiento(request, equipo_pk, pk):
    """
    Displays the details of a specific maintenance record.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    mantenimiento = get_object_or_404(Mantenimiento, pk=pk, equipo=equipo)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para ver este mantenimiento.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    # --- Generar URL del documento de mantenimiento ---
    def get_file_url(file_field, carpeta="pdfs"):
        if file_field and file_field.name:
            try:
                nombre_archivo_seguro = sanitize_filename(file_field.name)
                ruta = f"{carpeta}/{nombre_archivo_seguro}"
                if default_storage.exists(ruta):
                    return default_storage.url(ruta)
            except Exception as e:
                logger.error(f"Error al obtener URL para {file_field.name}: {e}")
        return None

    documento_mantenimiento_url = pdf_image_url(mantenimiento.documento_mantenimiento)

    context = {
        'equipo': equipo,
        'mantenimiento': mantenimiento,
        'documento_mantenimiento_url': documento_mantenimiento_url,
        'titulo_pagina': f'Detalle de Mantenimiento: {equipo.nombre}',
    }
    return render(request, 'core/detalle_mantenimiento.html', context)

# --- Vistas de Comprobaciones ---

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_comprobacion', raise_exception=True)
def añadir_comprobacion(request, equipo_pk):
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para añadir comprobaciones a este equipo.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        form = ComprobacionForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Validar límite de almacenamiento para archivos de comprobación
                if 'documento_comprobacion' in request.FILES:
                    archivo = request.FILES['documento_comprobacion']
                    try:
                        StorageLimitValidator.validate_storage_limit(equipo.empresa, archivo.size)
                    except ValidationError as e:
                        messages.error(request, str(e))
                        form = ComprobacionForm()
                        return render(request, 'core/añadir_comprobacion.html', {
                            'form': form,
                            'equipo': equipo,
                            'titulo_pagina': f'Añadir Comprobación a {equipo.nombre}',
                        })

                comprobacion = Comprobacion(
                    equipo=equipo,
                    fecha_comprobacion=form.cleaned_data['fecha_comprobacion'],
                    proveedor=form.cleaned_data['proveedor'],
                    nombre_proveedor=form.cleaned_data['nombre_proveedor'],
                    responsable=form.cleaned_data['responsable'],
                    resultado=form.cleaned_data['resultado'],
                    observaciones=form.cleaned_data['observaciones'],
                )

                # --- Manejo de archivo (PDF) ---
                if 'documento_comprobacion' in request.FILES:
                    archivo_subido = request.FILES['documento_comprobacion']
                    nombre_archivo = sanitize_filename(archivo_subido.name)
                    ruta_final = f"pdfs/{nombre_archivo}"
                    default_storage.save(ruta_final, archivo_subido)
                    comprobacion.documento_comprobacion = ruta_final

                comprobacion.save()

                messages.success(request, 'Comprobación añadida exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except Exception as e:
                logger.error(f"ERROR al guardar comprobación o archivo: {e}")
                messages.error(request, f'Hubo un error al guardar la comprobación: {e}')

    else:
        form = ComprobacionForm()

    return render(request, 'core/añadir_comprobacion.html', {
        'form': form,
        'equipo': equipo,
        'titulo_pagina': f'Añadir Comprobación para {equipo.nombre}'
    })
    
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_comprobacion', raise_exception=True)
def editar_comprobacion(request, equipo_pk, pk):
    """
    Handles editing an existing verification record.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    comprobacion = get_object_or_404(Comprobacion, pk=pk, equipo=equipo)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para editar esta comprobación.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        form = ComprobacionForm(request.POST, request.FILES, instance=comprobacion)
        if form.is_valid():
            try:
                comprobacion = form.save(commit=False)

                # --- Manejo del archivo (PDF) ---
                if 'documento_comprobacion' in request.FILES:
                    archivo_subido = request.FILES['documento_comprobacion']
                    nombre_archivo = sanitize_filename(archivo_subido.name)
                    ruta_final = f"pdfs/{nombre_archivo}"
                    default_storage.save(ruta_final, archivo_subido)
                    comprobacion.documento_comprobacion = ruta_final

                comprobacion.save()

                messages.success(request, 'Comprobación actualizada exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except Exception as e:
                logger.error(f"ERROR al actualizar comprobación o archivo: {e}")
                messages.error(request, f'Hubo un error al actualizar la comprobación: {e}')
    else:
        form = ComprobacionForm(instance=comprobacion)

    return render(request, 'core/editar_comprobacion.html', {
        'form': form,
        'equipo': equipo,
        'comprobacion': comprobacion,
        'titulo_pagina': f'Editar Comprobación para {equipo.nombre}'
    })
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.delete_comprobacion', raise_exception=True)
def eliminar_comprobacion(request, equipo_pk, pk):
    """
    Handles deleting a verification record.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)
    comprobacion = get_object_or_404(Comprobacion, pk=pk, equipo=equipo)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para eliminar esta comprobación.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if request.method == 'POST':
        try:
            comprobacion.delete()
            messages.success(request, 'Comprobación eliminada exitosamente.')
            return redirect('core:detalle_equipo', pk=equipo.pk)
        except Exception as e:
            messages.error(request, f'Error al eliminar la comprobación: {e}')
            logger.error(f"Error al eliminar comprobación {comprobacion.pk}: {e}") 
            return redirect('core:detalle_equipo', pk=equipo.pk)
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'la comprobación de {equipo.nombre}',
        'return_url_name': 'core:detalle_equipo', # URL a la que volver si se cancela
        'return_url_pk': equipo.pk, # PK para la URL de retorno
        'titulo_pagina': f'Eliminar Comprobación de {equipo.nombre}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


# --- Vistas de Baja de Equipo y Nueva Inactivación ---

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_bajaequipo', raise_exception=True)
@require_http_methods(["GET", "POST"]) # Asegúrate de permitir GET para la página de confirmación
def dar_baja_equipo(request, equipo_pk):
    equipo = get_object_or_404(Equipo, pk=equipo_pk)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para dar de baja este equipo.')
        return redirect('core:detalle_equipo', pk=equipo.pk)

    if equipo.estado == 'De Baja':
        messages.warning(request, f'El equipo "{equipo.nombre}" ya está dado de baja.')
        return redirect('core:detalle_equipo', pk=equipo.pk)
    
    if request.method == 'POST':
        form = BajaEquipoForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Crear el objeto manualmente
                baja_registro = BajaEquipo(
                    equipo=equipo,
                    fecha_baja=form.cleaned_data['fecha_baja'],
                    razon_baja=form.cleaned_data['razon_baja'],
                    observaciones=form.cleaned_data['observaciones'],
                    dado_de_baja_por=request.user,
                )
                
                # --- Manejo del archivo y subida a S3 ---
                if 'documento_baja' in request.FILES:
                    archivo_subido = request.FILES['documento_baja']
                    nombre_archivo = sanitize_filename(archivo_subido.name)
                    subir_archivo(nombre_archivo, archivo_subido)
                    # Guardar la ruta completa que coincida con donde se sube el archivo
                    baja_registro.documento_baja = f"pdfs/{nombre_archivo}"

                # Guardar en DB
                baja_registro.save()
                
                messages.success(request, f'Equipo "{equipo.nombre}" dado de baja exitosamente.')
                return redirect('core:detalle_equipo', pk=equipo.pk)

            except Exception as e:
                messages.error(request, f'Hubo un error al dar de baja el equipo: {e}')
                logger.error(f"Error al dar de baja equipo {equipo.pk}: {e}")
                return render(request, 'core/dar_baja_equipo.html', {
                    'form': form,
                    'equipo': equipo,
                    'titulo_pagina': f'Dar de Baja Equipo: {equipo.nombre}'
                })
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario de baja.')
    else:
        form = BajaEquipoForm()
    
    return render(request, 'core/dar_baja_equipo.html', {
        'form': form,
        'equipo': equipo,
        'titulo_pagina': f'Dar de Baja Equipo: {equipo.nombre}'
    })

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_equipo', raise_exception=True)
@require_http_methods(["GET", "POST"])
def inactivar_equipo(request, equipo_pk):
    """
    Inactiva un equipo (cambia su estado a 'Inactivo').
    Esto es para una pausa temporal, no una baja definitiva.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para inactivar este equipo.')
        return redirect('core:detalle_equipo', pk=equipo_pk)

    if equipo.estado == 'Inactivo':
        messages.info(request, f'El equipo "{equipo.nombre}" ya está inactivo.')
        return redirect('core:detalle_equipo', pk=equipo_pk)
    elif equipo.estado == 'De Baja':
        messages.error(request, f'El equipo "{equipo.nombre}" ha sido dado de baja de forma permanente y no puede ser inactivado.')
        return redirect('core:detalle_equipo', pk=equipo_pk)
    
    if request.method == 'POST':
        equipo.estado = 'Inactivo'
        # Poner a None las próximas fechas al inactivar
        equipo.proxima_calibracion = None
        equipo.proximo_mantenimiento = None
        equipo.proxima_comprobacion = None
        equipo.save(update_fields=['estado', 'proxima_calibracion', 'proximo_mantenimiento', 'proxima_comprobacion'])
        messages.success(request, f'Equipo "{equipo.nombre}" inactivado exitosamente.')
        return redirect('core:detalle_equipo', pk=equipo_pk)
    
    # Mostrar página de confirmación si es GET
    return render(request, 'core/confirmar_inactivacion.html', {
        'equipo': equipo,
        'titulo_pagina': f'Inactivar Equipo: {equipo.nombre}'
    })


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_equipo', raise_exception=True)
@require_http_methods(["GET", "POST"])
def activar_equipo(request, equipo_pk):
    """
    Activa un equipo (cambia su estado de 'Inactivo' o 'De Baja' a 'Activo').
    Si estaba 'De Baja', elimina el registro de BajaEquipo asociado.
    """
    equipo = get_object_or_404(Equipo, pk=equipo_pk)

    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para activar este equipo.')
        return redirect('core:detalle_equipo', pk=equipo_pk)

    if equipo.estado == 'Activo':
        messages.info(request, f'El equipo "{equipo.nombre}" ya está activo.')
        return redirect('core:detalle_equipo', pk=equipo_pk)

    if request.method == 'POST':
        if equipo.estado == 'De Baja':
            try:
                baja_registro = BajaEquipo.objects.get(equipo=equipo)
                baja_registro.delete() # Esto activará el equipo a través de la señal post_delete
                messages.success(request, f'Equipo "{equipo.nombre}" activado exitosamente y registro de baja eliminado.')
            except BajaEquipo.DoesNotExist:
                equipo.estado = 'Activo'
                equipo.save(update_fields=['estado'])
                messages.warning(request, f'Equipo "{equipo.nombre}" activado. No se encontró registro de baja asociado.')
            except Exception as e:
                messages.error(request, f'Error al activar el equipo y eliminar el registro de baja: {e}')
                logger.error(f"Error al activar equipo {equipo.pk} (De Baja): {e}")
                return redirect('core:detalle_equipo', pk=equipo.pk)
        
        elif equipo.estado == 'Inactivo':
            equipo.estado = 'Activo'
            # Es crucial recalcular las próximas fechas cuando un equipo pasa de Inactivo a Activo
            equipo.calcular_proxima_calibracion()
            equipo.calcular_proximo_mantenimiento()
            equipo.calcular_proxima_comprobacion()
            equipo.save()
            messages.success(request, f'Equipo "{equipo.nombre}" activado exitosamente.')
            
        return redirect('core:detalle_equipo', pk=equipo.pk)
    
    return render(request, 'core/confirmar_activacion.html', {
        'equipo': equipo,
        'titulo_pagina': f'Activar Equipo: {equipo.nombre}'
    })


# --- Vistas de Empresas ---

@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/')
def listar_empresas(request):
    """
    Lists all companies, with filtering and pagination (superuser only).
    """
    query = request.GET.get('q')
    empresas_list = Empresa.objects.all()

    if not request.user.is_superuser:
        empresas_list = empresas_list.filter(pk=request.user.empresa.pk) # Un usuario normal solo ve su propia empresa

    if query:
        empresas_list = empresas_list.filter(
            Q(nombre__icontains=query) |
            Q(nit__icontains=query) |
            Q(direccion__icontains=query) |
            Q(telefono__icontains=query) |
            Q(email__icontains=query)
        )

    # Las propiedades fecha_fin_plan_display y fecha_fin_plan_status
    # ya están implementadas como @property en el modelo Empresa


    paginator = Paginator(empresas_list, 10)
    page_number = request.GET.get('page')
    try:
        empresas = paginator.page(page_number)
    except PageNotAnInteger:
        empresas = paginator.page(1)
    except EmptyPage:
        empresas = paginator.page(paginator.num_pages)

    return render(request, 'core/listar_empresas.html', {'empresas': empresas, 'query': query, 'titulo_pagina': 'Listado de Empresas'})

@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/')
def añadir_empresa(request):
    """ Handles adding a new company (superuser only). """
    if request.method == 'POST':
        formulario = EmpresaForm(request.POST)
        if formulario.is_valid():
            try:
                empresa = formulario.save(commit=False)

                if 'logo_empresa' in request.FILES:
                    archivo_subido = request.FILES['logo_empresa']
                    nombre_archivo = sanitize_filename(archivo_subido.name)
                    ruta_s3 = f'empresas_logos/{nombre_archivo}'
                    default_storage.save(ruta_s3, archivo_subido)
                    empresa.logo_empresa = nombre_archivo

                empresa.save()
                messages.success(request, 'Empresa añadida exitosamente.')
                return redirect('core:listar_empresas')
            except Exception as e:
                messages.error(request, f'Hubo un error al añadir la empresa: {e}. Revisa el log para más detalles.')
                logger.error(f"Error al añadir empresa: {e}")
        else:
            messages.error(request, 'Hubo un error al añadir la empresa. Por favor, revisa los datos.')
    else:
        formulario = EmpresaForm()

    return render(request, 'core/añadir_empresa.html', {
        'formulario': formulario,
        'titulo_pagina': 'Añadir Nueva Empresa'
    })

@access_check # APLICAR ESTE DECORADOR
@login_required
def detalle_empresa(request, pk):
    """
    Displays the details of a specific company and its associated equipment.
    """
    empresa = get_object_or_404(Empresa, pk=pk)

    # Verificar permisos: solo superusuario o usuario de la misma empresa
    if not request.user.is_superuser and (not request.user.empresa or request.user.empresa.pk != empresa.pk):
        messages.error(request, 'No tienes permisos para ver esta empresa.')
        return redirect('core:access_denied')

    # Obtener los equipos asociados a esta empresa
    equipos_asociados = Equipo.objects.filter(empresa=empresa).order_by('codigo_interno')

    # Obtener los usuarios asociados a esta empresa
    usuarios_empresa = CustomUser.objects.filter(empresa=empresa).order_by('username') # <--- SE AÑADIÓ ESTA LÍNEA

    context = {
        'empresa': empresa,
        'equipos_asociados': equipos_asociados,
        'usuarios_empresa': usuarios_empresa, # <--- SE AÑADIÓ ESTA LÍNEA AL CONTEXTO
        'titulo_pagina': f'Detalle de Empresa: {empresa.nombre}'
    }
    return render(request, 'core/detalle_empresa.html', context)


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/')
def editar_empresa(request, pk):
    """
    Handles editing an existing company (superuser only).
    """
    empresa = get_object_or_404(Empresa, pk=pk)
    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            try:
                empresa = form.save(commit=False)
                
                if 'logo_empresa' in request.FILES:
                    archivo_subido = request.FILES['logo_empresa']
                    nombre_archivo = sanitize_filename(archivo_subido.name)
                    ruta_s3 = f'empresas_logos/{nombre_archivo}'
                    default_storage.save(ruta_s3, archivo_subido)
                    empresa.logo_empresa = ruta_s3
                    logger.info(f'Archivo subido a: {ruta_s3}')
                
                empresa.save()
                messages.success(request, 'Empresa actualizada exitosamente.')
                return redirect('core:detalle_empresa', pk=empresa.pk)
            except Exception as e:
                messages.error(request, f'Error al actualizar la empresa: {e}')
                logger.error(f"Error al editar empresa: {e}")
        else:
            messages.error(request, 'Hubo un error al actualizar la empresa. Por favor, revisa los datos.')
    else:
        form = EmpresaForm(instance=empresa)
    return render(request, 'core/editar_empresa.html', {'form': form, 'empresa': empresa, 'titulo_pagina': f'Editar Empresa: {empresa.nombre}'})

@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/')
def eliminar_empresa(request, pk):
    """
    Handles deleting a company (superuser only).
    """
    empresa = get_object_or_404(Empresa, pk=pk)
    if request.method == 'POST':
        try:
            empresa_nombre = empresa.nombre # Capturar el nombre antes de eliminar
            empresa.delete()
            messages.success(request, f'Empresa "{empresa_nombre}" eliminada exitosamente.')
            return redirect('core:listar_empresas')
        except Exception as e:
            messages.error(request, f'Error al eliminar la empresa: {e}')
            logger.error(f"Error al eliminar empresa {empresa.pk}: {e}")
            return redirect('core:listar_empresas')
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'la empresa "{empresa.nombre}"',
        'return_url_name': 'core:listar_empresas', # URL a la que volver si se cancela
        'return_url_pk': None, # No se necesita PK para la lista de empresas
        'titulo_pagina': f'Eliminar Empresa: {empresa.nombre}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_empresa', raise_exception=True) # Este permiso es apropiado para administrar usuarios de una empresa
def añadir_usuario_a_empresa(request, empresa_pk):
    """
    Vista para añadir un usuario existente a una empresa específica.
    Solo accesible por superusuarios o usuarios con permiso para cambiar empresas.
    """
    empresa = get_object_or_404(Empresa, pk=empresa_pk)
    titulo_pagina = f"Añadir Usuario a {empresa.nombre}"

    # REVISAR: Permiso: Superusuario o usuario asociado a la empresa (si la empresa es la del usuario)
    if not request.user.is_superuser and request.user.empresa != empresa:
        messages.error(request, 'No tienes permiso para añadir usuarios a esta empresa.')
        return redirect('core:detalle_empresa', pk=empresa.pk)

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        if user_id:
            try:
                user_to_add = CustomUser.objects.get(pk=user_id)
                if user_to_add.empresa and user_to_add.empresa != empresa:
                    messages.warning(request, f"El usuario '{user_to_add.username}' ya está asociado a la empresa '{user_to_add.empresa.nombre}'. Se ha reasignado a '{empresa.nombre}'.")
                
                user_to_add.empresa = empresa
                user_to_add.save()
                messages.success(request, f"Usuario '{user_to_add.username}' añadido exitosamente a '{empresa.nombre}'.")
                return redirect('core:detalle_empresa', pk=empresa.pk)
            except CustomUser.DoesNotExist:
                messages.error(request, "El usuario seleccionado no existe.")
            except Exception as e:
                messages.error(request, f"Error al añadir usuario: {e}")
                logger.debug(f" Error en añadir_usuario_a_empresa: {e}")
        else:
            messages.error(request, "Por favor, selecciona un usuario.")
    
    users_available = CustomUser.objects.filter(is_superuser=False).exclude(empresa=empresa)

    context = {
        'empresa': empresa,
        'users_available': users_available,
        'titulo_pagina': titulo_pagina,
    }
    return render(request, 'core/añadir_usuario_a_empresa.html', context)


# --- Vistas de Ubicación ---
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.view_ubicacion', raise_exception=True)
def listar_ubicaciones(request):
    """
    Lists all locations.
    """
    ubicaciones = Ubicacion.objects.all()
    # Filtrar por empresa si el usuario no es superusuario
    if not request.user.is_superuser and request.user.empresa:
        ubicaciones = ubicaciones.filter(empresa=request.user.empresa)
    elif not request.user.is_superuser and not request.user.empresa: # Usuario normal sin empresa asignada
        ubicaciones = Ubicacion.objects.none()

    return render(request, 'core/listar_ubicaciones.html', {'ubicaciones': ubicaciones, 'titulo_pagina': 'Listado de Ubicaciones'})

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_ubicacion', raise_exception=True)
def añadir_ubicacion(request):
    """
    Handles adding a new location.
    """
    if request.method == 'POST':
        # Pasar el request al formulario
        form = UbicacionForm(request.POST, request=request)
        if form.is_valid():
            ubicacion = form.save(commit=False)
            if not request.user.is_superuser and not ubicacion.empresa:
                ubicacion.empresa = request.user.empresa
            ubicacion.save()
            messages.success(request, 'Ubicación añadida exitosamente.')
            return redirect('core:listar_ubicaciones')
        else:
            messages.error(request, 'Hubo un error al añadir la ubicación. Por favor, revisa los datos.')
    else:
        # Pasar el request al formulario
        form = UbicacionForm(request=request)
    return render(request, 'core/añadir_ubicacion.html', {'form': form, 'titulo_pagina': 'Añadir Nueva Ubicación'})

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_ubicacion', raise_exception=True)
def editar_ubicacion(request, pk):
    """
    Handles editing an existing location.
    """
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    # Permiso: Superusuario o usuario asociado a la empresa de la ubicación
    if not request.user.is_superuser and request.user.empresa != ubicacion.empresa:
        messages.error(request, 'No tienes permiso para editar esta ubicación.')
        return redirect('core:listar_ubicaciones')

    if request.method == 'POST':
        # Pasar el request al formulario
        form = UbicacionForm(request.POST, instance=ubicacion, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ubicación actualizada exitosamente.')
            return redirect('core:listar_ubicaciones')
        else:
            messages.error(request, 'Hubo un error al actualizar la ubicación. Por favor, revisa los datos.')
    else:
        # Pasar el request al formulario
        form = UbicacionForm(instance=ubicacion, request=request)
    return render(request, 'core/editar_ubicacion.html', {'form': form, 'ubicacion': ubicacion, 'titulo_pagina': f'Editar Ubicación: {ubicacion.nombre}'})

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.delete_ubicacion', raise_exception=True)
def eliminar_ubicacion(request, pk):
    """
    Handles deleting a location.
    """
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    # Permiso: Superusuario o usuario asociado a la empresa de la ubicación
    if not request.user.is_superuser and request.user.empresa != ubicacion.empresa:
        messages.error(request, 'No tienes permiso para eliminar esta ubicación.')
        return redirect('core:listar_ubicaciones')

    if request.method == 'POST':
        try:
            nombre_ubicacion = ubicacion.nombre # Capturar el nombre antes de eliminar
            ubicacion.delete()
            messages.success(request, f'Ubicación "{nombre_ubicacion}" eliminada exitosamente.')
            return redirect('core:listar_ubicaciones')
        except Exception as e:
            messages.error(request, f'Error al eliminar la ubicación: {e}')
            logger.error(f"Error al eliminar ubicación {ubicacion.pk}: {e}")
            return redirect('core:listar_ubicaciones')
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'la ubicación "{ubicacion.nombre}"',
        'return_url_name': 'core:listar_ubicaciones', # URL a la que volver si se cancela
        'return_url_pk': None, # No se necesita PK para la lista de ubicaciones
        'titulo_pagina': f'Eliminar Ubicación: {ubicacion.nombre}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


# --- Vistas de Procedimiento ---
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.view_procedimiento', raise_exception=True)
def listar_procedimientos(request):
    """
    Lists all procedures, filtered by user's company or selected company for superusers.
    Also passes company format info.
    """
    user = request.user
    selected_company_id = request.GET.get('empresa_id') # Para superusuarios

    # Query inicial de todos los procedimientos
    procedimientos_list = Procedimiento.objects.all().order_by('codigo')

    current_company_format_info = None

    if not user.is_superuser:
        if user.empresa:
            procedimientos_list = procedimientos_list.filter(empresa=user.empresa)
            current_company_format_info = user.empresa
            selected_company_id = str(user.empresa.id) # Asegura que el filtro muestre la empresa del usuario
        else:
            procedimientos_list = Procedimiento.objects.none() # Usuario sin empresa no ve procedimientos
            messages.info(request, "No tienes una empresa asociada. Contacta al administrador para asignar una.")
    else: # Es superusuario
        if selected_company_id:
            try:
                current_company_format_info = Empresa.objects.get(pk=selected_company_id)
                procedimientos_list = procedimientos_list.filter(empresa_id=selected_company_id)
            except Empresa.DoesNotExist:
                messages.error(request, 'La empresa seleccionada no existe.')
                procedimientos_list = Procedimiento.objects.none()
        # Si superusuario y no selected_company_id, current_company_format_info queda None,
        # lo que significa que no se mostrará información de formato de empresa específica.
        # En este caso, el superusuario ve TODOS los procedimientos de TODAS las empresas.

    # Lógica de paginación
    paginator = Paginator(procedimientos_list, 10)
    page_number = request.GET.get('page')
    try:
        procedimientos = paginator.page(page_number)
    except PageNotAnInteger:
        procedimientos = paginator.page(1)
    except EmptyPage:
        procedimientos = paginator.page(paginator.num_pages)

    # Lista de empresas disponibles para el filtro (solo si es superusuario)
    empresas_disponibles = Empresa.objects.all().order_by('nombre') if user.is_superuser else Empresa.objects.none()


    context = {
        'procedimientos': procedimientos,
        'titulo_pagina': 'Listado de Procedimientos',
        'current_company_format_info': current_company_format_info,
        'is_superuser': user.is_superuser,
        'empresas_disponibles': empresas_disponibles,
        'selected_company_id': selected_company_id,
    }
    return render(request, 'core/listar_procedimientos.html', context)

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_procedimiento', raise_exception=True)
@require_http_methods(["GET", "POST"])
def añadir_procedimiento(request):
    """
    Handles adding a new procedure.
    """
    if request.method == 'POST':
        form = ProcedimientoForm(request.POST, request.FILES, request=request)
        if form.is_valid():
            try:
                # Validar límite de almacenamiento antes de subir archivo
                archivo_subido = request.FILES.get("documento_pdf")
                if archivo_subido:
                    # Obtener empresa del procedimiento
                    procedimiento_temp = form.save(commit=False)
                    empresa = procedimiento_temp.empresa

                    # Validar límite de almacenamiento
                    try:
                        StorageLimitValidator.validate_storage_limit(empresa, archivo_subido.size)
                    except ValidationError as e:
                        messages.error(request, str(e))
                        return render(request, 'core/añadir_procedimiento.html', {'form': form, 'titulo_pagina': 'Añadir Nuevo Procedimiento'})

                # Solo continuar si pasó la validación de límites
                # 1. obtener el archivo desde request.FILES
                archivo_subido = request.FILES["documento_pdf"]

                # 2. obtener y sanitizar el nombre del archivo
                nombre_archivo = sanitize_filename(archivo_subido.name)

                # 3. pasar a tu función (contenido puede ser directamente el objeto archivo)
                subir_archivo(nombre_archivo, archivo_subido)
                procedimiento = form.save(commit=False)
                # La lógica de empresa ya se maneja en el formulario
                procedimiento.save()
                messages.success(request, 'Procedimiento añadido exitosamente.')
                return redirect('core:listar_procedimientos')
            except ValidationError as ve:
                # Manejar específicamente errores de validación (límites)
                messages.error(request, str(ve))
                return render(request, 'core/añadir_procedimiento.html', {'form': form, 'titulo_pagina': 'Añadir Nuevo Procedimiento'})
            except Exception as e:
                messages.error(request, f'Hubo un error al añadir el procedimiento: {e}. Revisa el log para más detalles.')
                logger.error(f"Error al añadir procedimiento: {e}")
                return render(request, 'core/añadir_procedimiento.html', {'form': form, 'titulo_pagina': 'Añadir Nuevo Procedimiento'})
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = ProcedimientoForm(request=request) # Pasa el request al formulario para la lógica de empresa
    return render(request, 'core/añadir_procedimiento.html', {'form': form, 'titulo_pagina': 'Añadir Nuevo Procedimiento'})
    
@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_procedimiento', raise_exception=True)
@require_http_methods(["GET", "POST"])
def editar_procedimiento(request, pk):
    """
    Handles editing an existing procedure.
    """
    procedimiento = get_object_or_404(Procedimiento, pk=pk)

    # Permiso: Superusuario o usuario asociado a la empresa del procedimiento
    if not request.user.is_superuser and (not request.user.empresa or request.user.empresa != procedimiento.empresa):
        messages.error(request, 'No tienes permiso para editar este procedimiento.')
        return redirect('core:listar_procedimientos')

    if request.method == 'POST':
        form = ProcedimientoForm(request.POST, request.FILES, instance=procedimiento, request=request)
        if form.is_valid():
            try:
                # Validar límite de almacenamiento si se sube un nuevo archivo
                nuevo_archivo = request.FILES.get("documento_pdf")
                if nuevo_archivo:
                    # Validar límite de almacenamiento
                    try:
                        StorageLimitValidator.validate_storage_limit(procedimiento.empresa, nuevo_archivo.size)
                    except ValidationError as e:
                        messages.error(request, str(e))
                        return render(request, 'core/editar_procedimiento.html', {'form': form, 'procedimiento': procedimiento, 'titulo_pagina': f'Editar Procedimiento: {procedimiento.nombre}'})

                # La lógica de empresa ya se maneja en el formulario, solo guardar
                form.save()
                messages.success(request, 'Procedimiento actualizado exitosamente.')
                return redirect('core:listar_procedimientos')
            except ValidationError as ve:
                # Manejar específicamente errores de validación (límites)
                messages.error(request, str(ve))
                return render(request, 'core/editar_procedimiento.html', {'form': form, 'procedimiento': procedimiento, 'titulo_pagina': f'Editar Procedimiento: {procedimiento.nombre}'})
            except Exception as e:
                messages.error(request, f'Hubo un error al actualizar el procedimiento: {e}. Revisa el log para más detalles.')
                logger.error(f"Error al actualizar procedimiento: {e}")
                return render(request, 'core/editar_procedimiento.html', {'form': form, 'procedimiento': procedimiento, 'titulo_pagina': f'Editar Procedimiento: {procedimiento.nombre}'})
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = ProcedimientoForm(instance=procedimiento, request=request)
    return render(request, 'core/editar_procedimiento.html', {'form': form, 'procedimiento': procedimiento, 'titulo_pagina': f'Editar Procedimiento: {procedimiento.nombre}'})

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.delete_procedimiento', raise_exception=True)
@require_http_methods(["GET", "POST"])
def eliminar_procedimiento(request, pk):
    """
    Handles deleting a procedure.
    """
    procedimiento = get_object_or_404(Procedimiento, pk=pk)

    # Permiso: Superusuario o usuario asociado a la empresa del procedimiento
    if not request.user.is_superuser and (not request.user.empresa or request.user.empresa != procedimiento.empresa):
        messages.error(request, 'No tienes permiso para eliminar este procedimiento.')
        return redirect('core:listar_procedimientos')

    if request.method == 'POST':
        try:
            nombre_proc = procedimiento.nombre
            procedimiento.delete()
            messages.success(request, f'Procedimiento "{nombre_proc}" eliminado exitosamente.')
            return redirect('core:listar_procedimientos')
        except Exception as e:
            messages.error(request, f'Error al eliminar el procedimiento: {e}. Revisa el log para más detalles.')
            logger.error(f"Error al eliminar procedimiento {procedimiento.pk}: {e}")
            return redirect('core:listar_procedimientos')
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'el procedimiento "{procedimiento.nombre}"',
        'return_url_name': 'core:listar_procedimientos', # URL a la que volver si se cancela
        'return_url_pk': None, # No se necesita PK para la lista de procedimientos
        'titulo_pagina': f'Eliminar Procedimiento: {procedimiento.nombre}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


# --- Vistas de Proveedores (GENERAL) ---

@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.view_proveedor', raise_exception=True)
def listar_proveedores(request):
    """
    Lists all general providers, with filtering and pagination.
    """
    query = request.GET.get('q')
    tipo_servicio_filter = request.GET.get('tipo_servicio')

    proveedores_list = Proveedor.objects.all().order_by('nombre_empresa')

    if not request.user.is_superuser:
        if request.user.empresa:
            proveedores_list = proveedores_list.filter(empresa=request.user.empresa)
        else:
            proveedores_list = Proveedor.objects.none() # Usuario normal sin empresa asignada

    if query:
        proveedores_list = proveedores_list.filter(
            Q(nombre_empresa__icontains=query) |
            Q(nombre_contacto__icontains=query) |
            Q(correo_electronico__icontains=query) |
            Q(alcance__icontains=query) |
            Q(servicio_prestado__icontains=query)
        )
    
    if tipo_servicio_filter and tipo_servicio_filter != '': 
        proveedores_list = proveedores_list.filter(tipo_servicio=tipo_servicio_filter)


    paginator = Paginator(proveedores_list, 10)
    page_number = request.GET.get('page')
    try:
        proveedores = paginator.page(page_number)
    except PageNotAnInteger:
        proveedores = paginator.page(1)
    except EmptyPage:
        proveedores = paginator.page(paginator.num_pages)

    tipo_servicio_choices = Proveedor.TIPO_SERVICIO_CHOICES

    context = {
        'proveedores': proveedores,
        'query': query,
        'tipo_servicio_choices': tipo_servicio_choices,
        'tipo_servicio_filter': tipo_servicio_filter,
        'titulo_pagina': 'Listado de Proveedores',
    }
    return render(request, 'core/listar_proveedores.html', context)


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.add_proveedor', raise_exception=True)
def añadir_proveedor(request):
    """
    Handles adding a new general provider.
    """
    if request.method == 'POST':
        form = ProveedorForm(request.POST, request=request)
        if form.is_valid():
            proveedor = form.save(commit=False)
            # La lógica de empresa ya se maneja en el formulario
            proveedor.save()
            messages.success(request, 'Proveedor añadido exitosamente.')
            return redirect('core:listar_proveedores')
        else:
            messages.error(request, 'Hubo un error al añadir el proveedor. Por favor, revisa los datos.')
    else:
        form = ProveedorForm(request=request)

    return render(request, 'core/añadir_proveedor.html', {'form': form, 'titulo_pagina': 'Añadir Nuevo Proveedor'})


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.change_proveedor', raise_exception=True)
def editar_proveedor(request, pk):
    """
    Handles editing an existing general provider.
    """
    proveedor = get_object_or_404(Proveedor, pk=pk)

    if not request.user.is_superuser and (not request.user.empresa or proveedor.empresa != request.user.empresa):
        messages.error(request, 'No tienes permiso para editar este proveedor.')
        return redirect('core:listar_proveedores')

    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor, request=request)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.save()
            messages.success(request, 'Proveedor actualizado exitosamente.')
            return redirect('core:listar_proveedores')
        else:
            messages.error(request, 'Hubo un error al actualizar el proveedor. Por favor, revisa los datos.')
    else:
        form = ProveedorForm(instance=proveedor, request=request)
    return render(request, 'core/editar_proveedor.html', {'form': form, 'proveedor': proveedor, 'titulo_pagina': f'Editar Proveedor: {proveedor.nombre_empresa}'})


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.delete_proveedor', raise_exception=True)
def eliminar_proveedor(request, pk):
    """
    Handles deleting a general provider.
    """
    proveedor = get_object_or_404(Proveedor, pk=pk)

    if not request.user.is_superuser and (not request.user.empresa or proveedor.empresa != request.user.empresa):
        messages.error(request, 'No tienes permiso para eliminar este proveedor.')
        return redirect('core:listar_proveedores')

    if request.method == 'POST':
        try:
            nombre_proveedor = proveedor.nombre_empresa # Capturar el nombre antes de eliminar
            proveedor.delete()
            messages.success(request, f'Proveedor "{nombre_proveedor}" eliminado exitosamente.')
            return redirect('core:listar_proveedores')
        except Exception as e:
            messages.error(request, f'Error al eliminar el proveedor: {e}')
            logger.error(f"Error al eliminar proveedor {proveedor.pk}: {e}")
            return redirect('core:listar_proveedores')
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'el proveedor "{proveedor.nombre_empresa}"',
        'return_url_name': 'core:listar_proveedores', # URL a la que volver si se cancela
        'return_url_pk': None, # No se necesita PK para la lista de proveedores
        'titulo_pagina': f'Eliminar Proveedor: {proveedor.nombre_empresa}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.view_proveedor', raise_exception=True)
def detalle_proveedor(request, pk):
    """
    Displays the details of a specific general provider.
    """
    proveedor = get_object_or_404(Proveedor, pk=pk)

    if not request.user.is_superuser and (not request.user.empresa or proveedor.empresa != request.user.empresa):
        messages.error(request, 'No tienes permiso para ver este proveedor.')
        return redirect('core:listar_proveedores')

    context = {
        'proveedor': proveedor,
        'titulo_pagina': f'Detalle de Proveedor: {proveedor.nombre_empresa}',
    }
    return render(request, 'core/detalle_proveedor.html', context)


# --- Vistas de Usuarios ---

@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser, login_url='/core/access_denied/')
def listar_usuarios(request):
    """
    Lists all custom users, with filtering and pagination.
    """
    query = request.GET.get('q')
    usuarios_list = CustomUser.objects.all()

    if not request.user.is_superuser:
        if request.user.empresa:
            usuarios_list = usuarios_list.filter(empresa=request.user.empresa)
        else: # Usuario normal sin empresa
            usuarios_list = CustomUser.objects.none()

    if query:
        usuarios_list = usuarios_list.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(empresa__nombre__icontains=query)
        )

    paginator = Paginator(usuarios_list, 10)
    page_number = request.GET.get('page')
    try:
        usuarios = paginator.page(page_number)
    except PageNotAnInteger:
        usuarios = paginator.page(1)
    except EmptyPage:
        usuarios = paginator.page(paginator.num_pages)

    # Los permisos de exportación ahora se obtienen automáticamente
    # a través de la propiedad has_export_permission del modelo CustomUser

    return render(request, 'core/listar_usuarios.html', {'usuarios': usuarios, 'query': query, 'titulo_pagina': 'Listado de Usuarios'})


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser, login_url='/core/access_denied/')
def añadir_usuario(request, empresa_pk=None):
    """
    Handles adding a new custom user and assigning groups.
    """
    if request.method == 'POST':
        # Pasar el request al formulario
        form = CustomUserCreationForm(request.POST, request=request)
        if form.is_valid():
            user = form.save(commit=False)
            # Si el usuario que crea NO es superusuario y no asignó una empresa (porque estaba deshabilitada),
            # asegurarse de que la empresa del nuevo usuario sea la misma que la del usuario que crea.
            if not request.user.is_superuser and not user.empresa:
                user.empresa = request.user.empresa
            user.save()
            form.save_m2m() # Guarda las relaciones ManyToMany como los grupos y permisos
            messages.success(request, 'Usuario añadido exitosamente.')
            return redirect('core:listar_usuarios')
        else:
            messages.error(request, 'Hubo un error al añadir el usuario. Por favor, revisa los datos.')
    else:
        # Pasar el request al formulario
        form = CustomUserCreationForm(request=request)
        if empresa_pk:
            try:
                form.fields['empresa'].initial = Empresa.objects.get(pk=empresa_pk)
            except Empresa.DoesNotExist:
                messages.error(request, 'La empresa especificada no existe.')
                return redirect('core:listar_usuarios')

    return render(request, 'core/añadir_usuario.html', {'form': form, 'empresa_pk': empresa_pk, 'titulo_pagina': 'Añadir Nuevo Usuario'})


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser, login_url='/core/access_denied/')
def detalle_usuario(request, pk):
    """
    Displays the details of a specific custom user.
    """
    usuario = get_object_or_404(CustomUser, pk=pk)

    if not request.user.is_superuser and request.user.pk != usuario.pk:
        # Si no es superusuario y no es su propio perfil
        if not request.user.empresa or request.user.empresa != usuario.empresa:
            messages.error(request, 'No tienes permiso para ver usuarios de otras empresas.')
            return redirect('core:listar_usuarios')

    return render(request, 'core/detalle_usuario.html', {
        'usuario_a_ver': usuario,
        'titulo_pagina': f'Detalle de Usuario: {usuario.username}',
    })


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser, login_url='/core/access_denied/')
def editar_usuario(request, pk):
    """
    Handles editing an existing custom user and updating groups.
    """
    usuario_a_editar = get_object_or_404(CustomUser, pk=pk)

    # Permiso:
    # 1. Superusuario puede editar cualquiera.
    # 2. Staff puede editar usuarios de su misma empresa, PERO NO su propio perfil con esta vista.
    # 3. Si el usuario intenta editar su propio perfil, lo redirigimos a la vista de perfil de usuario.
    if request.user.pk == usuario_a_editar.pk:
        messages.info(request, "Estás editando tu propio perfil. Para cambiar tu contraseña o datos básicos, usa la opción específica en 'Mi Perfil'.")
        return redirect('core:perfil_usuario')
    
    if not request.user.is_superuser:
        if not request.user.is_staff or (request.user.empresa and request.user.empresa != usuario_a_editar.empresa):
            messages.error(request, 'No tienes permiso para editar este usuario.')
            return redirect('core:listar_usuarios')

    if request.method == 'POST':
        # Pasar el request al formulario
        form = CustomUserChangeForm(request.POST, instance=usuario_a_editar, request=request)
        if form.is_valid():
            user = form.save(commit=False)
            # Asegurar que la empresa no se cambie si el campo está deshabilitado para no superusuarios
            if not request.user.is_superuser:
                 user.empresa = usuario_a_editar.empresa # Mantener la empresa original
            user.save()
            form.save_m2m() # Guarda las relaciones ManyToMany como los grupos y permisos
            messages.success(request, f'Usuario "{user.username}" actualizado exitosamente.')
            return redirect('core:detalle_usuario', pk=usuario_a_editar.pk)
        else:
            messages.error(request, 'Hubo un error al actualizar el usuario. Por favor, revisa los datos.')
    else:
        # Pasar el request al formulario
        form = CustomUserChangeForm(instance=usuario_a_editar, request=request)

    return render(request, 'core/editar_usuario.html', {'form': form, 'usuario_a_editar': usuario_a_editar, 'titulo_pagina': f'Editar Usuario: {usuario_a_editar.username}'})


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/')
def eliminar_usuario(request, pk):
    """
    Handles deleting a custom user (superuser only).
    """
    usuario = get_object_or_404(CustomUser, pk=pk)

    if request.user.pk == usuario.pk:
        messages.error(request, 'No puedes eliminar tu propia cuenta de superusuario.')
        return redirect('core:listar_usuarios')

    # Si no es superusuario, no puede eliminar usuarios de otras empresas.
    # Este check ya está implícito por el user_passes_test a is_superuser.
    # Sin embargo, si un staff pudiera eliminar, la lógica sería similar a la de edición/detalle.
    if not request.user.is_superuser and request.user.empresa != usuario.empresa:
        messages.error(request, 'No tienes permiso para eliminar usuarios de otras empresas.')
        return redirect('core:listar_usuarios')

    if request.method == 'POST':
        try:
            username_to_delete = usuario.username # Capturar el nombre antes de eliminar
            usuario.delete()
            messages.success(request, f'Usuario "{username_to_delete}" eliminado exitosamente.')
            return redirect('core:listar_usuarios')
        except Exception as e:
            messages.error(request, f'Error al eliminar el usuario: {e}')
            logger.error(f"Error al eliminar usuario {usuario.pk}: {e}")
            return redirect('core:detalle_usuario', pk=usuario.pk)
    
    # CAMBIO: Contexto para la plantilla genérica de confirmación
    context = {
        'object_name': f'el usuario "{usuario.username}"',
        'return_url_name': 'core:detalle_usuario', # URL a la que volver si se cancela
        'return_url_pk': usuario.pk, # PK para la URL de retorno
        'titulo_pagina': f'Eliminar Usuario: {usuario.username}',
    }
    return render(request, 'core/confirmar_eliminacion.html', context)


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/') # Solo superusuarios pueden cambiar contraseñas de otros
def change_user_password(request, pk):
    """
    Handles changing another user's password (admin only).
    """
    user_to_change = get_object_or_404(CustomUser, pk=pk)

    if request.user.pk == user_to_change.pk:
        messages.warning(request, "No puedes cambiar tu propia contraseña desde esta sección. Usa 'Mi Perfil' -> 'Cambiar contraseña'.")
        return redirect('core:perfil_usuario') # Redirige a la vista de perfil para cambio de contraseña propio

    # Si un staff pudiera cambiar contraseñas (no superusuario), se añadiría una verificación de empresa.
    # if not request.user.is_superuser and request.user.empresa != user_to_change.empresa:
    #     messages.error(request, 'No tienes permiso para cambiar la contraseña de usuarios de otras empresas.')
    #     return redirect('core:listar_usuarios')


    if request.method == 'POST':
        form = PasswordChangeForm(user_to_change, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'La contraseña de "{user_to_change.username}" ha sido actualizada exitosamente.')
            return redirect('core:detalle_usuario', pk=user_to_change.pk)
        else:
            messages.error(request, 'Por favor, corrige los errores a continuación.')
    else:
        form = PasswordChangeForm(user_to_change)

    context = {
        'form': form,
        'user_to_change': user_to_change,
        'titulo_pagina': f'Cambiar Contraseña de {user_to_change.username}'
    }
    return render(request, 'core/change_user_password.html', context)


# --- Vistas de Informes ---

@access_check # APLICAR ESTE DECORADOR
@login_required
def informes(request):
    """
    Displays the reports and activities page.
    """
    user = request.user
    today = date.today()
    
    selected_company_id = request.GET.get('empresa_id')
    empresas_disponibles = Empresa.objects.all().order_by('nombre')
    
    equipos_queryset = Equipo.objects.all().select_related('empresa')

    if not user.is_superuser:
        if user.empresa:
            equipos_queryset = equipos_queryset.filter(empresa=user.empresa)
            selected_company_id = str(user.empresa.id)
        else:
            equipos_queryset = Equipo.objects.none()
            empresas_disponibles = Empresa.objects.none()

    if selected_company_id:
        equipos_queryset = equipos_queryset.filter(empresa_id=selected_company_id)

    # --- Actividades a Realizar (Listados Detallados) ---
    # Filtrar solo equipos activos y no de baja o inactivos
    equipos_activos_para_actividades = equipos_queryset.exclude(estado__in=['De Baja', 'Inactivo'])

    scheduled_activities = []

    for equipo in equipos_activos_para_actividades.filter(proxima_calibracion__isnull=False).order_by('proxima_calibracion'):
        if equipo.proxima_calibracion:
            days_remaining = (equipo.proxima_calibracion - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Calibración',
                'equipo': equipo,
                'fecha_programada': equipo.proxima_calibracion,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    mantenimientos_query = equipos_activos_para_actividades.filter(
        proximo_mantenimiento__isnull=False
    ).order_by('proximo_mantenimiento')

    for equipo in mantenimientos_query:
        if equipo.proximo_mantenimiento:
            days_remaining = (equipo.proximo_mantenimiento - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Mantenimiento',
                'equipo': equipo,
                'fecha_programada': equipo.proximo_mantenimiento,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    comprobaciones_query = equipos_activos_para_actividades.filter(
        proxima_comprobacion__isnull=False
    ).order_by('proxima_comprobacion')

    for equipo in comprobaciones_query:
        if equipo.proxima_comprobacion:
            days_remaining = (equipo.proxima_comprobacion - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Comprobación',
                'equipo': equipo,
                'fecha_programada': equipo.proxima_comprobacion,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    scheduled_activities.sort(key=lambda x: x['fecha_programada'] if x['fecha_programada'] else date.max)

    calibraciones_proximas_30 = [act for act in scheduled_activities if act['tipo'] == 'Calibración' and act['estado_vencimiento'] == 'Próxima' and act['dias_restantes'] <= 30 and act['dias_restantes'] > 15]
    calibraciones_proximas_15 = [act for act in scheduled_activities if act['tipo'] == 'Calibración' and act['estado_vencimiento'] == 'Próxima' and act['dias_restantes'] <= 15 and act['dias_restantes'] >= 0]
    calibraciones_vencidas = [act for act in scheduled_activities if act['tipo'] == 'Calibración' and act['estado_vencimiento'] == 'Vencida']

    mantenimientos_proximos_30 = [act for act in scheduled_activities if act['tipo'] == 'Mantenimiento' and act['estado_vencimiento'] == 'Próxima' and act['dias_restantes'] <= 30 and act['dias_restantes'] > 15]
    mantenimientos_proximos_15 = [act for act in scheduled_activities if act['tipo'] == 'Mantenimiento' and act['estado_vencimiento'] == 'Próxima' and act['dias_restantes'] <= 15 and act['dias_restantes'] >= 0]
    mantenimientos_vencidos = [act for act in scheduled_activities if act['tipo'] == 'Mantenimiento' and act['estado_vencimiento'] == 'Vencida']

    comprobaciones_proximos_30 = [act for act in scheduled_activities if act['tipo'] == 'Comprobación' and act['estado_vencimiento'] == 'Próxima' and act['dias_restantes'] <= 30 and act['dias_restantes'] > 15]
    comprobaciones_proximos_15 = [act for act in scheduled_activities if act['tipo'] == 'Comprobación' and act['estado_vencimiento'] == 'Próxima' and act['dias_restantes'] <= 15 and act['dias_restantes'] >= 0]
    comprobaciones_vencidas = [act for act in scheduled_activities if act['tipo'] == 'Comprobación' and act['estado_vencimiento'] == 'Vencida']

    # Información de paginación para ZIPs
    zip_info = None
    if selected_company_id:
        total_equipos, total_partes, equipos_por_zip = calcular_info_paginacion_zip(selected_company_id, user.is_superuser)

        # Crear lista de partes con información detallada
        partes_info = []
        for parte_num in range(1, total_partes + 1):
            inicio_equipo = (parte_num - 1) * equipos_por_zip + 1
            fin_equipo = min(parte_num * equipos_por_zip, total_equipos)
            partes_info.append({
                'numero': parte_num,
                'inicio_equipo': inicio_equipo,
                'fin_equipo': fin_equipo
            })

        zip_info = {
            'total_equipos': total_equipos,
            'total_partes': total_partes,
            'equipos_por_zip': equipos_por_zip,
            'requiere_paginacion': total_partes > 1,
            'partes_info': partes_info
        }

    context = {
        'titulo_pagina': 'Informes y Actividades',
        'is_superuser': user.is_superuser,
        'empresas_disponibles': empresas_disponibles,
        'selected_company_id': selected_company_id,
        'today': today,
        'zip_info': zip_info,  # Nueva información de paginación

        'calibraciones_proximas_30': calibraciones_proximas_30,
        'calibraciones_proximas_15': calibraciones_proximas_15,
        'calibraciones_vencidas': calibraciones_vencidas,

        'mantenimientos_proximos_30': mantenimientos_proximos_30,
        'mantenimientos_proximos_15': mantenimientos_proximos_15,
        'mantenimientos_vencidos': mantenimientos_vencidos,

        'comprobaciones_proximos_30': comprobaciones_proximos_30,
        'comprobaciones_proximos_15': comprobaciones_proximos_15,
        'comprobaciones_vencidas': comprobaciones_vencidas,
    }
    return render(request, 'core/informes.html', context)


def stream_file_to_zip(zip_file, file_path, zip_path):
    """Helper que hace streaming directo de archivo a ZIP sin cargar en memoria."""
    try:
        if default_storage.exists(file_path):
            with default_storage.open(file_path, 'rb') as f:
                # Usar streaming con chunks pequeños para reducir memoria
                with zip_file.open(zip_path, 'w') as zip_entry:
                    while True:
                        chunk = f.read(8192)  # 8KB chunks
                        if not chunk:
                            break
                        zip_entry.write(chunk)
            return True
    except Exception as e:
        logger.warning(f"Error streaming archivo {file_path}: {e}")
    return False


def calcular_info_paginacion_zip(empresa_id, is_superuser=False):
    """
    Calcula información de paginación para ZIPs basado en número de equipos.
    Returns: (total_equipos, total_partes, equipos_por_zip)
    """
    # OPTIMIZACIÓN: 35 equipos por ZIP (con hoja de vida PDF, nombres optimizados)
    EQUIPOS_POR_ZIP = 35

    total_equipos = Equipo.objects.filter(empresa_id=empresa_id).count()
    total_partes = (total_equipos + EQUIPOS_POR_ZIP - 1) // EQUIPOS_POR_ZIP if total_equipos > 0 else 1
    return total_equipos, total_partes, EQUIPOS_POR_ZIP


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser or u.has_perm('core.can_export_reports'), login_url='/core/access_denied/')
def generar_informe_zip(request):
    """
    Generates a ZIP file containing equipment reports and associated documents, including procedures.
    For companies with more than 50 equipments, it generates paginated ZIPs.

    Parameters:
    - parte (optional): Part number for paginated downloads (1, 2, 3, etc.)

    The ZIP structure includes:
    [Company Name]/
    ├── Equipos/
    │   ├── [Equipment Internal Code 1]/
    │   │   ├── Hoja_de_vida.pdf
    │   │   ├── Baja/ (Documento de Baja del Equipo)
    │   │   ├── Calibraciones/
    │   │   │   ├── Certificados/
    │   │   │   │   └── (Calibration PDFs)
    │   │   │   ├── Confirmaciones/
    │   │   │   │   └── (Confirmation PDFs)
    │   │   │   └── Intervalos/
    │   │   │       └── (Intervals PDFs)
    │   │   ├── Comprobaciones/
    │   │   │   └── (Verification PDFs)
    │   │   ├── Mantenimientos/
    │   │   │   └── (Maintenance PDFs)
    │   │   ├── Hoja_de_vida_General_Excel.xlsx
    │   │   └── Hoja_de_vida_Actividades_Excel.xlsx
    │   └── [Equipment Internal Code 2]/
    │       └── ...
    ├── Procedimientos/
    │   ├── [Procedure Code 1].pdf
    │   └── [Procedure Code 2].pdf
    ├── Listado_de_equipos.xlsx
    ├── Listado_de_proveedores.xlsx
    └── Listado_de_procedimientos.xlsx
    """
    selected_company_id = request.GET.get('empresa_id')
    parte_param = request.GET.get('parte', '1')  # Parámetro para paginación (por defecto parte 1)

    # Para usuarios normales, usar su empresa asignada
    if not selected_company_id and not request.user.is_superuser:
        if request.user.empresa:
            selected_company_id = str(request.user.empresa.id)
        else:
            messages.error(request, "No tiene una empresa asignada para generar el informe ZIP.")
            return redirect('core:informes')

    # Para superusuarios, empresa_id es obligatorio
    if not selected_company_id:
        messages.error(request, "Por favor, selecciona una empresa para generar el informe ZIP.")
        return redirect('core:informes')

    empresa = get_object_or_404(Empresa, pk=selected_company_id)

    # Validar parámetro de parte
    try:
        parte_numero = int(parte_param)
        if parte_numero < 1:
            parte_numero = 1
    except (ValueError, TypeError):
        parte_numero = 1

    # OPTIMIZACIÓN: Prefetch relacionados para evitar consultas N+1
    # PAGINACIÓN: Limitar equipos para evitar problemas de memoria en servidor (512MB)
    EQUIPOS_POR_ZIP = 35  # 35 equipos por ZIP (con hoja de vida PDF, nombres optimizados)

    equipos_empresa_total = Equipo.objects.filter(empresa=empresa).count()

    # Calcular offset para paginación
    offset = (parte_numero - 1) * EQUIPOS_POR_ZIP

    # OPTIMIZACIÓN: Prefetch simple sin slice para evitar QuerySet conflicts
    equipos_empresa = Equipo.objects.filter(empresa=empresa).select_related('empresa').prefetch_related(
        'calibraciones', 'mantenimientos', 'comprobaciones', 'baja_registro'
    ).order_by('codigo_interno')[offset:offset + EQUIPOS_POR_ZIP]

    # Calcular información de paginación
    total_partes = (equipos_empresa_total + EQUIPOS_POR_ZIP - 1) // EQUIPOS_POR_ZIP  # Redondeo hacia arriba
    equipos_en_esta_parte = len(equipos_empresa)  # Usar len() en lugar de count() para QuerySet con slice

    # Información para logs y filename con optimizaciones aplicadas
    if total_partes > 1:
        logger.info(f"Generando ZIP OPTIMIZADO parte {parte_numero}/{total_partes} para empresa {empresa.nombre}: equipos {offset + 1}-{offset + equipos_en_esta_parte} de {equipos_empresa_total} (sin Excel individuales)")
    else:
        logger.info(f"Generando ZIP OPTIMIZADO completo para empresa {empresa.nombre}: {equipos_en_esta_parte} equipos (sin Excel individuales)")

    proveedores_empresa = Proveedor.objects.filter(empresa=empresa).order_by('nombre_empresa')
    procedimientos_empresa = Procedimiento.objects.filter(empresa=empresa).order_by('codigo')

    # OPTIMIZACIÓN: Configurar compresión más eficiente y streaming de archivos

    zip_buffer = io.BytesIO()
    # OPTIMIZACIÓN: Compresión eficiente sin Excel individuales
    compresslevel = 2  # Compresión balanceada para todos los usuarios

    # OPTIMIZACIÓN: Cache nombre empresa para evitar accesos repetidos
    empresa_nombre = empresa.nombre

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, compresslevel=compresslevel) as zf:
        # OPTIMIZACIÓN: Un solo Excel consolidado con 4 hojas (Equipos, Proveedores, Procedimientos, Dashboard)
        excel_consolidado = _generate_consolidated_excel_content(equipos_empresa, proveedores_empresa, procedimientos_empresa)
        zf.writestr(f"{empresa_nombre}/Informe_Consolidado.xlsx", excel_consolidado)

        # 4. For each equipment, add its "Hoja de Vida" (PDF) and existing activity PDFs
        for equipo in equipos_empresa:
            # OPTIMIZACIÓN: Código interno con procesamiento mínimo pero necesario para identificación
            safe_codigo = equipo.codigo_interno.replace('/', '_').replace('\\', '_')
            equipo_folder = f"{empresa_nombre}/Equipos/{safe_codigo}"

            # HOJA DE VIDA PDF (requerida por usuario)
            try:
                hoja_vida_pdf_content = _generate_equipment_hoja_vida_pdf_content(request, equipo)
                zf.writestr(f"{equipo_folder}/Hoja_de_vida.pdf", hoja_vida_pdf_content)
            except Exception as e:
                logger.error(f"Error generating Hoja de Vida PDF for {equipo.codigo_interno}: {e}")
                zf.writestr(f"{equipo_folder}/Hoja_de_vida_ERROR.txt", f"Error: {e}")

            # OPTIMIZACIÓN: Excel individuales eliminados para optimizar memoria y velocidad
            # Toda la información está disponible en el archivo Informe_Consolidado.xlsx

            # --- Añadir Documento de Baja si existe ---
            try:
                # OPTIMIZACIÓN: Usar prefetch de baja_registro en lugar de consulta individual
                try:
                    baja_registro = equipo.baja_registro
                except BajaEquipo.DoesNotExist:
                    baja_registro = None
                if baja_registro and baja_registro.documento_baja:
                    baja_folder = f"{equipo_folder}/Baja"
                    file_name_in_zip = os.path.basename(baja_registro.documento_baja.name)
                    if stream_file_to_zip(zf, baja_registro.documento_baja.name, f"{baja_folder}/{file_name_in_zip}"):
                        logger.debug(f" Documento de baja '{file_name_in_zip}' añadido para equipo {equipo.codigo_interno}")
                    else:
                        logger.debug(f" No se pudo agregar documento de baja para equipo {equipo.codigo_interno}")
                else:
                    logger.debug(f" No se encontró documento de baja para equipo {equipo.codigo_interno} o no existe en storage.")
            except Exception as e:
                logger.error(f"Error adding decommission document for {equipo.codigo_interno} to zip: {e}")
                zf.writestr(f"{equipo_folder}/Baja/Documento_Baja_ERROR.txt", f"Error adding decommission document: {e}")


            # Add existing Calibration PDFs (Certificado, Confirmación, Intervalos)
            # OPTIMIZACIÓN: TODAS las calibraciones con nombres simples
            calibraciones = equipo.calibraciones.all()
            # OPTIMIZACIÓN: Nombres simples sin formateo complejo
            cal_idx = 1
            for cal in calibraciones:
                if cal.documento_calibracion:
                    try:
                        # Nombre simple: cal_1.pdf, cal_2.pdf, etc.
                        simple_name = f"cal_{cal_idx}.pdf"
                        if stream_file_to_zip(zf, cal.documento_calibracion.name, f"{equipo_folder}/Calibraciones/{simple_name}"):
                            cal_idx += 1
                    except Exception as e:
                        logger.error(f"Error adding calibration file to zip: {e}")

                if cal.confirmacion_metrologica_pdf:
                    try:
                        # Nombre simple: conf_1.pdf, conf_2.pdf, etc.
                        simple_name = f"conf_{cal_idx}.pdf"
                        if stream_file_to_zip(zf, cal.confirmacion_metrologica_pdf.name, f"{equipo_folder}/Calibraciones/{simple_name}"):
                            cal_idx += 1
                    except Exception as e:
                        logger.error(f"Error adding confirmation file to zip: {e}")

                if cal.intervalos_calibracion_pdf:
                    try:
                        # Nombre simple: int_1.pdf, int_2.pdf, etc.
                        simple_name = f"int_{cal_idx}.pdf"
                        if stream_file_to_zip(zf, cal.intervalos_calibracion_pdf.name, f"{equipo_folder}/Calibraciones/{simple_name}"):
                            cal_idx += 1
                    except Exception as e:
                        logger.error(f"Error adding intervals file to zip: {e}")

            # Add existing Maintenance PDFs
            # OPTIMIZACIÓN: TODOS los mantenimientos con nombres simples
            mantenimientos = equipo.mantenimientos.all()
            # OPTIMIZACIÓN: Nombres simples sin formateo complejo
            mant_idx = 1
            for mant in mantenimientos:
                if mant.documento_mantenimiento:
                    try:
                        # Nombre simple: mant_1.pdf, mant_2.pdf, etc.
                        simple_name = f"mant_{mant_idx}.pdf"
                        if stream_file_to_zip(zf, mant.documento_mantenimiento.name, f"{equipo_folder}/Mantenimientos/{simple_name}"):
                            mant_idx += 1
                    except Exception as e:
                        logger.error(f"Error adding maintenance file to zip: {e}")

            # Add existing Verification PDFs
            # OPTIMIZACIÓN: TODAS las comprobaciones con nombres simples
            comprobaciones = equipo.comprobaciones.all()
            # OPTIMIZACIÓN: Nombres simples sin formateo complejo
            comp_idx = 1
            for comp in comprobaciones:
                if comp.documento_comprobacion:
                    try:
                        # Nombre simple: comp_1.pdf, comp_2.pdf, etc.
                        simple_name = f"comp_{comp_idx}.pdf"
                        if stream_file_to_zip(zf, comp.documento_comprobacion.name, f"{equipo_folder}/Comprobaciones/{simple_name}"):
                            comp_idx += 1
                    except Exception as e:
                        logger.error(f"Error adding verification file to zip: {e}")
            
            # Add other equipment documents (if they exist and are PDF)
            equipment_docs = [
                (equipo.archivo_compra_pdf, 'compra'),
                (equipo.ficha_tecnica_pdf, 'ficha_tecnica'),
                (equipo.manual_pdf, 'manual'),
                (equipo.otros_documentos_pdf, 'otros_documentos')
            ]
            for doc_field, doc_type in equipment_docs:
                if doc_field:
                    try:
                        if doc_field.name.lower().endswith('.pdf'):
                            # Nombre simple: tipo_documento.pdf
                            nombre_descriptivo = f"{doc_type}.pdf"
                            if stream_file_to_zip(zf, doc_field.name, f"{equipo_folder}/{nombre_descriptivo}"):
                                logger.debug(f" Documento de equipo '{nombre_descriptivo}' añadido")
                            else:
                                logger.debug(f" Archivo no encontrado en storage (doc. equipo): {doc_field.name}")
                        else:
                            logger.debug(f" Archivo no es PDF o no encontrado: {doc_field.name}")
                    except Exception as e:
                        logger.error(f"Error adding equipment document {doc_field.name} to zip: {e}")

        # Add existing Procedure PDFs (NEW)
        for proc in procedimientos_empresa:
            if proc.documento_pdf:
                try:
                    safe_proc_code = proc.codigo.replace('/', '_').replace('\\', '_').replace(':', '_')
                    proc_folder = f"{empresa.nombre}/Procedimientos"
                    file_name_in_zip = os.path.basename(proc.documento_pdf.name)
                    # Usar código del procedimiento como prefijo para el nombre del archivo en el zip
                    zip_path = f"{proc_folder}/{safe_proc_code}_{file_name_in_zip}"
                    if stream_file_to_zip(zf, proc.documento_pdf.name, zip_path):
                        logger.debug(f" Documento de procedimiento '{file_name_in_zip}' añadido para procedimiento {proc.codigo}")
                    else:
                        logger.debug(f" Archivo de procedimiento no encontrado en storage: {proc.documento_pdf.name}")
                except Exception as e:
                    logger.error(f"Error adding procedure document {proc.documento_pdf.name} to zip: {e}")
                    zf.writestr(f"{empresa.nombre}/Procedimientos/Documento_Procedimiento_{proc.codigo}_ERROR.txt", f"Error adding procedure document: {e}")


    # OPTIMIZACIÓN: Mejorar respuesta HTTP para transferencia eficiente y reducir memoria
    zip_buffer.seek(0)
    zip_content = zip_buffer.getvalue()
    zip_size = len(zip_content)

    # Crear respuesta optimizada
    response = HttpResponse(zip_content, content_type='application/zip')

    # Headers optimizados para descarga
    filename_safe = empresa.nombre.replace(' ', '_').replace('/', '_').replace('\\', '_')
    if total_partes > 1:
        filename = f"Informes_{filename_safe}_Parte_{parte_numero}_de_{total_partes}.zip"
    else:
        filename = f"Informes_{filename_safe}.zip"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = zip_size
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'

    # Limpiar buffer para liberar memoria inmediatamente
    zip_buffer.close()
    del zip_buffer

    if total_partes > 1:
        logger.info(f"ZIP parte {parte_numero}/{total_partes} generado exitosamente para empresa {empresa.nombre}: {zip_size} bytes ({equipos_en_esta_parte} equipos)")
    else:
        logger.info(f"ZIP generado exitosamente para empresa {empresa.nombre}: {zip_size} bytes ({equipos_en_esta_parte} equipos)")

    return response


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser or u.has_perm('core.can_export_reports'), login_url='/core/access_denied/')
def generar_informe_dashboard_excel(request):
    """
    Genera un Excel consolidado con Dashboard (4 hojas: Equipos, Proveedores, Procedimientos, Dashboard).
    Versión independiente del ZIP para informes mensuales o consultas rápidas.
    """
    selected_company_id = request.GET.get('empresa_id')

    # Para usuarios normales, usar su empresa asignada
    if not selected_company_id and not request.user.is_superuser:
        if request.user.empresa:
            selected_company_id = str(request.user.empresa.id)
        else:
            messages.error(request, "No tiene una empresa asignada para generar el informe Excel.")
            return redirect('core:informes')

    # Para superusuarios, empresa_id es obligatorio
    if not selected_company_id:
        messages.error(request, "Por favor, selecciona una empresa para generar el informe Excel.")
        return redirect('core:informes')

    empresa = get_object_or_404(Empresa, pk=selected_company_id)

    # OPTIMIZACIÓN: Cargar todos los equipos para el dashboard completo
    equipos_empresa = Equipo.objects.filter(empresa=empresa).select_related('empresa').prefetch_related(
        'calibraciones', 'mantenimientos', 'comprobaciones', 'baja_registro'
    ).order_by('codigo_interno')

    proveedores_empresa = Proveedor.objects.filter(empresa=empresa).order_by('nombre_empresa')
    procedimientos_empresa = Procedimiento.objects.filter(empresa=empresa).order_by('codigo')

    logger.info(f"Generando Excel Dashboard para empresa {empresa.nombre}: {equipos_empresa.count()} equipos")

    try:
        # Generar Excel simplificado del dashboard (solo estadísticas)
        excel_content = _generate_dashboard_excel_content(equipos_empresa, empresa)

        # Crear respuesta HTTP
        response = HttpResponse(excel_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Headers para descarga
        filename_safe = empresa.nombre.replace(' ', '_').replace('/', '_').replace('\\', '_')
        response['Content-Disposition'] = f'attachment; filename="Dashboard_{filename_safe}.xlsx"'
        response['Content-Length'] = len(excel_content)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'

        logger.info(f"Excel Dashboard generado exitosamente para empresa {empresa.nombre}: {len(excel_content)} bytes")
        return response

    except Exception as e:
        logger.error(f"Error generando Excel Dashboard para empresa {empresa.nombre}: {e}")
        messages.error(request, f'Error al generar el informe Dashboard: {e}')
        return redirect('core:informes')


@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_superuser or u.has_perm('core.can_export_reports'), login_url='/core/access_denied/')
def informe_vencimientos_pdf(request):
    """
    Generates a PDF report of upcoming and overdue activities.
    """
    today = timezone.localdate()

    equipos_base_query = Equipo.objects.all()
    if not request.user.is_superuser and request.user.empresa:
        equipos_base_query = equipos_base_query.filter(empresa=request.user.empresa)
    elif not request.user.is_superuser and not request.user.empresa:
        equipos_base_query = Equipo.objects.none()

    # Excluir equipos "De Baja" y "Inactivo" para este informe
    equipos_base_query = equipos_base_query.exclude(estado__in=['De Baja', 'Inactivo'])

    calibraciones_query = equipos_base_query.filter(
        proxima_calibracion__isnull=False
    ).order_by('proxima_calibracion')

    scheduled_activities = []
    for equipo in calibraciones_query:
        if equipo.proxima_calibracion:
            days_remaining = (equipo.proxima_calibracion - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Calibración',
                'equipo': equipo,
                'fecha_programada': equipo.proxima_calibracion,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    mantenimientos_query = equipos_base_query.filter(
        proximo_mantenimiento__isnull=False
    ).order_by('proximo_mantenimiento')

    for equipo in mantenimientos_query:
        if equipo.proximo_mantenimiento:
            days_remaining = (equipo.proximo_mantenimiento - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Mantenimiento',
                'equipo': equipo,
                'fecha_programada': equipo.proximo_mantenimiento,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    comprobaciones_query = equipos_base_query.filter(
        proxima_comprobacion__isnull=False
    ).order_by('proxima_comprobacion')

    for equipo in comprobaciones_query:
        if equipo.proxima_comprobacion:
            days_remaining = (equipo.proxima_comprobacion - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Comprobación',
                'equipo': equipo,
                'fecha_programada': equipo.proxima_comprobacion,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    scheduled_activities.sort(key=lambda x: x['fecha_programada'] if x['fecha_programada'] else date.max)

    context = {
        'scheduled_activities': scheduled_activities,
        'today': today,
        'titulo_pagina': 'Informe de Vencimientos',
    }

    template_path = 'core/informe_vencimientos_pdf.html'
    
    try:
        pdf_file = _generate_pdf_content(request, template_path, context)
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="informe_vencimientos.pdf"'
        return response
    except Exception as e:
        messages.error(request, f'Tuvimos algunos errores al generar el PDF de vencimientos: {e}. Revisa los logs para más detalles.')
        logger.error(f"Error al generar informe_vencimientos_pdf: {e}")
        return redirect('core:informes')


@access_check # APLICAR ESTE DECORADOR
@login_required
def programmed_activities_list(request):
    """
    Lists all programmed activities.
    """
    today = timezone.localdate()
    scheduled_activities = []

    equipos_base_query = Equipo.objects.all()
    if not request.user.is_superuser and request.user.empresa:
        equipos_base_query = equipos_base_query.filter(empresa=request.user.empresa)
    elif not request.user.is_superuser and not request.user.empresa:
        equipos_base_query = Equipo.objects.none()

    # Excluir equipos "De Baja" y "Inactivo" para esta lista
    equipos_base_query = equipos_base_query.exclude(estado__in=['De Baja', 'Inactivo'])

    calibraciones_query = equipos_base_query.filter(
        proxima_calibracion__isnull=False
    ).order_by('proxima_calibracion')

    for equipo in calibraciones_query:
        if equipo.proxima_calibracion:
            days_remaining = (equipo.proxima_calibracion - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Calibración',
                'equipo': equipo,
                'fecha_programada': equipo.proxima_calibracion,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    mantenimientos_query = equipos_base_query.filter(
        proximo_mantenimiento__isnull=False
    ).order_by('proximo_mantenimiento')

    for equipo in mantenimientos_query:
        if equipo.proximo_mantenimiento:
            days_remaining = (equipo.proximo_mantenimiento - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Mantenimiento',
                'equipo': equipo,
                'fecha_programada': equipo.proximo_mantenimiento,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    comprobaciones_query = equipos_base_query.filter(
        proxima_comprobacion__isnull=False
    ).order_by('proxima_comprobacion')

    for equipo in comprobaciones_query:
        if equipo.proxima_comprobacion:
            days_remaining = (equipo.proxima_comprobacion - today).days
            estado_vencimiento = 'Vencida' if days_remaining < 0 else 'Próxima'
            scheduled_activities.append({
                'tipo': 'Comprobación',
                'equipo': equipo,
                'fecha_programada': equipo.proxima_comprobacion,
                'dias_restantes': days_remaining,
                'estado_vencimiento': estado_vencimiento
            })

    scheduled_activities.sort(key=lambda x: x['fecha_programada'] if x['fecha_programada'] else date.max)

    paginator = Paginator(scheduled_activities, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    return render(request, 'core/programmed_activities_list.html', {'page_obj': page_obj, 'titulo_pagina': 'Actividades Programadas'})


@access_check # APLICAR ESTE DECORADOR
@login_required
@permission_required('core.can_export_reports', raise_exception=True)
def exportar_equipos_excel(request):
    """
    Exports a general list of equipment to an Excel file.
    """
    equipos = Equipo.objects.all().select_related('empresa')
    if not request.user.is_superuser and request.user.empresa:
        equipos = equipos.filter(empresa=request.user.empresa)
    elif not request.user.is_superuser and not request.user.empresa:
        equipos = Equipo.objects.none()

    excel_content = _generate_general_equipment_list_excel_content(equipos)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_equipos.xlsx"'
    response.write(excel_content)
    return response

@access_check # APLICAR ESTE DECORADOR
@login_required
def generar_hoja_vida_pdf(request, pk):
    """
    Generates the "Hoja de Vida" (Life Sheet) PDF for a specific equipment.
    """
    equipo = get_object_or_404(Equipo, pk=pk)
    if not request.user.is_superuser and request.user.empresa != equipo.empresa:
        messages.error(request, 'No tienes permiso para generar la hoja de vida de este equipo.')
        return redirect('core:home')

    try:
        pdf_content = _generate_equipment_hoja_vida_pdf_content(request, equipo)
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="hoja_vida_{equipo.codigo_interno}.pdf"'
        return response
    except Exception as e:
        messages.error(request, f'Tuvimos algunos errores al generar el PDF: {e}. Revisa los logs para más detalles.')
        logger.error(f"Error al generar hoja_vida_pdf para equipo {equipo.pk}: {e}")
        return redirect('core:detalle_equipo', pk=equipo.pk)
        
@access_check # APLICAR ESTE DECORADOR
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser, login_url='/core/access_denied/') # Solo superusuarios o staff
@require_POST
@csrf_exempt # Necesario para AJAX POST requests si no manejas el CSRF token de otra forma en JS
def toggle_user_active_status(request):
    """
    Alterna el estado 'is_active' de un usuario vía AJAX POST.
    Espera JSON con {'user_id': <id>, 'is_active': <true/false>}.
    """
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        new_is_active_status = data.get('is_active') # Esperamos un booleano

        if user_id is None or new_is_active_status is None:
            return JsonResponse({'status': 'error', 'message': 'ID de usuario o estado activo no proporcionado.'}, status=400)

        user_to_toggle = get_object_or_404(CustomUser, pk=user_id)

        # Permisos: Superusuario puede alternar cualquiera. Staff solo puede alternar de su empresa.
        # Un staff NO puede desactivar a otro staff o superusuario si no es superusuario.
        if not request.user.is_superuser:
            # Si el usuario que intenta cambiar NO es superusuario:
            # 1. Debe ser staff.
            # 2. El usuario a cambiar debe pertenecer a la misma empresa que el staff.
            # 3. El staff NO puede desactivar a un superusuario.
            # 4. El staff NO puede desactivar a otro staff (si es is_staff) de su misma empresa.
            if not request.user.is_staff or (request.user.empresa and user_to_toggle.empresa != request.user.empresa):
                return JsonResponse({'status': 'error', 'message': 'No tienes permiso para modificar este usuario.'}, status=403)
            
            if user_to_toggle.is_superuser:
                return JsonResponse({'status': 'error', 'message': 'Un usuario staff no puede desactivar a un superusuario.'}, status=403)
            
            if user_to_toggle.is_staff and not request.user.is_superuser:
                return JsonResponse({'status': 'error', 'message': 'Un usuario staff no puede desactivar a otro staff.'}, status=403)

        # Prevenir que un superusuario se desactive a sí mismo (por seguridad general).
        # Esto es importante para evitar que se bloquee el único superusuario.
        if request.user.pk == user_to_toggle.pk and not new_is_active_status:
             return JsonResponse({'status': 'error', 'message': 'No puedes desactivar tu propia cuenta.'}, status=403)


        user_to_toggle.is_active = new_is_active_status
        user_to_toggle.save(update_fields=['is_active'])

        status_text = "activado" if new_is_active_status else "desactivado"
        messages.success(request, f'Usuario "{user_to_toggle.username}" {status_text} exitosamente.')

        return JsonResponse({'status': 'success', 'new_status': user_to_toggle.is_active, 'message': f'Usuario {user_to_toggle.username} {status_text}.'})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Formato JSON inválido.'}, status=400)
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Usuario no encontrado.'}, status=404)
    except Exception as e:
        logger.error(f"Error al alternar estado de usuario: {e}")
        return JsonResponse({'status': 'error', 'message': f'Error interno del servidor: {str(e)}'}, status=500)

@access_check
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/')
@require_POST
@csrf_exempt
def toggle_download_permission(request):
    """
    Grants or revokes download permissions for a user via AJAX POST.
    Expected JSON: {'user_id': <id>, 'grant_permission': <true/false>}.
    Only superusers can manage permissions.
    """
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        grant_permission = data.get('grant_permission')

        if user_id is None or grant_permission is None:
            return JsonResponse({'status': 'error', 'message': 'ID de usuario o estado de permiso no proporcionado.'}, status=400)

        user_to_modify = get_object_or_404(CustomUser, pk=user_id)

        # Get the export permission
        from django.contrib.contenttypes.models import ContentType
        content_type = ContentType.objects.get_for_model(Equipo)
        export_permission = Permission.objects.get(codename='can_export_reports', content_type=content_type)

        if grant_permission:
            # Grant permission
            user_to_modify.user_permissions.add(export_permission)
            action_message = f'Permiso de descarga otorgado a {user_to_modify.username}'
        else:
            # Revoke permission
            user_to_modify.user_permissions.remove(export_permission)
            action_message = f'Permiso de descarga revocado para {user_to_modify.username}'

        return JsonResponse({
            'status': 'success',
            'message': action_message
        })

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Formato JSON inválido.'}, status=400)
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Usuario no encontrado.'}, status=404)
    except Permission.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Permiso no encontrado.'}, status=404)
    except Exception as e:
        logger.error(f"Error al alternar permisos de descarga: {e}")
        return JsonResponse({'status': 'error', 'message': f'Error interno del servidor: {str(e)}'}, status=500)

@access_check
@login_required
def redirect_to_change_password(request, pk):
    """
    Redirección temporal para URLs antiguas de cambio de contraseña.
    Redirige /usuarios/<pk>/password/ a /usuarios/<pk>/cambiar_password/
    """
    return redirect('core:change_user_password', pk=pk)

@access_check # APLICAR ESTE DECORADOR
@require_POST
@csrf_exempt
def add_message(request):
    """
    Adds a Django message via an AJAX POST request.
    Expected JSON body: {"message": "Your message here", "tags": "success|info|warning|error"}
    """
    try:
        data = json.loads(request.body)
        message_text = data.get('message', 'Mensaje genérico.')
        message_tags = data.get('tags', 'info')

        if message_tags == 'success':
            messages.success(request, message_text)
        elif message_tags == 'error':
            messages.error(request, message_text)
        elif message_tags == 'warning':
            messages.warning(request, message_text)
        else:
            messages.info(request, message_text)
        
        return JsonResponse({'status': 'success', 'message': 'Mensaje añadido.'})
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def access_denied(request):
    """
    Renders the access denied page.
    """
    return render(request, 'core/access_denied.html', {'titulo_pagina': 'Acceso Denegado'})

@login_required
def cache_diagnostics(request):
    """Vista de diagnóstico temporal para verificar el estado del cache."""
    from django.db import connection
    from django.core.cache import cache
    from django.conf import settings
    from datetime import datetime

    # Verificar si la tabla existe
    table_exists = False
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables
                    WHERE table_name = 'sam_cache_table'
                );
            """)
            table_exists = cursor.fetchone()[0]
    except Exception:
        table_exists = False

    # Test de cache
    cache_works = False
    cache_error = None
    try:
        test_key = 'diagnostics_test_key'
        test_value = 'diagnostics_test_value_123'
        cache.set(test_key, test_value, 60)
        retrieved = cache.get(test_key)
        cache_works = (retrieved == test_value)
        if cache_works:
            cache.delete(test_key)
    except Exception as e:
        cache_error = str(e)

    # Información de configuración
    cache_config = settings.CACHES.get('default', {})
    cache_backend = cache_config.get('BACKEND', 'Unknown')
    cache_location = cache_config.get('LOCATION', 'Unknown')

    context = {
        'table_exists': table_exists,
        'cache_works': cache_works,
        'cache_error': cache_error,
        'cache_backend': cache_backend,
        'cache_location': cache_location,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return render(request, 'core/cache_diagnostics.html', context)


# ================================
# SISTEMA DE COLA PARA ZIP
# ================================

from django.http import JsonResponse
from django.db.models import Max
from django.core.files.storage import default_storage
from django.utils import timezone
from datetime import timedelta
import os
from .models import ZipRequest

from django.views.decorators.csrf import csrf_exempt

@access_check
@csrf_exempt  # Temporalmente para debugging
def trigger_zip_processing(request):
    """
    Endpoint para activar procesamiento de ZIP desde frontend.
    Diseñado para Render - procesa una solicitud y termina.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        from django.core.management import call_command
        from io import StringIO

        # Capturar output del comando
        output = StringIO()

        # Ejecutar procesamiento de una sola solicitud
        call_command('process_single_zip', '--cleanup-expired', stdout=output)

        result = output.getvalue()

        # Determinar si se procesó algo
        if '[PROCESANDO]' in result:
            return JsonResponse({
                'status': 'processing_started',
                'message': 'Solicitud en procesamiento',
                'output': result
            })
        elif '[COLA-VACÍA]' in result:
            return JsonResponse({
                'status': 'queue_empty',
                'message': 'No hay solicitudes pendientes',
                'output': result
            })
        else:
            return JsonResponse({
                'status': 'completed',
                'message': 'Procesamiento completado',
                'output': result
            })

    except Exception as e:
        logger.error(f'Error en trigger_zip_processing: {e}', exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Error procesando: {str(e)}'
        }, status=500)


@access_check
@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/core/access_denied/')
def manual_process_zip(request):
    """Vista manual para procesar ZIPs cuando el automático falla."""
    if request.method == 'POST':
        try:
            from django.core.management import call_command
            from io import StringIO

            # Capturar output del comando
            output = StringIO()

            # Ejecutar procesamiento de una sola solicitud
            call_command('process_single_zip', '--cleanup-expired', stdout=output)

            result = output.getvalue()

            return JsonResponse({
                'status': 'success',
                'message': 'Procesamiento ejecutado',
                'output': result
            })

        except Exception as e:
            logger.error(f'Error en manual_process_zip: {e}', exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': f'Error: {str(e)}'
            }, status=500)

    # GET request - mostrar interfaz
    pending_requests = ZipRequest.objects.filter(status='pending').order_by('position_in_queue')
    processing_requests = ZipRequest.objects.filter(status='processing')

    context = {
        'pending_requests': pending_requests,
        'processing_requests': processing_requests,
    }

    return render(request, 'core/manual_zip_processor.html', context)


@access_check
@login_required
@user_passes_test(lambda u: u.is_superuser or u.has_perm('core.can_export_reports'), login_url='/core/access_denied/')
@trial_check
def solicitar_zip(request):
    """Vista para solicitar un ZIP. Agrega usuario a la cola."""

    # Verificar si ya tiene una solicitud pendiente o procesándose
    existing = ZipRequest.objects.filter(
        user=request.user,
        status__in=['pending', 'processing']
    ).first()

    if existing:
        return JsonResponse({
            'status': 'already_pending',
            'request_id': existing.id,
            'position': existing.get_current_position(),
            'estimated_time': existing.get_estimated_wait_time(),
            'message': f'Ya tienes una solicitud en posición #{existing.get_current_position()}'
        })

    # Obtener número de parte desde parámetros
    parte_numero = int(request.GET.get('parte', 1))
    empresa_id = request.GET.get('empresa_id') if request.user.is_superuser else None

    # Determinar empresa
    if request.user.is_superuser and empresa_id:
        try:
            empresa = Empresa.objects.get(id=empresa_id)
        except Empresa.DoesNotExist:
            return JsonResponse({'error': 'Empresa no encontrada'}, status=400)
    else:
        empresa = request.user.empresa
        if not empresa:
            return JsonResponse({'error': 'Usuario sin empresa asignada'}, status=400)

    # Obtener próxima posición en cola
    max_position = ZipRequest.objects.aggregate(
        max_pos=Max('position_in_queue')
    )['max_pos'] or 0

    # Crear nueva solicitud
    zip_request = ZipRequest.objects.create(
        user=request.user,
        empresa=empresa,
        status='pending',
        position_in_queue=max_position + 1,
        parte_numero=parte_numero,
        expires_at=timezone.now() + timedelta(hours=6)  # Expira en 6 horas
    )

    return JsonResponse({
        'status': 'queued',
        'request_id': zip_request.id,
        'position': zip_request.get_current_position(),
        'estimated_time': zip_request.get_estimated_wait_time(),
        'message': f'Solicitud agregada a la cola en posición #{zip_request.get_current_position()}'
    })


@access_check
@login_required
def zip_status(request, request_id):
    """Vista para consultar el estado de una solicitud ZIP."""

    try:
        zip_req = ZipRequest.objects.get(id=request_id, user=request.user)

        # Verificar si expiró
        if zip_req.status == 'completed' and zip_req.expires_at and timezone.now() > zip_req.expires_at:
            zip_req.status = 'expired'
            zip_req.save()

        response_data = {
            'status': zip_req.status,
            'position': zip_req.get_current_position(),
            'estimated_time': zip_req.get_estimated_wait_time(),
            'created_at': zip_req.created_at.isoformat(),
            'parte_numero': zip_req.parte_numero,
            'request_id': zip_req.id,
            'empresa_nombre': zip_req.empresa.nombre if zip_req.empresa else 'Sin empresa'
        }

        if zip_req.status == 'completed' and zip_req.file_path:
            response_data['download_url'] = f'/core/download_zip/{zip_req.id}/'
            response_data['file_size'] = zip_req.file_size

        elif zip_req.status == 'failed':
            response_data['error_message'] = zip_req.error_message

        elif zip_req.status == 'processing':
            response_data['started_at'] = zip_req.started_at.isoformat() if zip_req.started_at else None

        return JsonResponse(response_data)

    except ZipRequest.DoesNotExist:
        return JsonResponse({'error': 'Solicitud no encontrada'}, status=404)


@access_check
@login_required
def download_zip(request, request_id):
    """Vista para descargar un ZIP completado."""

    try:
        zip_req = ZipRequest.objects.get(id=request_id, user=request.user)

        if zip_req.status != 'completed' or not zip_req.file_path:
            return JsonResponse({'error': 'Archivo no disponible'}, status=404)

        # Verificar si expiró
        if zip_req.expires_at and timezone.now() > zip_req.expires_at:
            zip_req.status = 'expired'
            zip_req.save()
            return JsonResponse({'error': 'Archivo expirado'}, status=410)

        # Verificar si el archivo existe
        if not default_storage.exists(zip_req.file_path):
            zip_req.status = 'expired'
            zip_req.save()
            return JsonResponse({'error': 'Archivo no encontrado'}, status=404)

        # Generar respuesta de descarga
        file_data = default_storage.open(zip_req.file_path).read()
        filename = f"Informe_{zip_req.empresa.nombre}_parte_{zip_req.parte_numero}.zip"

        response = HttpResponse(file_data, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(file_data)

        return response

    except ZipRequest.DoesNotExist:
        return JsonResponse({'error': 'Solicitud no encontrada'}, status=404)


@access_check
@login_required
def cancel_zip_request(request, request_id):
    """Vista para cancelar una solicitud ZIP pendiente."""

    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        zip_req = ZipRequest.objects.get(id=request_id, user=request.user)

        if zip_req.status == 'pending':
            zip_req.delete()
            return JsonResponse({'status': 'cancelled', 'message': 'Solicitud cancelada'})
        else:
            return JsonResponse({'error': 'No se puede cancelar, solicitud en proceso o completada'}, status=400)

    except ZipRequest.DoesNotExist:
        return JsonResponse({'error': 'Solicitud no encontrada'}, status=404)


@access_check
@login_required
def my_zip_requests(request):
    """Vista para mostrar las solicitudes ZIP del usuario."""

    # Obtener solicitudes del usuario (últimas 10)
    requests = ZipRequest.objects.filter(user=request.user).order_by('-created_at')[:10]

    # Limpiar solicitudes expiradas
    expired_ids = []
    for zip_req in requests:
        if (zip_req.status == 'completed' and zip_req.expires_at and
            timezone.now() > zip_req.expires_at):
            zip_req.status = 'expired'
            zip_req.save()
            expired_ids.append(zip_req.id)

    requests_data = []
    for zip_req in requests:
        data = {
            'id': zip_req.id,
            'status': zip_req.status,
            'status_display': zip_req.get_status_display(),
            'created_at': zip_req.created_at.strftime('%d/%m/%Y %H:%M'),
            'parte_numero': zip_req.parte_numero,
            'empresa': zip_req.empresa.nombre,
        }

        if zip_req.status == 'pending':
            data['position'] = zip_req.get_current_position()
            data['estimated_time'] = zip_req.get_estimated_wait_time()
        elif zip_req.status == 'completed':
            data['download_url'] = f'/core/download_zip/{zip_req.id}/'
            data['expires_at'] = zip_req.expires_at.strftime('%d/%m/%Y %H:%M') if zip_req.expires_at else None
        elif zip_req.status == 'failed':
            data['error_message'] = zip_req.error_message

        requests_data.append(data)

    return JsonResponse({'requests': requests_data})