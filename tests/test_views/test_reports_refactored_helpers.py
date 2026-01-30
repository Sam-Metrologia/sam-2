"""
Tests para las funciones helper refactorizadas en reports.py.

Objetivo: Aumentar coverage de reports.py de 34.58% a >60%.
Enfoque: Testar los 19 helpers creados durante refactorización.
"""
import pytest
from django.test import TestCase
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta
from django.utils import timezone

from core.models import Empresa, Equipo, Calibracion, Mantenimiento, Comprobacion
from core.views.reports import (
    # Helpers para consolidated Excel
    _add_professional_sheet_header,
    _add_sheet_headers,
    _add_equipos_sheet,
    _add_proveedores_sheet,
    _add_procedimientos_sheet,
    _add_dashboard_sheet,
    # Helpers para dashboard Excel
    _add_codigos_internos_section,
    _add_actividades_programadas_section,
    _add_actividades_correctivas_section,
    # Helpers para template Excel
    _add_template_header,
    _add_template_headers_row,
    _add_template_validations,
    _add_template_example_row,
    _apply_template_formatting,
    # Helpers para hoja de vida PDF
    _generate_hoja_vida_cache_key,
    _get_hoja_vida_activities,
    _generate_hoja_vida_charts,
    _get_hoja_vida_file_urls,
    _build_hoja_vida_context,
)


@pytest.mark.django_db
class TestConsolidatedExcelHelpers:
    """Tests para helpers de _generate_consolidated_excel_content."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup común."""
        self.empresa = Empresa.objects.create(
            nombre="Empresa Test Helpers",
            nit="900111222-3",
        )
        self.workbook = Workbook()
        self.generation_date = timezone.now()

    def test_add_professional_sheet_header(self):
        """Test: _add_professional_sheet_header agrega header correcto."""
        sheet = self.workbook.active

        _add_professional_sheet_header(
            sheet,
            "INFORMES GENERADOS POR SAM",
            'A1:E2',
            self.generation_date
        )

        # Verificar que se creó el header
        assert sheet['A1'].value == "INFORMES GENERADOS POR SAM"
        assert sheet['A3'].value.startswith("Generado el:")

        # Verificar merge
        assert 'A1:E2' in [str(merged) for merged in sheet.merged_cells.ranges]

    def test_add_sheet_headers(self):
        """Test: _add_sheet_headers agrega columnas con estilo."""
        sheet = self.workbook.active
        headers = ['Columna 1', 'Columna 2', 'Columna 3']

        _add_sheet_headers(sheet, headers, row_num=5)

        # Verificar headers
        assert sheet.cell(row=5, column=1).value == 'Columna 1'
        assert sheet.cell(row=5, column=2).value == 'Columna 2'
        assert sheet.cell(row=5, column=3).value == 'Columna 3'

        # Verificar estilo (font bold)
        assert sheet.cell(row=5, column=1).font.bold == True

    def test_add_equipos_sheet(self):
        """Test: _add_equipos_sheet crea hoja completa."""
        # Crear equipos de prueba
        equipos = []
        for i in range(3):
            equipo = Equipo.objects.create(
                codigo_interno=f"EQ-TEST-{i}",
                nombre=f"Equipo Test {i}",
                empresa=self.empresa,
                tipo_equipo="Equipo de Medición",
                estado="Activo",
            )
            equipos.append(equipo)

        equipos_queryset = Equipo.objects.filter(empresa=self.empresa)

        # Ejecutar helper
        sheet = _add_equipos_sheet(self.workbook, equipos_queryset, self.generation_date)

        # Verificar que la hoja existe
        assert sheet.title == "Equipos"

        # Verificar que hay headers
        assert sheet.cell(row=7, column=1).value is not None

        # Verificar que hay datos (al menos 1 equipo en fila 8)
        # Nota: No validamos orden específico ya que depende de DB
        assert sheet.cell(row=8, column=1).value is not None
        # Verificar que el codigo_interno es uno de los creados
        codigo_interno = sheet.cell(row=8, column=1).value
        assert codigo_interno in ["EQ-TEST-0", "EQ-TEST-1", "EQ-TEST-2"]

    def test_add_proveedores_sheet(self):
        """Test: _add_proveedores_sheet crea hoja vacía correctamente."""
        from core.models import Proveedor
        proveedores_queryset = Proveedor.objects.none()

        sheet = _add_proveedores_sheet(self.workbook, proveedores_queryset, self.generation_date)

        # Verificar que la hoja existe
        assert sheet.title == "Proveedores"

        # Verificar que tiene headers aunque no haya datos (fila 5)
        assert sheet.cell(row=5, column=1).value is not None

    def test_add_procedimientos_sheet(self):
        """Test: _add_procedimientos_sheet crea hoja vacía correctamente."""
        from core.models import Procedimiento
        procedimientos_queryset = Procedimiento.objects.none()

        sheet = _add_procedimientos_sheet(self.workbook, procedimientos_queryset, self.generation_date)

        # Verificar que la hoja existe
        assert sheet.title == "Procedimientos"

        # Verificar que tiene headers (fila 5)
        assert sheet.cell(row=5, column=1).value is not None

    def test_add_dashboard_sheet(self):
        """Test: _add_dashboard_sheet crea hoja con estadísticas."""
        # Crear equipos con diferentes estados
        for i in range(5):
            Equipo.objects.create(
                codigo_interno=f"EQ-DASH-{i}",
                nombre=f"Equipo Dashboard {i}",
                empresa=self.empresa,
                tipo_equipo="Equipo de Medición",
                estado="Activo" if i % 2 == 0 else "En Calibración",
            )

        equipos_queryset = Equipo.objects.filter(empresa=self.empresa)

        sheet = _add_dashboard_sheet(self.workbook, equipos_queryset, self.generation_date)

        # Verificar que la hoja existe
        assert sheet.title == "Dashboard"

        # Verificar que tiene título
        assert "INFORMES" in str(sheet['A1'].value)

        # Verificar que tiene estadísticas (row 7 aprox)
        assert sheet.cell(row=7, column=1).value == "Estado"


@pytest.mark.django_db
class TestDashboardExcelHelpers:
    """Tests para helpers de _generate_dashboard_excel_content."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup común."""
        self.empresa = Empresa.objects.create(
            nombre="Empresa Dashboard Test",
            nit="900333444-5",
        )
        self.workbook = Workbook()
        self.sheet = self.workbook.active

    def test_add_codigos_internos_section(self):
        """Test: _add_codigos_internos_section lista equipos."""
        # Crear 20 equipos
        equipos = []
        for i in range(20):
            equipo = Equipo.objects.create(
                codigo_interno=f"EQ-COD-{i:02d}",
                nombre=f"Equipo Código {i}",
                empresa=self.empresa,
                tipo_equipo="Equipo de Medición",
                estado="Activo",
            )
            equipos.append(equipo)

        # Ejecutar helper
        new_row = _add_codigos_internos_section(self.sheet, equipos, row=5)

        # Verificar que avanzó filas
        assert new_row > 5

        # Verificar título
        assert "CÓDIGOS" in str(self.sheet['A5'].value)

        # Verificar que solo muestra 15 (límite)
        assert "y 5 equipos más" in str(self.sheet.cell(row=new_row-1, column=1).value)

    def test_add_actividades_programadas_section(self):
        """Test: _add_actividades_programadas_section calcula stats."""
        today = timezone.now().date()

        # Crear equipos con fechas variadas
        for i in range(10):
            Equipo.objects.create(
                codigo_interno=f"EQ-ACT-{i}",
                nombre=f"Equipo Actividades {i}",
                empresa=self.empresa,
                tipo_equipo="Equipo de Medición",
                estado="Activo",
                proxima_calibracion=today + timedelta(days=i*10),
                proximo_mantenimiento=today + timedelta(days=i*8),
            )

        equipos_list = list(Equipo.objects.filter(empresa=self.empresa))

        # Ejecutar helper
        new_row = _add_actividades_programadas_section(
            self.sheet, equipos_list, self.empresa, row=5
        )

        # Verificar que avanzó filas
        assert new_row > 5

        # Verificar título
        assert "ACTIVIDADES PROGRAMADAS" in str(self.sheet['A5'].value)

    def test_add_actividades_correctivas_section(self):
        """Test: _add_actividades_correctivas_section sin datos."""
        # Sin crear mantenimientos correctivos

        new_row = _add_actividades_correctivas_section(self.sheet, self.empresa, row=5)

        # Verificar que avanzó filas
        assert new_row > 5

        # Verificar mensaje de "no hay correctivos"
        assert "No se registraron" in str(self.sheet.cell(row=7, column=1).value)


@pytest.mark.django_db
class TestTemplateExcelHelpers:
    """Tests para helpers de _generate_excel_template."""

    def test_add_template_header(self):
        """Test: _add_template_header agrega título."""
        workbook = Workbook()
        sheet = workbook.active

        _add_template_header(sheet)

        # Verificar título
        assert "PLANTILLA" in str(sheet['A1'].value)
        assert "SAM METROLOGÍA" in str(sheet['A1'].value)

        # Verificar información
        assert "Generado el:" in str(sheet['A4'].value)

    def test_add_template_headers_row(self):
        """Test: _add_template_headers_row agrega headers técnicos y legibles."""
        workbook = Workbook()
        sheet = workbook.active

        _add_template_headers_row(sheet)

        # Verificar header técnico (fila 6)
        assert sheet['A6'].value == "codigo_interno"
        assert sheet['B6'].value == "nombre"

        # Verificar header legible (fila 7)
        assert "Código Interno" in str(sheet['A7'].value)
        assert "Nombre del Equipo" in str(sheet['B7'].value)

    def test_add_template_validations(self):
        """Test: _add_template_validations agrega validaciones de datos."""
        workbook = Workbook()
        sheet = workbook.active

        _add_template_validations(sheet)

        # Verificar que se agregaron validaciones
        # (openpyxl guarda validations en sheet.data_validations)
        assert len(sheet.data_validations.dataValidation) >= 2

    def test_add_template_example_row(self):
        """Test: _add_template_example_row agrega fila de ejemplo."""
        workbook = Workbook()
        sheet = workbook.active

        _add_template_example_row(sheet)

        # Verificar fila de ejemplo (fila 8)
        assert "EQ-001" in str(sheet['A8'].value)
        assert "Balanza" in str(sheet['B8'].value)

    def test_apply_template_formatting(self):
        """Test: _apply_template_formatting ajusta anchos y congela paneles."""
        workbook = Workbook()
        sheet = workbook.active

        _apply_template_formatting(sheet)

        # Verificar anchos de columna
        assert sheet.column_dimensions['A'].width == 15
        assert sheet.column_dimensions['B'].width == 25

        # Verificar freeze panes
        assert sheet.freeze_panes == "A8"


@pytest.mark.django_db
class TestHojaVidaPDFHelpers:
    """Tests para helpers de _generate_equipment_hoja_vida_pdf_content."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup común."""
        self.empresa = Empresa.objects.create(
            nombre="Empresa HV Test",
            nit="900555666-7",
        )
        self.equipo = Equipo.objects.create(
            codigo_interno="EQ-HV-001",
            nombre="Equipo Hoja Vida",
            empresa=self.empresa,
            tipo_equipo="Equipo de Medición",
            estado="Activo",
        )

    def test_generate_hoja_vida_cache_key(self):
        """Test: _generate_hoja_vida_cache_key genera clave única."""
        cache_key = _generate_hoja_vida_cache_key(self.equipo)

        # Verificar formato
        assert cache_key.startswith(f"hoja_vida_{self.equipo.id}_")
        assert len(cache_key) > 30  # Debe incluir hash MD5

        # Verificar que es determinística
        cache_key2 = _generate_hoja_vida_cache_key(self.equipo)
        assert cache_key == cache_key2

    def test_get_hoja_vida_activities(self):
        """Test: _get_hoja_vida_activities obtiene actividades optimizadas."""
        # Crear actividades
        Calibracion.objects.create(
            equipo=self.equipo,
            fecha_calibracion=timezone.now().date(),
        )
        Mantenimiento.objects.create(
            equipo=self.equipo,
            fecha_mantenimiento=timezone.now().date(),
            tipo_mantenimiento='Preventivo',
        )

        # Ejecutar helper
        calibraciones, mantenimientos, comprobaciones = _get_hoja_vida_activities(self.equipo)

        # Verificar que retorna querysets
        assert calibraciones.count() == 1
        assert mantenimientos.count() == 1
        assert comprobaciones.count() == 0

    def test_generate_hoja_vida_charts_sin_datos(self):
        """Test: _generate_hoja_vida_charts sin datos retorna None."""
        calibraciones = Calibracion.objects.none()
        comprobaciones = Comprobacion.objects.none()

        grafica_conf, grafica_comp = _generate_hoja_vida_charts(calibraciones, comprobaciones)

        # Sin datos, no hay gráficas
        assert grafica_conf is None
        assert grafica_comp is None

    def test_build_hoja_vida_context(self):
        """Test: _build_hoja_vida_context construye diccionario completo."""
        calibraciones = Calibracion.objects.none()
        mantenimientos = Mantenimiento.objects.none()
        comprobaciones = Comprobacion.objects.none()

        file_urls = {
            'baja_registro': None,
            'logo_empresa_url': None,
            'imagen_equipo_url': None,
            'documento_baja_url': None,
            'archivo_compra_pdf_url': None,
            'ficha_tecnica_pdf_url': None,
            'manual_pdf_url': None,
            'otros_documentos_pdf_url': None,
        }

        charts = (None, None)

        # Ejecutar helper
        context = _build_hoja_vida_context(
            self.equipo,
            calibraciones,
            mantenimientos,
            comprobaciones,
            file_urls,
            charts
        )

        # Verificar estructura del contexto
        assert 'equipo' in context
        assert 'calibraciones' in context
        assert 'mantenimientos' in context
        assert 'comprobaciones' in context
        assert 'titulo_pagina' in context
        assert context['equipo'] == self.equipo
        assert "Hoja de Vida" in context['titulo_pagina']
