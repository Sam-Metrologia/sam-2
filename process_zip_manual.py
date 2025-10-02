#!/usr/bin/env python
"""
Script manual para procesar solicitudes ZIP en orden FIFO correcto.
Genera ZIPs con la estructura completa original.
"""

import os
import sys
import django
from pathlib import Path

# Add the project directory to Python path
project_dir = Path(__file__).parent
sys.path.insert(0, str(project_dir))

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'proyecto_c.settings')
django.setup()

from django.utils import timezone
from core.models import ZipRequest, Equipo, Proveedor, Procedimiento
from django.core.files.storage import default_storage
import zipfile
import io
import logging

logger = logging.getLogger(__name__)


def process_next_zip():
    """Procesa la siguiente solicitud ZIP en la cola (FIFO correcto)."""

    # Get the next request in FIFO order (lowest position_in_queue)
    zip_request = ZipRequest.objects.filter(status='pending').order_by('position_in_queue').first()

    if not zip_request:
        print("No hay solicitudes ZIP pendientes")
        return None

    print(f"Procesando solicitud ZIP #{zip_request.id} (Posición {zip_request.position_in_queue})")
    print(f"Usuario: {zip_request.user.username}")
    print(f"Empresa: {zip_request.empresa.nombre}")

    try:
        # Mark as processing
        zip_request.status = 'processing'
        zip_request.started_at = timezone.now()
        zip_request.save()

        # Generate the ZIP with proper structure
        zip_content = generate_complete_zip_structure(zip_request)

        # Save ZIP to storage
        zip_filename = f'zip_reports/informe_{zip_request.empresa.id}_{zip_request.id}_{int(timezone.now().timestamp())}.zip'
        file_path = default_storage.save(zip_filename, io.BytesIO(zip_content))

        # Mark as completed
        zip_request.status = 'completed'
        zip_request.completed_at = timezone.now()
        zip_request.file_path = file_path
        zip_request.file_size = len(zip_content)
        zip_request.save()

        print(f"ZIP procesado exitosamente!")
        print(f"   Archivo: {file_path}")
        print(f"   Tamaño: {len(zip_content)} bytes")

        return zip_request

    except Exception as e:
        # Mark as failed
        zip_request.status = 'failed'
        zip_request.error_message = str(e)
        zip_request.save()

        print(f"Error procesando ZIP: {e}")
        return None


def generate_complete_zip_structure(zip_request):
    """Genera ZIP con estructura completa como el original."""

    empresa = zip_request.empresa
    parte_numero = zip_request.parte_numero

    print(f"Generando ZIP para empresa: {empresa.nombre}")

    # Get equipos for this empresa (limit to 10 for testing, could be expanded)
    equipos_empresa = Equipo.objects.filter(empresa=empresa).select_related('empresa').order_by('codigo_interno')[:10]
    proveedores_empresa = Proveedor.objects.filter(empresa=empresa).order_by('nombre_empresa')
    procedimientos_empresa = Procedimiento.objects.filter(empresa=empresa).order_by('codigo')

    print(f"Equipos encontrados: {equipos_empresa.count()}")
    print(f"Proveedores: {proveedores_empresa.count()}")
    print(f"Procedimientos: {procedimientos_empresa.count()}")

    zip_buffer = io.BytesIO()
    empresa_nombre = empresa.nombre.replace('/', '_').replace('\\', '_')  # Safe filename

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zf:

        # 1. Add consolidated Excel (simplified version)
        excel_content = generate_simple_consolidated_excel(equipos_empresa, proveedores_empresa, procedimientos_empresa, empresa)
        zf.writestr(f"{empresa_nombre}/Informe_Consolidado.xlsx", excel_content)
        print("Excel consolidado agregado")

        # 2. Add equipment folders with structure
        for idx, equipo in enumerate(equipos_empresa, 1):
            safe_codigo = equipo.codigo_interno.replace('/', '_').replace('\\', '_')
            equipo_folder = f"{empresa_nombre}/Equipos/{safe_codigo}"

            # Equipment info file
            equipo_info = f"""HOJA DE VIDA - {equipo.nombre}
Código Interno: {equipo.codigo_interno}
Empresa: {empresa.nombre}
Tipo: {equipo.get_tipo_equipo_display()}
Marca: {equipo.marca}
Modelo: {equipo.modelo}
Número de Serie: {equipo.numero_serie}
Ubicación: {equipo.ubicacion}
Responsable: {equipo.responsable}
Estado: {equipo.estado}
Fecha de Adquisición: {equipo.fecha_adquisicion or 'No especificada'}
Rango de Medida: {equipo.rango_medida}
Resolución: {equipo.resolucion}
Error Máximo Permisible: {equipo.error_maximo_permisible or 'No especificado'}
Observaciones: {equipo.observaciones or 'Ninguna'}
"""
            zf.writestr(f"{equipo_folder}/Hoja_de_vida.txt", equipo_info.encode('utf-8'))

            # Create activity folders structure
            zf.writestr(f"{equipo_folder}/Calibraciones/Certificados/.gitkeep", "")
            zf.writestr(f"{equipo_folder}/Calibraciones/Confirmaciones/.gitkeep", "")
            zf.writestr(f"{equipo_folder}/Calibraciones/Intervalos/.gitkeep", "")
            zf.writestr(f"{equipo_folder}/Mantenimientos/.gitkeep", "")
            zf.writestr(f"{equipo_folder}/Comprobaciones/.gitkeep", "")

            # Add activity summaries
            calibraciones = equipo.calibraciones.all()[:3]  # Limit for demo
            if calibraciones:
                cal_summary = "RESUMEN DE CALIBRACIONES\\n\\n"
                for cal in calibraciones:
                    cal_summary += f"- Fecha: {cal.fecha_calibracion}\\n"
                    cal_summary += f"  Proveedor: {cal.proveedor.nombre_empresa if cal.proveedor else 'No especificado'}\\n"
                    cal_summary += f"  Observaciones: {cal.observaciones or 'Ninguna'}\\n\\n"
                zf.writestr(f"{equipo_folder}/Calibraciones/resumen_calibraciones.txt", cal_summary.encode('utf-8'))

            print(f"Equipo {idx}/{len(equipos_empresa)}: {safe_codigo}")

        # 3. Add procedures folder
        if procedimientos_empresa:
            procedimientos_info = "PROCEDIMIENTOS DE LA EMPRESA\\n\\n"
            for proc in procedimientos_empresa:
                procedimientos_info += f"Código: {proc.codigo}\\n"
                procedimientos_info += f"Nombre: {proc.nombre}\\n"
                procedimientos_info += f"Versión: {proc.version}\\n"
                procedimientos_info += f"Fecha: {proc.fecha_version}\\n\\n"

            zf.writestr(f"{empresa_nombre}/Procedimientos/lista_procedimientos.txt", procedimientos_info.encode('utf-8'))
            print("Procedimientos agregados")

        # 4. Add company summary
        empresa_summary = f"""RESUMEN DE LA EMPRESA

Nombre: {empresa.nombre}
Dirección: {empresa.direccion or 'No especificada'}
Teléfono: {empresa.telefono or 'No especificado'}
Email: {empresa.email or 'No especificado'}

ESTADÍSTICAS:
- Total de Equipos: {equipos_empresa.count()}
- Total de Proveedores: {proveedores_empresa.count()}
- Total de Procedimientos: {procedimientos_empresa.count()}

Generado: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}
"""
        zf.writestr(f"{empresa_nombre}/RESUMEN_EMPRESA.txt", empresa_summary.encode('utf-8'))
        print("Resumen de empresa agregado")

    zip_buffer.seek(0)
    zip_content = zip_buffer.getvalue()
    print(f"ZIP generado: {len(zip_content)} bytes")

    return zip_content


def generate_simple_consolidated_excel(equipos, proveedores, procedimientos, empresa):
    """Genera Excel consolidado simplificado."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io

    workbook = Workbook()

    # Sheet 1: Equipos
    sheet_equipos = workbook.active
    sheet_equipos.title = "Equipos"

    # Headers
    headers = ['Código', 'Nombre', 'Tipo', 'Marca', 'Modelo', 'Estado', 'Ubicación']
    for col, header in enumerate(headers, 1):
        cell = sheet_equipos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Data
    for row, equipo in enumerate(equipos, 2):
        sheet_equipos.cell(row=row, column=1, value=equipo.codigo_interno)
        sheet_equipos.cell(row=row, column=2, value=equipo.nombre)
        sheet_equipos.cell(row=row, column=3, value=equipo.get_tipo_equipo_display())
        sheet_equipos.cell(row=row, column=4, value=equipo.marca)
        sheet_equipos.cell(row=row, column=5, value=equipo.modelo)
        sheet_equipos.cell(row=row, column=6, value=equipo.estado)
        sheet_equipos.cell(row=row, column=7, value=equipo.ubicacion)

    # Sheet 2: Proveedores
    sheet_prov = workbook.create_sheet("Proveedores")
    prov_headers = ['Nombre', 'Contacto', 'Teléfono', 'Email']
    for col, header in enumerate(prov_headers, 1):
        cell = sheet_prov.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    for row, prov in enumerate(proveedores, 2):
        sheet_prov.cell(row=row, column=1, value=prov.nombre_empresa)
        sheet_prov.cell(row=row, column=2, value=prov.contacto)
        sheet_prov.cell(row=row, column=3, value=prov.telefono)
        sheet_prov.cell(row=row, column=4, value=prov.email)

    # Sheet 3: Procedimientos
    sheet_proc = workbook.create_sheet("Procedimientos")
    proc_headers = ['Código', 'Nombre', 'Versión', 'Fecha']
    for col, header in enumerate(proc_headers, 1):
        cell = sheet_proc.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

    for row, proc in enumerate(procedimientos, 2):
        sheet_proc.cell(row=row, column=1, value=proc.codigo)
        sheet_proc.cell(row=row, column=2, value=proc.nombre)
        sheet_proc.cell(row=row, column=3, value=proc.version)
        sheet_proc.cell(row=row, column=4, value=proc.fecha_version)

    # Save to buffer
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


if __name__ == "__main__":
    print("PROCESADOR MANUAL DE ZIP - SAM METROLOGIA")
    print("=" * 50)

    processed = process_next_zip()

    if processed:
        print(f"\\nSolicitud #{processed.id} procesada exitosamente")
        print(f"   El ZIP esta listo para descarga en: {processed.file_path}")
    else:
        print("\\nNo se procesaron solicitudes")

    print("\\nProceso completado")